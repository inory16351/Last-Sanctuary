# -*- coding: utf-8 -*-
"""성역(중앙 건물)의 <b>이름을 바꾼다</b> — 「중앙건물」 → 「성역의 심장부」
(2026-08-27 · 185절 · 유저 지시: *"중앙 건물을 '성역의 심장부'로 바꾸고 영어 이름은
Heart Of Sanctuary 로 바꾸고 테이블에도 연동해줘"*).

■ 어느 칸을 고치나 — <b>스트링 키 테이블 한 곳</b>이다
  「건물데이터시트」 `Construction` 시트의 <c>Const_name</c> 칸에는 <b>글자가 아니라 키</b>가
  들어 있다(<c>const_name_10001</c> · `convert_tables_to_string_keys.py` 가 그렇게 바꿔 뒀다).
  즉 <b>글자의 정본은 스트링 키 테이블</b>이고, 건물 시트는 그 키를 가리킬 뿐이다.
  → 그래서 건물 시트는 <b>손댈 것이 없다</b>. 여기 한 칸만 고치면 둘 다 따라온다.

■ ⚠⚠ 이 스크립트만 <b>이미 있는 키를 덮어쓴다</b>
  다른 `table_update_*` 는 «없는 키만 붙인다» 가 규칙이다(178-6절의 «표 전체를 훑는 치환» 이
  유저 번역 21칸을 덮은 사고 때문이다). 여기는 <b>유저가 이름을 바꾸라고 지시한</b> 자리라
  덮어쓰는 것이 목적이다 — 그래서 <b>이 키 하나만</b> 손대고, 바꾸기 전 값을 찍는다.

■ ⚠ 코드의 폴백도 같이 고쳐야 한다
  <c>NexusDefinitionSO.displayName</c> 이 이 키의 폴백이다. 규약은 «kr 은 코드의 폴백과
  한 글자도 다르지 않게» 이므로 <b>둘을 같이</b> 고친다(코드 쪽은 손으로 고쳤다).

■ 다음
    py -3 Tools/gen_string_table.py   →   py -3 Tools/link_string_keys.py
    py -3 Tools/check_string_keys.py --strict
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET = 'string'
DATA_ROW0 = 4

#: 키 → (새 kr, 새 en). ⚠ 여기 있는 키는 <b>덮어쓴다</b>.
#  ★ 영어는 유저가 적어 준 그대로다 — 「Heart Of Sanctuary」(가운데 O 가 대문자).
RENAME = {
    'const_name_10001': ('성역의 심장부', 'Heart Of Sanctuary'),
}


def norm(v):
    return '' if v is None else str(v).strip()


def field_index(ws):
    """2행의 필드명 → 열 번호. gen_string_table.py 와 같은 규약이다."""
    idx = {}
    for c in range(1, ws.max_column + 1):
        f = norm(ws.cell(2, c).value)
        if f:
            idx[f] = c
    return idx


def main():
    if not os.path.isfile(STRING_XLSX):
        raise SystemExit('⚠ 스트링 키 테이블이 없습니다: %s' % STRING_XLSX)

    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]
    idx = field_index(ws)
    c_key = idx.get('string_key', 1)
    c_kr = idx.get('kr', 2)
    c_en = idx.get('en', 3)

    where = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = norm(ws.cell(r, c_key).value)
        if k:
            where[k] = r

    changed = 0
    for key, (kr, en) in RENAME.items():
        r = where.get(key)
        if r is None:
            # ⚠ 조용히 넘기지 않는다 — 키가 없으면 «바꿨다» 는 보고가 거짓이 된다.
            print('  ✗ 표에 없는 키입니다: %s' % key)
            continue
        before = (norm(ws.cell(r, c_kr).value), norm(ws.cell(r, c_en).value))
        if before == (kr, en):
            print('  · 이미 같은 값입니다: %s' % key)
            continue
        ws.cell(r, c_kr).value = kr
        ws.cell(r, c_en).value = en
        print('  ~ %-22s  %s / %s   →   %s / %s'
              % (key, before[0], before[1], kr, en))
        changed += 1

    if not changed:
        print('바뀐 것이 없어 저장하지 않았습니다.')
        return 0

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    wb.save(STRING_XLSX)
    print('저장: %s (백업 .bak) · %d키 변경'
          % (os.path.basename(STRING_XLSX), changed))
    print('다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')
    return 0


if __name__ == '__main__':
    sys.exit(main())
