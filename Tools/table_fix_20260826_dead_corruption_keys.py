# -*- coding: utf-8 -*-
"""스트링 키 테이블에서 <b>죽은 `*_Corruption_*` 키</b>를 지운다 (2026-08-26 · 17차).

■ 무엇이 있었나 — 176-1절이 «흔적» 이라고만 적어 둔 그 사고의 잔해다
  언젠가 이 표에 <b>`erosion` → `Corruption` 전체 치환</b>이 돌았고, 그것이 <b>키 이름까지</b>
  바꿨다. 176절이 살아 있는 `help_erosion_*` 셋을 <b>되살렸지만</b>, 이름이 바뀐 쪽은
  <b>지우지 않아서</b> 표에 두 벌이 남았다. 게다가 그 일곱 줄이 <b>각각 두 번</b> 들어 있어
  판을 켤 때마다 콘솔에 이런 경고가 일곱 개 떴다:

      [String] 키가 중복됐습니다: 'help_Corruption_body' — 먼저 나온 값을 씁니다.

■ 어느 쪽이 살아 있나 — <b>원본 표가 정본이다</b> (실측)
      Last_Sanctuary_도움말테이블_Ver01.xlsx  Help·HelpStep·StringKeys →  help_erosion*
      Last_Sanctuary_유물테이블_Ver02.xlsx    DigOutcome              →  dig_erosion_*
      HelpTable.asset                         titleKey 등             →  help_erosion_*
  즉 <b>`*_Corruption_*` 을 읽는 곳은 한 곳도 없다</b>. 그런데도 그쪽에만 영어가 들어 있어서
  «번역은 다 됐는데 화면은 한국어» 가 됐다 — 유저가 본 것이 그것이다.

■ 안전장치 셋 — 하나라도 어긋나면 <b>아무것도 지우지 않는다</b>
  ① 지우려는 키마다 <b>`erosion` 짝이 표에 있는지</b> 확인한다.
  ② 그 짝의 <b>영어 칸이 비어 있지 않은지</b> 확인한다(비어 있으면 지우면 손실이다).
  ③ <b>프로젝트 어디에도</b> 그 키 이름이 적혀 있지 않은지 확인한다(코드·에셋·표).

■ 다음
    py -3 Tools/gen_string_table.py   →  py -3 Tools/link_string_keys.py
"""
import io
import os
import re
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

_PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET = 'string'
DATA_ROW0 = 4

#: 이 낱말이 <b>키 이름</b>에 들어 있으면 치환 사고의 잔해다.
DEAD = re.compile(r'_Corruption(_|$)')

#: 프로젝트에서 키 이름을 찾을 곳 (에셋·코드·내보낸 표)
SCAN_DIRS = [os.path.join(_PROJECT, 'Assets', '_Project')]
SCAN_EXT = ('.cs', '.asset', '.txt', '.json', '.tsv')


def live_name(key):
    """`*_Corruption_*` → `*_erosion_*` (짝의 이름)."""
    return DEAD.sub(lambda m: '_erosion' + m.group(1), key)


def scan_project(keys):
    """키 이름이 프로젝트 파일에 나타나는 자리를 모은다(내보낸 StringTable.txt 는 뺀다)."""
    hits = {}
    for root in SCAN_DIRS:
        for dirpath, _, files in os.walk(root):
            for f in files:
                if not f.endswith(SCAN_EXT):
                    continue
                p = os.path.join(dirpath, f)
                if os.path.basename(p) == 'StringTable.txt':
                    continue          # 이 표에서 만들어진 파일이라 근거가 못 된다
                try:
                    text = io.open(p, encoding='utf-8', errors='replace').read()
                except OSError:
                    continue
                for k in keys:
                    if k in text:
                        hits.setdefault(k, []).append(os.path.relpath(p, _PROJECT))
    return hits


def main():
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    rows = []          # (row, key)
    by_key = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        k = str(k).strip()
        rows.append((r, k))
        by_key.setdefault(k, []).append(r)

    dead_rows = [(r, k) for r, k in rows if DEAD.search(k)]
    dead_keys = sorted({k for _, k in dead_rows})

    print('[죽은 Corruption 키 정리] 표 %d줄 · 지울 후보 키 %d개(줄 %d개)'
          % (len(rows), len(dead_keys), len(dead_rows)))
    if not dead_rows:
        print('  지울 것이 없습니다.')
        return

    # ── 안전장치 ①② — erosion 짝이 있고, 그 영어가 채워져 있는가 ──────────
    problems = []
    for k in dead_keys:
        pair = live_name(k)
        pair_rows = by_key.get(pair)
        if not pair_rows:
            problems.append('%s → 짝(%s)이 표에 없다' % (k, pair))
            continue
        en = ws.cell(row=pair_rows[0], column=3).value
        if en is None or str(en).strip() == '':
            problems.append('%s → 짝(%s)의 영어 칸이 비어 있다' % (k, pair))

    # ── 안전장치 ③ — 프로젝트가 그 이름을 쓰고 있지 않은가 ────────────────
    hits = scan_project(dead_keys)
    for k, files in sorted(hits.items()):
        problems.append('%s → 프로젝트가 쓰고 있다: %s' % (k, ', '.join(files[:3])))

    if problems:
        wb.close()
        sys.exit('\n✗ 안전장치에 걸렸습니다 — 아무것도 지우지 않았습니다:\n  '
                 + '\n  '.join(problems))

    print('  안전장치 통과 — 짝이 모두 있고 영어가 채워져 있고, 쓰는 곳이 없습니다.')
    for k in dead_keys:
        print('    - %-42s (짝: %s · 줄 %s)'
              % (k, live_name(k), ', '.join(str(r) for r in by_key[k])))

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    for r, _ in sorted(dead_rows, reverse=True):      # 아래에서부터 지운다
        ws.delete_rows(r, 1)
    wb.save(STRING_XLSX)
    print('  줄 %d개 삭제 · 저장: %s (백업 .bak)'
          % (len(dead_rows), os.path.basename(STRING_XLSX)))
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
