# -*- coding: utf-8 -*-
"""이벤트 테이블(xlsx) → `EventDefinitionSO` 에셋.

**Ver013 (2026-08-21 개정)** — 유저 지시: *"이벤트 테이블 수정된거 읽어보고 다시 인게임에
구현"*. 표의 구조가 통째로 바뀌었다.

원본
----
``<볼트>/데이터 테이블/Last_Sanctuary_이벤트테이블_Ver013.xlsx``

시트 다섯 벌 중 <b>둘</b>만 에셋으로 옮긴다:

| 시트 | 옮기나 | 이유 |
|---|---|---|
| ``Event``       | ○ | 이벤트 43개. 에셋 하나가 이 한 행이다 |
| ``ChoiceGroup`` | ○ | 선택지 86행. <b>이벤트 에셋 안에</b> 넣는다(아래 ★) |
| ``Info``        | × | 사람이 읽는 문서다 |
| ``Condition``   | × | 조건 enum 의 <b>설명</b>이다. 판정은 코드가 한다 |
| ``RewardType``  | × | 보상 enum 의 <b>설명</b>이다. 적용은 `EventRewardService` 가 한다 |

★ **왜 선택지를 이벤트 안에 넣나** — 선택지 86행은 전부 ``choice_group_id`` 로 «어느
이벤트의 것인지» 가 정해져 있다(이벤트와 1:1). 에셋을 86개 만들면 «이 이벤트의 선택지» 를
찾는 일이 매번 전수 검색이 된다. 그래서 <b>43개 에셋</b>에 각자의 선택지만 담는다.
⚠ 표에 그룹이 있는데 ``Event`` 시트에 행이 없으면 그 선택지는 <b>버려진다</b> — 그때는
  «고아 그룹» 으로 알려 준다(조용히 사라지면 «선택지가 안 나온다» 를 못 찾는다).

**Ver012 에서 무엇이 바뀌었나** (Info 시트가 적어 둔 그대로)

| Ver012 | Ver013 |
|---|---|
| ``Dialogue`` 시트 (대사 사슬 168행) | 지워짐 → ``Event.event_script`` 한 칸 |
| ``EventType`` 시트 (5001/5002/5003) | 지워짐 → ``Event.trigger_cond`` enum 3종 |
| ``Switch`` 시트 | 지워짐 → ``Event.repeatable`` 불리언 하나 |
| ``event_value_01`` (타이머 길이) | 지워짐 — 코드가 아는 값이다 |
| ``event_value_02`` (가중치) | ``weight`` |
| 보상 = 타입 + 수치 | 보상 = 타입 + 수치 + <b>지속시간(초)</b> |

⚠⚠ **지속시간이 «초» 가 된 것이 가장 큰 변화다.** 옛 표는 «이벤트가 끝날 때까지» 라는
  상대값이었는데, 이벤트가 웨이브 <b>종료 시</b> 에 뜨게 바뀌면서 그 기준점이 사라졌다.

⚠ .asset YAML 에 **빈 줄을 넣으면 Unity 파서가 그 뒤 필드를 전부 무시한다**(진행상황 8절 3번).
  아래 :func:`yaml_str` 도 그래서 <b>줄바꿈을 이스케이프</b>한다 — 대사에 줄바꿈이 실제로
  들어 있다(``event_desc`` 는 네 줄짜리도 있다).

⚠ MCP 에는 SO 에셋을 다루는 도구가 없다 — 그래서 이 종류만 스크립트로 쓴다(59-2절).

사용법:  py -3 Tools/gen_event_assets.py
다음:    유니티에서 Assets/Refresh
"""

import hashlib
import io
import os
import re
import sys

import openpyxl

from vault_path import TABLE_DIR, PROJECT

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "Last_Sanctuary_이벤트테이블_Ver013.xlsx")
OUT_DIR = os.path.join(PROJECT, "Assets", "_Project", "Resources", "Events")

#: 표의 첫 세 줄은 «한글 제목 / 영문 키 / 자료형» 이다 — 값은 4행부터.
FIRST_ROW = 4

ASSET_META = """fileFormatVersion: 2
guid: {guid}
NativeFormatImporter:
  externalObjects: {{}}
  mainObjectFileID: 11400000
  userData:
  assetBundleName:
  assetBundleVariant:
"""

HEADER = """%YAML 1.1
%TAG !u! tag:unity3d.com,2011:
--- !u!114 &11400000
MonoBehaviour:
  m_ObjectHideFlags: 0
  m_CorrespondingSourceObject: {{fileID: 0}}
  m_PrefabInstance: {{fileID: 0}}
  m_PrefabAsset: {{fileID: 0}}
  m_GameObject: {{fileID: 0}}
  m_Enabled: 1
  m_EditorHideFlags: 0
  m_Script: {{fileID: 11500000, guid: {script_guid}, type: 3}}
  m_Name: {name}
  m_EditorClassIdentifier:
"""


def script_guid(rel_cs_path):
    """`.cs.meta` 에서 스크립트 guid 를 읽는다 (`sync_tables_to_assets.py` 와 같은 규칙)."""
    meta = os.path.join(PROJECT, "Assets", "_Project", "Scripts", rel_cs_path) + ".meta"
    with io.open(meta, encoding="utf-8") as f:
        for line in f:
            if line.startswith("guid:"):
                return line.split(":", 1)[1].strip()
    raise RuntimeError("guid 를 찾지 못했습니다: " + meta)


def guid_for(key):
    """경로에서 결정적으로 만든다 — 다시 돌려도 같은 guid 라 참조가 안 끊긴다."""
    return hashlib.md5(("LastSanctuary/" + key).encode("utf-8")).hexdigest()


def num(v, default=0):
    if v is None:
        return default
    try:
        f = float(v)
    except (TypeError, ValueError):
        return default
    return int(f) if f == int(f) else f


def text(v):
    return "" if v is None else str(v).strip()


def yaml_str(v):
    """
    YAML 한 줄 문자열. <b>줄바꿈을 ``\\n`` 으로 이스케이프</b>한다.

    ⚠ 대사와 ``event_desc`` 에는 <b>실제 줄바꿈</b>이 들어 있다. 그대로 쓰면 YAML 이
      «다음 필드» 로 읽거나 빈 줄이 생겨 <b>그 뒤 필드를 전부 잃는다</b>(8절 3번).
      Unity 는 큰따옴표 안의 ``\\n`` 을 줄바꿈으로 되돌려 준다 — TMP 에서도 줄이 바뀐다.
    """
    s = text(v)
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\\n")
    return '"%s"' % s


def read_sheet(wb, name):
    ws = wb[name]
    rows = []
    for r in range(FIRST_ROW, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if vals and vals[0] is not None:
            rows.append(vals)
    return rows


def safe_name(s):
    """파일 이름에 쓸 수 있게 다듬는다. 한글은 그대로 둔다(이 프로젝트의 다른 에셋과 같다)."""
    return re.sub(r"[^0-9A-Za-z가-힣_]+", "_", s).strip("_") or "Event"


def main():
    if not os.path.isfile(XLSX):
        raise SystemExit("⚠ 표를 찾지 못했습니다: %s" % XLSX)

    guid = script_guid(os.path.join("Events", "EventDefinitionSO.cs"))
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    events = read_sheet(wb, "Event")
    choices = read_sheet(wb, "ChoiceGroup")

    # 선택지를 그룹별로 모은다 — 표의 등장 순서를 <b>그대로</b> 유지한다
    # (choice_order 로 다시 정렬하는 것은 코드가 한다 — EventDefinitionSO.OrderedChoices).
    by_group = {}
    for c in choices:
        gid = int(num(c[0]))
        by_group.setdefault(gid, []).append(c)

    os.makedirs(OUT_DIR, exist_ok=True)

    # ★ 옛 에셋을 먼저 지운다 — 표에서 행이 빠졌을 때 «없어진 이벤트» 가 계속 뜨는 것을 막는다.
    #   ⚠ 이 폴더는 이 스크립트가 <b>통째로 소유</b>한다(사람이 여기에 파일을 두지 말 것).
    dropped = 0
    for f in os.listdir(OUT_DIR):
        if f.startswith("Event_") and (f.endswith(".asset") or f.endswith(".asset.meta")):
            os.remove(os.path.join(OUT_DIR, f))
            dropped += 1
    if dropped:
        print("  옛 에셋 %d개 지움" % dropped)

    #: 표에 적힐 수 있는 발동 조건 — 오타를 여기서 잡는다.
    #  ⚠ 코드(EventDefinitionSO.Trigger)도 같은 세 이름만 안다. 이름이 늘면 <b>두 곳</b>이다.
    KNOWN_CONDS = ("wave_end", "private_timer", "habitat_contact")

    used_groups = set()
    made = 0
    by_cond = {}
    for row in events:
        eid = int(num(row[0]))
        ename = text(row[1])
        cond = text(row[2])
        tval = int(num(row[3]))
        weight = int(num(row[4]))
        repeatable = int(num(row[5]))
        gid = int(num(row[6]))
        bg = text(row[7])
        script = text(row[8]) if len(row) > 8 else ""

        if cond not in KNOWN_CONDS:
            # ⚠ 지어내지 않는다 — 코드가 못 읽는 조건이면 그 이벤트는 <b>안 뜬다</b>.
            print("  ⚠ %s(%d) — 발동 조건 '%s' 은 코드가 모릅니다(%s). 이 이벤트는 안 뜹니다."
                  % (ename, eid, cond, " / ".join(KNOWN_CONDS)))
        by_cond[cond] = by_cond.get(cond, 0) + 1

        rows = by_group.get(gid, [])
        used_groups.add(gid)
        if not rows:
            print("  ⚠ %s(%d) — 선택지 그룹 %d 이 표에 없습니다. "
                  "선택지가 없으면 창을 닫을 수 없어 <b>못 쓰는 정의</b>가 됩니다." % (ename, eid, gid))

        asset = "Event_%d_%s" % (eid, safe_name(ename))
        body = [HEADER.format(script_guid=guid, name=asset)]
        body.append("  eventId: %d\n" % eid)
        body.append("  eventName: %s\n" % yaml_str(ename))
        body.append("  triggerCond: %s\n" % yaml_str(cond))
        body.append("  triggerValue: %d\n" % tval)
        body.append("  weight: %d\n" % weight)
        body.append("  repeatable: %d\n" % (1 if repeatable else 0))
        body.append("  choiceGroupId: %d\n" % gid)
        body.append("  eventBg: %s\n" % yaml_str(bg))
        body.append("  eventScript: %s\n" % yaml_str(script))

        if rows:
            body.append("  choices:\n")
            for c in rows:
                body.append("  - choiceId: %d\n" % int(num(c[1])))
                body.append("    choiceOrder: %d\n" % int(num(c[2])))
                body.append("    choiceText: %s\n" % yaml_str(c[3]))
                body.append("    resultScript: %s\n" % yaml_str(c[4]))
                body.append("    resultEffect: %s\n" % yaml_str(c[5]))
                body.append("    rewardType01: %s\n" % yaml_str(c[6]))
                body.append("    rewardValue01: %d\n" % int(num(c[7])))
                body.append("    rewardDuration01: %d\n" % int(num(c[8])))
                body.append("    rewardType02: %s\n" % yaml_str(c[9] if len(c) > 9 else None))
                body.append("    rewardValue02: %d\n" % int(num(c[10] if len(c) > 10 else None)))
                body.append("    rewardDuration02: %d\n" % int(num(c[11] if len(c) > 11 else None)))
        else:
            body.append("  choices: []\n")

        path = os.path.join(OUT_DIR, asset + ".asset")
        with io.open(path, "w", encoding="utf-8", newline="\n") as f:
            f.write("".join(body))
        with io.open(path + ".meta", "w", encoding="utf-8", newline="\n") as f:
            f.write(ASSET_META.format(guid=guid_for("Events/" + asset)))
        made += 1

    orphan = sorted(set(by_group) - used_groups)
    if orphan:
        print("  ⚠ 고아 선택지 그룹 %d개 — Event 시트에 행이 없어 <b>버려진다</b>: %s"
              % (len(orphan), orphan))

    print("  이벤트 에셋 %d개 · 선택지 %d행" % (made, len(choices)))
    print("  조건별: " + " · ".join("%s %d" % (k, v) for k, v in sorted(by_cond.items())))
    print("  다음: 유니티에서 Assets/Refresh")


if __name__ == "__main__":
    main()
