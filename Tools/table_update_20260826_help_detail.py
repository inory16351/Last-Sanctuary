# -*- coding: utf-8 -*-
"""도움말 표 — 「능력치 낱낱」·「정신 이상 낱낱」 두 항목을 새로 넣는다 (2026-08-26).

유저 지시: *"도움말에 정신이상과 능력치의 상세 설명을 볼 수 있는 별도 설명 추가
(문어체로 / 타 도움말 어투와 통일해서 / 스트링 키 테이블에도 추가)"*.

★ <b>«별도» 로 만든 이유</b> — 이미 있는 <c>help_stats</c>(능력치 읽는 법)와
  <c>help_mental_error</c>(정신 이상)는 <b>조언 카드</b>로도 뜨는 항목이다. 카드는
  서너 줄이 한계인데(창 크기가 그렇다), 유저가 원한 것은 «열두 칸이 각각 무엇을 정하는가» ·
  «열한 가지가 각각 무엇을 하는가» 라는 <b>목록</b>이다. 한 항목에 다 넣으면 카드가 넘치고,
  카드에 맞춰 줄이면 목록이 안 된다. 그래서 <b>백과에만 있는 항목</b>을 옆에 세운다
  (<c>trigger</c> 를 비워 두면 저절로 뜨지 않는다 — <c>help_camera</c>·<c>help_retreat</c> 가
  이미 그 모양이다).

★ <b>「함께 볼 것」으로 이어 붙인다</b> — 기존 항목에서 상세로, 상세에서 원래 이웃으로.
  <c>see_also</c> 는 한 칸뿐이라 사슬로 잇는다:
      help_stats → help_stats_detail → help_upgrade
      help_mental_error → help_mental_detail → help_erosion

⚠ 문구는 <b>이 표가 정본</b>이다(<c>help_string_merge.py</c> 의 ★★). 스트링 키 테이블에서
  고치면 다음 병합에 되돌아간다.
⚠ 어투는 기존 도움말과 맞췄다 — <b>문어체 «~습니다»</b> · 짧은 평서문 · 강조는 <c>&lt;b&gt;</c>.

사용법:  python Tools/table_update_20260826_help_detail.py
다음:    python Tools/help_string_merge.py
         python Tools/gen_string_table.py
         python Tools/link_string_keys.py
         python Tools/gen_help_assets.py
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HELP_XLSX = os.path.join(TABLE_DIR, "Last_Sanctuary_도움말테이블_Ver01.xlsx")

# ══════════════════════════════════════════════════════════════════════
#  ① Help 시트에 넣을 두 줄
# ══════════════════════════════════════════════════════════════════════
# 열: help_id | category | order | title_key | summary_key | body_key |
#     trigger | trigger_arg | priority | show_once | see_also | 비고(제목) | open_panel
NEW_ROWS = [
    # ★ order 는 이웃 바로 뒤 — help_stats 320 · help_mental_error 520
    ["help_stats_detail", "성장", 325,
     "help_stats_detail_title", "help_stats_detail_summary", "help_stats_detail_body",
     "", "", 3, 1, "help_upgrade", "능력치 낱낱", ""],
    ["help_mental_detail", "위험", 525,
     "help_mental_detail_title", "help_mental_detail_summary", "help_mental_detail_body",
     "", "", 3, 1, "help_erosion", "정신 이상 낱낱", ""],
]

# 기존 항목의 「함께 볼 것」을 상세 쪽으로 돌린다(위 ★).
RELINK = {
    "help_stats": "help_stats_detail",
    "help_mental_error": "help_mental_detail",
}

# ══════════════════════════════════════════════════════════════════════
#  ② StringKeys 시트에 넣을 문구
# ══════════════════════════════════════════════════════════════════════
# ⚠ 여기 적는 수치는 <b>표와 코드에서 확인한 것</b>이다:
#    · 능력치 상한 100        — BalanceConfigSO.statMax
#    · 치명타 피해 1.5배      — BalanceConfigSO.criticalDamageMultiplier
#    · 명중·크리는 원거리 전용 — CharacterUnit.HitChancePercent / CriticalAppliesToCurrentAttack
#    · 저항력은 강화로 안 오름 — StatBlock.IsGrowable
#    · 정신 이상 열한 가지     — 정신 이상 테이블.xlsx (id 40001~40011)
STRINGS = {
    "help_stats_detail_title": "능력치 낱낱",

    "help_stats_detail_summary":
        "열두 칸이 각각 무엇을 정하는지 적어 두었습니다.\n"
        "어느 칸을 키울지는 <b>성장 유형</b>으로 고릅니다.",

    "help_stats_detail_body":
        "<b>체력</b> — 최대 체력입니다. 이 값이 0 이 되면 쓰러집니다.\n"
        "<b>근거리 공격력</b> — 붙어서 때릴 때의 타격입니다.\n"
        "<b>원거리 공격력</b> — 떨어져서 쏠 때의 타격입니다.\n"
        "<b>마법</b> — 마법으로 칠 때의 타격입니다. 넓은 자리를 함께 칩니다.\n"
        "<b>회복력</b> — 동료를 되살릴 때의 회복량입니다.\n"
        "★ 위의 넷 중 <b>실제로 쓰이는 것은 하나</b>입니다. 전술 지침에서 고른 "
        "공격 유형이 그것을 정합니다. 고르지 않은 칸은 올려도 쓰이지 않습니다.\n"
        "\n"
        "<b>방어력</b> — 받는 피해를 줄입니다. 높을수록 한 대가 덜 아픕니다.\n"
        "<b>체력 재생</b> — 싸움이 끝나고 잠시 뒤부터 스스로 아뭅니다.\n"
        "<b>공격 속도</b> — 같은 시간에 몇 번 때리는지를 정합니다.\n"
        "<b>이동 속도</b> — 걷는 빠르기입니다. 물러설 때도 이 값이 씁니다.\n"
        "\n"
        "<b>명중률</b> — 빗나가지 않을 확률입니다.\n"
        "<b>크리티컬</b> — 급소를 칠 확률입니다. 급소는 피해가 <b>1.5배</b>입니다.\n"
        "⚠ 이 둘은 <b>원거리 공격에만</b> 걸립니다. 근거리·마법·회복은 언제나 맞고 "
        "급소가 뜨지 않습니다. 몇몇 타고난 능력이 그 문을 열어 주기도 합니다.\n"
        "\n"
        "<b>저항력</b> — 침식이 차는 속도와 빠지는 속도를 정합니다.\n"
        "⚠ 저항력만은 <b>강화로 오르지 않습니다</b>. 타고난 값 그대로 갑니다.\n"
        "\n"
        "능력치의 끝은 <b>100</b> 입니다. <b>영웅 각성</b>과 <b>유물</b>만 그 위를 넘습니다.\n"
        "성장 창에 적힌 숫자에는 유물과 각성 보너스가 이미 더해져 있습니다.",

    # ──────────────────────────────────────────────────────────────
    "help_mental_detail_title": "정신 이상 낱낱",

    "help_mental_detail_summary":
        "침식이 100 에 닿으면 열한 가지 중 하나가 걸립니다.\n"
        "무엇이 걸리는지는 고를 수 없습니다.",

    "help_mental_detail_body":
        "침식이 100 에 닿는 순간 하나를 뽑습니다. 걸리고 나면 침식은 절반 남짓으로 "
        "내려가지만 0 으로 돌아가지는 않습니다. 그래서 두 번째는 더 빨리 옵니다.\n"
        "\n"
        "■ 스스로를 해치는 것\n"
        "<b>자해</b> — 그 자리에서 최대 체력의 4분의 1을 잃습니다.\n"
        "<b>피학</b> — 45초 동안 스스로 체력이 줄어듭니다.\n"
        "<b>이기심</b> — 90초 동안 동료의 치유를 받지 못합니다. 스스로 아무는 것만 남습니다.\n"
        "\n"
        "■ 말을 듣지 않는 것\n"
        "<b>혼란</b> — 30초 동안 <b>동료를 공격</b>합니다.\n"
        "<b>공포</b> — 45초 동안 싸움을 거부하고 성역 쪽으로 물러납니다.\n"
        "<b>광분</b> — 90초 동안 전술 지침을 버리고 앞으로 뛰쳐나갑니다.\n"
        "\n"
        "■ 옆 사람에게 옮는 것\n"
        "<b>우울</b> — 곁의 동료들에게 침식을 옮깁니다.\n"
        "<b>역겨움</b> — 40초 동안 곁의 동료들의 체력을 갉아 냅니다.\n"
        "\n"
        "■ 오히려 도움이 되는 것\n"
        "<b>진정</b> — 곁의 동료들의 침식을 덜어 줍니다.\n"
        "<b>각성</b> — 120초 동안 모든 능력치가 올라갑니다.\n"
        "<b>고조</b> — 에너지를 쓰지 않고 강화됩니다. 다음 강화 비용은 그만큼 오릅니다.\n"
        "\n"
        "좋은 셋이 섞여 있어도 <b>기대할 것은 못 됩니다</b>. 나쁜 여덟이 훨씬 자주 나옵니다.\n"
        "침식을 아예 안 쌓는 방법은 없습니다. <b>누구에게 언제 쌓게 할지</b>를 고르는 일입니다.\n"
        "전방에 오래 세워 둔 동료일수록 빨리 찹니다. 정비 시간에 뒤로 빼 두면 그만큼 빠집니다.",
}


def main():
    wb = openpyxl.load_workbook(HELP_XLSX)

    # ── ① Help 시트 ──────────────────────────────────────────────
    ws = wb["Help"]
    existing = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}

    row = ws.max_row + 1
    added = 0
    for values in NEW_ROWS:
        if values[0] in existing:
            print(f"  건너뜀 — {values[0]} 은(는) 이미 있습니다")
            continue
        for c, v in enumerate(values, start=1):
            ws.cell(row, c).value = v
        print(f"  Help 추가 — {values[0]} ({values[11]})")
        row += 1
        added += 1

    relinked = 0
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if key in RELINK:
            before = ws.cell(r, 11).value
            ws.cell(r, 11).value = RELINK[key]
            print(f"  see_also — {key}: {before} → {RELINK[key]}")
            relinked += 1

    # ── ② StringKeys 시트 ────────────────────────────────────────
    ks = wb["StringKeys"]
    have = {}
    for r in range(2, ks.max_row + 1):
        k = ks.cell(r, 1).value
        if k:
            have[k] = r

    note = {
        "help_stats_detail_title": "성장 · 제목",
        "help_stats_detail_summary": "성장 · 조언 카드",
        "help_stats_detail_body": "성장 · 백과 본문",
        "help_mental_detail_title": "위험 · 제목",
        "help_mental_detail_summary": "위험 · 조언 카드",
        "help_mental_detail_body": "위험 · 백과 본문",
    }

    krow = ks.max_row + 1
    for key, text in STRINGS.items():
        r = have.get(key, krow)
        if key not in have:
            krow += 1
        ks.cell(r, 1).value = key
        ks.cell(r, 2).value = text
        ks.cell(r, 4).value = note.get(key, "")
        print(f"  StringKeys — {key} ({len(text)}자)")

    wb.save(HELP_XLSX)
    print(f"\n저장 완료 — Help {added}줄 추가 · see_also {relinked}줄 수정 · "
          f"StringKeys {len(STRINGS)}키")
    print("다음: help_string_merge.py → gen_string_table.py → link_string_keys.py → gen_help_assets.py")


if __name__ == "__main__":
    main()
