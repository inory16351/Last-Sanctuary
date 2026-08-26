# -*- coding: utf-8 -*-
"""유저가 «스트링 키 테이블» 에서 직접 다듬은 도움말 문구를 <b>도움말 표로 되올린다</b>
(2026-08-26 · 11차).

■ 왜 필요한가 — 규칙이 «도움말 표가 정본» 이기 때문이다
  `help_string_merge.py` 는 <b>도움말 표 → 스트링 키 테이블</b> 한 방향으로 덮어쓴다
  (그 파일 머리글: "help_ 로 시작하는 키는 도움말 표가 정본"). 그래서 스트링 키 테이블에서
  고친 도움말 문구는 <b>다음 병합에 조용히 되돌아간다</b>.
  유저가 2026-08-26 에 「능력치 낱낱」·「정신 이상 낱낱」 여섯 칸을 스트링 키 테이블에서
  다시 썼다(영어 용어도 erosion → <b>Corruption</b> 으로 통일했다). 그 문구가 정본이므로
  <b>도움말 표에 되올려</b> 두 표를 같게 만든다 — 안 그러면 언젠가 사라진다.

■ 무엇을 하나
  스트링 키 테이블 `string` 시트의 kr·en 을 읽어 도움말 표 `StringKeys` 시트의 같은 키에
  <b>그대로 덮어쓴다</b>. 대상은 아래 KEYS 여섯 개뿐이다 — 다른 도움말 문구는 손대지 않는다.

■ 다음 순서 (도움말 문구를 고친 뒤에는 늘 이 셋)
    py -3 Tools/help_string_merge.py          도움말 표 → 스트링 키 테이블
    py -3 Tools/gen_string_table.py           스트링 키 테이블 → StringTable.txt
    py -3 Tools/link_string_keys.py           하이퍼링크 재생성
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
HELP_XLSX = os.path.join(TABLE_DIR, 'Last_Sanctuary_도움말테이블_Ver01.xlsx')

#: 유저가 스트링 키 테이블에서 다시 쓴 칸. 이 여섯만 되올린다.
KEYS = [
    'help_stats_detail_title', 'help_stats_detail_summary', 'help_stats_detail_body',
    'help_mental_detail_title', 'help_mental_detail_summary', 'help_mental_detail_body',
]

# 스트링 키 테이블은 3행 헤더 + 4행부터 데이터다(gen_string_table.py 의 DATA_ROW0).
STRING_DATA_ROW0 = 4


def read_string_table():
    wb = openpyxl.load_workbook(STRING_XLSX, data_only=True)
    ws = wb['string']
    out = {}
    for row in ws.iter_rows(min_row=STRING_DATA_ROW0, values_only=True):
        key = row[0]
        if key is None or str(key).strip() == '':
            continue
        out.setdefault(str(key).strip(), (row[1] or '', row[2] or ''))
    return out


def main():
    src = read_string_table()

    missing = [k for k in KEYS if k not in src]
    if missing:
        sys.exit('스트링 키 테이블에 없는 키: ' + ', '.join(missing))

    wb = openpyxl.load_workbook(HELP_XLSX)
    ws = wb['StringKeys']

    changed = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if key is None:
            continue
        key = str(key).strip()
        if key not in KEYS:
            continue
        seen.add(key)
        kr, en = src[key]
        before = (ws.cell(row=r, column=2).value or '', ws.cell(row=r, column=3).value or '')
        if before == (kr, en):
            continue
        ws.cell(row=r, column=2).value = kr
        ws.cell(row=r, column=3).value = en
        changed.append(key)

    absent = [k for k in KEYS if k not in seen]
    if absent:
        sys.exit('도움말 표 StringKeys 시트에 없는 키: ' + ', '.join(absent))

    if not changed:
        print('바뀐 칸이 없습니다 — 두 표가 이미 같습니다.')
        return

    bak = HELP_XLSX + '.bak'
    shutil.copyfile(HELP_XLSX, bak)
    wb.save(HELP_XLSX)

    print('[스트링 → 도움말] 되올린 칸 %d개' % len(changed))
    for k in changed:
        print('  ~', k, '|', str(src[k][0])[:50].replace('\n', '/'))
    print('  저장:', os.path.basename(HELP_XLSX), ' (백업 %s)' % os.path.basename(bak))
    print('  다음: py -3 Tools/help_string_merge.py  →  gen_string_table.py  →  link_string_keys.py')


if __name__ == '__main__':
    main()
