# -*- coding: utf-8 -*-
"""도움말의 <b>분류 이름</b>과 <b>짚어 주기 단계 글</b>을 스트링 키로 옮긴다 (2026-08-27 · 184절).

■ 무엇이 빠져 있었나 — 183절이 «화면에 뜨는 한글 중 표를 안 거치는 것 = 0» 이라고
  셌는데도 도움말이 영어로 안 바뀌는 자리가 <b>둘</b> 남아 있었다. 둘 다 <b>코드가 아니라
  에셋</b>에 한글이 들어 있어 그 전수 조사(코드의 리터럴을 세는 방식)에 안 잡혔다.

    ① `HelpEntry.category` — 백과 위쪽의 <b>탭 여섯</b>·상세의 분류 칸·조언 카드의 머리표.
       「기본 · 전투 · 성장 · 지휘 · 위험 · 운영」이 표에서 <b>글자 그대로</b> 와서 그대로 찍혔다.
    ② `HelpStepRow.stepText` — 「자세히 보기」가 빨간 테두리로 짚어 주는 <b>단계 글 44줄</b>.
       제목·요약·본문은 진작 키를 거치는데 <b>이 44줄만</b> 표에 한글로 박혀 있었다.

  ⚠ ①은 «표시용 이름» 인 동시에 «분류를 가르는 식별자» 다(`CollectByCategory` 가 이 값으로
    묶는다). 그래서 <b>`category` 는 그대로 두고 `category_key` 를 따로 더한다</b> —
    식별자를 번역하면 언어를 바꾼 순간 «그 분류에 속한 항목이 하나도 없다» 가 된다.

■ 무엇을 하나
    Help 시트       : `category_key` 열을 더하고 채운다
    HelpStep 시트   : `step_text_key` 열을 더하고 채운다 (키 = <help_id>_step<step_order>)
    StringKeys 시트 : 위의 키 6 + 44 = <b>50줄</b>을 kr·en 과 함께 붙인다

  ★ 단계 글의 <b>한글은 HelpStep 시트에서 그대로 옮겨 온다</b> — 손으로 다시 치지 않는다.
    영어만 이 파일이 들고 있다.
  ★ <b>여러 번 돌려도 안전하다</b> — 이미 열이 있으면 값만 다시 쓰고, StringKeys 에 이미
    있는 키는 <b>덮어쓰지 않는다</b>(유저가 다듬은 영어를 되돌리지 않기 위해서다).

■ 다음 순서
    py -3 Tools/help_string_merge.py     도움말 표 → 스트링 키 테이블
    py -3 Tools/gen_string_table.py      스트링 키 테이블 → StringTable.txt
    py -3 Tools/link_string_keys.py      하이퍼링크 재생성
    py -3 Tools/gen_help_assets.py       도움말 표 → HelpTable.asset
"""
import os
import shutil
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

HELP_XLSX = os.path.join(TABLE_DIR, 'Last_Sanctuary_도움말테이블_Ver01.xlsx')

HELP_ROW0 = 2      # 도움말 표는 머리글 한 줄이다

#: 분류 이름 → (스트링 키, 영어). ⚠ 왼쪽 한글은 <b>식별자</b>라 절대 바꾸지 않는다.
CATEGORY = {
    '기본': ('help_cat_basic',   'Basics'),
    '전투': ('help_cat_combat',  'Combat'),
    '성장': ('help_cat_growth',  'Growth'),
    '지휘': ('help_cat_command', 'Command'),
    '위험': ('help_cat_danger',  'Danger'),
    '운영': ('help_cat_system',  'System'),
}

#: (help_id, step_order) → 영어. 한글은 HelpStep 시트에서 읽어 온다.
STEP_EN = {
    ('help_awaken', 1):
        "An awakened ally's name turns gold. They show in the same color in the roster.",
    ('help_awaken', 2):
        "Whether they have awakened, and what conditions are left, is written here.",
    ('help_awaken', 3):
        "Awakening adds a bonus to these fields, and they can rise above 100.",

    ('help_create', 1):
        "Press this button and one new ally arrives. The number beside it is the Energy "
        "it costs right now.",

    ('help_energy', 1):
        "This is the Energy you have. Summoning or upgrading allies is paid out of this number.",

    ('help_epic', 1):
        "Epics you have found are listed here, along with the level needed to handle them.",
    ('help_epic', 2):
        "Choose which squad to send here. Only the squads you give the order to will go.\n"
        "You can send up to two squads to slay a neutral epic monster.",
    ('help_epic', 3):
        "Guidance on what you should choose right now appears here.",

    ('help_erosion', 1):
        "The purple gauge is Corruption. It fills while they fight and falls on its own "
        "once the fighting ends.",
    ('help_erosion', 2):
        "Among the stats, Resistance decides how fast Corruption fills and how fast it drains.",

    ('help_mental_error', 1):
        "When this gauge fills to the end, one Mental Breakdown appears.",
    ('help_mental_error', 2):
        "The name of the Mental Breakdown they are under is written here. "
        "It is shown in the roster as well.",

    ('help_rally', 1):
        "Every squad slot has a button to set a rally point and one to clear it.",
    ('help_rally', 2):
        "After pressing Set, click a spot on the ground. Hold the flag for about a second "
        "and drag to move it.",

    ('help_relic_equip', 1):
        "Every relic you have found collects here. Higher grades come to the top.",
    ('help_relic_equip', 2):
        "What the relic you picked does for you is written here.",
    ('help_relic_equip', 3):
        "This button equips it. It goes to the ally you picked in the roster on the left.",
    ('help_relic_equip', 4):
        "If someone is already wearing that relic, their name appears here.",

    ('help_retreat', 1):
        "Drag this slider to set the retreat threshold. They pull back on their own when "
        "that much Health is left.",
    ('help_retreat', 2):
        "This is the threshold you set. Low for the front line and high for the back line "
        "works well.",
    ('help_retreat', 3):
        "Choose what they do after pulling back.",

    ('help_save', 1):
        "This button records where you are now. There is only one save slot.",
    ('help_save', 2):
        "Saves and leaves for the lobby. Use Continue in the lobby to pick this run back up.",
    ('help_save', 3):
        "The time of the last save is written here.",

    ('help_speed', 1):
        "These buttons pick the game speed. The speed in effect is shown brightly.",
    ('help_speed', 2):
        "This is the fastest, 8x. It is handy for skipping through Preparation, "
        "when there are no enemies.",
    ('help_speed', 3):
        "This button, or P, stops time. Even while stopped you can open windows and "
        "change directives.",

    ('help_squad', 1):
        "This button makes a new squad. You can have up to six squads.",
    ('help_squad', 2):
        "The squads you make line up here as slots. Click an ally in the roster on the left "
        "to assign them.",

    ('help_stats', 1):
        "These are the twelve stats. The numbers here already include relic and "
        "Awakening bonuses.",
    ('help_stats', 2):
        "These are the stats this ally mainly uses. They rise more often, and by more, "
        "when you upgrade.",
    ('help_stats', 3):
        "These are innate special abilities. Upgrading does not change them.",
    ('help_stats', 4):
        "These are the relics equipped right now. Stats from relics are added past the 100 cap.",

    ('help_tactics', 1):
        "Choose whether they stand at the front, middle, or rear. The front line takes the "
        "enemy head-on, and the rear fights from the edge of its range.",
    ('help_tactics', 2):
        "Choose how they attack.",
    ('help_tactics', 3):
        "Choose who they strike first when there are several enemies.",
    ('help_tactics', 4):
        "Choose how they move when the enemy comes in a wave.",
    ('help_tactics', 5):
        "Gathers the directives you chose into one view. This alone tells you the "
        "current setup.",

    ('help_upgrade', 1):
        "This button is Upgrade. Every press raises the level of the ally you picked by 1.",
    ('help_upgrade', 2):
        "This is the Energy one upgrade costs. The value grows as the level rises.",
    ('help_upgrade', 3):
        "This is the level right now. It goes up every time you upgrade.",
    ('help_upgrade', 4):
        "When the level rises, some of these stat fields rise with it. Which ones is "
        "decided at that moment.",

    ('help_wave', 1):
        "This field tells you which phase you are in. It reads Preparation, Advance, "
        "Battle, or Enraged.",
    ('help_wave', 2):
        "This is the time left. When it reaches 0 during Preparation, the enemy comes.",
}

HEAD_FILL = PatternFill('solid', fgColor='FF1F3B4D')
HEAD_FONT = Font(bold=True, color='FFFFFFFF')


def norm(v):
    return '' if v is None else str(v).strip()


def field_col(ws, name):
    """1행에서 필드명의 열 번호. 없으면 None."""
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(1, c).value) == name:
            return c
    return None


def ensure_column(ws, name, width):
    """그 필드가 없으면 <b>맨 뒤에</b> 머리글을 만들어 준다. 있으면 그 열을 그대로 쓴다."""
    c = field_col(ws, name)
    if c is not None:
        return c, False

    c = ws.max_column + 1
    cell = ws.cell(1, c)
    cell.value = name
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions[cell.column_letter].width = width
    return c, True


def step_key(help_id, order):
    """단계의 스트링 키. ⚠ help_id 가 이미 `help_` 로 시작하므로 접두사 규칙을 지킨다."""
    return '%s_step%d' % (help_id, order)


def main():
    if not os.path.isfile(HELP_XLSX):
        raise SystemExit('⚠ 도움말 표가 없습니다: %s' % HELP_XLSX)

    wb = openpyxl.load_workbook(HELP_XLSX)

    # ── ① Help 시트 · category_key ────────────────────────────────────
    ws = wb['Help']
    c_cat = field_col(ws, 'category')
    if c_cat is None:
        raise SystemExit("⚠ Help 시트에 'category' 열이 없습니다.")
    c_key, made = ensure_column(ws, 'category_key', 22.0)

    unknown, filled = set(), 0
    for r in range(HELP_ROW0, ws.max_row + 1):
        if not norm(ws.cell(r, 1).value):
            continue
        cat = norm(ws.cell(r, c_cat).value)
        if cat not in CATEGORY:
            unknown.add(cat)
            continue
        ws.cell(r, c_key).value = CATEGORY[cat][0]
        filled += 1

    print('[Help] category_key 열 %s · %d줄 채움'
          % ('신설' if made else '이미 있음', filled))
    if unknown:
        # ⚠ 조용히 넘기지 않는다 — 표에 분류가 늘면 여기 CATEGORY 도 늘려야 한다.
        print('  ⚠ 이 파일이 모르는 분류 %d개 — 키가 비어 화면에 한글이 그대로 남습니다: %s'
              % (len(unknown), ' · '.join(sorted(unknown))))

    # ── ② HelpStep 시트 · step_text_key ───────────────────────────────
    ws = wb['HelpStep']
    c_id = field_col(ws, 'help_id')
    c_ord = field_col(ws, 'step_order')
    c_txt = field_col(ws, 'step_text')
    if None in (c_id, c_ord, c_txt):
        raise SystemExit('⚠ HelpStep 시트의 열 이름이 예전과 다릅니다.')
    c_skey, made = ensure_column(ws, 'step_text_key', 26.0)

    steps, missing_en = [], []
    for r in range(HELP_ROW0, ws.max_row + 1):
        hid = norm(ws.cell(r, c_id).value)
        if not hid:
            continue
        order = int(round(float(ws.cell(r, c_ord).value or 0)))
        # ⚠ 앞뒤 공백만 다듬고 <b>줄바꿈은 그대로</b> 둔다 — 접는 일은
        #   gen_string_table.py 가 TSV 로 내보낼 때 한다(help_string_merge.py 의 ⚠).
        kr = (ws.cell(r, c_txt).value or '').strip()

        key = step_key(hid, order)
        ws.cell(r, c_skey).value = key

        en = STEP_EN.get((hid, order))
        if not en:
            missing_en.append(key)
        steps.append((key, kr, en or '', '%s · 짚어 주기 %d단계' % (hid, order)))

    print('[HelpStep] step_text_key 열 %s · %d줄'
          % ('신설' if made else '이미 있음', len(steps)))
    if missing_en:
        print('  ⚠ 영어가 없는 단계 %d개 — 영어에서도 한글이 그대로 뜹니다: %s'
              % (len(missing_en), ' · '.join(missing_en)))

    # ── ③ StringKeys 시트 · 키 50줄 ───────────────────────────────────
    ws = wb['StringKeys']
    where = {}
    for r in range(HELP_ROW0, ws.max_row + 1):
        k = norm(ws.cell(r, 1).value)
        if k:
            where[k] = r

    rows = [(key, cat, en, '백과 분류 탭')
            for cat, (key, en) in CATEGORY.items()] + steps

    added, skipped = 0, 0
    write_row = ws.max_row + 1
    for key, kr, en, note in rows:
        if key in where:
            # ★ 덮어쓰지 않는다 — 유저가 다듬은 문구를 되돌릴 이유가 없다.
            skipped += 1
            continue
        ws.cell(write_row, 1).value = key
        ws.cell(write_row, 2).value = kr
        ws.cell(write_row, 3).value = en
        ws.cell(write_row, 4).value = note
        ws.cell(write_row, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(write_row, 3).alignment = Alignment(wrap_text=True, vertical='top')
        where[key] = write_row
        write_row += 1
        added += 1

    print('[StringKeys] 새 키 %d개 붙임 · 이미 있어 건너뛴 키 %d개' % (added, skipped))

    shutil.copy2(HELP_XLSX, HELP_XLSX + '.bak')
    wb.save(HELP_XLSX)
    print('저장: %s  (백업 %s)'
          % (os.path.basename(HELP_XLSX), os.path.basename(HELP_XLSX) + '.bak'))
    print('다음: py -3 Tools/help_string_merge.py  →  gen_string_table.py  →  '
          'link_string_keys.py  →  gen_help_assets.py')
    return 0


if __name__ == '__main__':
    sys.exit(main())
