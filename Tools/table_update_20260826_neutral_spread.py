# -*- coding: utf-8 -*-
"""중립 사냥값을 <b>전 구간에 나눈다</b> — 종양 두더지 쏠림을 걷어낸다 (2026-08-26, 2차).

유저 지시
---------
  *"지금 그리고 종양 두더지 사냥 보상이 과도하게 높은데 전 구간에 나눠서 경제 밸런스를
    잡아줘 중립 몬스터 사냥값"*

★★★ 무엇이 잘못돼 있었나 — <b>거리로만 곡선을 만들려다 한 종에 다 실렸다</b>
──────────────────────────────────────────────────────────────────────
UI-67 은 «먼 종의 보상만 올리면 후반에만 수입이 오른다» 는 생각으로
1003(종양 두더지)과 에픽 넷을 **×3.0** 했다. 곡선의 <b>모양</b>은 맞았는데
<b>한 종이 그 곡선을 통째로 짊어졌다</b>.

    지속 수급 상한 = maxAlive ÷ respawnSeconds × 마리당 에너지

    1001 종양 거미   43 / 9초  ×   9 =    43 E/s
    1002 종양귀      33 / 13초 ×  46 =   116 E/s
    1004 고르도네    33 / 13초 ×  46 =   116 E/s
    1003 종양 두더지 19 / 25초 × 795 =   604 E/s   ← 다음 종의 5.2배

두더지는 체력이 <b>14</b> 다 — 1002(8)·1004(15)와 같은 급이다. «멀다» 는 것 말고는
비싼 이유가 없는데 마리당 값이 중반 종의 <b>17배</b> 였다. 그래서 26~30 구간이
<b>한 판 중립 수입의 58%</b> 를 차지했고, 그 구간은 사실상 «두더지 밭에 원정
보내는 것» 하나로 굴러갔다.

★ 구간별 사냥 구성 모델 (옛 수입에서 되짚었다 — 오차 구간별 ±4.3% · 총합 +1.2%)
──────────────────────────────────────────────────────────────────────
    구간      1001   1002+1004   1003   에픽      →  시트의 중립수입/웨이브
    1~5        20        0         0    0.0            176
    6~10       39        0         0    0.0            352
    11~15      20       18         0    0.0           1013
    16~20      20       46         0    0.0           2291
    21~25      20       40         3    0.5           5282
    26~30      20       40        10    1.2          12588

  이 구성으로 옛 값을 넣으면 시트가 재현된다 — <b>그래서 이 모델로 새 값을 고를 수 있다</b>.

★ 새 사냥값 — <b>총 수입은 그대로, 실리는 자리만 옮긴다</b>
──────────────────────────────────────────────────────────────────────
    1001 종양 거미     6~12   →   11~17    (평균  9 →  14)
    1002 종양귀       34~57   →   62~98    (평균 46 →  80)
    1004 고르도네     34~57   →   62~98    (평균 46 →  80)
    1003 종양 두더지 600~990  →  260~420   (평균 795 → 340)   ← −57%
    1101~1104 에픽    그대로 (유저가 문제 삼지 않았고, 20분 리스폰·레벨 제한이 이미 잠근다)

    지속 수급이 <b>단조롭게</b> 오른다:  67 → 203 → 258 E/s
    마리당 값도 «멀리 갈수록 비싸다» 가 유지된다: 14 → 80 → 340

    구간 몫    1~5   6~10  11~15  16~20  21~25  26~30
      옛       0.8%  1.6%   4.7%  10.6%  24.3%  58.0%
      새       1.3%  2.5%   7.9%  18.2%  25.9%  44.1%

    총 중립 수입 108,510 → 108,555 (+0.04%) — 착지점 «w30 12명 Lv35» 는 그대로다.

⚠⚠ <b>대가 — 중반이 세진다.</b> 후반에 오던 돈이 중반으로 오므로 파티 레벨이
    w15 11.9→15.0 · w20 20.4→25.9 · w25 24.4→28.5 로 오른다(w30 은 35.2 로 같다).
    16~25 웨이브의 <b>잡몹 부하가 그만큼 내려간다</b> — 이 스크립트가 시트의 부하 열을
    다시 계산해 그 사실을 남긴다. 너무 싱거우면 그때 그 구간 몬스터를 잡을 것.
    (몬스터 수치를 여기서 손대지 않는 것은 UI-67 의 유저 확정 ② 를 그대로 지킨 것이다.)

⚠ 이 표들은 <b>사람이 엑셀에서 고쳐 나가는 문서</b>다 — 통째로 다시 굽지 않고 바꿀 칸만 고친다.
⚠ <b>Excel COM 으로 쓴다</b> — openpyxl 로 저장하면 하이퍼링크·주석·수식 캐시가 날아간다(136-4절).

사용법:  python Tools/table_update_20260826_neutral_spread.py
다음:    python Tools/sync_tables_to_assets.py     (중립 에셋에 반영)
         Unity 에서 Assets/Refresh
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

NEUTRAL_XLSX = os.path.join(TABLE_DIR, "임시용 중립 몬스터.xlsx")
FORMULA_XLSX = os.path.join(TABLE_DIR, "능력치 및 공식 정리.xlsx")

# ══════════════════════════════════════════════════════════════════════
#  ① 중립 보상 — mon_id → (min_energy, max_energy)
# ══════════════════════════════════════════════════════════════════════
NEW_ENERGY = {
    1001: (11, 17),          # 평균  9 →  14
    1002: (62, 98),          # 평균 46 →  80
    1004: (62, 98),          # 평균 46 →  80
    1003: (260, 420),        # 평균 795 → 340   ★ 이번 지시의 본체
    # 1101~1104 에픽은 그대로 둔다.
}

# ══════════════════════════════════════════════════════════════════════
#  ② 구간별 사냥 구성 — (구간 상한, (1001, 1002+1004, 1003, 에픽))
# ══════════════════════════════════════════════════════════════════════
MIX = [
    (5,  (20,  0,  0, 0.0)),
    (10, (39,  0,  0, 0.0)),
    (15, (20, 18,  0, 0.0)),
    (20, (20, 46,  0, 0.0)),
    (25, (20, 40,  3, 0.5)),
    (30, (20, 40, 10, 1.2)),
]

EPIC_AVG = 2250.0        # 1101~1103 의 평균 (1104 는 3150 이지만 26~30 에만 섞인다)


def avg(mon_id):
    lo, hi = NEW_ENERGY[mon_id]
    return (lo + hi) / 2.0


def mix_for(wave):
    for upto, m in MIX:
        if wave <= upto:
            return m
    return MIX[-1][1]


def neutral_income(wave):
    n1, n2, n3, ne = mix_for(wave)
    return n1 * avg(1001) + n2 * avg(1002) + n3 * avg(1003) + ne * EPIC_AVG


# ══════════════════════════════════════════════════════════════════════
#  ③ 성장 모델 — UI-67 의 계수를 <b>그대로</b> 쓴다 (두 벌이 되면 안 된다)
# ══════════════════════════════════════════════════════════════════════
FOCUS_PER_LEVEL = 2.62
PLAIN_PER_LEVEL = 1.62
STAT_BASE_FOCUS = 10
STAT_BASE_PLAIN = 5
STAT_MAX = 100

HP_PER_STAT = 6.1832
HP_BASE = 90

DPS_A, DPS_B, DPS_C = 0.01961, 3.329, -9.50
EFFECTIVE = 0.647
BATTLE_SECONDS = 120

CREATE_BASE, CREATE_STEP = 200, 150
FREE_CHARACTERS = 4

UP_BASE, UP_STEP = 40, 10
UP_STEEP_LEVEL, UP_STEEP_RATE, UP_ROUND = 30, 1.35, 10

WAVE_KILL_ENERGY = 10

SHEET = "웨이브 부하"
FIRST_ROW = 7
COL = dict(wave=1, lv=2, people=3, focus=4, plain=5, hp=6,
           hp_scale=7, atk_scale=8, count=9, mob_hp=10, mob_hit=11,
           hits=12, hit_ratio=13, dps=14, load=15,
           boss=16, boss_raw=17, boss_hp=18, boss_goal=19, neutral=20)


def upgrade_cost(level):
    linear = UP_BASE + UP_STEP * level
    if level < UP_STEEP_LEVEL:
        return linear
    at_start = UP_BASE + UP_STEP * UP_STEEP_LEVEL
    cost = at_start * (UP_STEEP_RATE ** (level - UP_STEEP_LEVEL))
    return int(round(cost / UP_ROUND) * UP_ROUND)


def run_model(rows, neutral):
    """UI-67 의 run_model 과 <b>같은 순서</b>다 — ① 충원 ② 최저 레벨부터 강화 ③ 기록 ④ 번다."""
    energy = 0.0
    levels = []
    created = 0
    out = []

    for r, nt in zip(rows, neutral):
        while len(levels) < r["people"]:
            if len(levels) < FREE_CHARACTERS:
                levels.append(0)
                continue
            cost = CREATE_BASE + CREATE_STEP * created
            if energy < cost:
                break
            energy -= cost
            created += 1
            levels.append(0)

        guard = 0
        while levels and guard < 500000:
            guard += 1
            i = min(range(len(levels)), key=lambda j: levels[j])
            c = upgrade_cost(levels[i])
            if energy < c:
                break
            energy -= c
            levels[i] += 1

        out.append(sum(levels) / len(levels) if levels else 0.0)

        energy += r["count"] * 2 * WAVE_KILL_ENERGY
        energy += nt

    return out


def read_rows():
    wb = openpyxl.load_workbook(FORMULA_XLSX, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=FIRST_ROW, max_row=ws.max_row, values_only=True):
        if r[0] is None:
            continue
        rows.append(dict(
            wave=int(r[0]), lv=float(r[1]), people=int(r[2]),
            focus=float(r[3]), plain=float(r[4]), hp=float(r[5]),
            count=int(r[8]), mob_hp=float(r[9]), mob_hit=float(r[10]),
            hits=float(r[11]), hit_ratio=float(r[12]), dps=float(r[13]),
            load=float(r[14]), neutral=float(r[19]),
        ))
    return rows


def main():
    rows = read_rows()
    print(f"「{SHEET}」 {len(rows)}행 읽음")

    # ── 구성 모델이 옛 시트를 재현하는지 먼저 확인한다 ──────────────
    old_vals = {1001: 9.0, 1002: 45.5, 1004: 45.5, 1003: 795.0}
    print("\n구성 모델 검산 (옛 사냥값 → 시트의 중립수입)")
    seen = set()
    for r in rows:
        band = next(u for u, _ in MIX if r["wave"] <= u)
        if band in seen:
            continue
        seen.add(band)
        n1, n2, n3, ne = mix_for(r["wave"])
        m = n1 * old_vals[1001] + n2 * old_vals[1002] + n3 * old_vals[1003] + ne * EPIC_AVG
        print(f"  ~{band:2d}웨이브  시트 {r['neutral']:7.0f}  모델 {m:7.0f}  "
              f"({(m / r['neutral'] - 1) * 100:+.1f}%)")

    # ── 새 수입 곡선 ────────────────────────────────────────────────
    neutral = [neutral_income(r["wave"]) for r in rows]
    old_total = sum(r["neutral"] for r in rows)
    new_total = sum(neutral)
    print(f"\n중립 총 수입 {old_total:,.0f} → {new_total:,.0f}  ({(new_total/old_total-1)*100:+.2f}%)")

    print("\n구간 몫")
    seen = set()
    for r, nt in zip(rows, neutral):
        band = next(u for u, _ in MIX if r["wave"] <= u)
        if band in seen:
            continue
        seen.add(band)
        print(f"  ~{band:2d}웨이브  {r['neutral']:7.0f} → {nt:7.0f} /웨이브   "
              f"몫 {r['neutral']*5/old_total*100:5.1f}% → {nt*5/new_total*100:5.1f}%")

    new_lv = run_model(rows, neutral)

    # ── 파생 칸 다시 계산 (UI-67 과 같은 식) ────────────────────────
    plan = []
    for r, lv, nt in zip(rows, new_lv, neutral):
        n = round(lv)
        focus = min(STAT_MAX, STAT_BASE_FOCUS + FOCUS_PER_LEVEL * n)
        plain = min(STAT_MAX, STAT_BASE_PLAIN + PLAIN_PER_LEVEL * n)
        hp = round(HP_BASE + HP_PER_STAT * (focus - STAT_BASE_FOCUS))

        per = DPS_A * focus * focus + DPS_B * focus + DPS_C
        dps = round(r["people"] * per)

        total_mob_hp = r["count"] * 2 * r["mob_hp"]
        load = total_mob_hp / (dps * EFFECTIVE * BATTLE_SECONDS) if dps > 0 else 0.0

        ratio = r["hit_ratio"] * (r["hp"] / hp) if hp > 0 else 0.0
        hits = r["hits"] * (hp / r["hp"]) if r["hp"] > 0 else 0.0

        plan.append(dict(
            wave=r["wave"], lv=round(lv, 1), focus=round(focus, 1), plain=round(plain, 1),
            hp=hp, dps=dps, load=round(load, 2),
            ratio=round(ratio, 3), hits=round(hits, 1), neutral=round(nt),
        ))

    print("\n웨이브 | 평균Lv (전) | 캐릭터HP | 파티DPS | 잡몹부하 (전) | 중립수입 (전)")
    for p, r in zip(plan, rows):
        if p["wave"] % 5 == 0 or p["wave"] <= 2:
            print(f"{p['wave']:6d} | {p['lv']:5.1f} ({r['lv']:4.1f}) | {p['hp']:8d} | "
                  f"{p['dps']:7d} | {p['load']:5.2f} ({r['load']:4.2f}) | "
                  f"{p['neutral']:6d} ({int(r['neutral'])})")

    write_excel(plan)


def write_excel(plan):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        # ── ① 중립 보상 ────────────────────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(NEUTRAL_XLSX))
        ws = wb.Worksheets("neutrality_mon")
        touched = 0
        row = 4                      # 3행이 자료형 줄, 4행부터 자료. id A · min_energy J · max_energy K
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            mon_id = int(v)
            if mon_id in NEW_ENERGY:
                lo, hi = NEW_ENERGY[mon_id]
                before = (ws.Cells(row, 10).Value, ws.Cells(row, 11).Value)
                ws.Cells(row, 10).Value = lo
                ws.Cells(row, 11).Value = hi
                print(f"  중립 {mon_id}: {int(before[0])}~{int(before[1])} → {lo}~{hi}")
                touched += 1
            row += 1
        wb.Save()
        wb.Close()
        print(f"임시용 중립 몬스터.xlsx — {touched}종 갱신")

        # ── ② 「웨이브 부하」 + 「계수」 ─────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(FORMULA_XLSX))
        ws = wb.Worksheets(SHEET)

        for i, p in enumerate(plan):
            r = FIRST_ROW + i
            assert int(ws.Cells(r, COL["wave"]).Value) == p["wave"], \
                f"{r}행이 웨이브 {p['wave']} 가 아니다"
            ws.Cells(r, COL["lv"]).Value = p["lv"]
            ws.Cells(r, COL["focus"]).Value = p["focus"]
            ws.Cells(r, COL["plain"]).Value = p["plain"]
            ws.Cells(r, COL["hp"]).Value = p["hp"]
            ws.Cells(r, COL["hits"]).Value = p["hits"]
            ws.Cells(r, COL["hit_ratio"]).Value = p["ratio"]
            ws.Cells(r, COL["dps"]).Value = p["dps"]
            ws.Cells(r, COL["load"]).Value = p["load"]
            ws.Cells(r, COL["neutral"]).Value = p["neutral"]

        note = ("※ 2026-08-26 개정 2차: 중립 사냥값을 «전 구간에» 나눴다(유저 지시 — 종양 두더지 보상 과다). "
                "1003 은 지속 수급이 604 E/s 로 다음 종의 5.2배였다 → 마리당 795→340. "
                "대신 1001 9→14 · 1002·1004 46→80. 총 중립 수입은 그대로(108,5xx)라 착지점 «w30 12명 Lv35» 도 그대로다. "
                "구간 몫 0.8/1.6/4.7/10.6/24.3/58.0% → 1.3/2.5/7.9/18.2/25.9/44.1%. "
                "⚠ 대가로 16~25웨이브 파티 레벨이 4~5 오르고 그만큼 잡몹 부하가 내려간다 — 싱거우면 그 구간 몬스터를 잡을 것.")
        ws.Cells(5, 1).Value = note

        coef = wb.Worksheets("계수")
        r = coef.UsedRange.Rows.Count + 1
        for label, value, where, why in [
            ("■ 2026-08-26 (2차) — 중립 사냥값 재분배", "", "", ""),
            ("종양 거미 1001 평균 보상", 14, "임시용 중립 몬스터.min/max_energy 11~17",
             "초반 구간이 판 전체 중립 수입의 0.8% 뿐이라 «전 구간에 나눠서» 지시에 따라 소폭 올렸다"),
            ("종양귀 1002 · 고르도네 1004 평균 보상", 80, "임시용 중립 몬스터 62~98",
             "중반 구간의 주 수입원. 지속 수급 116 → 203 E/s"),
            ("종양 두더지 1003 평균 보상", 340, "임시용 중립 몬스터 260~420",
             "★ 이번 지시의 본체. 체력 14 로 중반 종과 같은 급인데 마리당 값이 17배였다 — 지속 수급 604 → 258 E/s"),
            ("구간 몫 (1~5 / 6~10 / 11~15 / 16~20 / 21~25 / 26~30)",
             "1.3 / 2.5 / 7.9 / 18.2 / 25.9 / 44.1 %", "위 세 값에서 파생",
             "옛 0.8/1.6/4.7/10.6/24.3/58.0% — 한 판 수입의 절반 이상이 마지막 다섯 웨이브에 몰려 있었다"),
        ]:
            coef.Cells(r, 1).Value = label
            if value != "":
                coef.Cells(r, 2).Value = value
            if where:
                coef.Cells(r, 3).Value = where
            if why:
                coef.Cells(r, 4).Value = why
            r += 1

        wb.Save()
        wb.Close()
        print("능력치 및 공식 정리.xlsx — 「웨이브 부하」 9열 × 30행 · 「계수」 5행 추가")
    finally:
        excel.Quit()


if __name__ == "__main__":
    main()
