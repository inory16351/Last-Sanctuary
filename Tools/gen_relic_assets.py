# -*- coding: utf-8 -*-
"""유물 테이블(xlsx) → `RelicDefinitionSO` 에셋 + `RelicDigTableSO` + <b>임시 아이콘</b>.

원본: ``<볼트>/데이터 테이블/Last_Sanctuary_유물테이블_Ver01.xlsx``
결과: ``Assets/_Project/Resources/Relics/*.asset``
      ``Assets/_Project/Resources/RelicIcons/*.png``   ← 임시 이미지(유저 지시 11번)

★ 이 스크립트가 <b>표와 게임 사이의 유일한 다리</b>다. 값을 바꾸려면 표를 고치고 다시 돌린다.

⚠ MCP 에는 SO 에셋을 다루는 도구가 없다 — 그래서 이 종류만 스크립트로 쓴다
  (`gen_event_assets.py` 가 같은 이유로 같은 방식이다. 씬 오브젝트는 MCP 로 만든다).

⚠ .asset YAML 에 <b>빈 줄을 넣으면 유니티 파서가 그 뒤 필드를 전부 무시한다</b>
  (진행상황 8절 3번). 아래 :func:`yaml_str` 이 줄바꿈을 이스케이프하는 이유다.

★ <b>guid 는 경로에서 결정적으로</b> 만든다 — 다시 돌려도 같은 guid 라 참조가 안 끊긴다
  (이 프로젝트의 모든 생성 스크립트가 같은 규칙).

임시 아이콘
-----------
원화가 없으므로 <b>등급 색 바탕 + 유물마다 다른 문양</b>을 그린다. 문양은 유물 ID 로
결정하므로 <b>같은 유물은 언제나 같은 그림</b>이다(다시 돌려도 안 바뀐다).
원화가 오면 같은 파일 이름으로 덮으면 된다 — 표의 ``icon`` 칸이 그 이름이다.

사용법:  py -3 Tools/gen_relic_assets.py
다음:    유니티에서 Assets/Refresh
"""

import hashlib
import io
import math
import os
import sys

import openpyxl
from PIL import Image, ImageDraw

from vault_path import TABLE_DIR, PROJECT

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "Last_Sanctuary_유물테이블_Ver01.xlsx")
OUT_DIR = os.path.join(PROJECT, "Assets", "_Project", "Resources", "Relics")
ICON_DIR = os.path.join(PROJECT, "Assets", "_Project", "Resources", "RelicIcons")

#: 표의 첫 세 줄은 «한글 제목 / 영문 키 / 자료형» 이다 — 값은 4행부터.
FIRST_ROW = 4

#: 아이콘 한 변(px). 화면에서 48~64px 로 쓰므로 두 배로 굽는다.
ICON_SIZE = 128

# ── 표의 enum → C# enum 정수값. <b>RelicDefinitionSO.cs 와 반드시 같아야 한다.</b> ──
GRADE = {"common": 1, "rare": 2, "epic": 3}
SOURCE = {"dig_monster": 1, "dig": 2, "boss": 3}
EFFECT = {
    "relic_hp_up": 1, "relic_melee_atk_up": 2, "relic_ranged_atk_up": 3,
    "relic_magic_atk_up": 4, "relic_def_up": 5, "relic_resist_up": 6,
    "relic_regen_up": 7, "relic_cure_up": 8, "relic_accuracy_up": 9,
    "relic_critical_up": 10, "relic_atk_spd_up": 11, "relic_move_spd_up": 12,
    "relic_lifesteal": 20, "relic_thorns": 21, "relic_kill_energy": 22,
    "relic_kill_heal": 23, "relic_low_hp_def_up": 24, "relic_revive_once": 25,
    "relic_erosion_slow": 30, "relic_vision_up": 31, "relic_dig_speed": 32,
}
GRADE_RGB = {"common": (184, 196, 207), "rare": (111, 195, 232), "epic": (216, 155, 255)}

ASSET_META = """fileFormatVersion: 2
guid: {guid}
NativeFormatImporter:
  externalObjects: {{}}
  mainObjectFileID: 11400000
  userData:
  assetBundleName:
  assetBundleVariant:
"""

PNG_META = """fileFormatVersion: 2
guid: {guid}
TextureImporter:
  internalIDToNameTable: []
  externalObjects: {{}}
  serializedVersion: 13
  mipmaps:
    mipMapMode: 0
    enableMipMap: 0
    sRGBTexture: 1
    linearTexture: 0
    fadeOut: 0
    borderMipMap: 0
    mipMapsPreserveCoverage: 0
    alphaTestReferenceValue: 0.5
    mipMapFadeDistanceStart: 1
    mipMapFadeDistanceEnd: 3
  bumpmap:
    convertToNormalMap: 0
    externalNormalMap: 0
    heightScale: 0.25
    normalMapFilter: 0
    flipGreenChannel: 0
  isReadable: 0
  streamingMipmaps: 0
  streamingMipmapsPriority: 0
  vTOnly: 0
  ignoreMipmapLimit: 0
  grayScaleToAlpha: 0
  generateCubemap: 6
  cubemapConvolution: 0
  seamlessCubemap: 0
  textureFormat: 1
  maxTextureSize: 2048
  textureSettings:
    serializedVersion: 2
    filterMode: 1
    aniso: 1
    mipBias: 0
    wrapU: 1
    wrapV: 1
    wrapW: 1
  nPOTScale: 0
  lightmap: 0
  compressionQuality: 50
  spriteMode: 1
  spriteExtrude: 1
  spriteMeshType: 1
  alignment: 0
  spritePivot: {{x: 0.5, y: 0.5}}
  spritePixelsToUnits: 100
  spriteBorder: {{x: 0, y: 0, z: 0, w: 0}}
  spriteGenerateFallbackPhysicsShape: 1
  alphaUsage: 1
  alphaIsTransparency: 1
  spriteTessellationDetail: -1
  textureType: 8
  textureShape: 1
  singleChannelComponent: 0
  flipbookRows: 1
  flipbookColumns: 1
  maxTextureSizeSet: 0
  compressionQualitySet: 0
  textureFormatSet: 0
  ignorePngGamma: 0
  applyGammaDecoding: 0
  swizzle: 50462976
  cookieLightType: 0
  platformSettings:
  - serializedVersion: 3
    buildTarget: DefaultTexturePlatform
    maxTextureSize: 2048
    resizeAlgorithm: 0
    textureFormat: -1
    textureCompression: 1
    compressionQuality: 50
    crunchedCompression: 0
    allowsAlphaSplitting: 0
    overridden: 0
    ignorePlatformSupport: 0
    androidETC2FallbackOverride: 0
    forceMaximumCompressionQuality_BC6H_BC7: 0
  spriteSheet:
    serializedVersion: 2
    sprites: []
    outline: []
    physicsShape: []
    bones: []
    spriteID: {sprite_id}
    internalID: 0
    vertices: []
    indices:
    edges: []
    weights: []
    secondaryTextures: []
  spritePackingTag:
  pSDRemoveMatte: 0
  pSDShowRemoveMatteOption: 0
  userData:
  assetBundleName:
  assetBundleVariant:
"""

HEADER = """%YAML 1.1
%TAG !u! tag:unity3d.com,2011:
--- !u!114 &11400000
MonoBehaviour:
  m_ObjectHideFlags: 0
  m_CorrespondingSourceObject: {{fileID: 0}}
  m_PrefabInstance: {{fileID: 0}}
  m_PrefabAsset: {{fileID: 0}}
  m_GameObject: {{fileID: 0}}
  m_Enabled: 1
  m_EditorHideFlags: 0
  m_Script: {{fileID: 11500000, guid: {script_guid}, type: 3}}
  m_Name: {name}
  m_EditorClassIdentifier:
"""

FOLDER_META = """fileFormatVersion: 2
guid: {guid}
folderAsset: yes
DefaultImporter:
  externalObjects: {{}}
  userData:
  assetBundleName:
  assetBundleVariant:
"""


def script_guid(rel_cs_path):
    """`.cs.meta` 에서 스크립트 guid 를 읽는다."""
    meta = os.path.join(PROJECT, "Assets", "_Project", "Scripts", rel_cs_path) + ".meta"
    if not os.path.isfile(meta):
        raise SystemExit(
            "⚠ %s 이(가) 없습니다.\n"
            "   유니티가 아직 새 스크립트를 임포트하지 않았습니다 — 에디터를 한 번 띄우거나\n"
            "   node Tools/mcp_unity_cli.js execute_menu_item '{\"menuPath\":\"Assets/Refresh\"}'"
            % meta)
    with io.open(meta, encoding="utf-8") as f:
        for line in f:
            if line.startswith("guid:"):
                return line.split(":", 1)[1].strip()
    raise SystemExit("guid 를 찾지 못했습니다: " + meta)


def guid_for(key):
    """경로에서 결정적으로 만든다 — 다시 돌려도 같은 guid 라 참조가 안 끊긴다."""
    return hashlib.md5(("LastSanctuary/" + key).encode("utf-8")).hexdigest()


def yaml_str(v):
    """⚠ 줄바꿈을 이스케이프한다 — 빈 줄이 들어가면 유니티가 뒤를 전부 버린다(맨 위 ⚠)."""
    s = "" if v is None else str(v)
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    s = s.replace("\r\n", "\\n").replace("\n", "\\n").replace("\r", "\\n")
    return '"%s"' % s


def num(v, default=0):
    if v is None:
        return default
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return default


def fnum(v, default=0.0):
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def rows_of(ws):
    """4행부터 «첫 칸이 비지 않은» 행만."""
    keys = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(FIRST_ROW, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None or str(first).startswith("—"):
            continue
        out.append({k: ws.cell(r, i + 1).value for i, k in enumerate(keys) if k})
    return out


def write(path, text):
    with io.open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write(text)


def ensure_folder_meta(path):
    meta = path + ".meta"
    if not os.path.isfile(meta):
        rel = os.path.relpath(path, PROJECT).replace("\\", "/")
        write(meta, FOLDER_META.format(guid=guid_for(rel)))


# ──────────────────────────────────────────────────────────────────────────
# 임시 아이콘
#
# ★ <b>유물마다 다른 문양</b>을 ID 로 정한다 — 같은 유물은 언제나 같은 그림이다.
#   면역 반응을 떠올리게 하는 도형 다섯을 돌려 쓴다:
#     0 세포(원 + 핵) · 1 항체(Y) · 2 결정(마름모) · 3 나선(고리 셋) · 4 가시(별)
# ──────────────────────────────────────────────────────────────────────────

def draw_icon(path, relic_id, grade, name_seed):
    s = ICON_SIZE
    img = Image.new("RGBA", (s, s), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)

    base = GRADE_RGB.get(grade, GRADE_RGB["common"])
    dark = tuple(max(0, int(c * 0.30)) for c in base)
    mid = tuple(int(c * 0.62) for c in base)

    # 등급 색 바탕 — 둥근 사각형 + 안쪽 테두리
    pad = 6
    d.rounded_rectangle([pad, pad, s - pad, s - pad], radius=18,
                        fill=dark + (235,), outline=base + (255,), width=4)
    d.rounded_rectangle([pad + 9, pad + 9, s - pad - 9, s - pad - 9], radius=12,
                        outline=mid + (170,), width=2)

    h = int(hashlib.md5(("%d/%s" % (relic_id, name_seed)).encode("utf-8")).hexdigest(), 16)
    shape = h % 5
    cx = cy = s // 2
    ink = base + (255,)
    glow = base + (110,)

    if shape == 0:                     # 세포 — 원 + 핵
        d.ellipse([cx - 34, cy - 34, cx + 34, cy + 34], outline=ink, width=6)
        d.ellipse([cx - 13, cy - 13, cx + 13, cy + 13], fill=ink)
        for k in range(6):
            a = h % 60 + k * 60
            x = cx + int(26 * math.cos(math.radians(a)))
            y = cy + int(26 * math.sin(math.radians(a)))
            d.ellipse([x - 4, y - 4, x + 4, y + 4], fill=glow)
    elif shape == 1:                   # 항체 — Y
        d.line([cx, cy + 34, cx, cy - 2], fill=ink, width=9)
        d.line([cx, cy - 2, cx - 26, cy - 32], fill=ink, width=9)
        d.line([cx, cy - 2, cx + 26, cy - 32], fill=ink, width=9)
        d.ellipse([cx - 30, cy - 40, cx - 18, cy - 28], fill=glow)
        d.ellipse([cx + 18, cy - 40, cx + 30, cy - 28], fill=glow)
    elif shape == 2:                   # 결정 — 마름모
        d.polygon([(cx, cy - 36), (cx + 26, cy), (cx, cy + 36), (cx - 26, cy)],
                  outline=ink, width=6)
        d.polygon([(cx, cy - 16), (cx + 12, cy), (cx, cy + 16), (cx - 12, cy)], fill=ink)
    elif shape == 3:                   # 나선 — 고리 셋
        for k, r in enumerate((34, 24, 14)):
            off = (k - 1) * 5
            d.ellipse([cx - r + off, cy - r, cx + r + off, cy + r],
                      outline=ink if k % 2 == 0 else glow, width=5)
    else:                              # 가시 — 별
        for k in range(8):
            a = math.radians(h % 45 + k * 45)
            d.line([cx, cy, cx + int(38 * math.cos(a)), cy + int(38 * math.sin(a))],
                   fill=ink if k % 2 == 0 else glow, width=6)
        d.ellipse([cx - 10, cy - 10, cx + 10, cy + 10], fill=ink)

    img.save(path)


def main():
    if not os.path.isfile(XLSX):
        raise SystemExit("⚠ 표가 없습니다: %s\n   먼저 py -3 Tools/gen_relic_table.py" % XLSX)

    relic_guid = script_guid("Relics/RelicDefinitionSO.cs")
    table_guid = script_guid("Relics/RelicDigTableSO.cs")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    relics = rows_of(wb["Relic"])
    outcomes = rows_of(wb["DigOutcome"])
    drops = rows_of(wb["Drop"])

    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(ICON_DIR, exist_ok=True)
    ensure_folder_meta(OUT_DIR)
    ensure_folder_meta(ICON_DIR)

    # ── 아이콘 먼저 — 에셋이 guid 로 가리켜야 하므로 ────────────────────
    icon_guid = {}
    for r in relics:
        key = str(r.get("icon") or ("relic_%s" % r["relic_id"]))
        if key in icon_guid:
            continue
        png = os.path.join(ICON_DIR, key + ".png")
        draw_icon(png, num(r["relic_id"]), str(r.get("grade") or "common"), key)
        rel = os.path.relpath(png, PROJECT).replace("\\", "/")
        g = guid_for(rel)
        icon_guid[key] = g
        write(png + ".meta", PNG_META.format(guid=g, sprite_id=guid_for(rel + "#sprite")))

    # ── 유물 에셋 ──────────────────────────────────────────────────────
    unknown_effect = []
    made = 0
    for r in relics:
        rid = num(r["relic_id"])
        if rid <= 0:
            continue
        name = "Relic_%d" % rid
        grade = str(r.get("grade") or "common").strip()
        eff = str(r.get("effect_type") or "").strip()
        if eff not in EFFECT:
            unknown_effect.append((rid, eff))
        key = str(r.get("icon") or "")

        body = HEADER.format(script_guid=relic_guid, name=name)
        body += "  relicId: %d\n" % rid
        body += "  relicName: %s\n" % yaml_str(r.get("relic_name"))
        body += "  grade: %d\n" % GRADE.get(grade, 0)
        body += "  effectType: %d\n" % EFFECT.get(eff, 0)
        body += "  value01: %d\n" % num(r.get("value_01"))
        body += "  value02: %d\n" % num(r.get("value_02"))
        body += "  relicDesc: %s\n" % yaml_str(r.get("relic_desc"))
        body += "  relicFlavor: %s\n" % yaml_str(r.get("relic_flavor"))
        body += "  source: %d\n" % SOURCE.get(str(r.get("source") or "").strip(), 0)
        body += "  sourceId: %d\n" % num(r.get("source_id"))
        body += "  dropWeight: %d\n" % max(0, num(r.get("drop_weight"), 10))
        body += "  iconKey: %s\n" % yaml_str(key)
        if key in icon_guid:
            # ⚠ 스프라이트는 텍스처 안의 <b>서브 에셋</b>이다 — fileID 21300000 이 규약이다.
            body += "  icon: {fileID: 21300000, guid: %s, type: 3}\n" % icon_guid[key]
        else:
            body += "  icon: {fileID: 0}\n"

        path = os.path.join(OUT_DIR, name + ".asset")
        write(path, body)
        rel = os.path.relpath(path, PROJECT).replace("\\", "/")
        write(path + ".meta", ASSET_META.format(guid=guid_for(rel)))
        made += 1

    # ── 발굴·드랍 표 에셋 ──────────────────────────────────────────────
    body = HEADER.format(script_guid=table_guid, name="RelicDigTable")
    body += "  outcomes:\n"
    for o in outcomes:
        body += "  - outcomeType: %s\n" % yaml_str(o.get("outcome_type"))
        body += "    weight: %d\n" % max(0, num(o.get("weight")))
        body += "    value01: %d\n" % num(o.get("value_01"))
        body += "    value02: %d\n" % num(o.get("value_02"))
        body += "    outcomeDesc: %s\n" % yaml_str(o.get("outcome_desc"))
        body += "    outcomeScript: %s\n" % yaml_str(o.get("outcome_script"))
    body += "  drops:\n"
    for dr in drops:
        body += "  - killSource: %s\n" % yaml_str(dr.get("kill_source"))
        body += "    grade: %d\n" % GRADE.get(
            {"일반": "common", "레어": "rare", "에픽": "epic"}.get(
                str(dr.get("grade") or "").strip(), str(dr.get("grade") or "").strip()), 0)
        body += "    percent: %s\n" % ("%g" % fnum(dr.get("percent")))

    path = os.path.join(OUT_DIR, "RelicDigTable.asset")
    write(path, body)
    rel = os.path.relpath(path, PROJECT).replace("\\", "/")
    write(path + ".meta", ASSET_META.format(guid=guid_for(rel)))

    # ── 보고 ───────────────────────────────────────────────────────────
    print("[유물 에셋]")
    print("  유물 %d개  →  %s" % (made, os.path.relpath(OUT_DIR, PROJECT)))
    print("  임시 아이콘 %d장  →  %s" % (len(icon_guid), os.path.relpath(ICON_DIR, PROJECT)))
    print("  발굴 결과 %d줄 · 드랍 %d줄  →  RelicDigTable.asset" % (len(outcomes), len(drops)))
    if unknown_effect:
        print("  ⚠ 코드에 없는 효과 타입 %d개 — 장착해도 아무 일이 없습니다:" % len(unknown_effect))
        for rid, eff in unknown_effect:
            print("      %d  %s" % (rid, eff))
    print("  다음: 유니티에서 Assets/Refresh")


if __name__ == "__main__":
    main()
