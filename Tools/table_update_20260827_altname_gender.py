# -*- coding: utf-8 -*-
"""대체 이름 표에 <b>성별</b> 칸을 만들고 이름 40개를 남/여로 가른다 (2026-08-27).

유저 지시
---------
  *"이후에 2번째 생성부터 캐릭터 이름 랜덤으로 들어가는 거 남녀 구분해서
    남캐는 남자 이름 여캐는 여자이름으로 들어가는 시스템으로 만들어줘"*

181-5절이 만든 «다른 이름 주머니» 는 <b>성별을 안 가렸다</b> — 남성 인물이
「오필리아」로, 여성 인물이 「테오도르」로 다시 태어날 수 있었다.
이 표에 성별 칸을 만들어 <b>주머니를 둘로</b> 나눈다.

    AltName 시트  4열에 끼워 넣는다(비고는 5열로 밀린다)
      gender   string   male / female

★ <b>왜 «비고» 앞에 끼우나</b> — 비고는 늘 맨 뒤라는 것이 이 프로젝트 표들의 결이다.
  데이터 칸이 비고 뒤로 가면 다음 사람이 «비고 다음에도 데이터가 있나?» 를 매번 확인해야 한다.
  ⚠ 그래서 <b>위치로 읽는 코드가 있으면 깨진다</b> — `gen_alt_name_table.py` 를 같은 판에
    고쳤다(그 스크립트도 4열을 성별로 읽는다). 다른 곳에서는 이 표를 읽지 않는다.

★★ <b>이름을 어떻게 갈랐나</b> — 라틴/게르만 어형의 통상적인 성별을 따랐다.
  -a / -ia / -ette / -ine 로 끝나는 것은 여성(라비니아 · 오데트 · 사비네),
  -on / -an / -or / -el 로 끝나는 것은 남성(에리온 · 파비안 · 테오도르 · 가리엘) 쪽이다.
  ⚠ 어느 쪽으로도 읽히는 이름이 몇 있다(아셀 · 메이런 · 페릴 · 솔레인) — 아래 표에
    그렇게 적어 뒀다. <b>바꾸고 싶으면 이 파일이 아니라 표를 고치면 된다</b>
    (`gen_alt_name_table.py` 는 표를 정본으로 읽는다).

남 19 · 여 21 — 인물은 남 5 · 여 9 이므로 어느 쪽도 먼저 동나지 않는다.

⚠ <b>Excel COM 으로 쓴다</b>(136-4절 · 이 PC 에 LibreOffice 가 없다).

다음:  python Tools/gen_alt_name_table.py    (성별표를 Unity 로 내보낸다)
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "대체 이름 테이블.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")
SHEET = "AltName"
DATA_ROW0 = 4
GENDER_COL = 4          # 비고 앞

HEADER = ("성별", "gender", "string")

# alt_name_id → (gender, 근거)   ★ 이름의 어형을 따랐다
ALT_GENDER = {
    1:  ("male",   "아드리엘 / Adriel — -iel 계 천사명, 남성형"),
    2:  ("female", "노아네 / Noane"),
    3:  ("female", "레이린 / Reilin"),
    4:  ("female", "미르카 / Mirka — 슬라브계 여성형 -ka"),
    5:  ("female", "세이하 / Seiha"),
    6:  ("male",   "오르넬 / Ornel"),
    7:  ("male",   "유리안 / Yurian — Julian 계"),
    8:  ("female", "이레아 / Irea — -ea 여성형"),
    9:  ("male",   "카일런 / Kailen"),
    10: ("female", "타비아 / Tavia — Octavia 계"),
    11: ("male",   "페릴 / Peril — ⚠ 중성적. 남성으로 뒀다"),
    12: ("male",   "하르윈 / Harwin — -win 게르만 남성형"),
    13: ("male",   "가리엘 / Gariel — -iel 계"),
    14: ("male",   "나비스 / Navis"),
    15: ("male",   "데인 / Dane"),
    16: ("female", "라비니아 / Lavinia"),
    17: ("male",   "마르첼 / Marcel"),
    18: ("male",   "바이런 / Byron"),
    19: ("female", "사비네 / Sabine"),
    20: ("female", "아셀 / Asel — ⚠ 중성적. 여성으로 뒀다"),
    21: ("male",   "에리온 / Erion"),
    22: ("female", "오필리아 / Ophelia"),
    23: ("female", "율리아 / Julia"),
    24: ("male",   "제피르 / Zephyr"),
    25: ("female", "카린 / Karin"),
    26: ("male",   "테오도르 / Theodor"),
    27: ("male",   "파비안 / Fabian"),
    28: ("female", "할리아 / Halia"),
    29: ("female", "그웬 / Gwen"),
    30: ("female", "니콜라 / Nicola — 여성형으로 읽었다"),
    31: ("male",   "도리안 / Dorian"),
    32: ("female", "류시아 / Lucia"),
    33: ("female", "메이런 / Meiren — ⚠ 중성적. 여성으로 뒀다"),
    34: ("female", "베르타 / Berta"),
    35: ("female", "솔레인 / Solein — ⚠ 중성적. 여성으로 뒀다"),
    36: ("male",   "아르덴 / Arden"),
    37: ("male",   "엘로이 / Eloi"),
    38: ("female", "오데트 / Odette"),
    39: ("female", "이자벨 / Isabel"),
    40: ("male",   "카시엘 / Casiel — -iel 계"),
}

VALID = {"male", "female"}


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_대체이름성별")
    os.makedirs(folder, exist_ok=True)
    shutil.copy2(XLSX, os.path.join(folder, os.path.basename(XLSX)))
    print("백업: " + folder)


def read_layout():
    """(필드명 → 열, 마지막 열, id → 행)."""
    ws = openpyxl.load_workbook(XLSX, data_only=True)[SHEET]
    names = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None and str(v).strip():
            names[str(v).strip()] = c
    rows = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            continue
        rows[int(v)] = r
    return names, ws.max_column, rows


def main():
    bad = {i: g for i, (g, _) in ALT_GENDER.items() if g not in VALID}
    if bad:
        sys.exit("✗ male/female 가 아닌 값이 있습니다: %s" % bad)

    names, last, rows = read_layout()
    print(f"「{SHEET}」 지금 칸 {last}개 · 이름 {len(rows)}개")

    missing = sorted(i for i in rows if i not in ALT_GENDER)
    if missing:
        sys.exit("✗ 성별을 안 적은 이름이 있습니다: id %s\n"
                 "  이 파일의 ALT_GENDER 에 줄을 더하고 다시 돌리세요." % missing)

    backup()

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(XLSX))
        ws = wb.Worksheets(SHEET)

        if "gender" in names:
            col = names["gender"]
            print(f"  칸이 이미 있습니다 — {col}열")
        else:
            ws.Columns(GENDER_COL).Insert()          # 비고가 5열로 밀린다
            for i, v in enumerate(HEADER, start=1):
                ws.Cells(i, GENDER_COL).Value = v
            ws.Columns(GENDER_COL).ColumnWidth = 10
            col = GENDER_COL
            print(f"  칸 신설 — {col}열  {HEADER[0]} / {HEADER[1]} / {HEADER[2]}"
                  "  (비고는 뒤로 밀림)")

        counts = {"male": 0, "female": 0}
        for alt_id, r in sorted(rows.items()):
            g, why = ALT_GENDER[alt_id]
            ws.Cells(r, col).Value = g
            counts[g] += 1
            print(f"  {alt_id:>2}: {g:<6} — {why}")

        wb.Save()
        wb.Close()
    finally:
        excel.Quit()

    print(f"\n대체 이름 테이블.xlsx — 남 {counts['male']} · 여 {counts['female']}"
          f" (합 {sum(counts.values())})")
    print("다음: python Tools/gen_alt_name_table.py")


if __name__ == "__main__":
    main()
