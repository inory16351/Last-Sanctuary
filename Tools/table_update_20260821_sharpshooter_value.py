# -*- coding: utf-8 -*-
"""★★ 「명사수」(80038)의 <b>밸류 구멍</b>을 메운다 (2026-08-21).

125-4절이 «적어만 뒀다» 고 남긴 그 구멍이다:

    Skill_Type.desc  "영구적으로 세라피엘이 <b>20</b>의 크리티컬 확률을 획득합니다…"
    Skill 80038      value_01 ~ value_06 이 <b>전부 0</b>

이 프로젝트의 규약은 «<b>수치는 밸류 칸, 문장은 자리표시</b>» 다(다른 41개 스킬이 전부
그렇다). 문장에 숫자를 박아 두면 코드가 밸류를 읽으므로 <b>그대로 구현하면 +0</b> 이 걸린다 —
파일만 봐서는 멀쩡해 보이는 종류의 사고다(119-1절의 «스킬 배정이 한 칸씩 밀려 있던» 것과 같다).

무엇을 하나
-----------
① `캐릭터 테이블.xlsx` / `Skill` 시트 80038 행의 <b>밸류타입_01 = 20</b>
② `스트링 키 테이블.xlsx` 의 `skill_type_desc_Sharpshooter` 에서 «20의» → «{value_01}의»

★ ①은 <b>Excel COM</b> 으로 쓴다 — 이 표에는 하이퍼링크가 154칸 있고(51-11절) openpyxl 로
  저장하면 전부 날아간다(64-2·69-10·125-3절과 같은 이유).
★ ②는 <b>openpyxl</b> 로 쓴다 — 스트링 키 테이블은 `gen_string_table.py` 가 매번 새 통합문서로
  통째로 다시 쓰는 파일이라 지켜야 할 서식이 없다.

⚠ <b>멱등하다</b> — 이미 20 이면/이미 자리표시면 건드리지 않는다.
⚠ 정의문의 괄호 문장(«120의 크리티컬 확률이면 100% 치명타»)은 <b>손대지 않는다</b> —
  그건 이 스킬의 수치가 아니라 <b>전역 규칙</b>을 설명하는 문장이고, 그 규칙을 구현할지는
  별개 사안이다(아래 ⚠⚠).

⚠⚠ <b>남는 것</b> — «크리티컬 확률 120 이면 원거리 공격이 100% 치명타» 라는 문장은
  <b>아직 구현되지 않았다</b>. 지금 코드의 치명타는 확률 굴림이고 «120 이상이면 확정» 이라는
  갈래가 없다. 그건 <b>이 스킬만의 규칙이 아니라 전투 공식</b>이라 밸런스 담당이 정할 일이다.

실행:  python Tools/table_update_20260821_sharpshooter_value.py
다음:  python Tools/gen_string_table.py && python Tools/gen_character_assets.py
"""

import datetime
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

CHAR_XLSX = "캐릭터 테이블.xlsx"
STRING_XLSX = "스트링 키 테이블.xlsx"
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")
FIRST_DATA_ROW = 4

SKILL_ID = 80038
CRITICAL = 20
DESC_KEY = "skill_type_desc_Sharpshooter"
DESC_FROM = "20의 크리티컬 확률"
DESC_TO = "{value_01}의 크리티컬 확률"


def check_locks(files):
    locked = [f for f in files if os.path.isfile(os.path.join(TABLE_DIR, "~$" + f))]
    if locked:
        raise SystemExit("[!] 엑셀에서 열려 있습니다 — 닫고 다시 실행하세요:\n   "
                         + "\n   ".join(locked))


def backup(files, tag):
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_ROOT, stamp + "_" + tag)
    os.makedirs(dst, exist_ok=True)
    for f in files:
        src = os.path.join(TABLE_DIR, f)
        if os.path.isfile(src):
            shutil.copy2(src, os.path.join(dst, f))
    print("백업:", dst)


def find_col_com(ws, field, max_col=32):
    for c in range(1, max_col + 1):
        v = ws.Cells(2, c).Value
        if v is not None and str(v).strip() == field:
            return c
    return 0


def fix_value(excel):
    """① 캐릭터 테이블 Skill 시트의 밸류타입_01 = 20 (Excel COM · 하이퍼링크 보존)."""
    path = os.path.join(TABLE_DIR, CHAR_XLSX)
    wb = excel.Workbooks.Open(path)
    changed = 0
    try:
        ws = wb.Worksheets("Skill")
        last = ws.UsedRange.Rows.Count
        c_id = find_col_com(ws, "skill_id")
        c_v1 = find_col_com(ws, "value_01")
        if not (c_id and c_v1):
            raise SystemExit("[!] Skill 시트에서 skill_id / value_01 컬럼을 못 찾았습니다.")

        row = 0
        for r in range(FIRST_DATA_ROW, last + 1):
            v = ws.Cells(r, c_id).Value
            if v is None:
                continue
            try:
                same = int(float(v)) == SKILL_ID
            except (TypeError, ValueError):
                same = False
            if same:
                row = r
                break
        if not row:
            raise SystemExit("[!] Skill 시트에 %d 행이 없습니다." % SKILL_ID)

        cur = ws.Cells(row, c_v1).Value
        cur = int(float(cur)) if cur is not None else 0
        if cur == CRITICAL:
            print("  · value_01 이 이미 %d 입니다 — 건드리지 않습니다." % CRITICAL)
        elif cur != 0:
            print("  · value_01 이 이미 %d 입니다(0 이 아님) — 건드리지 않습니다." % cur)
        else:
            ws.Cells(row, c_v1).Value = CRITICAL
            print("  Skill %d행 value_01 <- %d" % (row, CRITICAL))
            changed += 1

        wb.Save()
    finally:
        wb.Close()
    return changed


def fix_desc():
    """② 스트링 키 테이블의 정의문에서 숫자를 자리표시로 (openpyxl)."""
    path = os.path.join(TABLE_DIR, STRING_XLSX)
    wb = openpyxl.load_workbook(path)
    ws = wb["string"]

    c_key = c_kr = 0
    for c in range(1, 12):
        v = ws.cell(row=2, column=c).value
        if v is None:
            continue
        if str(v).strip() == "string_key":
            c_key = c
        elif str(v).strip() == "kr":
            c_kr = c
    if not (c_key and c_kr):
        raise SystemExit("[!] string 시트에서 컬럼을 못 찾았습니다.")

    changed = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        key = ws.cell(row=r, column=c_key).value
        if key is None or str(key).strip() != DESC_KEY:
            continue
        kr = ws.cell(row=r, column=c_kr).value or ""
        if DESC_TO in kr:
            print("  · 정의문이 이미 자리표시입니다 — 건드리지 않습니다.")
        elif DESC_FROM not in kr:
            print("  [!] 정의문에서 \"%s\" 를 못 찾았습니다 — 문장이 바뀌었는지 확인하세요." % DESC_FROM)
        else:
            ws.cell(row=r, column=c_kr).value = kr.replace(DESC_FROM, DESC_TO, 1)
            print("  %s kr: \"%s\" -> \"%s\"" % (DESC_KEY, DESC_FROM, DESC_TO))
            changed += 1
        break
    else:
        raise SystemExit("[!] %s 키가 없습니다 — gen_string_table.py 를 먼저 돌리세요." % DESC_KEY)

    if changed:
        wb.save(path)
    return changed


def main():
    print("[명사수(80038) 밸류 구멍 메우기]")
    check_locks([CHAR_XLSX, STRING_XLSX])
    backup([CHAR_XLSX, STRING_XLSX], "명사수_밸류")

    import win32com.client as win32
    # ★ DispatchEx — 유저가 열어둔 엑셀과 섞이지 않게 새 프로세스를 띄운다(125-3절).
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        n1 = fix_value(excel)
    finally:
        excel.Quit()

    n2 = fix_desc()
    print("-> 표 %d칸 · 정의문 %d칸" % (n1, n2))
    if n1 or n2:
        print("   다음: python Tools/gen_string_table.py "
              "&& python Tools/gen_character_assets.py")


if __name__ == "__main__":
    main()
