# -*- coding: utf-8 -*-
"""스트링 키 테이블의 <b>영어(en) 칸</b>을 채운다 — 신규 3인 (2026-08-21).

유저 지시: *"볼트 폴더의 테이블이랑 에셋 이미지, 일러스트 이미지 확인해서 엘리시아랑
Seraphiel, Cyan 이렇게 총 3종 캐릭터 인게임에 구현해줘"*

★★ <b>왜 이 스크립트가 반드시 필요한가</b>
==========================================
`gen_character_assets.py` 는 <b>영어 이름으로 에셋 파일 이름을 만든다</b>
(`Character_9012_Elysia.asset`) 그리고 그 이름의 <b>경로 해시로 guid 를 고정</b>한다.
그래서 en 칸이 비면 <b>죽는다</b>(조용히 넘어가지 않는다):

    스트링 키 테이블에 character_name_9012 의 영어(en) 이름이 없습니다.
    에셋 파일 이름에 쓰이므로 비면 guid 가 바뀌어 참조가 끊깁니다

★ <b>수집기는 en 을 만들 수 없다</b> — `gen_string_table.py` 는 표의 <b>한국어 리터럴</b>을
  긁어 모으는 도구고, 영어는 <b>사람이 짓는 것</b>이다(51-1절 규칙). 그래서 손으로 적는다.

★ <b>openpyxl 로 쓴다</b>(Excel COM 이 아니다) — 이 파일은 `gen_string_table.py` 가
  <b>매번 새 통합문서로 통째로 다시 쓰는</b> 파일이라(그쪽 `wb.save(OUT_XLSX)`) 지켜야 할
  셀 주석·하이퍼링크가 애초에 없다. ⚠ <b>원본 표(캐릭터 테이블)는 반대다</b> — 그쪽은
  하이퍼링크가 있어 `convert_tables_to_string_keys.py` 가 Excel COM 을 쓴다.

⚠ <b>이미 값이 있으면 건드리지 않는다</b> — 사람이 다듬은 번역을 덮으면 안 된다
  (`gen_string_table.py` 의 merge 규칙과 같은 이유).

실행 순서
---------
    python Tools/gen_string_table.py                    # ① 한국어를 수집한다
    python Tools/table_update_20260821_new_characters_en.py   # ② 영어를 채운다  ← 이 파일
    python Tools/convert_tables_to_string_keys.py       # ③ 표의 리터럴을 키로 바꾼다
    python Tools/gen_string_table.py                    # ④ StringTable.txt 재생성
    python Tools/gen_character_assets.py                # ⑤ 캐릭터·패시브 에셋 생성
"""

import os
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "스트링 키 테이블.xlsx")
FIRST_DATA_ROW = 4

#: 키 → 영어 문구.
#:
#: ★ <b>이름</b>은 표의 «인게임 에셋» 칸이 정본이다 — 표가 `Char_Asset_Elysia` ·
#:   `Char_Asset_Seraphiel` · `Char_Asset_Cyan` 이라고 적어 두었으므로 영어 표기가
#:   이미 정해져 있다. 지어낸 것이 아니다.
#: ★ <b>칭호</b>는 한국어를 그대로 옮겼다(기존 열셋과 같은 결 — 「눈먼 파수꾼」
#:   The Blind Watchman · 「빛의 궁수」 The Light Archer …).
#: ★ <b>스킬 이름</b>은 표의 `skill_type`(enum)이 이미 영어다 — 그것을 문장 표기로
#:   되돌린 것이라 새로 짓는 부분이 없다(`Strong_mind` → "Strong Mind").
EN = {
    # ── 인물 ────────────────────────────────────────────────────────────
    "character_name_9012": "Elysia",
    "character_name_9013": "Seraphiel",
    "character_name_9014": "Cyan",
    "character_title_9012": "The Unbreakable Shield",
    "character_title_9013": "The Silent Gunshot",
    "character_title_9014": "The Reaper of Dawn",

    # ── 엘리시아 9012 ───────────────────────────────────────────────────
    "skill_name_80034": "Strong Mind",
    "skill_name_80035": "The Legion's Shield",
    "skill_name_80036": "Blessing of Four Wings",

    # ── 세라피엘 9013 ───────────────────────────────────────────────────
    "skill_name_80037": "Evasive Maneuver",
    "skill_name_80038": "Sharpshooter",
    "skill_name_80039": "Declaration of the End",

    # ── 시안 9014 ───────────────────────────────────────────────────────
    "skill_name_80040": "Soul Absorption",
    "skill_name_80041": "The Reaper's Scythe",
    "skill_name_80042": "Breaking Through Limits",
}


def find_col(ws, field, max_col=12):
    """2행(필드명)에서 컬럼 번호를 찾는다. 없으면 0."""
    for c in range(1, max_col + 1):
        v = ws.cell(row=2, column=c).value
        if v is not None and str(v).strip() == field:
            return c
    return 0


def main():
    if not os.path.isfile(XLSX):
        raise SystemExit("[!] 스트링 키 테이블이 없습니다: " + XLSX)
    if os.path.isfile(os.path.join(TABLE_DIR, "~$스트링 키 테이블.xlsx")):
        raise SystemExit("[!] 엑셀에서 열려 있습니다 — 닫고 다시 실행하세요.")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["string"]
    c_key = find_col(ws, "string_key")
    c_kr = find_col(ws, "kr")
    c_en = find_col(ws, "en")
    if not (c_key and c_kr and c_en):
        raise SystemExit("[!] string 시트에서 컬럼(string_key/kr/en)을 못 찾았습니다.")

    print("[스트링 키 테이블 — 영어 채우기]")
    changed = 0
    found = set()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        key = ws.cell(row=r, column=c_key).value
        key = str(key).strip() if key is not None else ""
        if key not in EN:
            continue
        found.add(key)
        cur = ws.cell(row=r, column=c_en).value
        cur = str(cur).strip() if cur is not None else ""
        want = EN[key]
        if cur == want:
            print("  · %-24s 이미 같음" % key)
            continue
        if cur:
            print("  · %-24s 이미 \"%s\" — 건드리지 않습니다" % (key, cur))
            continue
        ws.cell(row=r, column=c_en).value = want
        print("  %-24s en <- %s" % (key, want))
        changed += 1

    missing = sorted(set(EN) - found)
    if missing:
        print("  [!] 표에 아직 없는 키 %d개 — `gen_string_table.py` 를 먼저 돌리세요:" % len(missing))
        for k in missing:
            print("      ", k)
        raise SystemExit(1)

    if changed:
        wb.save(XLSX)
        print("  -> %d칸 저장" % changed)
    else:
        print("  -> 바뀐 것이 없습니다(멱등).")


if __name__ == "__main__":
    main()
