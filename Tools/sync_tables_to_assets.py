# -*- coding: utf-8 -*-
"""데이터 테이블(xlsx) → Unity ScriptableObject 에셋 동기화.

`gen_character_assets.py` 가 캐릭터·패시브 스킬을 맡는다면, 이 스크립트는 **나머지 전부**를 맡는다:
정신 이상 11종 · 웨이브 몬스터(잡몹·중간보스·최종보스) · 포탑/중앙건물 · 웨이브 구성표.

진행상황 54절(다른 PC 의 테이블 재밸런싱)이 "게임 미반영" 으로 남긴 항목들을 실제로 반영하는
파이프라인이다. 54-10절이 손으로 옮겨야 한다고 적어둔 것을 스크립트로 만든 것 -
표가 정본이고 이 스크립트가 그대로 옮긴다. **값을 손으로 옮겨 적지 않는다.**

⚠ 왜 스크립트인가 (MCP 가 아니라) - MCP 에는 **ScriptableObject 에셋(.asset)을 다루는 도구가
  없다.** `update_component` 는 씬의 GameObject 컴포넌트만 만진다. 그래서 씬 오브젝트는 MCP 로,
  에셋은 이 스크립트로 갈라 놓는다(진행상황 5절·8절부터 이 프로젝트가 쓰는 방식이고,
  33-5절의 `gen_character_assets.py` 도 같은 이유로 스크립트다).

⚠ .asset YAML 에 **빈 줄을 넣으면 Unity 파서가 그 뒤 필드를 전부 무시한다**(8절 3번).
  이 스크립트는 기존 파일의 필드 값만 **한 줄씩 치환**하고 구조는 건드리지 않는다 -
  전체를 다시 쓰면 내가 모르는 필드(나중에 추가된 것)를 날릴 수 있기 때문이다.
"""
import os
import re
import sys

# 콘솔이 cp949 라 한글 출력에서 죽는다 - 출력만 UTF-8 로 바꾼다(파일 내용과 무관).
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass
import hashlib
import openpyxl

from vault_path import TABLE_DIR as VAULT   # ★ PC 마다 다른 볼트 위치를 찾아준다(2026-08-15)
_PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS = os.path.join(_PROJECT, 'Assets', '_Project')

XLSX_MENTAL = os.path.join(VAULT, '정신 이상 테이블.xlsx')
XLSX_WAVE_MON = os.path.join(VAULT, '웨이브 몬스터 테이블.xlsx')
XLSX_WAVE = os.path.join(VAULT, '웨이브테이블.xlsx')
XLSX_BUILDING = os.path.join(VAULT, 'Last_Sanctuary_건물데이터시트_Ver05.xlsx')
XLSX_NEUTRAL = os.path.join(VAULT, '임시용 중립 몬스터.xlsx')

SCRIPT_GUID_MONSTER = '5dbe527860d1cbe42a3efae9fd5cb4b2'   # MonsterDefinitionSO.cs


def script_guid(rel_cs_path):
    """`.cs.meta` 에서 스크립트 guid 를 읽는다.

    ⚠ 위 `SCRIPT_GUID_MONSTER` 처럼 상수로 박아두면 스크립트를 옮기거나 다시 만들었을 때
      조용히 어긋난다(에셋이 `Missing Script` 가 된다). 새로 쓰는 코드는 이 함수를 쓴다.
    """
    meta = os.path.join(ASSETS, 'Scripts', rel_cs_path) + '.meta'
    with open(meta, encoding='utf-8') as f:
        for line in f:
            if line.startswith('guid:'):
                return line.split(':', 1)[1].strip()
    raise RuntimeError('guid 를 찾지 못했습니다: ' + meta)

ASSET_META = """fileFormatVersion: 2
guid: {guid}
NativeFormatImporter:
  externalObjects: {{}}
  mainObjectFileID: 11400000
  userData:
  assetBundleName:
  assetBundleVariant:
"""

HEADER = """%YAML 1.1
%TAG !u! tag:unity3d.com,2011:
--- !u!114 &11400000
MonoBehaviour:
  m_ObjectHideFlags: 0
  m_CorrespondingSourceObject: {{fileID: 0}}
  m_PrefabInstance: {{fileID: 0}}
  m_PrefabAsset: {{fileID: 0}}
  m_GameObject: {{fileID: 0}}
  m_Enabled: 1
  m_EditorHideFlags: 0
  m_Script: {{fileID: 11500000, guid: {script_guid}, type: 3}}
  m_Name: {name}
  m_EditorClassIdentifier:
"""


def guid_for(key):
    """gen_character_assets.py 와 같은 규칙 - 경로에서 결정적으로 만든다(다시 돌려도 같은 guid)."""
    return hashlib.md5(('LastSanctuary/' + key).encode('utf-8')).hexdigest()


def num(v, default=0):
    if v is None:
        return default
    try:
        f = float(v)
    except (TypeError, ValueError):
        return default
    return int(f) if f == int(f) else f


def read_sheet(path, sheet, first_row=4):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = []
    for r in range(first_row, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if vals and vals[0] is not None:
            rows.append(vals)
    return rows


def read_rows(path, sheet, first_row=4):
    """`{필드명: 값}` 사전 목록으로 읽는다 — <b>컬럼 위치에 의존하지 않는다.</b>

    ⚠ 왜 필요했나: 2026-08-13 에 표에서 영어 이름 컬럼(`character_name_EG` 등)을 지우자
      그 뒤 컬럼이 전부 한 칸씩 밀렸고, `row[8]` 처럼 위치로 읽던 코드가 조용히 엉뚱한
      값을 읽었다. 65-2절이 "컬럼은 맨 뒤에만 붙인다"고 정한 것과 같은 뿌리의 문제인데,
      <b>삭제는 그 규약으로도 못 막는다.</b> 컬럼이 바뀔 여지가 있는 시트는 이 함수를 쓸 것.

    필드명(2행)의 앞뒤 공백은 없앤다 — 이 표에는 ' resistance' 처럼 앞 공백이 붙은 헤더가
    실제로 있다.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]

    names = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None and str(v).strip():
            names[c] = str(v).strip()

    rows = []
    for r in range(first_row, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        rows.append({name: ws.cell(r, c).value for c, name in names.items()})
    return rows


def patch_fields(path, changes, label, add_missing=()):
    """에셋 파일의 `  key: value` 줄만 치환한다. 없는 키는 조용히 건너뛴다(보고만 한다).

    `add_missing` 에 넣은 키는 <b>없으면 파일 끝에 새로 만든다</b>. C# 에 필드를 새로
    추가한 첫 실행에는 에셋 YAML 에 그 줄이 아예 없어서(이 파일들은 사람/스크립트가
    쓴 것이라 Unity 가 기본값을 채워 넣어준 적이 없다) 그냥 두면 영원히 반영되지 않는다.

    ⚠ .asset YAML 에 <b>빈 줄을 넣으면 Unity 가 그 뒤 필드를 전부 무시한다</b>(8절 3번) -
      덧붙이기 전에 꼬리 빈 줄을 지운다(write_int_list 와 같은 규칙).
    """
    if not os.path.exists(path):
        print('  ! 없는 파일:', path)
        return 0
    with open(path, encoding='utf-8') as f:
        text = f.read()

    hit = 0
    appended = []
    for key, value in changes.items():
        pattern = re.compile(r'^(\s*%s:).*$' % re.escape(key), re.M)
        if not pattern.search(text):
            if key in add_missing:
                text = text.rstrip('\n') + '\n  %s: %s\n' % (key, value)
                appended.append(key)
                hit += 1
            else:
                print('  ! %s: 필드 %r 가 없습니다 (건너뜀)' % (label, key))
            continue
        text, n = pattern.subn(lambda m: '%s %s' % (m.group(1), value), text, count=1)
        hit += n

    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(text)
    print('  %s: %d개 필드 갱신%s'
          % (label, hit, ' (신규 %s)' % ', '.join(appended) if appended else ''))
    return hit


# ---------------------------------------------------------------------------
# 1) 정신 이상 11종 - 54-6절 (after_erosion 45~55 로 낮춤 · 가중치 재배분)
# ---------------------------------------------------------------------------
def sync_mental_errors():
    print('[정신 이상]')
    folder = os.path.join(ASSETS, 'Resources', 'MentalErrors')
    by_id = {}
    for name in os.listdir(folder):
        if not name.endswith('.asset'):
            continue
        with open(os.path.join(folder, name), encoding='utf-8') as f:
            m = re.search(r'mentalErrorId:\s*(\d+)', f.read())
        if m:
            by_id[int(m.group(1))] = os.path.join(folder, name)

    total = 0
    weight_sum = 0.0
    for row in read_sheet(XLSX_MENTAL, 'mental_error'):
        mid = int(row[0])
        if mid not in by_id:
            print('  ! id %d 에 해당하는 에셋이 없습니다' % mid)
            continue
        weight_sum += float(num(row[6]))
        total += patch_fields(by_id[mid], {
            'value01': num(row[2]),
            'value02': num(row[3]),
            'durationSeconds': num(row[4]),
            'afterErosion': int(num(row[5])),
            'activationProbability': num(row[6]),
        }, 'MentalError %d' % mid)

    # 29-3절: 이 값은 "판정 확률"이 아니라 추첨 가중치이고 합이 1.00 이어야 한다.
    print('  가중치 합 = %.4f %s' % (weight_sum, '(OK)' if abs(weight_sum - 1.0) < 1e-6 else '(⚠ 1.00 아님!)'))
    return total


# ---------------------------------------------------------------------------
# 2) 웨이브 몬스터 - 잡몹 2종 + 최종보스 갱신, 중간보스 2종 신규 생성
# ---------------------------------------------------------------------------
# 표의 monster_id → 에셋 파일명. 파일이 없으면 <b>이 스크립트가 만든다</b>(create_monster_asset).
#
# ★ 2026-08-18 — 중간보스 2종(110001·110002)이 <b>빠졌다</b>. 유저 지시로 중간보스를 없애고
#   보스를 「5웨이브마다 단탈리온/말파스 교대」로 바꿨다. 말파스(120002)가 새로 들어온다.
MONSTER_ASSET_BY_ID = {
    100001: 'Monster_HellFang',
    100002: 'Monster_SoulArcher',
    120001: 'Monster_Dantalian',
    120002: 'Monster_Malphas',
    120003: 'Monster_Kasinoma',
    120004: 'Monster_Laryngeal',   # ★ 라린길 신규 (2026-08-19) — 표만 있고 원화(스킨)는 아직 없다
}

# 삭제된 중간보스 에셋 — 남아 있으면 스포너 슬롯에 다시 끌려 들어갈 수 있어서 지운다.
REMOVED_MONSTER_ASSETS = (
    'Monster_MidBoss_BloodMark',
    'Monster_MidBoss_VoidWhisper',
)

# 중간보스 - 54-4절. tier=MidBoss(1).
#
# ⚠ **전투 파라미터를 지어내지 않는다** - 같은 공격 타입의 잡몹 에셋에서 그대로 물려받는다
#   (`inherit` 이 그 에셋 이름이다). 유저 지시 4번이 "중간 보스 인게임 모션은 임시로 일반
#   몬스터 스킨을 그대로 쓰는 것이니 신경 쓰지 말고 냅둬" 였고, 표에도 사거리·인식범위 칸이
#   없다. 표에 있는 것(능력치·체력보정)만 표에서 가져오고 나머지는 물려받는 것이 정확하다.
#
# ⚠ **외형(template)은 이 에셋에 넣을 수 없다** - ScriptableObject 는 씬 오브젝트를 참조할 수
#   없다(진행상황 5절). 스포너의 슬롯이 템플릿을 지정하는 구조이므로 씬에서 MCP 로 연결한다.
# ⚠ 에셋 이름은 **표의 `character_name_EG`** 를 따른다(2026-08-13). 예전에는 물려받는
#   잡몹 이름(`Monster_MidBoss_HellFang`)을 그대로 썼는데, 그러면 하이라키·에셋 목록에서
#   중간보스가 잡몹으로 보인다. 표에 영어 이름 컬럼이 아예 없어서 그랬던 것이라
#   컬럼을 만들고(BloodMark · VoidWhisper) 그 값으로 개명했다.
MID_BOSS = {}

# 새로 만드는 몬스터의 <b>표에 없는</b> 값. {monster_id: {필드: 값}}
#
# ⚠ **표에 있는 것은 여기 적지 않는다** — 체력·공격·방어·콜라이더 등은 전부 `first_Stat`
#   에서 온다. 여기 있는 것은 표에 <b>칸 자체가 없는</b> 항목뿐이다(사거리·인식범위·발판).
#
# ⚠ **외형(template)은 이 에셋에 넣을 수 없다** — ScriptableObject 는 씬 오브젝트를 참조할
#   수 없다(진행상황 5절). 스포너 슬롯이 템플릿을 지정하므로 씬에서 MCP 로 연결한다.
NEW_MONSTER_DEFAULTS = {
    120002: dict(
        # 원거리 보스라 인식·사거리를 단탈리온(근접 2.0)보다 길게 잡는다.
        # 표의 Skill 시트에서 저주광선이 10타일까지 뻗으므로 평타 사거리를 그 안쪽에 둔다.
        detectRange=12,
        attackRange=7.0,
        footprintTiles=1,
        # 스킨이 없을 때만 쓰는 폴백(콜라이더 상자가 있으면 안 쓰인다).
        bodyWidthTiles=2,
        bodyHeightTiles=3,
        spriteScale=0.75,
    ),
    120003: dict(
        # ★ 근접 보스 — 단탈리온과 같은 사거리(2.0)를 쓴다. 다만 <b>인식은 더 넓다</b>:
        #    「이끌리는 혈취」가 지름 20타일 안의 적에게 돌진하는 기술이라, 인식이 10 이면
        #    보스가 <b>자기 스킬 사거리의 절반도 못 보고</b> 서 있게 된다. 스킬의 반지름
        #    (10타일)에 맞춰 12 로 둔다.
        detectRange=12,
        attackRange=2.0,
        footprintTiles=1,
        bodyWidthTiles=2,
        bodyHeightTiles=3,
        spriteScale=0.75,
    ),
}

# 크기 검산용 - 이 몬스터가 실제로 쓰는 스킨(원화). renderHeightTiles(표) 로 스케일을
# 잡았을 때 나오는 가로가 표의 render_width_tiles 와 맞는지 확인하는 데만 쓴다(아래 참조).
# 중간보스는 전용 원화가 없어 잡몹 스킨을 그대로 쓴다(63절) - 그래서 같은 경로다.
SKIN_FOR_MONSTER = {
    100001: 'MonsterSkins/HellFang/Skin_HellFang',
    100002: 'MonsterSkins/SoulArcher/Skin_SoulArcher',
    120001: 'MonsterSkins/Dantalian/Skin_Dantalian',
    120002: 'MonsterSkins/Malphas/Skin_Malphas',
    120003: 'MonsterSkins/Kasinoma/Skin_Kasinoma',
}


def skin_content_tiles(skin_rel_path):
    """스킨 에셋의 `contentSizeTiles`(measure_skin_tiles.py 가 알파 경계로 실측해 적어둔 값)."""
    path = os.path.join(ASSETS, 'Resources', skin_rel_path + '.asset')
    if not os.path.exists(path):
        return None
    with open(path, encoding='utf-8') as f:
        m = re.search(r'contentSizeTiles:\s*\{x:\s*([0-9.]+),\s*y:\s*([0-9.]+)\}', f.read())
    return (float(m.group(1)), float(m.group(2))) if m else None


def report_collider_fit_skin(name, skin_rel, box_w, box_h):
    """
    <see cref="report_collider_fit"/> 와 같은 계산인데, 스킨 경로를 <b>직접</b> 받는다.

    중립 몬스터는 표의 `mon_skin` 이 스킨을 정하므로 `SKIN_FOR_MONSTER` 같은 코드 상수가
    필요 없다 — 표가 정본이라는 이 파일의 원칙에 오히려 더 맞는다.
    """
    art = skin_content_tiles(skin_rel)
    if art is None or art[0] <= 0 or art[1] <= 0 or box_w <= 0 or box_h <= 0:
        return
    s = min(box_w / art[0], box_h / art[1])
    print('    %s 콜라이더 표 %.1f x %.1f -> 실제 %.2f x %.2f (원화 %.2f x %.2f 타일)'
          % (name, box_w, box_h, art[0] * s, art[1] * s, art[0], art[1]))


def report_collider_fit(mid, name, box_w, box_h):
    """
    표의 콜라이더 상자에 그림을 맞추면 실제로 어떤 크기가 나오는지 보여준다(계산만, 기록 안 함).

    게임에서 벌어지는 일과 **같은 계산**이다(`CharacterAnimator.ResolveScale`):
      배율 = min(상자가로 / 원화가로, 상자세로 / 원화세로)   ← 상자 안에 들어가는 최대(contain)
    그 배율로 그린 크기가 곧 **재설정된 콜라이더**다. 표 값과 한 축은 같고 다른 축은 조금 작다.

    ⚠ 이 값을 에셋에 적어두지 않는다 - 원화(스킨)를 바꾸면 결과도 같이 바뀌어야 하므로
      **런타임에 계산**하는 것이 정본이다. 여기서는 유저가 표를 채울 때 참고하도록 찍어만 준다.
    """
    art = skin_content_tiles(SKIN_FOR_MONSTER.get(mid, ''))
    if art is None or art[0] <= 0 or art[1] <= 0 or box_w <= 0 or box_h <= 0:
        return
    s = min(box_w / art[0], box_h / art[1])
    print('    %s 콜라이더 표 %.1f x %.1f -> 실제 %.2f x %.2f (원화 비율 %.2f:1)'
          % (name, box_w, box_h, art[0] * s, art[1] * s, art[0] / art[1]))

# 능력치 → 공속/이속 치환은 게임이 인스펙터 값을 그대로 쓰므로(몬스터는 StatMoveSpeedTiles 0)
# 표의 공속·이속 스탯을 38-1절 공식으로 미리 풀어서 넣는다.
def aspd_from_stat(s):
    return round(0.6 + 3.0 * s / (s + 50.0), 3)


def mspd_from_stat(s):
    return round(2.1 + 3.9 * s / (s + 50.0), 3)


def boss_title_ids():
    """칭호가 실제로 적혀 있는 몬스터 id 집합.

    체력바에 칭호를 띄우려면 정의 에셋에 `titleKey` 가 있어야 한다(2026-08-13). 키는
    id 로 조립하므로(`boss_title_<id>`) <b>칸에 뭐가 들어 있든</b> 상관없다 - 한국어
    문구여도(중간보스) 이미 키로 바뀐 값이어도(최종보스) 똑같이 동작한다. 여기서는
    "그 칸이 비어 있지 않은가" 만 본다.

    ⚠ 칸이 비어 있으면 키를 안 넣는다 - 없는 키를 넣어두면 StringTable 조회가 매번
      실패해 조용히 빈 문자열이 되고, 인스펙터만 보면 "칭호가 있는데 왜 안 뜨지"가 된다.
    """
    ids = set()
    # ★ 2026-08-18 — `wave_mid_boss` 시트가 사라졌다(중간보스 삭제). 최종보스 시트만 본다.
    for sheet in ('wave_top_boss',):
        # ⚠ 위치가 아니라 <b>필드명</b>으로 읽는다 - 2026-08-13 에 영어 이름 컬럼을 지우면서
        #   뒤 컬럼이 한 칸씩 밀렸다(read_rows 주석 참조).
        for row in read_rows(XLSX_WAVE_MON, sheet):
            if row.get('boss_title') not in (None, ''):
                ids.add(int(num(row.get('monster_id'))))
    return ids


# ── 공격 계열 + 명중·치명 (2026-08-15 신설) ─────────────────────────────
#
# ★ 예전에는 `attackStat` 한 칸에 `max(melee_atk, ranged_atk)` 를 넣어
#   <b>표의 두 칸을 하나로 접고</b>, `accuracy`·`critical` 은 <b>통째로 버렸다</b>
#   (담을 필드가 C# 쪽에 없었다). 그래서 표에 `critical: 8` 이라고 적힌 최종보스가
#   실제로는 치명타를 한 번도 내지 않았다.
#
#   이제 `MonsterDefinitionSO` 에 칸이 생겼으므로 표의 값을 <b>그대로</b> 옮긴다.
#   어느 칸을 쓸지는 런타임의 `UnitCombat.AttackType` 이 고른다
#   (`MonsterUnit.AttackStatType` — 캐릭터와 같은 규칙).
#
# ⚠ 명중·치명은 <b>원거리 공격 유형에만</b> 적용된다(유저 확정 2026-08-15).
#   근거리·마법 몬스터에 값이 들어가 있어도 판정에 쓰이지 않는다 — 표를 그대로
#   옮기되, 그 사실을 알고 볼 것.
NEW_STAT_FIELDS = ('rangedAttackStat', 'magicStat', 'cureStat',
                   'accuracyStat', 'criticalStat', 'resistanceStat')


def attack_stat_fields(r):
    """웨이브 `first_Stat` 한 줄 → 공격 계열·명중·치명·저항 필드 묶음.

    컬럼 위치(0부터): 2 melee_atk · 3 ranged_atk · 4 accuracy · 5 critical
                      6 magic · 7 cure · 12 resistance
    """
    return {
        'attackStat': int(num(r[2])),
        'rangedAttackStat': int(num(r[3])),
        'accuracyStat': int(num(r[4])),
        'criticalStat': int(num(r[5])),
        'magicStat': int(num(r[6])) if len(r) > 6 else 0,
        'cureStat': int(num(r[7])) if len(r) > 7 else 0,
        'resistanceStat': int(num(r[12], 50)) if len(r) > 12 else 50,
    }


def top_boss_rows():
    """`wave_top_boss` 시트를 id 로 색인. 공격 타입·이름·칭호가 여기 있다."""
    return {int(num(row.get('monster_id'))): row
            for row in read_rows(XLSX_WAVE_MON, 'wave_top_boss')}


def create_monster_asset(path, mid, asset, r, top):
    """
    <b>에셋 파일이 아예 없을 때</b> 표 값으로 새로 만든다 (2026-08-18, 말파스).

    예전에는 중간보스 전용 생성 코드가 따로 있었는데(잡몹 에셋에서 값을 물려받는 구조),
    중간보스가 없어지면서 그 코드를 이 함수가 대체한다. <b>물려받지 않는다</b> —
    표에 있는 값은 표에서, 표에 <b>칸이 없는</b> 값만 `NEW_MONSTER_DEFAULTS` 에서 가져온다.

    만들기만 한다 — 그 다음 `patch_fields` 가 매번 표를 다시 덮으므로 여기서 적는 값은
    <b>첫 생성 순간의 뼈대</b>일 뿐이다.
    """
    d = NEW_MONSTER_DEFAULTS.get(mid, {})
    atk = attack_stat_fields(r)

    body = HEADER.format(script_guid=SCRIPT_GUID_MONSTER, name=asset)
    body += "  monsterId: %d\n" % mid
    body += "  nameKey: monster_name_%d\n" % mid
    body += "  displayName: %s\n" % asset
    body += "  titleKey: %s\n" % (('boss_title_%d' % mid) if top.get('boss_title') else '')
    body += "  tier: 2\n"                       # MonsterTier.MainBoss
    body += "  template: {fileID: 0}\n"         # 스포너 슬롯이 지정한다 (위 ⚠ 참조)
    body += "  hpStat: %d\n" % int(num(r[1]))
    body += "  attackStat: %d\n" % atk['attackStat']
    body += "  defenseStat: %d\n" % int(num(r[10]))
    body += "  regenStat: %d\n" % int(num(r[11]))
    for _f in NEW_STAT_FIELDS:
        body += "  %s: %d\n" % (_f, atk[_f])
    body += "  hpPercent: %d\n" % int(num(r[13], 100))
    body += "  attackType: %d\n" % attack_type_value(top.get('atk_type'))
    body += "  detectRange: %s\n" % d.get('detectRange', 10)
    body += "  attackRange: %s\n" % d.get('attackRange', 2.0)
    body += "  attacksPerSecond: %s\n" % aspd_from_stat(num(r[8]))
    body += "  moveSpeedTiles: %s\n" % mspd_from_stat(num(r[9]))
    body += "  footprintTiles: %s\n" % d.get('footprintTiles', 1)
    body += "  bodyWidthTiles: %s\n" % d.get('bodyWidthTiles', 2)
    body += "  bodyHeightTiles: %s\n" % d.get('bodyHeightTiles', 3)
    body += "  spriteScale: %s\n" % d.get('spriteScale', 0.75)
    body += "  renderHeightTiles: 0\n"
    body += "  colliderWidthTiles: %s\n" % (num(r[15]) if len(r) > 15 else 0)
    body += "  colliderHeightTiles: %s\n" % (num(r[14]) if len(r) > 14 else 0)
    body += "  bossSkillIds: []\n"              # sync_boss_skills 가 표에서 채운다

    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(body)
    mp = path + '.meta'
    if not os.path.exists(mp):
        with open(mp, 'w', encoding='utf-8', newline='\n') as f:
            f.write(ASSET_META.format(guid=guid_for('Data/Units/%s.asset' % asset)))
    print('  %s 신규 생성 (id %d)' % (asset, mid))


def remove_deleted_assets(folder):
    """중간보스처럼 <b>없어진</b> 정의 에셋을 지운다. .meta 도 같이 지워야 한다."""
    for asset in REMOVED_MONSTER_ASSETS:
        for suffix in ('.asset', '.asset.meta'):
            p = os.path.join(folder, asset + suffix)
            if os.path.exists(p):
                os.remove(p)
                print('  삭제 %s%s' % (asset, suffix))


def sync_monsters():
    print('[웨이브 몬스터]')
    stats = {int(r[0]): r for r in read_sheet(XLSX_WAVE_MON, 'first_Stat')}
    tops = top_boss_rows()
    titled = boss_title_ids()
    folder = os.path.join(ASSETS, 'Data', 'Units')
    total = 0

    remove_deleted_assets(folder)

    for mid, asset in MONSTER_ASSET_BY_ID.items():
        r = stats.get(mid)
        if r is None:
            print('  ! 표에 id %d 가 없습니다' % mid)
            continue

        path = os.path.join(folder, asset + '.asset')
        if not os.path.exists(path):
            create_monster_asset(path, mid, asset, r, tops.get(mid, {}))

        box_h = num(r[14]) if len(r) > 14 else 0     # collider_height_tiles
        box_w = num(r[15]) if len(r) > 15 else 0     # collider_width_tiles
        changes = {
            # 표의 id 를 에셋에 적어둔다 (2026-08-18) — 웨이브 표의 boss_monster_id 로
            # 스포너가 정의를 찾을 수 있어야 한다(MonsterDefinitionSO.monsterId).
            'monsterId': mid,
            'hpStat': int(num(r[1])),
            'defenseStat': int(num(r[10])),
            'regenStat': int(num(r[11])),
            'hpPercent': int(num(r[13], 100)),
            'attacksPerSecond': aspd_from_stat(num(r[8])),
            # ★ 이동속도도 표에서 온다 (2026-08-13). 예전에는 <b>중간보스만</b> 표를 따르고
            #   잡몹·최종보스는 에셋에 손으로 적힌 값을 그대로 뒀다 - 그래서 최종보스에
            #   공식과 무관한 1.4 가 박혀 있었고(잡몹 2.2), 보스가 자기 호위대보다 느렸다.
            #   유저 지시 "보스 이동 속도 수정(증가) -> 너무 느림" 을 표에서 고칠 수 있게
            #   여기로 끌어온다. 공식은 38-1절의 mspd_from_stat 하나뿐이다.
            'moveSpeedTiles': mspd_from_stat(num(r[9])),
            # 콜라이더 상자 - 표가 정본이다(2026-08-13). 그림을 이 상자 안에 비율 유지로
            # 맞추고 콜라이더를 다시 그 그림 크기로 맞추는 것은 CharacterAnimator 가 한다.
            # 계산 결과를 에셋에 적어두지 않는 이유: 원화를 바꾸면 결과도 바뀌어야 한다.
            'colliderWidthTiles': box_w,
            'colliderHeightTiles': box_h,
            # 세로 전용 폴백은 비운다 - 콜라이더 상자가 있으면 안 쓰이는데 값이 남아 있으면
            # 인스펙터에서 어느 쪽이 적용되는지 헷갈린다. 필드 자체는 지우지 않는다(U-D3).
            'renderHeightTiles': 0,
        }
        changes.update(attack_stat_fields(r))
        # 칭호 - 표에 칭호가 적힌 몬스터에만 키를 넣는다(위 boss_title_ids 주석).
        if mid in titled:
            changes['titleKey'] = 'boss_title_%d' % mid

        # ★ 일러스트 (2026-08-18) - `wave_top_boss.illust` 그대로. 잡몹은 그 시트에 행이
        #   없어 빈 문자열이 되고, 그러면 클릭해도 초상화 창이 안 뜬다(중립과 같은 규칙).
        changes['illustName'] = str(tops.get(mid, {}).get('illust') or '').strip()

        total += patch_fields(os.path.join(folder, asset + '.asset'), changes, asset,
                              add_missing=('titleKey', 'monsterId', 'illustName') + NEW_STAT_FIELDS)
        report_collider_fit(mid, asset, box_w, box_h)

    # --- 중간보스 2종 신규 (2026-08-18 이후 MID_BOSS 는 비어 있어 돌지 않는다) ---
    for mid, spec in MID_BOSS.items():
        r = stats.get(mid)
        if r is None:
            print('  ! 표에 중간보스 id %d 가 없습니다' % mid)
            continue
        rel = 'Data/Units/%s.asset' % spec['asset']
        path = os.path.join(folder, spec['asset'] + '.asset')

        # 전투 파라미터는 같은 타입의 잡몹에서 물려받는다 (표에 없는 값을 지어내지 않는다).
        with open(os.path.join(folder, spec['inherit'] + '.asset'), encoding='utf-8') as f:
            base = f.read()

        def inherited(key, default):
            m = re.search(r'^\s*%s:\s*(\S+)\s*$' % re.escape(key), base, re.M)
            return m.group(1) if m else default

        body = HEADER.format(script_guid=SCRIPT_GUID_MONSTER, name=spec['asset'])
        body += "  nameKey: monster_name_%d\n" % mid
        body += "  displayName: %s\n" % spec['name']
        # 칭호(2026-08-13) - 표 wave_mid_boss 의 boss_title 칸이 채워진 중간보스만.
        # 체력바가 이 키로 스트링 테이블을 조회한다(BossHealthPanel.NameLine).
        if mid in titled:
            body += "  titleKey: boss_title_%d\n" % mid
        body += "  tier: 1\n"                       # MonsterTier.MidBoss
        body += "  template: {fileID: 0}\n"         # 스포너 슬롯이 지정한다 (위 ⚠ 참조)
        body += "  hpStat: %d\n" % int(num(r[1]))
        # 공격 계열 4칸 + 명중·치명·저항 (2026-08-15) — 잡몹과 <b>같은 헬퍼</b>를 쓴다.
        # 여기만 따로 적으면 표에 컬럼이 늘 때 한쪽만 반영되어 갈라진다.
        atk = attack_stat_fields(r)
        body += "  attackStat: %d\n" % atk['attackStat']
        body += "  defenseStat: %d\n" % int(num(r[10]))
        body += "  regenStat: %d\n" % int(num(r[11]))
        for _f in NEW_STAT_FIELDS:
            body += "  %s: %d\n" % (_f, atk[_f])
        body += "  hpPercent: %d\n" % int(num(r[13], 100))
        body += "  attackType: %s\n" % inherited('attackType', '0')
        body += "  detectRange: %s\n" % inherited('detectRange', '7')
        body += "  attackRange: %s\n" % inherited('attackRange', '1.2')
        body += "  attacksPerSecond: %s\n" % aspd_from_stat(num(r[8]))
        body += "  moveSpeedTiles: %s\n" % mspd_from_stat(num(r[9]))
        body += "  footprintTiles: %s\n" % inherited('footprintTiles', '1')
        # 콜라이더 상자 - <b>표의 collider_width/height_tiles 칸에서 온다</b>(2026-08-13, 65절).
        # ⚠ 한때 이 값을 이 스크립트 안 MID_BOSS 딕셔너리에 리터럴로 박아뒀었다 -
        #   "값을 손으로 옮겨 적지 않는다"는 이 파일 맨 위 원칙을 정확히 어긴 것이었다.
        #   표에 컬럼이 생겨서 다른 몬스터와 똑같이 읽어온다.
        # bodyWidth/Height·spriteScale 은 스킨이 없을 때만 쓰는 폴백이라 그대로 둔다.
        box_h = num(r[14], 3) if len(r) > 14 else 3
        box_w = num(r[15], 0) if len(r) > 15 else 0
        body += "  bodyWidthTiles: 2\n"
        body += "  bodyHeightTiles: 2\n"
        body += "  spriteScale: 2\n"
        body += "  colliderWidthTiles: %s\n" % box_w
        body += "  colliderHeightTiles: %s\n" % box_h

        with open(path, 'w', encoding='utf-8', newline='\n') as f:
            f.write(body)
        mp = path + '.meta'
        if not os.path.exists(mp):
            with open(mp, 'w', encoding='utf-8', newline='\n') as f:
                f.write(ASSET_META.format(guid=guid_for(rel)))
        report_collider_fit(mid, spec['asset'], box_w, box_h)
        print('  %s 생성 (id %d, HP스탯 %s, 보정 %s%%)'
              % (spec['asset'], mid, int(num(r[1])), int(num(r[13], 100))))
        total += 1
    return total


# ---------------------------------------------------------------------------
# 2-b) 보스 스킬 - Skill 시트 → BossSkillSO 에셋 + 보스 정의의 bossSkillIds (2026-08-13)
#
# 66-7절이 "표·문구·아트는 준비돼 있는데 발동시키는 코드가 없다"로 남겨둔 미결 111번을
# 실제로 잇는 부분이다. **수치를 코드에 적지 않는다** - 표의 Skill 시트가 정본이다.
#
# ⚠ 스킬 에셋은 **전체를 다시 쓴다**(다른 에셋처럼 줄 치환이 아니라). 표에 줄이 늘면
#   에셋 개수 자체가 늘어야 하고, 이 에셋에는 사람이 손으로 넣는 값이 하나도 없어서
#   덮어써도 잃을 것이 없다. guid 는 경로에서 결정적으로 만들므로 다시 돌려도 그대로다.
# ---------------------------------------------------------------------------
def sync_boss_skills():
    print('[보스 스킬]')
    folder = os.path.join(ASSETS, 'Resources', 'BossSkills')
    os.makedirs(folder, exist_ok=True)
    guid = script_guid(os.path.join('Combat', 'BossSkillSO.cs'))

    made = []
    # ★★ 2026-08-18 - <b>위치가 아니라 필드명으로 읽는다</b>(read_rows).
    #
    #   여기가 <b>조용히 3칸씩 밀려 있었다.</b> 2026-08-13 에 이 코드를 쓸 때 Skill 시트는
    #   9칸(… value_03 · cool_time · skill_icon · skill_explain)이었고, 침식·시전시간·
    #   범위모양을 <b>맨 뒤에</b> 붙이는 규약으로 row[9]/row[10]/row[11] 을 읽게 했다.
    #   그 뒤 표가 <b>value_04·05·06 을 value_03 바로 뒤</b>로 넣고 침식을
    #   `mentalerror_damage` 라는 자기 컬럼으로 옮기면서 위치가 전부 어긋났다:
    #
    #       coolTime   ← value_04       → 단탈리온 두 스킬 모두 <b>0</b>
    #       value04    ← cool_time
    #       rangeType  ← skill_explain  (문자열이 그대로 들어가 있었다)
    #
    #   `coolTime == 0` 이면 `BossSkillSO.IsUsable` 이 false 라 <b>스킬이 한 번도
    #   발동하지 않는다.</b> 즉 단탈리온은 광역기를 쓰도록 배선돼 있었을 뿐 실제로는
    #   평타만 때리고 있었다. read_rows 의 주석이 경고하던 바로 그 사고다.
    for row in read_rows(XLSX_WAVE_MON, 'Skill'):
        raw_id = row.get('skill_id')
        if raw_id in (None, ''):
            continue
        sid = int(num(raw_id))
        name = 'BossSkill_%d' % sid
        rel = 'Resources/BossSkills/%s.asset' % name

        body = HEADER.format(script_guid=guid, name=name)
        body += '  skillId: %d\n' % sid
        body += '  nameKey: %s\n' % (row.get('skill_name') or '')
        body += "  displayName: ''\n"          # 문구는 스트링 테이블이 정본이다
        body += '  skillType: %s\n' % (row.get('skill_type') or '')
        body += '  explainKey: %s\n' % (row.get('skill_explain') or '')
        for i in range(1, 7):
            body += '  value%02d: %s\n' % (i, num(row.get('value_%02d' % i)))
        # 침식은 이제 <b>자기 컬럼</b>이다 — value_04 를 침식으로 읽지 않는다.
        body += '  erosionValue: %s\n' % num(row.get('mentalerror_damage'))
        body += '  coolTime: %s\n' % num(row.get('cool_time'))
        #   cast_time  : 이 스킬의 연출 길이(초). 0 이면 BossSkillCaster 의 전역 기본값.
        #   range_type : 범위 모양. 비어 있으면 Line(직사각형, 조준 방향 자유각).
        body += '  castSeconds: %s\n' % num(row.get('cast_time'))
        body += '  rangeType: %s\n' % (row.get('range_type') or '')
        # 「구속」 표시 이름 (2026-08-19) — 비어 있으면 UnitCombat 의 기본값("구속")을 쓴다.
        body += '  statusNameKey: %s\n' % (row.get('status_name') or '')

        path = os.path.join(folder, name + '.asset')
        with open(path, 'w', encoding='utf-8', newline='\n') as f:
            f.write(body)
        mp = path + '.meta'
        if not os.path.exists(mp):
            with open(mp, 'w', encoding='utf-8', newline='\n') as f:
                f.write(ASSET_META.format(guid=guid_for(rel)))

        made.append(sid)
        print('  %s: %s · v1~v6 %s/%s/%s/%s/%s/%s · 침식 +%s · 쿨 %s초 · 시전 %s초 · %s'
              % (name, row.get('skill_type'),
                 num(row.get('value_01')), num(row.get('value_02')), num(row.get('value_03')),
                 num(row.get('value_04')), num(row.get('value_05')), num(row.get('value_06')),
                 num(row.get('mentalerror_damage')), num(row.get('cool_time')),
                 num(row.get('cast_time')), row.get('range_type') or 'Line'))

    # 최종보스 시트의 boss_skill_1~3 을 그 보스의 정의 에셋으로 옮긴다.
    # 표에 없는 몬스터(잡몹·중간보스)는 스킬 칸 자체가 없으므로 건드리지 않는다.
    # ⚠ 필드명으로 읽는다 - 영어 이름 컬럼 삭제(2026-08-13)로 위치가 밀렸다.
    for row in read_rows(XLSX_WAVE_MON, 'wave_top_boss'):
        mid = int(num(row.get('monster_id')))
        asset = MONSTER_ASSET_BY_ID.get(mid)
        if asset is None:
            print('  ! 최종보스 id %d 에 해당하는 에셋 이름을 모릅니다' % mid)
            continue

        ids = [int(num(row.get(k))) for k in ('boss_skill_1', 'boss_skill_2', 'boss_skill_3')
               if int(num(row.get(k))) > 0]
        missing = [i for i in ids if i not in made]
        if missing:
            print('  ! %s 의 스킬 %s 가 Skill 시트에 없습니다' % (asset, missing))

        write_int_list(os.path.join(ASSETS, 'Data', 'Units', asset + '.asset'),
                       'bossSkillIds', ids, asset)
    return len(made)


# ---------------------------------------------------------------------------
# 2-c) 중립 몬스터 스킬 - 중립 표의 Skill 시트 → BossSkillSO 에셋 + 정의의 skillIds
#      (2026-08-15, 카르시노스)
#
# ★ 웨이브 쪽과 <b>같은 폴더</b>(Resources/BossSkills)에 쓴다. id 대역이 겹치지 않고
#   (웨이브 130001~ · 중립 2001~), 게임 쪽 로더(BossSkillCaster)가 폴더 하나만 읽으면
#   되기 때문이다 — 폴더를 나누면 캐스터에 "어느 폴더를 볼지" 분기가 생긴다.
#
# ⚠⚠ <b>위치가 아니라 필드명으로 읽는다.</b> 두 표의 Skill 시트는 <b>컬럼 순서가 다르다</b>:
#     웨이브: … value_03 · cool_time · skill_icon · skill_explain · value_04 · cast_time
#     중립  : … value_03 · value_04 · value_05 · cool_time · skill_icon · skill_explain · cast_time
#   위 sync_boss_skills() 처럼 인덱스로 읽으면 <b>쿨타임 자리에서 value_04 를 읽는다</b>.
#   같은 함수를 재사용하지 않고 따로 둔 이유가 이것이다.
# ---------------------------------------------------------------------------
def sync_neutral_skills():
    print('[중립 몬스터 스킬]')
    if not os.path.exists(XLSX_NEUTRAL):
        print('  ! 표가 없습니다:', XLSX_NEUTRAL)
        return 0

    sheets = openpyxl.load_workbook(XLSX_NEUTRAL, read_only=True).sheetnames
    if 'Skill' not in sheets:
        print('  ! Skill 시트가 없습니다')
        return 0

    folder = os.path.join(ASSETS, 'Resources', 'BossSkills')
    os.makedirs(folder, exist_ok=True)
    guid = script_guid(os.path.join('Combat', 'BossSkillSO.cs'))

    made = []
    for row in read_rows(XLSX_NEUTRAL, 'Skill'):
        raw_id = row.get('skill_id')
        if raw_id in (None, ''):
            continue
        sid = int(num(raw_id))
        name = 'BossSkill_%d' % sid
        rel = 'Resources/BossSkills/%s.asset' % name

        body = HEADER.format(script_guid=guid, name=name)
        body += '  skillId: %d\n' % sid
        body += '  nameKey: %s\n' % (row.get('skill_name') or '')
        body += "  displayName: ''\n"
        body += '  skillType: %s\n' % (row.get('skill_type') or '')
        body += '  explainKey: %s\n' % (row.get('skill_explain') or '')
        for i in range(1, 6):
            body += '  value%02d: %s\n' % (i, num(row.get('value_%02d' % i)))
        # ★ 2026-08-19 — 중립 Skill 시트에도 `mentalerror_damage` 칸이 생겼다.
        #   그전까지 이 값을 안 써서 중립 에픽의 침식은 <b>표에 없는 코드 기본값 0</b> 이었다.
        #   설계상 0 이 맞지만(29절: 중립 사냥은 침식을 올리지 않는다) 이제 <b>표가 그 0 을
        #   말한다</b> — 올리려면 코드가 아니라 표를 고친다.
        #   ⚠ 칸이 없는 옛 표로도 돌아야 하므로 기본값 0 을 준다.
        body += '  erosionValue: %s' % num(row.get('mentalerror_damage'), 0) + chr(10)
        body += '  coolTime: %s\n' % num(row.get('cool_time'))
        body += '  castSeconds: %s\n' % num(row.get('cast_time'))
        body += '  rangeType: %s\n' % (row.get('range_type') or '')
        # 「구속」 표시 이름 (2026-08-19) — 아니사킬 「거대한 위협 포효」는 "기절"을 쓴다.
        body += '  statusNameKey: %s\n' % (row.get('status_name') or '')

        path = os.path.join(folder, name + '.asset')
        with open(path, 'w', encoding='utf-8', newline='\n') as f:
            f.write(body)
        mp = path + '.meta'
        if not os.path.exists(mp):
            with open(mp, 'w', encoding='utf-8', newline='\n') as f:
                f.write(ASSET_META.format(guid=guid_for(rel)))

        made.append(sid)
        print('  %s: %s · v1~v5 %s/%s/%s/%s/%s · 쿨 %s초 · 시전 %s초 · %s'
              % (name, row.get('skill_type'),
                 num(row.get('value_01')), num(row.get('value_02')), num(row.get('value_03')),
                 num(row.get('value_04')), num(row.get('value_05')),
                 num(row.get('cool_time')), num(row.get('cast_time')),
                 row.get('range_type') or 'Line'))

    # neutrality_mon 의 mon_skill_1·2 → 정의 에셋의 skillIds
    folder_def = os.path.join(ASSETS, 'Resources', 'NeutralMonsters')
    for row in read_rows(XLSX_NEUTRAL, 'neutrality_mon'):
        mid = int(num(row.get('mon_id')))
        asset = NEUTRAL_ASSET_BY_ID.get(mid)
        if asset is None:
            continue

        ids = [int(num(row.get(k))) for k in ('mon_skill_1', 'mon_skill_2')
               if int(num(row.get(k))) > 0]
        if not ids:
            continue                    # 스킬 없는 종은 건드리지 않는다

        missing = [i for i in ids if i not in made]
        if missing:
            print('  ! %s 의 스킬 %s 가 Skill 시트에 없습니다' % (asset, missing))

        write_int_list(os.path.join(folder_def, asset + '.asset'), 'skillIds', ids, asset)
    return len(made)


def write_int_list(path, key, values, label):
    """`key:` 아래의 정수 배열을 통째로 갈아끼운다. 필드가 없으면 파일 끝에 덧붙인다.

    ⚠ .asset YAML 에 빈 줄을 넣으면 Unity 가 그 뒤 필드를 전부 무시한다(8절 3번) -
      덧붙이기 전에 꼬리 빈 줄을 지운다.
    """
    if not os.path.exists(path):
        print('  ! 없는 파일:', path)
        return
    with open(path, encoding='utf-8') as f:
        lines = f.readlines()

    block = ['  %s: []\n' % key] if not values else \
            ['  %s:\n' % key] + ['  - %d\n' % v for v in values]

    start = None
    for i, line in enumerate(lines):
        if re.match(r'\s*%s:' % re.escape(key), line):
            start = i
            break

    if start is None:
        while lines and lines[-1].strip() == '':
            lines.pop()
        lines.extend(block)
    else:
        end = start + 1
        while end < len(lines) and re.match(r'\s*-\s', lines[end]):
            end += 1
        lines[start:end] = block

    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        f.writelines(lines)
    print('  %s: %s = %s' % (label, key, values or '없음'))


# ---------------------------------------------------------------------------
# 2-c) 중립 몬스터 - `임시용 중립 몬스터.xlsx` → NeutralMonster_1~3.asset (2026-08-13)
#
# 유저 지시: "중립 몬스터에도 퍼스트 스탯 시트 추가해서 넣어줘 ... 2, 3번째 중립 몬스터
# 개체량을 조금 늘려 스폰 주기도 조절하고 ... 후반에 멀리까지 가서 사냥할 당위성이 생기게
# 조절해서 테이블에 기입하고 게임내에 넣어".
#
# 그전까지 이 표는 <b>게임에 전혀 반영되지 않았다</b> - 22절에서 손으로 옮겨 적은 뒤로
# 파이프라인이 없었다(54절이 다른 표에 대해 지적한 "엑셀만, 게임 미반영" 과 같은 상황).
# 이제 다른 표와 똑같이 여기서 옮긴다.
#
# ⚠ `first_Stat` 시트가 능력치의 정본이고, `neutrality_mon` 은 식별·등장범위·보상·개체수를
#   맡는다. 두 시트 모두 <b>필드명으로</b> 읽는다(read_rows) - 컬럼이 늘거나 줄어도 안 깨진다.
# ---------------------------------------------------------------------------
# ★★ 2026-08-19 — 에픽몬스터 id 를 1000번대에서 1100번대로 옮긴 표 개정 반영
#   (커밋 "중립몹 ID 수정": "에픽몬스터 ID를 1000번대에서 1100번대로 수정").
#
# 에셋 <b>파일 이름은 그대로 두고</b>(하이라키·Resources 경로가 바뀌면 참조가 끊긴다),
# 그 파일이 표의 <b>어느 id</b>를 지금 나타내는지만 다시 잇는다 — 크리처와 파일의 관계는
# 안 바뀌었다(카르시노스 = NeutralMonster_4, 아니사킬 = NeutralMonster_5, 고르도네 =
# NeutralMonster_6, 셋 다 그대로), <b>표의 번호만</b> 카르시노스 1004→1101, 아니사킬
# 1005→1102 로 올라갔고 고르도네는 1006→1004 로 "일반" 구간에 내려왔다.
#
# ⚠ 예전에는 <b>파일이 존재하면</b> 다시 안 만든다는 이유로 아래 sync_neutral_monsters()가
#   monId·nameKey 를 patch 하지 않았다 — 그래서 표의 id 가 바뀌어도 에셋 안의 낡은 id·
#   nameKey(예: mon_name_1005)가 그대로 남아 StringTable 조회가 실패하고 화면에 에셋
#   파일명(예: "NeutralMonster_5")이 그대로 떴다. 지금부터는 changes 에 monId·nameKey 도
#   넣어 이런 재넘버링이 다시 있어도 스크립트 한 번으로 따라간다.
NEUTRAL_ASSET_BY_ID = {
    1001: 'NeutralMonster_1',
    1002: 'NeutralMonster_2',
    1003: 'NeutralMonster_3',
    1004: 'NeutralMonster_6',      # 고르도네 — 1006 에서 "일반" 구간 1004 로 내려옴
    1101: 'NeutralMonster_4',      # 카르시노스 — 1004 에서 "에픽" 구간 1101 로 올라감
    1102: 'NeutralMonster_5',      # 아니사킬 — 1005 에서 "에픽" 구간 1102 로 올라감
    1103: 'NeutralMonster_7',      # ★ 바리올라 신규 (2026-08-19)
}

# 표의 atk_type 문자열 → TacticalAttackType 의 <b>정수값</b>(YAML 은 enum 을 정수로 쓴다).
ATTACK_TYPE_VALUE = {'melee': 0, 'ranged': 1, 'magic': 2, 'heal': 3}


def attack_type_value(raw):
    """표의 `atk_type` 칸 → enum 정수. 못 알아보면 근거리(0). 대소문자·공백에 관대하다."""
    if raw is None:
        return 0
    return ATTACK_TYPE_VALUE.get(str(raw).strip().lower(), 0)


# 중립 정의에 2026-08-15 로 새로 생긴 필드들 — 기존 에셋 YAML 에는 줄이 없으므로
# `add_missing` 으로 넘겨야 첫 실행에 실제로 기록된다(patch_fields 주석 참조).
NEUTRAL_NEW_FIELDS = (
    'rangedAttackStat', 'magicStat', 'cureStat',
    'accuracyStat', 'criticalStat', 'resistanceStat',
    'attackType', 'groupMaking', 'groupMember', 'packRetaliate',
    'epic', 'habitatRadiusTiles', 'habitatChaseTiles', 'habitatIdleSlackTiles',
    'titleKey', 'colliderWidthTiles', 'colliderHeightTiles',
    'illustName', 'skinAssetName',
    # 서식지 바닥 타일 (habitat_design 시트) — 2026-08-15 신설
    'habitatTileAsset',
)


def skin_species(raw):
    """
    표의 `mon_skin` → <b>종 이름</b>.

    표에는 웨이브의 `ingame_asset`(``Char_Asset_HellFang``)과 같은 감각으로
    ``Carcinos_asset`` 처럼 적힌다. 게임이 쓰는 것은 종 이름뿐이고, 경로 규약
    (``MonsterSkins/<종>/Skin_<종>``)은 코드가 안다 —
    <see cref="NeutralMonsterDefinitionSO.SkinResourcePath"/>.

    꼬리표(`_asset`)와 머리표(`Char_Asset_`) 둘 다 떼어 준다. 사람이 손으로 적는 칸이라
    두 형식 중 무엇이 와도 통하게 한다.
    """
    s = str(raw or '').strip()
    if not s:
        return ''
    if s.startswith('Char_Asset_'):
        s = s[len('Char_Asset_'):]
    if s.endswith('_asset'):
        s = s[:-len('_asset')]
    return s

# 에픽의 서식지 <b>범위</b> 값(반경·추격거리·유휴여유) — 2026-08-19 부터 `habitat_design`
# 시트의 컬럼이다(유저 지시: "아니사킬 청크값" → 118-1절에서 남긴 미결 항목).
#
# ⚠ <b>여전히 「처음 한 번」만 심는다</b> — 표가 정본이 된 것은 <b>씨앗값의 출처</b>일 뿐이고,
#   유저 지시("타일 계산 값들은 에딧에서 수정할 수 있도록")는 안 바뀐다. 에셋에 이미 그
#   필드가 있으면(=한 번이라도 태어난 뒤) 표를 다시 돌려도 <b>절대 덮어쓰지 않는다</b>
#   (아래 seed_only 참조) — 매직 넘버가 하드코딩 상수에서 표 컬럼으로 옮겨간 것뿐,
#   에디터에서 조정한 값을 지키는 규칙은 그대로다.
#
# 표에 아직 이 종의 범위 값이 없으면(컬럼 신설 전에 만들어진 미래의 에픽 등) 이 기본값으로
# 떨어진다 — 카르시노스·아니사킬이 지금까지 쓰던 값과 같다.
EPIC_HABITAT_SEED_FALLBACK = {
    'habitatRadiusTiles': 14,
    'habitatChaseTiles': 8,
    'habitatIdleSlackTiles': 1,
}

#: `habitat_design` 컬럼명 → 에셋 필드명.
HABITAT_RANGE_FIELDS = {
    'habitat_radius_tiles': 'habitatRadiusTiles',
    'habitat_chase_tiles': 'habitatChaseTiles',
    'habitat_idle_slack_tiles': 'habitatIdleSlackTiles',
}


def sync_neutral_monsters():
    print('[중립 몬스터]')
    if not os.path.exists(XLSX_NEUTRAL):
        print('  ! 표가 없습니다:', XLSX_NEUTRAL)
        return 0

    sheets = openpyxl.load_workbook(XLSX_NEUTRAL, read_only=True).sheetnames

    stats = {int(num(r.get('mon_id'))): r for r in read_rows(XLSX_NEUTRAL, 'first_Stat')} \
        if 'first_Stat' in sheets else {}
    if not stats:
        print('  ! first_Stat 시트가 없습니다 - 능력치는 기존 값을 유지합니다')

    # 서식지 바닥 타일 + 범위 값 — <b>별도 시트</b>(habitat_design)에 있다. 종당 한 줄이고
    # 지금은 에픽만 채워져 있다. 빈 종은 타일 값이 '' 라 게임이 아무것도 안 그린다.
    #
    # ⚠ 범위 값(반경·추격거리·유휴여유)은 <b>씨앗으로만</b> 쓴다 — 위 HABITAT_RANGE_FIELDS
    #   주석 참조. 표에 그 종 줄 자체가 없거나 칸이 비면 EPIC_HABITAT_SEED_FALLBACK 로 뗀다.
    habitat = {}
    habitat_range = {}
    if 'habitat_design' in sheets:
        for r in read_rows(XLSX_NEUTRAL, 'habitat_design'):
            mid = r.get('mon_id')
            if mid in (None, ''):
                continue
            mid = int(num(mid))
            habitat[mid] = str(r.get('habitat_tile_asset') or '').strip()

            seed = dict(EPIC_HABITAT_SEED_FALLBACK)
            for col, field in HABITAT_RANGE_FIELDS.items():
                v = r.get(col)
                if v not in (None, ''):
                    seed[field] = num(v)
            habitat_range[mid] = seed

    # ★ 2026-08-15 부터 중립 정의는 <b>Resources</b> 에 산다 — 스포너가 폴더를 통째로 읽어
    #   자동 등록하기 때문이다(씬 슬롯을 손으로 만들 필요가 없어졌다).
    #   캐릭터가 Resources/Characters 를 쓰는 것과 같은 구조다.
    folder = os.path.join(ASSETS, 'Resources', 'NeutralMonsters')
    os.makedirs(folder, exist_ok=True)
    total = 0

    for row in read_rows(XLSX_NEUTRAL, 'neutrality_mon'):
        mid = int(num(row.get('mon_id')))
        asset = NEUTRAL_ASSET_BY_ID.get(mid)
        if asset is None:
            print('  ! 중립 id %d 에 해당하는 에셋 이름을 모릅니다' % mid)
            continue

        # ★ 등장 범위는 2026-08-13 부터 최소/최대 두 칸이다(유저 지시: "중립 몬스터 등장
        #   범위 최대 최소 범위 ... 넥서스 기준 타일 범위로 360도 원형").
        #   값은 <b>지름</b>이다(유저 확정: "넥서스를 중심에 두고 지름 15의 원에서부터 99의
        #   원까지 - 반지름이 아니라 지름 기준"). 절반으로 나누는 것은 게임 쪽
        #   (NeutralMonsterDefinitionSO.MinDistanceFromNexus)이 하므로 여기서는 표 값을
        #   그대로 옮긴다 - 인스펙터에 표와 같은 숫자가 보여야 대조가 된다.
        #   표에 옛 `spawn_range` 한 칸만 있는 경우를 대비해 폴백을 남긴다(뜻이 같다).
        legacy = num(row.get('spawn_range'), 0)
        is_epic = str(row.get('mon_type') or '').strip().lower() == 'epic'

        changes = {
            # ★ 2026-08-19 — id·이름 키도 매번 따라간다(위 NEUTRAL_ASSET_BY_ID 주석 참조).
            #   예전에는 이 둘이 <b>생성될 때만</b> 적히고 그 뒤로는 손대지 않아서, 표에서
            #   id 가 바뀌면(에픽 1000번대 → 1100번대) 에셋 안에는 존재하지 않는 키
            #   (`mon_name_1005`)가 남아 StringTable 조회가 실패하고 화면에 에셋 파일명이
            #   그대로 떴다("NeutralMonster_5"). 표가 정본이므로 매번 다시 쓴다.
            'monId': mid,
            'nameKey': str(row.get('mon_name') or 'mon_name_%d' % mid).strip(),
            'spawnRangeMinTiles': num(row.get('spawn_range_min'), legacy),
            'spawnRangeMaxTiles': num(row.get('spawn_range_max'), 0),
            'minEnergy': int(num(row.get('min_energy'))),
            'maxEnergy': int(num(row.get('max_energy'))),
            'maxAlive': int(num(row.get('max_alive'))),
            'respawnSeconds': num(row.get('respawn_seconds')),

            # ★ 공격 유형이 표에서 온다(2026-08-15). 예전에는 스포너가 근거리로 못박아
            #   `atk_type: ranged` 인 1002 도 붙어서 싸웠다.
            'attackType': attack_type_value(row.get('atk_type')),

            # ★ 무리 3칸 (2026-08-15 재정의)
            #
            # ⚠ `atk_take` 는 <b>선공 여부가 아니다.</b> 표의 한글 헤더가 처음부터
            #   "동료 협공 여부" 였는데 71절이 선공으로 읽어 `aggressive` 에 넣고 있었다.
            #   유저 확정(2026-08-15): <b>중립은 전부 비선공</b>, 웨이브는 전부 선공 —
            #   선공 여부는 표에 없고 종류가 정한다. 그래서 `aggressive` 필드는 없어졌고
            #   이 칸은 원래 뜻대로 <b>무리 반격 여부</b>로 쓴다.
            'groupMaking': 1 if int(num(row.get('group_making'))) else 0,
            'groupMember': int(num(row.get('group_member'))),
            'packRetaliate': 1 if int(num(row.get('atk_take'))) else 0,

            # ★ 에픽 = 서식지를 갖는 보스형 (표 mon_type)
            'epic': 1 if is_epic else 0,

            # ★ 외형 4칸 (2026-08-15 유저가 표에 신설) — 웨이브 몬스터와 같은 이름·같은 뜻
            #   콜라이더는 그림을 그 상자 안에 맞추는 데 쓰인다(61·66절).
            'colliderHeightTiles': num(row.get('collider_height_tiles'), 0),
            'colliderWidthTiles': num(row.get('collider_width_tiles'), 0),
            'illustName': str(row.get('mon_illust') or '').strip(),

            # mon_skin 은 표에 `Carcinos_asset` 처럼 적힌다(웨이브의 `ingame_asset` 과 같은 형식).
            # 게임은 <b>종 이름</b>만 필요하므로 꼬리표를 뗀다 →
            #   Carcinos_asset → Carcinos → Resources/MonsterSkins/Carcinos/Skin_Carcinos
            'skinAssetName': skin_species(row.get('mon_skin')),

            # 서식지 바닥 타일 묶음 이름 (habitat_design 시트). 게임은
            # `Resources/HabitatTiles/<이름>/` 을 통째로 읽는다.
            'habitatTileAsset': habitat.get(mid, ''),
        }

        # 칭호 — 표에 적힌 종만 키를 넣는다.
        #
        # ★ 2026-08-15 수정 — <b>칸에 적힌 키를 그대로 쓴다</b>.
        #   예전에는 웨이브 보스를 따라 `'mon_title_%d' % mid` 로 <b>유추한</b> 키를 넣었는데,
        #   중립 1004 의 `mon_title` 칸은 <b>`epic_boss_title_1004`</b> 라고 적혀 있다.
        #   그래서 에셋에는 실재하지 않는 `mon_title_1004` 가 들어가고, StringTable 조회가
        #   매번 실패해 <b>칭호가 조용히 빈 문자열</b>이 됐다(보스 체력바에 이름만 떴다).
        #   `looks_like_key` 와 같은 판정으로, 이미 키면 그 값을, 한국어 리터럴이면
        #   생성기(`gen_string_table.py`)가 만들 이름을 넣는다.
        title_cell = str(row.get('mon_title') or '').strip()
        if title_cell:
            is_key = (title_cell.isascii() and '_' in title_cell
                      and not any(ch.isspace() for ch in title_cell))
            changes['titleKey'] = title_cell if is_key else 'mon_title_%d' % mid

        st = stats.get(mid)
        if st is not None:
            changes.update({
                'hpStat': int(num(st.get('hp'), 1)),
                'attackStat': int(num(st.get('melee_atk'))),
                'defenseStat': int(num(st.get('def'))),
                'regenStat': int(num(st.get('hp_recovery'))),
                # 공속·이속은 웨이브 몬스터와 같은 38-1절 공식으로 미리 풀어서 넣는다.
                'attacksPerSecond': aspd_from_stat(num(st.get('atk_speed'))),
                'moveSpeedTiles': mspd_from_stat(num(st.get('movement_speed'))),

                # ★ 나머지 공격 계열 + 명중·치명·저항 (2026-08-15 신설)
                #   ⚠ 명중·치명은 <b>원거리 유형에만</b> 적용된다 — 근거리 종에 값이
                #     들어 있어도 판정에 쓰이지 않는다.
                'rangedAttackStat': int(num(st.get('ranged_atk'))),
                'magicStat': int(num(st.get('magic'))),
                'cureStat': int(num(st.get('cure'))),
                'accuracyStat': int(num(st.get('accuracy'), 50)),
                'criticalStat': int(num(st.get('critical'))),
                'resistanceStat': int(num(st.get('resistance'), 50)),
            })

        path = os.path.join(folder, asset + '.asset')

        # 표에 새로 생긴 종은 에셋 파일 자체가 없다 — 뼈대를 먼저 만든다(1004 가 그렇다).
        if not os.path.exists(path):
            create_neutral_asset(path, asset, mid, row)

        # 서식지 범위 값은 <b>에디터에서 조정</b>하는 값이라 매번 덮어쓰지 않는다.
        # 아직 필드 자체가 없을 때(= 처음 만들어질 때)만 표(또는 폴백)의 씨앗값을 심는다.
        if is_epic:
            with open(path, encoding='utf-8') as f:
                cur = f.read()
            seed = habitat_range.get(mid, EPIC_HABITAT_SEED_FALLBACK)
            for k, v in seed.items():
                if not re.search(r'^\s*%s:' % re.escape(k), cur, re.M):
                    changes[k] = v

        total += patch_fields(path, changes, asset,
                              add_missing=('maxAlive', 'respawnSeconds',
                                           'spawnRangeMinTiles', 'spawnRangeMaxTiles')
                                          + NEUTRAL_NEW_FIELDS)
        # ⚠ 2026-08-16 부터 등장 범위는 <b>정사각형</b>이다(유저 확정) — 표 값은 "한 변",
        #   실제 판정은 그 절반(반변, 체비셰프 거리)이다. 로그 문구도 그에 맞췄다.
        print('    %s: %s · 등장 정사각 변 %s~%s타일(반변 %.1f~%.1f) · 에너지 %s~%s · '
              '비선공%s · 최대 %s마리 · 재생성 %s초'
              % (asset,
                 '에픽(서식지)' if is_epic else '일반',
                 changes['spawnRangeMinTiles'], changes['spawnRangeMaxTiles'],
                 changes['spawnRangeMinTiles'] / 2.0, changes['spawnRangeMaxTiles'] / 2.0,
                 changes['minEnergy'], changes['maxEnergy'],
                 (' · 무리 %d마리%s' % (changes['groupMember'],
                                       '·무리반격' if changes['packRetaliate'] else ''))
                 if changes['groupMaking'] else '',
                 changes['maxAlive'], changes['respawnSeconds']))

        # 스킨이 지정된 종만 — 표의 상자에 원화를 맞추면 실제로 몇 타일이 되는지 보여준다.
        species = changes['skinAssetName']
        if species:
            report_collider_fit_skin(asset, 'MonsterSkins/%s/Skin_%s' % (species, species),
                                     changes['colliderWidthTiles'],
                                     changes['colliderHeightTiles'])
    return total


def create_neutral_asset(path, asset, mid, row):
    """
    표에 새로 생긴 중립 종의 에셋 <b>뼈대</b>를 만든다. 값 채우기는 곧바로 이어지는
    `patch_fields` 가 하므로, 여기서는 <b>모든 필드가 한 줄씩 존재</b>하게만 해준다.

    ⚠ `.asset` YAML 에 <b>빈 줄을 넣으면 안 된다</b> — Unity 파서가 그 뒤 필드를 전부
      무시한다(진행상황 8절 3번). 이 파일 맨 위 원칙과 같다.

    ⚠ `template` 은 `{fileID: 0}` 으로 둔다 — ScriptableObject 는 <b>씬 오브젝트를
      참조할 수 없다</b>(5절). 스포너의 `spawnTable` 슬롯에서 씬 쪽으로 연결한다.
    """
    guid = script_guid('Units/NeutralMonsterDefinitionSO.cs')
    body = HEADER.format(script_guid=guid, name=asset)
    body += "  monId: %d\n" % mid
    body += "  nameKey: %s\n" % (str(row.get('mon_name') or 'mon_name_%d' % mid).strip())
    body += "  displayName: %s\n" % asset          # 스트링 테이블이 정본이라 폴백용일 뿐이다
    body += "  template: {fileID: 0}\n"
    # 나머지 값은 patch_fields 가 채운다 — 여기서는 줄만 만들어 둔다.
    for line in ('minEnergy: 0', 'maxEnergy: 0', 'attackStat: 0', 'hpStat: 1',
                 'defenseStat: 0', 'regenStat: 0', 'detectRange: 6', 'attackRange: 1.2',
                 'attacksPerSecond: 0.7', 'moveSpeedTiles: 1.8', 'leashRangeTiles: 6'):
        body += "  %s\n" % line

    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(body)

    mp = path + '.meta'
    if not os.path.exists(mp):
        with open(mp, 'w', encoding='utf-8', newline='\n') as f:
            f.write(ASSET_META.format(
                guid=guid_for('Resources/NeutralMonsters/%s.asset' % asset)))
    print('    + %s 신규 생성' % asset)


# ---------------------------------------------------------------------------
# 3) 건물 - 54-7절 (포탑 상향 · 중앙건물은 이미 게임 값과 같다)
# ---------------------------------------------------------------------------
def sync_buildings():
    print('[건물]')
    const = {int(r[0]): r for r in read_sheet(XLSX_BUILDING, 'Construction')}
    atk = {int(r[0]): r for r in read_sheet(XLSX_BUILDING, 'ATK')}

    turret = const.get(10002)
    if turret is None:
        print('  ! Construction 시트에 10002(포탑)가 없습니다')
        return 0

    a = atk.get(int(num(turret[9])))    # ATK_ID
    changes = {
        'hp': int(num(turret[4])),
        'defenseStat': int(num(turret[7])),
        'attackStat': int(num(turret[8])),
        'buildSeconds': int(num(turret[11])),
        'maxCount': int(num(turret[3])),
    }
    if a is not None:
        changes['attacksPerSecond'] = num(a[2])
        changes['attackRange'] = num(a[3])

    n = patch_fields(os.path.join(ASSETS, 'Data', 'Buildings', 'Building_Turret.asset'),
                     changes, 'Building_Turret')

    # 중앙건물(10001) - 게임은 능력치 치환(체력스탯 100 × 보정 250% = 2,600)으로 같은 값을
    # 이미 만들고 있다(54-7절이 "시트를 게임에 맞췄다"). 표의 HP/DEF 와 실제 값이 같은지만 검사한다.
    core = const.get(10001)
    if core is not None:
        want_hp, want_def = int(num(core[4])), int(num(core[7]))
        nexus = os.path.join(ASSETS, 'Data', 'Combat', 'NexusDefinition.asset')
        with open(nexus, encoding='utf-8') as f:
            t = f.read()
        hp_stat = int(re.search(r'hpStat:\s*(\d+)', t).group(1))
        hp_pct = int(re.search(r'hpPercent:\s*(\d+)', t).group(1))
        d = int(re.search(r'defenseStat:\s*(\d+)', t).group(1))
        actual = round((40 + hp_stat * 10) * hp_pct / 100)
        ok = (actual == want_hp and d == want_def)
        print('  중앙건물: 표 HP %d/DEF %d ↔ 게임 %d/DEF %d %s'
              % (want_hp, want_def, actual, d, '(일치)' if ok else '(⚠ 불일치)'))
    return n


# ---------------------------------------------------------------------------
# 4) 웨이브 구성표 - 54-5절 (중간보스 수량 컬럼 신설 + 5·15웨이브 잡몹 감소)
# ---------------------------------------------------------------------------
def sync_waves():
    print('[웨이브 구성표]')
    path = os.path.join(ASSETS, 'Data', 'Wave', 'WaveDefinitions.asset')
    with open(path, encoding='utf-8') as f:
        text = f.read()

    # 증원(reinforce*)은 표에 없는 값이다 - 27절이 코드/에셋 쪽에서 정한 것이므로
    # 표를 반영할 때 **덮어쓰지 않고 기존 값을 그대로 유지**한다(54-8절과 같은 취지).
    keep = {}
    for m in re.finditer(r'- waveNumber: (\d+)(.*?)(?=\n  - waveNumber:|\Z)', text, re.S):
        blk = m.group(2)
        keep[int(m.group(1))] = (
            re.search(r'reinforceIntervalSeconds: ([\d.]+)', blk).group(1),
            re.search(r'reinforceCount: (\d+)', blk).group(1),
        )

    # ★ 2026-08-18 - <b>위치가 아니라 이름으로 읽는다</b>(read_rows).
    #   이날 `mid_boss_mon_num` 컬럼을 지우고 `boss_monster_id` 를 새로 만들었는데,
    #   예전처럼 r[6]/r[7] 로 읽으면 <b>컬럼이 한 칸씩 밀려 조용히 엉뚱한 값</b>이 들어간다
    #   (read_rows 의 주석이 적어둔 2026-08-13 사고와 같은 것이다).
    rows = read_rows(XLSX_WAVE, 'Sheet2')
    out = ['  waves:']
    for r in rows:
        wn = int(num(r.get('wave_num')))
        ri, rc = keep.get(wn, ('0', '0'))
        out.append('  - waveNumber: %d' % wn)
        out.append('    meleeCount: %d' % int(num(r.get('melee_mon_num'))))
        out.append('    rangedCount: %d' % int(num(r.get('ranged_mon_num'))))
        out.append('    bossCount: %d' % int(num(r.get('boss_mon_num'))))
        # boss_monster_id(2026-08-18 신설) - 그 웨이브에 나올 보스의 표 id.
        # midBossCount 를 대체한다: 유저 지시로 중간보스가 없어지고 보스가 2종
        # (단탈리온·말파스)이 되면서 "어느 보스인지" 를 표가 정해야 했다.
        out.append('    bossMonsterId: %d' % int(num(r.get('boss_monster_id'))))
        out.append('    statPercent: %d' % round(float(num(r.get('wave_mon_abil_per'))) * 100))
        out.append('    reinforceIntervalSeconds: %s' % ri)
        out.append('    reinforceCount: %s' % rc)
        # spawn_group_size - 2026-08-13 신설. 포탈 한 곳에서 한 번에 나오는 마리 수.
        # 0/1 이면 예전처럼 한 마리씩 나온다(동작 불변).
        out.append('    spawnGroupSize: %d' % int(num(r.get('spawn_group_size'), 1)))

    head = text[:text.index('  waves:')]
    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(head + '\n'.join(out) + '\n')

    boss_waves = ['%d(%d)' % (int(num(r.get('wave_num'))), int(num(r.get('boss_monster_id'))))
                  for r in rows if int(num(r.get('boss_mon_num'))) > 0]
    print('  웨이브 %d개 기록 · 보스가 나오는 웨이브(보스 id): %s'
          % (len(rows), ' '.join(boss_waves) or '없음'))
    return len(rows)


if __name__ == '__main__':
    sync_mental_errors()
    sync_monsters()
    sync_boss_skills()
    sync_neutral_monsters()
    # ⚠ 중립 정의(에셋)가 만들어진 <b>뒤에</b> 돌아야 한다 — skillIds 를 그 파일에 쓴다.
    sync_neutral_skills()
    sync_buildings()
    sync_waves()
    print('\n완료 - Unity 에서 Assets/Refresh 를 실행할 것.')
