# -*- coding: utf-8 -*-
"""도움말 표(xlsx) → `HelpTableSO` 에셋 (2026-08-24).

원본: ``<볼트>/데이터 테이블/Last_Sanctuary_도움말테이블_Ver01.xlsx`` 의 ``Help`` 시트
결과: ``Assets/_Project/Resources/Data/Help/HelpTable.asset``

진행상황 140-6절이 «아직 안 한 것» 2번으로 적어 둔 그 일이다. 다만 그 절은
`sync_tables_to_assets.py` 에 시트를 <b>더하라</b>고 적어 뒀는데, <b>따로 뺐다</b> —
그 스크립트는 «기존 에셋의 필드 값만 한 줄씩 치환» 하는 방식이고(자기 맨 위 ⚠),
도움말은 <b>줄 수가 표에서 결정되는 목록 하나</b>라 «통째로 다시 쓰는» 방식이 맞다
(`gen_relic_assets.py` 의 대사표와 같은 모양이다).

★ 이 스크립트가 <b>표와 게임 사이의 유일한 다리</b>다. 구조를 바꾸려면 표를 고치고 다시 돌린다.
★ <b>문구는 여기로 오지 않는다</b> — 스트링 키만 옮긴다. 문구는
  `help_string_merge.py` → `gen_string_table.py` 가 스트링 테이블로 나른다.
  그래서 <b>문구를 다듬을 때 이 스크립트를 다시 돌릴 필요가 없다</b>.

⚠ .asset YAML 에 <b>빈 줄을 넣으면 유니티 파서가 그 뒤 필드를 전부 무시한다</b>(8절 3번).
⚠ 표의 `trigger` 를 코드가 못 알아보면 그 항목은 <b>백과에만</b> 남는다(저절로 안 뜬다).
  조용히 넘기지 않고 <b>세어서 알린다</b>.

사용법:  py -3 Tools/gen_help_assets.py
다음:    유니티에서 Assets/Refresh
"""

import hashlib
import io
import os
import sys

import openpyxl

from vault_path import TABLE_DIR, PROJECT

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "Last_Sanctuary_도움말테이블_Ver01.xlsx")
OUT_DIR = os.path.join(PROJECT, "Assets", "_Project", "Resources", "Data", "Help")
TSV = os.path.join(PROJECT, "Assets", "_Project", "Resources", "Data", "StringTable.txt")

SHEET = "Help"
STEP_SHEET = "HelpStep"
FIRST_ROW = 2      # 도움말 표는 머리글 <b>한 줄</b>이다 (다른 데이터 테이블은 세 줄이다)

# ── 표의 trigger → C# enum 정수값. <b>HelpTableSO.cs 의 HelpTrigger 와 반드시 같아야 한다.</b> ──
TRIGGER = {
    "NewRunFirstPreparation": 1,
    "EnergyGained": 2,
    "CanCreateCharacter": 3,
    "BattleStarted": 4,
    "EnrageStarted": 5,
    "AllyDied": 6,
    "FirstMiss": 7,
    "BossWaveSpawned": 8,
    "CharacterUpgraded": 9,
    "HeroAwakened": 10,
    "RelicDigMarkAppeared": 11,
    "RelicObtained": 12,
    "SquadCreated": 13,
    "RallyPointCreated": 14,
    "TacticsChanged": 15,
    # 16 = BuildModeEntered — 지웠다 (2026-08-25 · 포탑 건설 기능이 없어졌다).
    #      번호는 비워 둔다: 다시 쓰면 이미 구운 에셋의 16 이 엉뚱한 계기가 된다.
    "ErosionReached": 17,
    "MentalErrorTriggered": 18,
    "NeutralKilled": 19,
    "EpicNeutralFound": 20,
    "EventStarted": 21,
    "AutoSaved": 22,
    "GameSpeedChanged": 23,

    # ★★★ 허드 액션 버튼의 «첫 클릭» (2026-08-25 신설 · HelpService.InterceptFirstUse)
    "ActionCreate": 24,
    "ActionUpgrade": 25,
    "ActionSquad": 26,
    "ActionTactics": 27,
    "ActionSubjugate": 28,
    "ActionRelic": 29,
    "ActionSettings": 30,
}

HEADER = """%YAML 1.1
%TAG !u! tag:unity3d.com,2011:
--- !u!114 &11400000
MonoBehaviour:
  m_ObjectHideFlags: 0
  m_CorrespondingSourceObject: {{fileID: 0}}
  m_PrefabInstance: {{fileID: 0}}
  m_PrefabAsset: {{fileID: 0}}
  m_GameObject: {{fileID: 0}}
  m_Enabled: 1
  m_EditorHideFlags: 0
  m_Script: {{fileID: 11500000, guid: {script_guid}, type: 3}}
  m_Name: {name}
  m_EditorClassIdentifier:
"""

ASSET_META = """fileFormatVersion: 2
guid: {guid}
NativeFormatImporter:
  externalObjects: {{}}
  mainObjectFileID: 11400000
  userData:
  assetBundleName:
  assetBundleVariant:
"""

FOLDER_META = """fileFormatVersion: 2
guid: {guid}
folderAsset: yes
DefaultImporter:
  externalObjects: {{}}
  userData:
  assetBundleName:
  assetBundleVariant:
"""


def guid_for(key):
    """경로에서 결정적으로 만든다 — 다시 돌려도 같은 guid 라 참조가 안 끊긴다."""
    return hashlib.md5(("LastSanctuary/" + key).encode("utf-8")).hexdigest()


def script_guid(rel_cs_path):
    meta = os.path.join(PROJECT, "Assets", "_Project", "Scripts", rel_cs_path) + ".meta"
    if not os.path.isfile(meta):
        raise SystemExit(
            "⚠ %s 이(가) 없습니다.\n"
            "   유니티가 아직 새 스크립트를 임포트하지 않았습니다 —\n"
            "   node Tools/mcp_unity_cli.js execute_menu_item '{\"menuPath\":\"Assets/Refresh\"}'"
            % meta)
    with io.open(meta, encoding="utf-8") as f:
        for line in f:
            if line.startswith("guid:"):
                return line.split(":", 1)[1].strip()
    raise SystemExit("⚠ guid 를 찾지 못했습니다: " + meta)


def yaml_str(v):
    """⚠ 줄바꿈을 이스케이프한다 — 빈 줄이 들어가면 유니티가 뒤를 전부 버린다(맨 위 ⚠)."""
    s = "" if v is None else str(v)
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    s = s.replace("\r\n", "\\n").replace("\n", "\\n").replace("\r", "\\n")
    return '"%s"' % s


def num(v, default=0):
    if v is None:
        return default
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return default


def norm(v):
    return "" if v is None else str(v).strip()


def write(path, text):
    with io.open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write(text)


def sheet_rows(wb, name, required=True):
    """1행이 필드명이고 2행부터 값이다. 첫 칸이 빈 줄은 건너뛴다."""
    if name not in wb.sheetnames:
        if required:
            raise SystemExit("⚠ '%s' 시트가 없습니다 — 시트: %s" % (name, wb.sheetnames))
        return None

    ws = wb[name]
    fields = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(FIRST_ROW, ws.max_row + 1):
        if not norm(ws.cell(r, 1).value):
            continue
        rows.append({fields[c - 1]: ws.cell(r, c).value
                     for c in range(1, ws.max_column + 1) if fields[c - 1]})
    return rows


def read_rows():
    """
    (항목, 단계) 를 돌려준다.

    ★ <b>단계 시트는 없어도 된다</b> — 있으면 「자세히 보기」가 빨간 테두리로 화면을 짚고,
      없으면 예전처럼 백과를 연다. 없다고 굽기를 멈추면 표를 나눠 만드는 동안 작업이 막힌다.
    """
    if not os.path.isfile(XLSX):
        raise SystemExit("⚠ 도움말 표가 없습니다: %s" % XLSX)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    return sheet_rows(wb, SHEET), sheet_rows(wb, STEP_SHEET, required=False)


def string_keys():
    """내보낸 TSV 에 실제로 있는 키 — 표의 키가 다 있는지 검사하는 데 쓴다."""
    if not os.path.isfile(TSV):
        return None
    keys = set()
    with io.open(TSV, encoding="utf-8") as f:
        for line in f:
            line = line.strip("\r\n")
            if not line or line.startswith("#"):
                continue
            k = line.split("\t")[0].strip()
            if k and k != "string_key":
                keys.add(k)
    return keys


def main():
    rows, steps = read_rows()
    os.makedirs(OUT_DIR, exist_ok=True)

    folder_meta = OUT_DIR + ".meta"
    if not os.path.isfile(folder_meta):
        rel = os.path.relpath(OUT_DIR, PROJECT).replace("\\", "/")
        write(folder_meta, FOLDER_META.format(guid=guid_for(rel)))

    ids = {norm(r.get("help_id")) for r in rows}
    known = string_keys()

    bad_trigger, missing_key, bad_see_also = [], [], []
    no_category_key, no_step_key = [], []
    by_category = {}

    #: help_id → 「자세히 보기」가 열어야 하는 창. 단계 검산이 이 값을 쓴다.
    panel_of = {norm(r.get("help_id")): norm(r.get("open_panel")) for r in rows}

    body = HEADER.format(script_guid=script_guid("Help/HelpTableSO.cs"), name="HelpTable")
    body += "  entries:\n"

    for r in rows:
        hid = norm(r.get("help_id"))
        cat = norm(r.get("category"))
        trig = norm(r.get("trigger"))
        see = norm(r.get("see_also"))

        if trig and trig not in TRIGGER:
            bad_trigger.append((hid, trig))
        if see and see not in ids:
            bad_see_also.append((hid, see))

        # ⚠ category_key 도 함께 본다 (2026-08-27 · 184절) — 이 키가 비거나 표에 없으면
        #   백과의 <b>탭 여섯</b>이 영어로 안 바뀐다. 폴백이 한글 분류명이라 «조용히
        #   한국어로 남는» 실패라서, 세지 않으면 아무도 모른다.
        for field in ("title_key", "summary_key", "body_key", "category_key"):
            key = norm(r.get(field))
            if key and known is not None and key not in known:
                missing_key.append((hid, key))
        if not norm(r.get("category_key")):
            no_category_key.append(hid)

        by_category[cat] = by_category.get(cat, 0) + 1

        body += "  - helpId: %s\n" % yaml_str(hid)
        body += "    category: %s\n" % yaml_str(cat)
        body += "    categoryKey: %s\n" % yaml_str(norm(r.get("category_key")))
        body += "    order: %d\n" % num(r.get("order"))
        body += "    titleKey: %s\n" % yaml_str(norm(r.get("title_key")))
        body += "    summaryKey: %s\n" % yaml_str(norm(r.get("summary_key")))
        body += "    bodyKey: %s\n" % yaml_str(norm(r.get("body_key")))
        body += "    trigger: %d\n" % TRIGGER.get(trig, 0)
        body += "    triggerArg: %d\n" % num(r.get("trigger_arg"))
        body += "    priority: %d\n" % max(1, min(3, num(r.get("priority"), 2)))
        body += "    showOnce: %d\n" % (1 if num(r.get("show_once"), 1) else 0)
        body += "    seeAlso: %s\n" % yaml_str(see)
        body += "    openPanelPath: %s\n" % yaml_str(norm(r.get("open_panel")))

    # ── 화면에서 짚어 주기 (HelpStep 시트) ─────────────────────────────
    #
    # ⚠ Help 시트에 없는 항목을 가리키는 단계는 <b>영영 안 뜬다</b> — 그 help_id 의
    #   「자세히 보기」에 도달할 방법이 없다. 조용히 넘기지 않고 세어서 알린다.
    bad_step, targeted, scattered = [], 0, []
    by_entry = {}
    body += "  steps:\n"
    for st in (steps or []):
        hid = norm(st.get("help_id"))
        target = norm(st.get("target_path"))
        if hid not in ids:
            bad_step.append((hid, num(st.get("step_order"))))
        if target:
            targeted += 1
            by_entry.setdefault(hid, []).append(target)

        # ⚠ 단계 글의 스트링 키 (2026-08-27 · 184절). 비거나 표에 없으면 그 단계는
        #   <b>영어에서도 한글로</b> 뜬다 — 폴백이 stepText 라 조용히 실패한다.
        step_key = norm(st.get("step_text_key"))
        if not step_key:
            no_step_key.append("%s %s" % (hid, num(st.get("step_order"))))
        elif known is not None and step_key not in known:
            missing_key.append((hid, step_key))

        body += "  - helpId: %s\n" % yaml_str(hid)
        body += "    stepOrder: %d\n" % num(st.get("step_order"))
        body += "    targetPath: %s\n" % yaml_str(target)
        body += "    stepText: %s\n" % yaml_str(norm(st.get("step_text")))
        body += "    stepTextKey: %s\n" % yaml_str(step_key)

    # ★★ <b>한 항목의 단계는 «한 UI 안» 에서만 머문다</b> (2026-08-24 유저 지시로 세운 규칙).
    #   창을 여는 항목은 그 창 안만, 창이 없는 항목은 늘 보이는 HUD 하나 안만 짚어야 한다.
    #   ⚠ 이 검산이 «어수선함» 을 막는 장치다 — 사람이 표를 늘리면 반드시 다시 흩어진다.
    #   ⚠ 표를 손으로 고쳤을 때도 잡히도록 <b>굽는 쪽에서 한 번 더</b> 본다
    #     (`table_update_*_help_rewrite.py` 의 검산과 <b>같은 규칙</b>이다).
    for hid, targets in by_entry.items():
        panel = panel_of.get(hid, "")
        if panel:
            out = [t for t in targets if not t.startswith(panel + "/")]
            if out:
                scattered.append("%s → %s 를 여는데 창 밖을 짚습니다: %s"
                                 % (hid, panel, " · ".join(out)))
        else:
            roots = {"/".join(t.split("/")[:2]) for t in targets}
            if len(roots) > 1:
                scattered.append("%s → 단계가 여러 UI 를 건너뜁니다: %s"
                                 % (hid, " · ".join(sorted(roots))))

    for hid, panel in panel_of.items():
        if panel and hid not in by_entry:
            scattered.append("%s → 여는 창(%s)만 있고 짚을 단계가 없습니다 (빈 창만 뜹니다)"
                             % (hid, panel))

    path = os.path.join(OUT_DIR, "HelpTable.asset")
    write(path, body)
    rel = os.path.relpath(path, PROJECT).replace("\\", "/")
    write(path + ".meta", ASSET_META.format(guid=guid_for(rel)))

    # ── 보고 ───────────────────────────────────────────────────────────
    print("[도움말 에셋]")
    print("  항목 %d개  →  %s" % (len(rows), os.path.relpath(path, PROJECT)))
    print("  분류 %s" % " · ".join("%s %d" % (k, v) for k, v in by_category.items()))
    print("  계기가 붙은 항목 %d개 · 백과 전용 %d개"
          % (sum(1 for r in rows if norm(r.get("trigger")) in TRIGGER),
             sum(1 for r in rows if not norm(r.get("trigger")))))

    if steps is None:
        print("  · HelpStep 시트가 없습니다 — 「자세히 보기」 버튼이 아무 항목에도 안 뜹니다")
    else:
        tour_ids = {norm(s.get("help_id")) for s in steps}
        opens = sum(1 for v in panel_of.values() if v)
        print("  짚어 주기 %d단계 — 화면을 가리키는 단계 %d개 · 글만 보여주는 단계 %d개"
              % (len(steps), targeted, len(steps) - targeted))
        print("  「자세히 보기」가 뜨는 항목 %d개 (그중 창을 여는 것 %d개) · "
              "안 뜨는 항목 %d개 (규칙·개념)"
              % (len(tour_ids), opens, len(rows) - len(tour_ids)))

    if scattered:
        print("  ⚠ «한 UI 안» 규칙을 어긴 항목 %d개 — 안내가 어수선해집니다:" % len(scattered))
        for line in scattered:
            print("      " + line)

    if bad_step:
        print("  ⚠ Help 시트에 없는 항목의 단계 %d개 — 영영 안 뜹니다:" % len(bad_step))
        for hid, order in bad_step:
            print("      %-20s %d" % (hid, order))

    if bad_trigger:
        print("  ⚠ 코드가 못 알아보는 계기 %d개 — 그 항목은 백과에만 남습니다:" % len(bad_trigger))
        for hid, t in bad_trigger:
            print("      %-20s [%s]" % (hid, t))
    if bad_see_also:
        print("  ⚠ 없는 항목을 가리키는 see_also %d개 — 「함께 볼 것」이 안 뜹니다:" % len(bad_see_also))
        for hid, s in bad_see_also:
            print("      %-20s → %s" % (hid, s))
    if missing_key:
        print("  ⚠ 스트링 테이블에 없는 키 %d개 — 화면에 <b>키 이름이 그대로</b> 뜹니다:"
              % len(missing_key))
        for hid, k in missing_key:
            print("      %-20s %s" % (hid, k))
    # ⚠⚠ 아래 둘은 «키 이름이 뜬다» 가 아니라 <b>한글이 그대로 뜬다</b> — 한국어로 보면
    #   멀쩡해 보여서 <b>영어로 켜 보기 전에는 아무도 모른다</b>. 그래서 따로 센다(184절).
    if no_category_key:
        print("  ⚠ category_key 가 빈 항목 %d개 — 백과 탭이 <b>영어에서도 한글</b>입니다:"
              % len(no_category_key))
        print("      " + " · ".join(no_category_key))
    if no_step_key:
        print("  ⚠ step_text_key 가 빈 단계 %d개 — 짚어 주기 글이 <b>영어에서도 한글</b>입니다:"
              % len(no_step_key))
        print("      " + " · ".join(no_step_key))

    if known is None:
        print("  · StringTable.txt 가 없어 키 검사를 못 했습니다 "
              "(py -3 Tools/gen_string_table.py 를 먼저 돌리세요)")
    if not (bad_trigger or bad_see_also or missing_key or bad_step or scattered
            or no_category_key or no_step_key):
        print("  ✓ 검산 통과 — 계기 · see_also · 스트링 키 · 짚어 주기 단계 · «한 UI 안» 규칙 모두 맞습니다")
    print("  다음: 유니티에서 Assets/Refresh")
    return 0


if __name__ == "__main__":
    sys.exit(main())
