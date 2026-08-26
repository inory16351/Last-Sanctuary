# -*- coding: utf-8 -*-
"""웨이브 잡몹 처치 보상을 <b>표의 min~max 에서 굴리게</b> 한다 (2026-08-26, 6차).

유저 지시
---------
  *"왜 자원값이 확정이 됐음? 테이블 보고 랜덤값으로 해"*

★★★ 무엇이 «확정» 이었나
──────────────────────────────────────────────────────────────────────
중립 몬스터는 처음부터 표의 <b>min~max_energy</b> 에서 굴렸다
(`NeutralMonsterUnit.RollEnergyReward`). 그런데 <b>웨이브 잡몹</b>은
`ResourceManager.energyPerMonsterKill` <b>한 값(10)</b> 이었다 —
종을 가리지 않고 «언제나 정확히 10» 이고, <b>표에 칸조차 없었다</b>.

★ 새 칸 — `웨이브 몬스터 테이블.xlsx` ▸ `wave_nom` 에 <b>두 칸</b>을 <b>맨 뒤에</b> 붙인다
  (65-2절 «컬럼은 맨 뒤에만» · 38-2절 «행 번호로 참조하는 시트가 있다»).

    F  획득 최소 에너지   min_energy   int
    G  획득 최대 에너지   max_energy   int

★ 값 — <b>체력 비례</b>(5차 중립 규칙과 같은 결) · <b>평균은 10 을 지킨다</b>
──────────────────────────────────────────────────────────────────────
    id      이름          체력   min~max   평균     근거
    100001  지옥 송곳니     7     9~13      11      7 × 1.54
    100002  영혼 사수       6     7~11       9      6 × 1.54
                                          ─────
                              두 종 평균    10      ← 옛 고정값과 같다

  ★★ <b>평균을 10 으로 맞춘 것이 핵심이다.</b> 「웨이브 부하」 시트의 성장 모델은
    «웨이브당 처치 수 × 2 × 10» 으로 수입을 세고 있고(`WAVE_KILL_ENERGY = 10`),
    두 종이 <b>같은 수</b>로 나온다(`근=원 마리` 열). 평균이 10 이면
    <b>경제 곡선이 한 칸도 안 움직인다</b> — 이번 변경은 «같은 평균에 폭을 준 것» 이다.
  ⚠ 보스는 칸을 <b>비워 둔다</b> — 0~0 이면 코드가 «표에 없음» 으로 읽고 예전처럼
    에너지를 주지 않는다(보스의 보상은 유물이다).

⚠ <b>Excel COM 으로 쓴다</b>(136-4절). 백업은 `_백업/<시각>_잡몹에너지/`.

사용법:  python Tools/table_update_20260826_wave_mob_energy.py
다음:    python Tools/sync_tables_to_assets.py     ·  Unity 에서 Assets/Refresh
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "웨이브 몬스터 테이블.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")
SHEET = "wave_nom"

# 헤더 세 줄 — (1행 한글, 2행 필드명, 3행 자료형)
NEW_COLUMNS = [
    ("획득 최소 에너지", "min_energy", "int"),
    ("획득 최대 에너지", "max_energy", "int"),
]

# monster_id → (min, max).  ★ 두 종의 평균이 10 이어야 한다(위 주석).
ENERGY = {
    100001: (9, 13),        # 지옥 송곳니 — 체력 7
    100002: (7, 11),        # 영혼 사수  — 체력 6
}


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_잡몹에너지")
    os.makedirs(folder, exist_ok=True)
    shutil.copy2(XLSX, os.path.join(folder, os.path.basename(XLSX)))
    print("백업: " + folder)


def read_layout():
    """지금 시트의 «필드명 → 열 번호» 와 마지막 열을 읽는다."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    names = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None and str(v).strip():
            names[str(v).strip()] = c
    return names, ws.max_column


def main():
    names, last = read_layout()
    print(f"「{SHEET}」 지금 칸 {last}개 — {', '.join(names)}")

    missing = [spec for spec in NEW_COLUMNS if spec[1] not in names]
    if not missing:
        print("두 칸이 이미 있다 — 값만 다시 쓴다(멱등).")

    backup()

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(XLSX))
        ws = wb.Worksheets(SHEET)

        col = {}
        nxt = last + 1
        for kr, field, typ in NEW_COLUMNS:
            if field in names:
                col[field] = names[field]
                continue
            ws.Cells(1, nxt).Value = kr
            ws.Cells(2, nxt).Value = field
            ws.Cells(3, nxt).Value = typ
            col[field] = nxt
            print(f"  칸 신설 — {nxt}열  {kr} / {field} / {typ}")
            nxt += 1

        touched = 0
        row = 4
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            mid = int(v)
            if mid in ENERGY:
                lo, hi = ENERGY[mid]
                ws.Cells(row, col["min_energy"]).Value = lo
                ws.Cells(row, col["max_energy"]).Value = hi
                print(f"  잡몹 {mid}: 에너지 {lo}~{hi} (평균 {(lo + hi) / 2:.1f})")
                touched += 1
            row += 1

        wb.Save()
        wb.Close()
        print(f"웨이브 몬스터 테이블.xlsx — {touched}종 갱신")
    finally:
        excel.Quit()

    print("\n평균 검산 — " +
          " · ".join(f"{m} {(v[0] + v[1]) / 2:.1f}" for m, v in ENERGY.items()) +
          f"  →  두 종 평균 {sum((v[0] + v[1]) / 2 for v in ENERGY.values()) / len(ENERGY):.1f}"
          "  (모델의 WAVE_KILL_ENERGY = 10)")


if __name__ == "__main__":
    main()
