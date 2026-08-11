# -*- coding: utf-8 -*-
"""
「데이터 테이블/능력치 및 공식 정리.xlsx」 개정 스크립트 (2026-08-11).

두 가지를 한다.
  1. 비어 있던 값을 채운다 — 계수 시트에 없던 폴백·생성 랜덤 범위 4개,
     '공식' 시트의 빈 계산 예시(C10), 수식으로 오인돼 #NAME? 이 뜨던 칸(C22),
     '능력치' 시트의 빈 비고 3칸.
  2. 공속·이속 공식을 **점근(diminishing returns) 형태**로 바꾼다.
     기존 `기본 + 능력치 × 계수` + 하드 상한은 공속 능력치 40 / 이속 36 에서
     상한에 닿아 그 위로는 능력치가 **전혀 반영되지 않았다**. 능력치 상한이 100 인데
     40 부터 죽는 값이라 강화 13회쯤부터 두 능력치가 쓸모없어진다.
     새 공식은 상한에 도달하지 않고 계속 증가하므로 100 이든 200 이든 반영된다.

⚠️ 왜 행을 새로 끼워넣지 않는가 — '캐릭터 실효값'·'성장 시뮬레이션' 시트가
   `계수!$B$33` 처럼 **절대 행 번호**로 계수를 참조한다. openpyxl 은 행을 끼워넣어도
   수식을 따라 고쳐주지 않으므로 참조가 조용히 어긋난다. 그래서 쓰이지 않게 되는
   '공속 계수'·'이속 계수' 행을 **같은 자리에서 '반감점' 행으로 용도 변경**했다.

⚠️ 저장 후 반드시 Excel COM 으로 재계산한다 — openpyxl 로 저장하면 수식의 캐시값이
   날아가 `data_only=True` 로 읽을 때 전부 None 이 된다(진행상황 33-0절).
   이 환경에 LibreOffice 는 없고 Excel 은 있다.
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

XLSX = r"C:\Project\라스트 생츄어리\데이터 테이블\능력치 및 공식 정리.xlsx"

# 새 공속·이속 상수 (코드의 BalanceConfig.asset 과 반드시 같아야 한다)
APS_BASE, APS_LIMIT, APS_HALF = 0.6, 3.6, 50
MOV_BASE, MOV_LIMIT, MOV_HALF = 2.1, 6.0, 50

YELLOW = PatternFill("solid", fgColor="FFF2CC")


def find_row(ws, name, col="A"):
    """A열 이름으로 행 번호를 찾는다. 못 찾으면 바로 죽는다 — 조용히 엉뚱한 칸을 고치면 안 된다."""
    for r in range(1, ws.max_row + 1):
        v = ws[f"{col}{r}"].value
        if isinstance(v, str) and v.strip() == name:
            return r
    raise KeyError(f"'{name}' 행을 {ws.title} 시트에서 찾지 못했습니다")


def revise_coefficients(ws):
    """계수 시트 — 공속·이속 계수 행을 반감점으로 용도 변경하고 폴백 값을 덧붙인다."""
    r_aps_base = find_row(ws, "공속 기본(회/초)")
    r_aps_coef = find_row(ws, "공속 계수(회/초)")
    r_aps_max = find_row(ws, "공속 상한(회/초)")
    r_mov_base = find_row(ws, "이속 기본(타일/초)")
    r_mov_coef = find_row(ws, "이속 계수(타일/초)")
    r_mov_max = find_row(ws, "이속 상한(타일/초)")

    ws[f"A{r_aps_base - 1}"] = "■ 공격 속도  (→ 실수 유지, 2026-08-11 점근 공식으로 개정)"
    ws[f"B{r_aps_base}"] = APS_BASE
    ws[f"D{r_aps_base}"] = "공격속도 0 일 때 초당 공격 횟수"

    # 계수 → 반감점 (같은 행을 재사용한다. 위 주석 참조)
    ws[f"A{r_aps_coef}"] = "공속 반감점(능력치)"
    ws[f"B{r_aps_coef}"] = APS_HALF
    ws[f"C{r_aps_coef}"] = "attacksPerSecondHalfStat"
    ws[f"D{r_aps_coef}"] = ("이 능력치에서 기본과 한계의 정확히 중간이 된다. "
                            "작을수록 초반에 빨리 오르고 뒤가 완만해진다")
    ws[f"A{r_aps_max}"] = "공속 한계(회/초)"
    ws[f"B{r_aps_max}"] = APS_LIMIT
    ws[f"C{r_aps_max}"] = "attacksPerSecondMax"
    ws[f"D{r_aps_max}"] = ("점근 한계 — 능력치를 무한히 올려도 이 값에 닿지 않는다. "
                           "예전처럼 잘라내는 상한이 아니다")

    ws[f"A{r_mov_base - 1}"] = "■ 이동 속도  (→ 실수 유지, 2026-08-11 점근 공식으로 개정)"
    ws[f"B{r_mov_base}"] = MOV_BASE
    ws[f"D{r_mov_base}"] = "이동속도 0 일 때. 웨이브 몬스터는 2.2"

    ws[f"A{r_mov_coef}"] = "이속 반감점(능력치)"
    ws[f"B{r_mov_coef}"] = MOV_HALF
    ws[f"C{r_mov_coef}"] = "moveSpeedHalfStat"
    ws[f"D{r_mov_coef}"] = "이 능력치에서 기본과 한계의 정확히 중간이 된다"
    ws[f"A{r_mov_max}"] = "이속 한계(타일/초)"
    ws[f"B{r_mov_max}"] = MOV_LIMIT
    ws[f"C{r_mov_max}"] = "moveSpeedMax"
    ws[f"D{r_mov_max}"] = "점근 한계 — 닿지 않는다"

    # ---- 비어 있던 값 4개를 맨 아래에 새 절로 붙인다 (행 삽입 금지 원칙 때문에 append) ----
    r = ws.max_row + 2
    rows = [
        ("■ 생성 시 랜덤 범위 (캐릭터 테이블에 없는 인물)", None, None, None),
        ("생성 랜덤 최소", 1, "initialStatMin",
         "테이블에 정의되지 않은 캐릭터는 각 능력치를 이 범위에서 균등 랜덤으로 받는다"),
        ("생성 랜덤 최대", 10, "initialStatMax",
         "엘린·비기오르·프레이야는 이 롤을 쓰지 않고 테이블 고정값을 받는다"),
        ("■ 폴백 (공속·이속 능력치가 없는 유닛 — 몬스터 · 포탑 · 넥서스)", None, None, None),
        ("폴백 공속(회/초)", 1, "attacksPerSecond",
         "몬스터 정의(MonsterDefinitionSO)에 값이 없을 때만 쓰인다"),
        ("폴백 이속(타일/초)", 3, "moveSpeedTilesPerSecond",
         "같은 이유. 웨이브 몬스터는 정의에 2.2 가 들어있어 이 값을 안 쓴다"),
    ]
    for name, val, field, desc in rows:
        ws[f"A{r}"] = name
        if val is not None:
            ws[f"B{r}"] = val
            ws[f"B{r}"].fill = YELLOW
            ws[f"C{r}"] = field
            ws[f"D{r}"] = desc
        else:
            ws[f"A{r}"].font = Font(bold=True)
        r += 1

    return dict(aps_base=r_aps_base, aps_half=r_aps_coef, aps_max=r_aps_max,
                mov_base=r_mov_base, mov_half=r_mov_coef, mov_max=r_mov_max)


def revise_formula_sheet(ws):
    """'공식' 시트 — 공속·이속 설명을 새 식으로 바꾸고 빈 칸 두 개를 채운다."""
    r = find_row(ws, "초당 공격 횟수")
    ws[f"B{r}"] = ("공속기본 + (공속한계 − 공속기본) × 공격속도 ÷ (공격속도 + 공속반감점)\n"
                   f"= {APS_BASE} + {round(APS_LIMIT - APS_BASE, 2)} × 공격속도 ÷ (공격속도 + {APS_HALF})")
    ws[f"C{r}"] = ("엘린 공속 3\n"
                   f"= {APS_BASE} + {round(APS_LIMIT - APS_BASE, 2)} × 3 ÷ 53\n"
                   f"= {APS_BASE} + 0.170\n"
                   f"= {round(APS_BASE + (APS_LIMIT - APS_BASE) * 3 / 53, 3)}회/초")
    ws[f"E{r}"] = ("★ 상한으로 잘라내지 않는다 — 능력치가 100 이든 200 이든 계속 오르며 "
                   f"{APS_LIMIT}회/초에 닿기만 한다. 예전 식은 공속 40 에서 상한에 닿아 "
                   "그 위로는 능력치가 아무 일도 하지 않았다")

    r = find_row(ws, "초당 이동 타일")
    ws[f"B{r}"] = ("이속기본 + (이속한계 − 이속기본) × 이동속도 ÷ (이동속도 + 이속반감점)\n"
                   f"= {MOV_BASE} + {round(MOV_LIMIT - MOV_BASE, 2)} × 이동속도 ÷ (이동속도 + {MOV_HALF})")
    ws[f"C{r}"] = ("엘린 이속 9\n"
                   f"= {MOV_BASE} + {round(MOV_LIMIT - MOV_BASE, 2)} × 9 ÷ 59\n"
                   f"= {MOV_BASE} + 0.595\n"
                   f"= {round(MOV_BASE + (MOV_LIMIT - MOV_BASE) * 9 / 59, 3)}타일/초")
    ws[f"E{r}"] = ("공격 속도와 같은 이유. 웨이브 몬스터가 2.2타일/초라 그보다 빠른지가 중요합니다 "
                   "(이속 1 이면 2.18 로 몬스터보다 느리다 — 개정 전과 같다)")

    # 비어 있던 계산 예시
    r = find_row(ws, "피해 감소율(%)")
    ws[f"C{r}"] = ("비기오르 방어 8 (피해 배율 0.862)\n"
                   "= (1 − 0.862) × 100\n"
                   "= 13.8\n"
                   "→ 반올림 14%")

    # '=1.5 * 1.37' 로 시작해 Excel 이 수식으로 읽어 #NAME? 이 떴던 칸
    r = find_row(ws, "실제 침식 누적/초")
    ws[f"C{r}"] = ("엘린 저항 13 (상승 배율 1.37)\n"
                   "1.5 × 1.37\n"
                   "= 2.055/초")

    # 하단 주석 — 개정 이력을 남긴다
    r = ws.max_row + 1
    ws[f"A{r}"] = ("【2026-08-11 2차 개정 — 공속·이속】 예전에는 `기본 + 능력치 × 계수` 로 오르다가 "
                   "상한에서 뚝 잘렸습니다. 상한이 공속은 능력치 40, 이속은 36 에서 걸려 "
                   "능력치 상한(100)의 절반도 못 쓰고 죽는 값이었습니다. "
                   "이제 `기본 + (한계 − 기본) × 능력치 ÷ (능력치 + 반감점)` 으로 계산해 "
                   "능력치가 100을 넘어도 계속 반영되고, 한계값에는 닿지 않습니다. "
                   "기존 캐릭터 값은 거의 그대로입니다 (엘린 공속 0.78 → 0.77 / 이속 2.82 → 2.70).")
    ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")


def revise_effective_sheet(ws, k):
    """'캐릭터 실효값' — 공속·이속 행의 수식을 새 식으로 갈아끼운다."""
    for label, base, half, limit in [("초당 공격 횟수", k["aps_base"], k["aps_half"], k["aps_max"]),
                                     ("초당 이동 타일", k["mov_base"], k["mov_half"], k["mov_max"])]:
        r = find_row(ws, label)
        stat_row = find_row(ws, "공격 속도" if "공격" in label else "이동속도")
        for col in ("B", "C", "D"):
            s = f"{col}{stat_row}"          # 같은 시트 위쪽의 원시 능력치 칸
            ws[f"{col}{r}"] = (f"=계수!$B${base}+(계수!$B${limit}-계수!$B${base})"
                               f"*{s}/({s}+계수!$B${half})")
        ws[f"E{r}"] = ("기본 + (한계 − 기본) × 능력치 ÷ (능력치 + 반감점)")


def add_speed_curve_sheet(wb, k):
    """
    새 시트 '속도 곡선' — 개정한 두 공식이 능력치 0~200 에서 어떻게 움직이는지 보여준다.

    새 시트로 만든 이유: 기존 시트에 표를 끼워넣으면 절대 행 참조가 어긋난다(파일 상단 주석).
    """
    name = "속도 곡선"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = "공속·이속 곡선 — 능력치가 100을 넘어도 계속 오르는지 눈으로 확인하는 표"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = ("※ 개정 전 열은 옛 공식(기본 + 능력치 × 계수, 상한에서 절단)을 그대로 계산한 것입니다. "
                "회색으로 굳어지는 지점이 상한에 닿아 능력치가 죽는 구간입니다.")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    head = ["능력치", "공속 (개정 후)", "공속 (개정 전)", "이속 (개정 후)", "이속 (개정 전)", "비고"]
    for i, h in enumerate(head):
        c = ws.cell(row=4, column=1 + i, value=h)
        c.font = Font(bold=True)

    notes = {
        0: "능력치 0 = 기본값",
        3: "엘린 공속",
        5: "프레이야 이속",
        9: "엘린 이속",
        36: "← 옛 이속이 상한에 닿아 죽던 지점",
        40: "← 옛 공속이 상한에 닿아 죽던 지점",
        50: "반감점 — 기본과 한계의 정확히 중간",
        100: "능력치 상한. 개정 후는 여기서도 계속 오른다",
        200: "상한을 풀었을 때에도 한계에 닿지 않는다",
    }
    stats = [0, 1, 3, 5, 9, 10, 20, 30, 36, 40, 50, 60, 75, 100, 150, 200]
    r = 5
    for s in stats:
        ws.cell(row=r, column=1, value=s)
        ws.cell(row=r, column=2,
                value=f"=계수!$B${k['aps_base']}+(계수!$B${k['aps_max']}-계수!$B${k['aps_base']})"
                      f"*A{r}/(A{r}+계수!$B${k['aps_half']})")
        ws.cell(row=r, column=3, value=f"=MIN(3,0.6+A{r}*0.06)")
        ws.cell(row=r, column=4,
                value=f"=계수!$B${k['mov_base']}+(계수!$B${k['mov_max']}-계수!$B${k['mov_base']})"
                      f"*A{r}/(A{r}+계수!$B${k['mov_half']})")
        ws.cell(row=r, column=5, value=f"=MIN(5,2.1+A{r}*0.08)")
        if s in notes:
            ws.cell(row=r, column=6, value=notes[s])
        r += 1

    for col, w in zip("ABCDEF", (10, 16, 16, 16, 16, 46)):
        ws.column_dimensions[col].width = w


def revise_stat_sheet(ws):
    """'능력치' 시트 — 비어 있던 비고 3칸을 채우고 속도 2종의 비고를 개정 내용으로 바꾼다."""
    fills = {
        "원거리 공격력": "전술이 원거리일 때만 쓰인다 — 다른 전술이면 놀게 된다(미결 3번)",
        "마법": "전술이 마법일 때만. 엘린은 마법 8 이라 마법 전술에서 가장 강하다",
        "회복력": "전술이 회복일 때 아군을 살리는 양. 체력 재생(hp_recovery)과 다른 능력치다",
        "공격 속도": "실수 유지. 2026-08-11 점근 공식으로 개정 — 100을 넘어도 반영된다",
        "이동속도": "실수 유지. 2026-08-11 점근 공식으로 개정 — 100을 넘어도 반영된다",
    }
    for name, note in fills.items():
        r = find_row(ws, name, col="B")
        ws[f"I{r}"] = note


def revise_open_issues(ws):
    """'미결 사항' — 해소된 항목을 갱신하고 새로 생긴 것을 붙인다."""
    r = find_row(ws, "신규 계수는 제안값", col="B")
    ws[f"C{r}"] = ("명중(85+0.3) · 치명(0.8, 상한 60%) 은 기획 테이블에 근거가 없어 제안한 값입니다. "
                   "체력·공격·방어 곡선은 개정 전과 똑같이 두었습니다(밸런스가 흔들리지 않게). "
                   "공속·이속은 2026-08-11 점근 공식으로 다시 잡았습니다(아래 8번).")
    r = ws.max_row + 1
    for num, title, body, scope in [
        (8, "공속·이속 한계·반감점이 제안값",
         "점근 한계(공속 3.6회/초 · 이속 6타일/초)와 반감점(둘 다 능력치 50)은 "
         "기존 캐릭터 값이 거의 안 변하도록 역산한 제안값입니다. "
         "능력치 100 에서 공속 2.60회/초 · 이속 4.70타일/초가 됩니다.", "밸런스 전반"),
        (9, "능력치 상한 100 은 그대로",
         "공식은 100을 넘겨도 반영되지만 statMax 가 100 이라 강화로는 100 을 넘길 수 없습니다. "
         "상한을 올릴지는 별도 결정이 필요합니다.", "성장 밸런스"),
    ]:
        ws[f"A{r}"] = num
        ws[f"B{r}"] = title
        ws[f"C{r}"] = body
        ws[f"D{r}"] = scope
        ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def recalc_with_excel(path):
    """openpyxl 저장으로 날아간 수식 캐시값을 Excel COM 으로 되살린다(진행상황 33-0절)."""
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(path))
        wb.Application.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        excel.Quit()


def main():
    wb = openpyxl.load_workbook(XLSX)
    k = revise_coefficients(wb["계수"])
    revise_formula_sheet(wb["공식"])
    revise_effective_sheet(wb["캐릭터 실효값"], k)
    revise_stat_sheet(wb["능력치"])
    revise_open_issues(wb["미결 사항"])
    add_speed_curve_sheet(wb, k)
    wb.save(XLSX)
    print("openpyxl 저장 완료 — 계수 행:", k)
    recalc_with_excel(XLSX)
    print("Excel 재계산·저장 완료")


if __name__ == "__main__":
    main()
