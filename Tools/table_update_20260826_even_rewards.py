# -*- coding: utf-8 -*-
"""중립 보상을 <b>체력(수고)에 비례</b>시킨다 — «두더지가 아직도 너무 좋다» (2026-08-26, 4차).

유저 지시
---------
  *"좀 더 몬스터 보상이 고르게 들어가야 할거 같은데 아직도 두더지가 너무 좋은데"*

★★★ 값을 «손으로 고르는 것» 을 그만두고 <b>규칙</b>을 만들었다
──────────────────────────────────────────────────────────────────────
3차까지는 «구간 수입» 을 맞추려고 종별 값을 손으로 골랐다. 그래서 두더지 하나가
체력 14 짜리인데 마리당 340(중반 종의 7.4배)이라는 <b>설명할 수 없는 값</b>을 계속 들고 있었다.
이번에는 순서를 뒤집었다 — <b>먼저 «마리당 얼마가 옳은가» 를 규칙으로 정하고</b>, 그 결과가
구간 수입과 착지점을 만족하는지 확인했다.

    마리당 에너지 = <b>체력</b> × 단가 × <b>거리 계수</b>
      · 단가        일반 <b>2.5</b> E/체력 · 에픽 <b>4.0</b> E/체력(부대 하나가 붙는 «사건» 값)
      · 거리 계수   15~99 타일 ×1.0 · 100~199 ×1.3 · 200~320 ×1.8

    id    이름           체력    거리      전(3차)      후        검산
    1001  종양 거미        4    15~99        10        10      4×2.5×1.0 = 10
    1002  종양귀           8    100~199      46        26      8×2.5×1.3 = 26
    1004  고르도네        15    100~199      46        49     15×2.5×1.3 = 49
    1003  종양 두더지     14    200~320     340        63     14×2.5×1.8 = 63    ← <b>−81%</b>
    1101  카르키노스    1365    에픽       5500      5450   1365×4.0     = 5460
    1102  아니사킬      2035    에픽       5500      8150
    1103  바리올라      2715    에픽       5500     10850
    1104  폴리르        4024    에픽       7700     16100

★ <b>같은 거리인 1002 와 1004 가 갈린다</b> — 체력이 8 과 15 다. 3차까지는 둘을 같은 값으로
  묶어 뒀는데, 그것 자체가 «고르지 않은» 것이었다(고르도네가 두 배 단단한데 값이 같았다).

★★ <b>두더지는 «마리당» 을 내리고 «개체수» 를 올렸다</b>
──────────────────────────────────────────────────────────────────────
마리당만 내리면 두더지는 <b>가면 안 되는 곳</b>이 된다 — 초당 수급이 고르도네의 0.39배로
떨어져 «원정 갈 이유» 가 사라지고, 그러면 21~30 구간 수입 모델(웨이브당 두더지 10마리)이
<b>실제로 일어나지 않는 가정</b>이 된다. 그래서 서식 밀도를 같이 손봤다:

    1003  동시 최대 개체수 19 → <b>26</b> · 재생성 주기 25초 → <b>16초</b>

    지속 수급 E/s (= 동시 개체수 ÷ 재생성 ×  마리당)
      전(3차)   1001  48 · 1002  117 · 1004 117 · 1003 <b>258</b>   ← 두더지가 최고
      후        1001  48 · 1002   66 · 1004 124 · 1003 <b>102</b>   ← 고르도네 다음

    마리당 비교 (두더지 ÷ 중반 두 종 평균)   7.4배  →  <b>1.7배</b>
    체력이 비슷한 고르도네와 비교하면       7.4배  →  <b>1.3배</b>(거리 계수 1.8/1.3 만큼)

⚠ 개체수를 올려도 <b>수입 모델은 안 변한다</b> — 웨이브당 10마리는 처음부터 «공급» 이 아니라
  «원정 왕복에 드는 시간» 이 정한 수였다(19마리 ÷ 25초면 한 웨이브에 90마리가 지나간다).
  밀도는 «가서 잡을 만한 곳» 으로 만드는 몫이다.

★ 결과 — 착지점은 그대로, 곡선은 기획서 요구선에 더 가까워졌다
──────────────────────────────────────────────────────────────────────
    파티 평균 Lv        w5   w10   w15   w20   w25   w30
      기획서 요구선      5    10    13    18    23   (30)
      2차              6.0  11.2  15.0  25.9  28.5  35.2
      3차              5.0   9.8  12.1  20.6  25.2  35.0
      <b>4차(이번)</b>   5.0   9.8  11.6  <b>19.4</b>  <b>24.6</b>  34.9

    총 중립 수입 108,920 → 107,258 (−1.5%) — «w30 12명 Lv35» 착지점 유지
    구간 몫 1.0/1.9/5.1/11.5/26.0/55.3% → 0.9/1.8/4.1/9.0/25.7/58.4%

⚠ <b>후반 몫은 오히려 커졌다</b>(55.3 → 58.4%). 이것은 피할 수 없다 — Lv30 부터 강화 비용이
  등비 1.35 로 뛰므로 «w30 에 Lv35» 라는 착지점 자체가 <b>후반에 큰 수입을 요구한다</b>.
  달라진 것은 <b>그 몫을 누가 지느냐</b>다: 2차까지는 체력 14 짜리 두더지가 졌고, 이제는
  체력 1365~4024 · 방어력 30~40 · 스킬 둘셋을 가진 <b>에픽 넷</b>이 진다. 「밸런스 기획서」의
  «후반부는 … 에픽 몬스터의 보상을 적극적으로 활용할 것을 요구하는 구간» 이 그 문장이다.

⚠ 이 표들은 <b>사람이 엑셀에서 고쳐 나가는 문서</b>다 — 통째로 다시 굽지 않고 바꿀 칸만 고친다.
⚠ <b>Excel COM 으로 쓴다</b> — openpyxl 로 저장하면 하이퍼링크·주석·수식 캐시가 날아간다(136-4절).

사용법:  python Tools/table_update_20260826_even_rewards.py
다음:    python Tools/sync_tables_to_assets.py     (중립 에셋에 반영)
         Unity 에서 Assets/Refresh
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

NEUTRAL_XLSX = os.path.join(TABLE_DIR, "임시용 중립 몬스터.xlsx")
FORMULA_XLSX = os.path.join(TABLE_DIR, "능력치 및 공식 정리.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")

# ══════════════════════════════════════════════════════════════════════
#  ① 규칙 — 마리당 에너지 = 체력 × 단가 × 거리 계수
# ══════════════════════════════════════════════════════════════════════
UNIT_NORMAL = 2.5           # 일반 종 E/체력
UNIT_EPIC = 4.0             # 에픽 E/체력 (부대 하나가 붙는 «사건» 이라 1.6배)

HP = {1001: 4, 1002: 8, 1003: 14, 1004: 15,
      1101: 1365, 1102: 2035, 1103: 2715, 1104: 4024}

DIST = {1001: 1.0, 1002: 1.3, 1004: 1.3, 1003: 1.8}      # 등장 거리 계수

SPREAD = 0.2                # min/max = 평균 ±20% (표의 기존 관례)

# 반올림 자리 — 값이 클수록 굵게 끊는다(표를 사람이 읽는다)
def round_to(v):
    if v < 100:
        return int(round(v))
    if v < 1000:
        return int(round(v / 10.0) * 10)
    return int(round(v / 100.0) * 100)


def energy_of(mon_id):
    """(min, max, 평균) — 규칙에서 바로 나온다."""
    unit = UNIT_EPIC if mon_id >= 1100 else UNIT_NORMAL
    avg = HP[mon_id] * unit * DIST.get(mon_id, 1.0)
    lo = round_to(avg * (1.0 - SPREAD))
    hi = round_to(avg * (1.0 + SPREAD))
    return lo, hi, (lo + hi) / 2.0


NEW_ENERGY = {m: energy_of(m)[:2] for m in HP}

# ══════════════════════════════════════════════════════════════════════
#  ② 두더지 서식 밀도 — 마리당을 내린 만큼 «가서 잡을 만한 곳» 으로 만든다
#     mon_id → (max_alive, respawn_seconds)
# ══════════════════════════════════════════════════════════════════════
NEW_DENSITY = {
    1003: (26, 16),         # 19마리 / 25초 → 26마리 / 16초
}

# ══════════════════════════════════════════════════════════════════════
#  ③ 구간별 사냥 구성 — 2차·3차와 <b>같은 표</b> (두 벌이 되면 안 된다)
#     (구간 상한, (1001, 1002+1004, 1003, 에픽))
# ══════════════════════════════════════════════════════════════════════
MIX = [
    (5,  (20,  0,  0, 0.0)),
    (10, (39,  0,  0, 0.0)),
    (15, (20, 18,  0, 0.0)),
    (20, (20, 46,  0, 0.0)),
    (25, (20, 40,  3, 0.5)),
    (30, (20, 40, 10, 1.2)),
]

# ══════════════════════════════════════════════════════════════════════
#  ④ 성장 모델 — UI-67·2차·3차의 계수를 <b>그대로</b> 쓴다
# ══════════════════════════════════════════════════════════════════════
FOCUS_PER_LEVEL = 2.62
PLAIN_PER_LEVEL = 1.62
STAT_BASE_FOCUS = 10
STAT_BASE_PLAIN = 5
STAT_MAX = 100

HP_PER_STAT = 6.1832
HP_BASE = 90

DPS_A, DPS_B, DPS_C = 0.01961, 3.329, -9.50
EFFECTIVE = 0.647
BATTLE_SECONDS = 120

CREATE_BASE, CREATE_STEP = 200, 150
FREE_CHARACTERS = 4

UP_BASE, UP_STEP = 40, 10
UP_STEEP_LEVEL, UP_STEEP_RATE, UP_ROUND = 30, 1.35, 10

WAVE_KILL_ENERGY = 10

SHEET = "웨이브 부하"
FIRST_ROW = 7
COL = dict(wave=1, lv=2, people=3, focus=4, plain=5, hp=6,
           hp_scale=7, atk_scale=8, count=9, mob_hp=10, mob_hit=11,
           hits=12, hit_ratio=13, dps=14, load=15,
           boss=16, boss_raw=17, boss_hp=18, boss_goal=19, neutral=20)


def upgrade_cost(level):
    linear = UP_BASE + UP_STEP * level
    if level < UP_STEEP_LEVEL:
        return linear
    at_start = UP_BASE + UP_STEP * UP_STEEP_LEVEL
    cost = at_start * (UP_STEEP_RATE ** (level - UP_STEEP_LEVEL))
    return int(round(cost / UP_ROUND) * UP_ROUND)


def mix_for(wave):
    for upto, m in MIX:
        if wave <= upto:
            return m
    return MIX[-1][1]


def avg_of(mon_id):
    return energy_of(mon_id)[2]


def neutral_income(wave):
    n1, n2, n3, ne = mix_for(wave)
    mid = (avg_of(1002) + avg_of(1004)) / 2.0
    epic = (avg_of(1101) + avg_of(1102) + avg_of(1103)) / 3.0
    return n1 * avg_of(1001) + n2 * mid + n3 * avg_of(1003) + ne * epic


def run_model(rows, neutral):
    """UI-67 의 run_model 과 <b>같은 순서</b>다 — ① 충원 ② 최저 레벨부터 강화 ③ 기록 ④ 번다."""
    energy = 0.0
    levels = []
    created = 0
    out = []

    for r, nt in zip(rows, neutral):
        while len(levels) < r["people"]:
            if len(levels) < FREE_CHARACTERS:
                levels.append(0)
                continue
            cost = CREATE_BASE + CREATE_STEP * created
            if energy < cost:
                break
            energy -= cost
            created += 1
            levels.append(0)

        guard = 0
        while levels and guard < 500000:
            guard += 1
            i = min(range(len(levels)), key=lambda j: levels[j])
            c = upgrade_cost(levels[i])
            if energy < c:
                break
            energy -= c
            levels[i] += 1

        out.append(sum(levels) / len(levels) if levels else 0.0)

        energy += r["count"] * 2 * WAVE_KILL_ENERGY
        energy += nt

    return out


def read_rows():
    wb = openpyxl.load_workbook(FORMULA_XLSX, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=FIRST_ROW, max_row=ws.max_row, values_only=True):
        if r[0] is None:
            continue
        rows.append(dict(
            wave=int(r[0]), lv=float(r[1]), people=int(r[2]),
            focus=float(r[3]), plain=float(r[4]), hp=float(r[5]),
            count=int(r[8]), mob_hp=float(r[9]), mob_hit=float(r[10]),
            hits=float(r[11]), hit_ratio=float(r[12]), dps=float(r[13]),
            load=float(r[14]), neutral=float(r[19]),
        ))
    return rows


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_보상고르게")
    os.makedirs(folder, exist_ok=True)
    for f in (NEUTRAL_XLSX, FORMULA_XLSX):
        shutil.copy2(f, os.path.join(folder, os.path.basename(f)))
    print("백업: " + folder)
    return folder


def main():
    print("규칙 — 마리당 = 체력 × 단가 × 거리계수 "
          f"(일반 {UNIT_NORMAL} · 에픽 {UNIT_EPIC} E/체력)")
    alive = {1001: (43, 9), 1002: (33, 13), 1004: (33, 13), 1003: (19, 25)}
    alive.update(NEW_DENSITY)
    for m in sorted(HP):
        lo, hi, av = energy_of(m)
        line = f"  {m}  체력 {HP[m]:5d}  거리 ×{DIST.get(m, 1.0):.1f}  →  {lo:6d}~{hi:6d} (평균 {av:8.0f})"
        if m in alive:
            a, rs = alive[m]
            line += f"   지속 {a / rs * av:6.0f} E/s  ({a}마리 ÷ {rs}초)"
        print(line)

    rows = read_rows()
    neutral = [neutral_income(r["wave"]) for r in rows]
    old_total = sum(r["neutral"] for r in rows)
    new_total = sum(neutral)
    print(f"\n중립 총 수입 {old_total:,.0f} → {new_total:,.0f}  "
          f"({(new_total / old_total - 1) * 100:+.2f}%)")

    print("\n구간 몫")
    seen = set()
    for r, nt in zip(rows, neutral):
        band = next(u for u, _ in MIX if r["wave"] <= u)
        if band in seen:
            continue
        seen.add(band)
        print(f"  ~{band:2d}웨이브  {r['neutral']:7.0f} → {nt:7.0f} /웨이브   "
              f"몫 {r['neutral'] * 5 / old_total * 100:5.1f}% → {nt * 5 / new_total * 100:5.1f}%")

    new_lv = run_model(rows, neutral)

    plan = []
    for r, lv, nt in zip(rows, new_lv, neutral):
        n = round(lv)
        focus = min(STAT_MAX, STAT_BASE_FOCUS + FOCUS_PER_LEVEL * n)
        plain = min(STAT_MAX, STAT_BASE_PLAIN + PLAIN_PER_LEVEL * n)
        hp = round(HP_BASE + HP_PER_STAT * (focus - STAT_BASE_FOCUS))

        per = DPS_A * focus * focus + DPS_B * focus + DPS_C
        dps = round(r["people"] * per)

        total_mob_hp = r["count"] * 2 * r["mob_hp"]
        load = total_mob_hp / (dps * EFFECTIVE * BATTLE_SECONDS) if dps > 0 else 0.0

        ratio = r["hit_ratio"] * (r["hp"] / hp) if hp > 0 else 0.0
        hits = r["hits"] * (hp / r["hp"]) if r["hp"] > 0 else 0.0

        plan.append(dict(
            wave=r["wave"], lv=round(lv, 1), focus=round(focus, 1), plain=round(plain, 1),
            hp=hp, dps=dps, load=round(load, 2),
            ratio=round(ratio, 3), hits=round(hits, 1), neutral=round(nt),
        ))

    print("\n웨이브 | 평균Lv (전) | 캐릭터HP | 파티DPS | 잡몹부하 (전) | 중립수입 (전)")
    for p, r in zip(plan, rows):
        if p["wave"] % 5 == 0 or p["wave"] in (11, 16, 21, 26):
            print(f"{p['wave']:6d} | {p['lv']:5.1f} ({r['lv']:4.1f}) | {p['hp']:8d} | "
                  f"{p['dps']:7d} | {p['load']:5.2f} ({r['load']:4.2f}) | "
                  f"{p['neutral']:6d} ({int(r['neutral'])})")

    backup()
    write_excel(plan)


def write_excel(plan):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        # ── ① 중립 보상 + 두더지 밀도 ──────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(NEUTRAL_XLSX))
        ws = wb.Worksheets("neutrality_mon")
        touched = 0
        row = 4                      # 3행이 자료형 줄, 4행부터 자료
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            mon_id = int(v)
            if mon_id in NEW_ENERGY:
                lo, hi = NEW_ENERGY[mon_id]
                before = (ws.Cells(row, 10).Value, ws.Cells(row, 11).Value)
                ws.Cells(row, 10).Value = lo      # J min_energy
                ws.Cells(row, 11).Value = hi      # K max_energy
                print(f"  중립 {mon_id}: 에너지 {int(before[0])}~{int(before[1])} → {lo}~{hi}")
                touched += 1
            if mon_id in NEW_DENSITY:
                a, rs = NEW_DENSITY[mon_id]
                was = (ws.Cells(row, 15).Value, ws.Cells(row, 16).Value)
                ws.Cells(row, 15).Value = a       # O max_alive
                ws.Cells(row, 16).Value = rs      # P respawn_seconds
                print(f"  중립 {mon_id}: 밀도 {int(was[0])}마리/{int(was[1])}초 → {a}마리/{rs}초")
            row += 1
        wb.Save()
        wb.Close()
        print(f"임시용 중립 몬스터.xlsx — {touched}종 갱신")

        # ── ② 「웨이브 부하」 + 「계수」 ─────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(FORMULA_XLSX))
        ws = wb.Worksheets(SHEET)

        for i, p in enumerate(plan):
            r = FIRST_ROW + i
            assert int(ws.Cells(r, COL["wave"]).Value) == p["wave"], \
                f"{r}행이 웨이브 {p['wave']} 가 아니다"
            ws.Cells(r, COL["lv"]).Value = p["lv"]
            ws.Cells(r, COL["focus"]).Value = p["focus"]
            ws.Cells(r, COL["plain"]).Value = p["plain"]
            ws.Cells(r, COL["hp"]).Value = p["hp"]
            ws.Cells(r, COL["hits"]).Value = p["hits"]
            ws.Cells(r, COL["hit_ratio"]).Value = p["ratio"]
            ws.Cells(r, COL["dps"]).Value = p["dps"]
            ws.Cells(r, COL["load"]).Value = p["load"]
            ws.Cells(r, COL["neutral"]).Value = p["neutral"]

        note = ("※ 2026-08-26 개정 4차(유저 지시 — 보상이 더 고르게 · 두더지가 아직도 좋다): "
                "마리당 에너지를 «손으로 고른 값» 에서 «규칙» 으로 바꿨다 — 마리당 = 체력 × 단가 × 거리계수 "
                "(단가 일반 2.5 · 에픽 4.0 E/체력 · 거리 ×1.0/×1.3/×1.8). "
                "그래서 1001 10 · 1002 26 · 1004 49 · 1003 63 · 에픽 5450/8150/10850/16100 이 된다. "
                "★ 종양 두더지는 340 → 63(−81%). 체력 14 로 고르도네(15)와 같은 급이니 값도 같은 급이어야 한다. "
                "대신 동시 개체수 19→26 · 재생성 25→16초 로 밀도를 올려 «원정 갈 값» 을 지속 수급으로 준다 "
                "(E/s: 1001 48 · 1002 66 · 1004 124 · 1003 102). "
                "총 중립 수입 107,258(-1.5%) · 착지점 «w30 12명 Lv35» 유지 · 파티 Lv 5.0/9.8/11.6/19.4/24.6/34.9 "
                "(기획서 요구선 5/10/13/18/23 에 3차보다 가깝다). "
                "⚠ 후반 몫은 55.3→58.4% 로 오히려 커졌다 — Lv30 부터 등비 1.35 강화벽이 있어 «w30 Lv35» 착지점 자체가 "
                "후반 수입을 요구한다. 달라진 것은 그 몫을 체력 14 짜리 두더지가 지지 않고 에픽 넷이 진다는 것이다. "
                "⚠ 몬스터 스펙(체력·공격력)은 한 칸도 고치지 않았다.")
        ws.Cells(5, 1).Value = note

        coef = wb.Worksheets("계수")
        r = coef.UsedRange.Rows.Count + 1
        rows_out = [
            ("■ 2026-08-26 (4차) — 보상을 «체력 비례» 규칙으로", "", "", ""),
            ("중립 보상 규칙", "체력 × 단가 × 거리계수", "임시용 중립 몬스터.min/max_energy",
             "값을 손으로 고르는 것을 그만뒀다 — 두더지 340(체력 14) 처럼 설명할 수 없는 값이 생기던 자리다"),
            ("단가 — 일반 / 에픽", "2.5 / 4.0 E/체력", "위 규칙의 계수",
             "에픽은 부대 하나가 붙는 «사건» 이라 1.6배. 방어력 30~40 · 스킬 둘셋이 그 근거다"),
            ("거리 계수 — 15~99 / 100~199 / 200~320 타일", "1.0 / 1.3 / 1.8", "위 규칙의 계수",
             "원정 왕복 시간의 값. 두더지의 «먼 곳» 프리미엄은 이 1.8 뿐이다(전에는 사실상 7.4배였다)"),
        ]
        for m in (1001, 1002, 1004, 1003, 1101, 1102, 1103, 1104):
            lo, hi, av = energy_of(m)
            rows_out.append((f"{m} 마리당 평균 보상", int(av),
                             f"임시용 중립 몬스터 {lo}~{hi}",
                             f"체력 {HP[m]} × {UNIT_EPIC if m >= 1100 else UNIT_NORMAL} × {DIST.get(m, 1.0)}"))
        rows_out += [
            ("종양 두더지 1003 서식 밀도", "26마리 / 16초", "임시용 중립 몬스터 max_alive·respawn_seconds",
             "마리당을 −81% 했으니 밀도로 되돌려 준다 — 안 그러면 원정 갈 이유가 없어져 수입 모델이 거짓이 된다"),
            ("구간 몫 (1~5 / 6~10 / 11~15 / 16~20 / 21~25 / 26~30)",
             "0.9 / 1.8 / 4.1 / 9.0 / 25.7 / 58.4 %", "위 값들에서 파생",
             "3차 1.0/1.9/5.1/11.5/26.0/55.3% — 후반 몫을 이제 에픽 넷이 진다"),
            ("파티 평균 Lv (w15 / w20 / w25 / w30)", "11.6 / 19.4 / 24.6 / 34.9",
             "성장 모델(Tools/table_update_20260826_even_rewards.py)",
             "기획서 요구선 13/18/23/(30) · 착지점 Lv35 유지 · 3차 12.1/20.6/25.2/35.0"),
        ]
        for label, value, where, why in rows_out:
            coef.Cells(r, 1).Value = label
            if value != "":
                coef.Cells(r, 2).Value = value
            if where:
                coef.Cells(r, 3).Value = where
            if why:
                coef.Cells(r, 4).Value = why
            r += 1

        wb.Save()
        wb.Close()
        print(f"능력치 및 공식 정리.xlsx — 「웨이브 부하」 9열 × 30행 · 「계수」 {len(rows_out)}행 추가")
    finally:
        excel.Quit()


if __name__ == "__main__":
    main()
