# -*- coding: utf-8 -*-
"""전투 연출의 <b>「빗나감」</b> 한 칸을 스트링 키 테이블에 넣는다
(2026-08-27 · 184절 · 유저 리포트: *"빗나감도 영어 번역 넣어야함"*).

■ 왜 여기만 빠졌나
  `DamageNumberFx` 는 178-5절에 이미 <c>LocalizeLabels</c> 를 갖췄고 「영웅 각성!」 둘은
  그때 표로 옮겨졌다. 그런데 「빗나감」은 <b>인스펙터 칸이 아니라 호출문에 박힌 리터럴</b>
  (<c>Show(target, "빗나감", …)</c>)이었다 — 그 절이 훑은 것은 <c>[SerializeField] string</c>
  이라 <b>이 한 줄만</b> 그물을 빠져나갔다.
  → 코드 쪽은 인스펙터 칸(<c>missText</c>)으로 바꿔 두었다. 여기서는 <b>표에 키를 넣는다</b>.

■ ⚠ 왜 `table_update_20260827_log_lines.py` 로 못 넣나
  그 스크립트는 <c>extract_string_fallbacks.py</c> 로 <b>코드에서 kr 을 읽어</b> 넣는데,
  이 자리는 폴백이 <b>변수</b>(<c>HudTheme.T("ui_fx_miss", missText)</c>)라 읽어낼 수 없다.
  그쪽의 `KR_EXPLICIT` 과 같은 처지다 — 그래서 kr 을 <b>여기 손으로</b> 적는다.
  ⚠ 아래 kr 은 <c>DamageNumberFx.missText</c> 의 기본값과 <b>한 글자도 다르면 안 된다</b>.

■ 다음
    py -3 Tools/gen_string_table.py   →   py -3 Tools/link_string_keys.py
    py -3 Tools/check_string_keys.py --strict
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET = 'string'
DATA_ROW0 = 4
SOURCE = '연출 문구 이관(2026-08-27)'

#: 키 → (kr, en, 비고). kr 은 코드의 폴백과 <b>같아야</b> 한다.
ROWS = {
    # ★ 「MISS」 가 아니라 「Miss」 로 적는다 — 이 프로젝트의 연출 글자는 전부 문장식이고
    #   («영웅 각성!» → «Hero Awakening!»), 대문자만 쓰면 그 하나만 튄다.
    'ui_fx_miss': ('빗나감', 'Miss', '빗나갔을 때 머리 위에 뜨는 글자'),

    # ── 성역(중앙 건물)의 초상화 (유저 리포트: *"중앙건물도 번역 안됨 — 초상화 UI"*)
    #  ⚠ <b>이름은 새 키를 만들지 않았다</b> — 「건설」 표의 const_name_10001(중앙건물/Nexus)을
    #    그대로 쓴다. 여기 있는 것은 <b>칭호</b> 한 줄뿐이다.
    'ui_nexus_title': ('마지막 성역', 'The Last Sanctuary', '성역 초상화의 칭호 줄'),
}


def main():
    if not os.path.isfile(STRING_XLSX):
        raise SystemExit('⚠ 스트링 키 테이블이 없습니다: %s' % STRING_XLSX)

    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    where, last_row = {}, DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k is None or str(k).strip() == '':
            continue
        where[str(k).strip()] = r
        last_row = r

    added = []
    for key, (kr, en, note) in ROWS.items():
        if key in where:
            # ⚠ 덮어쓰지 않는다 — 178-6절의 «표 전체를 훑는 치환» 사고와 같은 규칙이다.
            print('  · 이미 있어 건너뜁니다: %s' % key)
            continue
        last_row += 1
        ws.cell(last_row, 1).value = key
        ws.cell(last_row, 2).value = kr
        ws.cell(last_row, 3).value = en
        ws.cell(last_row, 4).value = SOURCE
        ws.cell(last_row, 5).value = note
        added.append(key)

    if not added:
        print('바뀐 것이 없어 저장하지 않았습니다.')
        return 0

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    wb.save(STRING_XLSX)
    print('저장: %s (백업 .bak) · %d키 추가 — %s'
          % (os.path.basename(STRING_XLSX), len(added), ' · '.join(added)))
    print('다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')
    return 0


if __name__ == '__main__':
    sys.exit(main())
