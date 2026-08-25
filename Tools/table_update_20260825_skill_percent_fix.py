# -*- coding: utf-8 -*-
"""정의문 두 곳에 <b>빠져 있던 `%`</b> 를 넣는다 (2026-08-25).

> 유저 확인: *"정의문 두곳에 % 넣어줘 % 맞음"*

150절이 «단위를 안 밝힌 두 곳» 으로 올려 둔 것을 유저가 확정한 것이다:

  · 80030 「완성되지 못한 고귀함」  아르세니아 마법 * {value_04}      (350)
  · 80039 「종말의 선언」          원거리 공격력 * {value_03}         (110)

★ <b>«* {value}%» 관례를 그대로 따른다</b> — 같은 표의 80029 「성스러운 축복」이 이미
  «아르세니아의 마법 * {value_03}% 의 데미지» 로 적혀 있다. 새 표기를 만들지 않고
  <b>이 표가 이미 쓰는 모양</b>에 맞춘다.

⚠⚠ <b>80039 에는 `*` 가 둘이다.</b> 앞의 것은 «{value_01}(가로) * {value_02}(세로)» 로
   <b>곱셈이 아니라 크기 표기</b>다. 거기에 `%` 를 붙이면 «가로 6% x 세로 2%» 가 된다.
   그래서 <b>문자열 전체를 바꾸지 않고</b> 아래처럼 <b>정확한 조각만</b> 갈아 끼운다.

★ 150절이 «×» 로 남겨 둔 <b>상세 설명 두 줄도 같이</b> 고친다 — 정의문만 고치면
  두 문장이 서로 다른 말을 하게 된다(같은 스킬인데 한쪽은 배수, 한쪽은 %).

사용법:  py -3 Tools/table_update_20260825_skill_percent_fix.py
다음:    py -3 Tools/gen_string_table.py
         py -3 Tools/link_string_keys.py
"""

import io
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, "스트링 키 테이블.xlsx")
SHEET = "string"
ROW0 = 4

#: 키 → (찾을 조각, 바꿀 조각). ⚠ <b>조각만</b> 바꾼다 — 위 ⚠⚠ 참조.
EDITS = {
    # ── 정의문 (Skill_Type.desc) — 유저가 확정한 두 곳 ──
    "skill_type_desc_Unfinished_nobility": (
        "아르세니아 마법 * {value_04} 데미지를",
        "아르세니아 마법 * {value_04}% 데미지를",
    ),
    "skill_type_desc_Declaration_of_the_End": (
        "원거리 공격력 * {value_03}의 피해를",
        "원거리 공격력 * {value_03}%의 피해를",
    ),

    # ── 상세 설명 — 150절이 «×» 로 남겨 둔 것을 같은 뜻으로 맞춘다 ──
    "skill_detail_80030": (
        "마법 × {value_04}의 피해를",
        "마법의 {value_04}% 피해를",
    ),
    "skill_detail_80039": (
        "원거리 공격력 × {value_03}의 피해를",
        "원거리 공격력의 {value_03}% 피해를",
    ),
}


def main():
    if not os.path.exists(STRING_XLSX):
        raise SystemExit(f"[실패] 스트링 키 테이블을 찾지 못했습니다: {STRING_XLSX}")

    shutil.copy2(STRING_XLSX, STRING_XLSX + ".bak")
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    done, missing = [], list(EDITS)
    for r in range(ROW0, ws.max_row + 1):
        key = str(ws.cell(r, 1).value or "").strip()
        if key not in EDITS:
            continue
        missing.remove(key)

        old_piece, new_piece = EDITS[key]
        text = str(ws.cell(r, 2).value or "")

        if new_piece in text:
            print(f"   = {key} — 이미 되어 있습니다")
            continue
        if old_piece not in text:
            # ⚠ 조용히 넘기지 않는다 — 표가 바뀌었는데 못 찾은 것일 수 있다.
            print(f"   ✗ {key} — 바꿀 조각을 찾지 못했습니다:\n       «{old_piece}»")
            continue

        # ★ count=1 — 같은 조각이 여러 번 나오면 <b>첫 번째만</b>. 지금은 하나뿐이다.
        ws.cell(r, 2).value = text.replace(old_piece, new_piece, 1)
        done.append(key)
        print(f"   ✓ {key}")
        print(f"       {old_piece}  →  {new_piece}")

    try:
        wb.save(STRING_XLSX)
    except PermissionError:
        raise SystemExit(f"[실패] 엑셀에서 열려 있습니다 — 닫고 다시 돌리십시오:\n  {STRING_XLSX}")

    print(f"\n[% 보정] 바꾼 칸 {len(done)}개")
    if missing:
        print(f"⚠ 스트링 키 테이블에 없는 키: {', '.join(missing)}")


if __name__ == "__main__":
    main()
