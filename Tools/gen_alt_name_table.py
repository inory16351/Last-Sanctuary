# -*- coding: utf-8 -*-
"""「대체 이름 테이블.xlsx」를 만들고 스트링 키 테이블에 병합한다 (2026-08-26 · 유저 지시:
*"같은 캐릭터가 두번째로 등장할때는 랜덤한 다른 이름을 가지고 태어나게 해 … 랜덤하게 배정할
다른이름은 테이블 따로 만들어서 관리해 주고, 스트링 키에도 영어 / 한국어 이름 추가해줘"*).

■ 왜 표를 따로 두나
  대체 이름은 <b>인물 정의와 짝이 없는 «이름 주머니»</b> 다. 캐릭터 테이블에 칸을 만들면
  «누구의 두 번째 이름» 처럼 보이지만, 실제로는 <b>누구에게든 갈 수 있는 이름</b>이다.
  그래서 표를 따로 두고, 코드는 <b>스트링 키</b>(`character_altname_N`)만 본다.

■ 코드가 «몇 개인지» 를 모른다
  `CharacterAltNames` 는 `character_altname_1` 부터 <b>빈 번호가 나올 때까지</b> 훑는다.
  그러니 이 표에 줄을 더하고 이 스크립트를 다시 돌리면 <b>코드를 고치지 않아도</b> 늘어난다.
  ⚠ 그래서 <b>id 는 1 부터 «구멍 없이»</b> 이어야 한다 — 중간이 비면 거기서 목록이 끊긴다.
     이 스크립트가 그 검사를 한다.

■ 이 스크립트가 하는 일
  ① 표가 없으면 <b>새로 만든다</b>(아래 NAMES 로). 이미 있으면 <b>표를 정본으로 읽는다</b> —
     사람이 다듬은 이름을 덮지 않는다(`--force` 로만 다시 만든다 · .bak 을 남긴다).
  ② 표의 (id, 한국어, 영어)를 스트링 키 테이블에 <b>없는 키만</b> 덧붙인다.
  ③ id 가 1..N 연속인지 검사하고, 아니면 실패로 끝낸다.

■ 성별 (2026-08-27 · 유저 지시: *"남캐는 남자 이름 여캐는 여자이름으로"*)
  표에 <b>gender</b> 칸(male/female)이 있고, 이 스크립트가 그것만 따로
  <b>Assets/_Project/Resources/Data/AltNameGender.txt</b>(TSV)로 내보낸다.
  ⚠ <b>스트링 키 테이블에 넣지 않는 이유</b> — 성별은 화면에 나가는 «문구» 가 아니다.
    번역할 것이 없는 값을 스트링 표에 넣으면 «영어 빈칸» 검사에 매번 걸리고,
    죽은 키인지 산 키인지도 구분이 안 된다(182-5절에서 겪은 종류의 함정).
  ★ 그래서 <b>이 스크립트 하나가</b> 표에서 두 곳으로 내보낸다 —
    문구는 스트링 표로, 성별은 Unity 로. 원본이 하나라 어긋날 수가 없다.

■ 다음
    py -3 Tools/gen_string_table.py   →  py -3 Tools/link_string_keys.py
"""
import os
import shutil
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, '대체 이름 테이블.xlsx')
STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')

# ★ 스크립트 위치에서 역산한다 — 프로젝트 경로를 박아 두면 PC 마다 어긋난다
#   (gen_character_assets.py 의 ⚠ 주석과 같은 사고를 막는 장치).
PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR_UNITY = os.path.join(PROJECT_DIR, 'Assets', '_Project', 'Resources', 'Data')
OUT_TSV = os.path.join(OUT_DIR_UNITY, 'AltNameGender.txt')

SHEET = 'AltName'
INFO = '읽기'
KEY_PREFIX = 'character_altname_'
SOURCE = '대체 이름 표'

HEADER_KR = ['대체이름id', '한국어', '영어', '성별', '비고']
HEADER_FIELD = ['alt_name_id', 'kr', 'en', 'gender', 'note']
HEADER_TYPE = ['int', 'string', 'string', 'string', '-']

#: 표에 적는 성별 글자. C# 쪽 CharacterGender enum 과 짝이고
#: CharacterGenderText.Parse 가 같은 변환을 한다.
GENDERS = ('male', 'female')
DATA_ROW0 = 4          # 스트링 키 테이블과 같은 3행 헤더 규약

FONT = 'Arial'

#: 처음 표를 만들 때 넣는 이름들. 기존 인물 이름의 표기 규칙을 따랐다
#: (한국어는 소리대로 · 영어는 라틴 표기 — 엘린/Elin · 시카리아/Sicaria · 카이론/Chiron).
NAMES = [
    ('아드리엘', 'Adriel', 'male'),    ('노아네', 'Noane', 'female'),
    ('레이린', 'Reilin', 'female'),    ('미르카', 'Mirka', 'female'),
    ('세이하', 'Seiha', 'female'),     ('오르넬', 'Ornel', 'male'),
    ('유리안', 'Yurian', 'male'),      ('이레아', 'Irea', 'female'),
    ('카일런', 'Kailen', 'male'),      ('타비아', 'Tavia', 'female'),
    ('페릴', 'Peril', 'male'),         ('하르윈', 'Harwin', 'male'),
    ('가리엘', 'Gariel', 'male'),      ('나비스', 'Navis', 'male'),
    ('데인', 'Dane', 'male'),          ('라비니아', 'Lavinia', 'female'),
    ('마르첼', 'Marcel', 'male'),      ('바이런', 'Byron', 'male'),
    ('사비네', 'Sabine', 'female'),    ('아셀', 'Asel', 'female'),
    ('에리온', 'Erion', 'male'),       ('오필리아', 'Ophelia', 'female'),
    ('율리아', 'Julia', 'female'),     ('제피르', 'Zephyr', 'male'),
    ('카린', 'Karin', 'female'),       ('테오도르', 'Theodor', 'male'),
    ('파비안', 'Fabian', 'male'),      ('할리아', 'Halia', 'female'),
    ('그웬', 'Gwen', 'female'),        ('니콜라', 'Nicola', 'female'),
    ('도리안', 'Dorian', 'male'),      ('류시아', 'Lucia', 'female'),
    ('메이런', 'Meiren', 'female'),    ('베르타', 'Berta', 'female'),
    ('솔레인', 'Solein', 'female'),    ('아르덴', 'Arden', 'male'),
    ('엘로이', 'Eloi', 'male'),        ('오데트', 'Odette', 'female'),
    ('이자벨', 'Isabel', 'female'),    ('카시엘', 'Casiel', 'male'),
]

INFO_LINES = [
    ('■ 이 표는 무엇인가', True),
    ('같은 인물이 <이번 판에 두 번째로> 등장할 때 받는 «다른 이름» 주머니다.', False),
    ('죽은 인물이 다시 생성되면 다른 이름으로 태어나 «다른 사람» 처럼 보인다(엔딩 명단도 그 이름).', False),
    ('', False),
    ('■ 규칙', True),
    ('· id 는 1 부터 «구멍 없이» 이어야 한다 — 코드가 1 부터 훑다가 빈 번호에서 멈춘다.', False),
    ('· 한 판에서 같은 이름은 두 번 쓰이지 않는다. 이름이 다 떨어지면 원래 이름으로 태어난다.', False),
    ('· 화면에 나가는 문구는 <스트링 키 테이블>의 character_altname_<id> 다(한국어·영어).', False),
    ('· 성별(gender)은 male / female 둘 중 하나여야 한다 — 비면 그 이름은 아무에게도 안 간다.', False),
    ('· 인물의 성별은 <캐릭터 테이블>의 gender 칸이다. 같은 성별끼리만 이름이 배정된다.', False),
    ('', False),
    ('■ 이름을 더하거나 고치는 방법', True),
    ('1) 이 시트에 줄을 더한다(id 는 다음 번호 · 성별을 반드시 적는다).', False),
    ('2) py -3 Tools/gen_alt_name_table.py   → 스트링 키 테이블에 새 키가 붙고,', False),
    ('                                          성별이 Resources/Data/AltNameGender.txt 로 나간다', False),
    ('3) py -3 Tools/gen_string_table.py     → StringTable.txt 로 내보낸다', False),
    ('4) py -3 Tools/link_string_keys.py     → 하이퍼링크 재생성', False),
    ('★ 코드는 «몇 개인지» 를 모른다 — 표가 정본이라 코드를 고칠 필요가 없다.', False),
    ('', False),
    ('■ 표기 규칙', True),
    ('한국어는 소리대로, 영어는 라틴 표기. 기존 인물 이름과 같은 결로 짓는다', False),
    ('(엘린/Elin · 비기오르/Bigior · 시카리아/Sicaria · 카이론/Chiron · 아르세니아/Arsenia).', False),
]


def style_header(ws, headers):
    fill = PatternFill('solid', fgColor='DDEBF7')
    for row_idx, row in enumerate(headers, start=1):
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = value
            c.font = Font(name=FONT, bold=row_idx == 1)
            c.fill = fill
            c.alignment = Alignment(vertical='center')
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.freeze_panes = 'A4'


def create():
    wb = openpyxl.Workbook()
    info = wb.active
    info.title = INFO
    info.column_dimensions['A'].width = 104
    for i, (text, bold) in enumerate(INFO_LINES, start=1):
        c = info.cell(row=i, column=1)
        c.value = text
        c.font = Font(name=FONT, bold=bold)
        c.alignment = Alignment(vertical='center', wrap_text=True)

    ws = wb.create_sheet(SHEET)
    style_header(ws, [HEADER_KR, HEADER_FIELD, HEADER_TYPE])
    for i, (kr, en, gender) in enumerate(NAMES, start=1):
        r = DATA_ROW0 + i - 1
        ws.cell(row=r, column=1).value = i
        ws.cell(row=r, column=2).value = kr
        ws.cell(row=r, column=3).value = en
        ws.cell(row=r, column=4).value = gender
        for col in range(1, len(HEADER_FIELD) + 1):
            ws.cell(row=r, column=col).font = Font(name=FONT)
    wb.save(XLSX)
    print('  표 생성:', os.path.basename(XLSX), '· 이름 %d개' % len(NAMES))


def read():
    """표를 읽는다 — 칸은 <b>위치가 아니라 2행의 필드명</b>으로 찾는다.
    2026-08-27 에 성별 칸을 «비고» 앞에 끼워 넣었는데, 위치로 읽었다면 그때
    비고를 성별로 읽고도 «생성 완료» 만 찍혔을 것이다
    (gen_character_assets.py 가 2026-08-13 에 실제로 겪은 사고다)."""
    ws = openpyxl.load_workbook(XLSX, data_only=True)[SHEET]

    col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is not None and str(v).strip():
            col[str(v).strip()] = c
    for need in ('alt_name_id', 'kr', 'en'):
        if need not in col:
            sys.exit('표의 2행에 «%s» 칸이 없습니다 — 헤더가 깨졌습니다.' % need)

    def cell(r, field):
        c = col.get(field)
        return ws.cell(row=r, column=c).value if c else None

    rows = []
    for r in range(DATA_ROW0, ws.max_row + 1):
        raw = cell(r, 'alt_name_id')
        if raw is None or str(raw).strip() == '':
            continue
        try:
            alt_id = int(raw)
        except (TypeError, ValueError):
            sys.exit('id 가 숫자가 아닙니다: %r (행 %d)' % (raw, r))
        kr = (cell(r, 'kr') or '')
        en = (cell(r, 'en') or '')
        # ★ 성별 칸이 아예 없는 옛 표도 읽힌다 — 그때는 빈 문자열이고, 아래 검사가 잡는다.
        gender = str(cell(r, 'gender') or '').strip().lower()
        rows.append((alt_id, str(kr).strip(), str(en).strip(), gender))
    return rows


def merge_into_string_table(rows):
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb['string']

    keys = {}
    last_row = DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        keys[str(k).strip()] = r
        last_row = r

    added = []
    for alt_id, kr, en, _gender in rows:
        key = KEY_PREFIX + str(alt_id)
        if key in keys:
            continue
        last_row += 1
        ws.cell(row=last_row, column=1).value = key
        ws.cell(row=last_row, column=2).value = kr
        ws.cell(row=last_row, column=3).value = en
        ws.cell(row=last_row, column=4).value = SOURCE
        ws.cell(row=last_row, column=5).value = '두 번째 등장 인물의 이름 주머니'
        keys[key] = last_row
        added.append((key, kr, en))

    if added:
        shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
        wb.save(STRING_XLSX)
    return added


def write_gender_tsv(rows):
    """성별만 따로 Unity 로 내보낸다 — <c>Resources/Data/AltNameGender.txt</c>.

    형식·이유는 스트링 테이블과 같다(TSV · <c>#</c> 주석 · 헤더 한 줄 ·
    확장자가 .txt 인 것은 Unity 가 .tsv 를 TextAsset 으로 안 읽기 때문).
    읽는 쪽은 <c>Units/CharacterAltNames.EnsureGenders</c> 다.
    """
    os.makedirs(OUT_DIR_UNITY, exist_ok=True)

    lines = ['# Last Sanctuary 대체 이름 성별 — Tools/gen_alt_name_table.py 가 생성한다. 직접 고치지 말 것.',
             '# 원본: 데이터 테이블/대체 이름 테이블.xlsx (AltName 시트의 gender 칸)',
             '# 문구(한국어·영어)는 여기 없다 — 스트링 표의 character_altname_<id> 가 정본이다.',
             'alt_name_key\tgender']
    for alt_id, _kr, _en, gender in sorted(rows, key=lambda r: r[0]):
        lines.append('%s%d\t%s' % (KEY_PREFIX, alt_id, gender))

    with open(OUT_TSV, 'w', encoding='utf-8', newline='\n') as f:
        f.write('\n'.join(lines) + '\n')

    write_meta(OUT_TSV)
    return OUT_TSV


def write_meta(path):
    """.meta 가 없으면 만든다. guid 는 경로에서 결정적으로 뽑아 재실행에도 같게 한다
    (gen_string_table.py 와 같은 방식). 이미 있으면 건드리지 않는다."""
    import hashlib
    meta = path + '.meta'
    if os.path.exists(meta):
        return
    rel = os.path.relpath(path, PROJECT_DIR).replace('\\', '/')
    guid = hashlib.md5(rel.encode('utf-8')).hexdigest()
    with open(meta, 'w', encoding='utf-8', newline='\n') as f:
        f.write('fileFormatVersion: 2\n')
        f.write('guid: %s\n' % guid)
        f.write('TextScriptImporter:\n')
        f.write('  externalObjects: {}\n')
        f.write('  userData: \n')
        f.write('  assetBundleName: \n')
        f.write('  assetBundleVariant: \n')


def main():
    force = '--force' in sys.argv

    if not os.path.exists(XLSX):
        create()
    elif force:
        shutil.copyfile(XLSX, XLSX + '.bak')
        create()
        print('  ⚠ --force 로 다시 만들었습니다(.bak 남김)')
    else:
        print('  표가 이미 있습니다 — 표를 정본으로 읽습니다(다시 만들려면 --force).')

    rows = read()
    if not rows:
        sys.exit('표에 이름이 없습니다.')

    # ★ id 가 1..N 연속인가 — 구멍이 있으면 코드가 거기서 목록을 끊는다.
    ids = sorted(r[0] for r in rows)
    expected = list(range(1, len(ids) + 1))
    if ids != expected:
        missing = [i for i in expected if i not in ids]
        sys.exit('✗ id 가 1 부터 연속이 아닙니다 — 코드가 빈 번호에서 멈춥니다.\n'
                 '  빠진 번호: %s' % (missing[:20] or '(중복이 있습니다)'))

    empty = [r[0] for r in rows if not r[1] or not r[2]]
    if empty:
        sys.exit('✗ 한국어/영어가 빈 줄이 있습니다: id %s' % empty[:20])

    # ★★ 성별 (2026-08-27) — 비었거나 오타면 <b>그 이름은 아무에게도 안 간다</b>.
    #   런타임은 «모르는 성별» 을 조용히 건너뛰므로(경고는 남긴다), 여기서 미리 막는다.
    wrong = [(r[0], r[3]) for r in rows if r[3] not in GENDERS]
    if wrong:
        sys.exit('✗ 성별이 %s 가 아닌 줄이 있습니다 (id, 값): %s\n'
                 '  표의 gender 칸을 채우세요 — 비면 그 이름은 뽑히지 않습니다.'
                 % (' / '.join(GENDERS), wrong[:20]))

    added = merge_into_string_table(rows)
    tsv = write_gender_tsv(rows)

    counts = {g: sum(1 for r in rows if r[3] == g) for g in GENDERS}
    print('[대체 이름] 표 %d줄 (%s) · 스트링 표에 덧붙인 키 %d개'
          % (len(rows), ' · '.join('%s %d' % (g, counts[g]) for g in GENDERS), len(added)))
    for k, kr, en in added[:60]:
        print('    +', k, '|', kr, '|', en)
    if not added:
        print('    (스트링 표가 이미 표와 같습니다)')
    print('  성별 ->', tsv)
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
