# -*- coding: utf-8 -*-
"""언어 고르기 창의 <b>제목 키</b>를 표에 넣는다 (2026-09-02).

■ 무엇이 빠져 있었나
  「언어추가」 커밋이 신설한 <c>LanguagePickerPopup</c> 이
      <c>StringTable.Get("ui_settings_language", "언어 / Language")</c>
  를 부르는데 <b>그 키가 표에 없다</b>. `check_string_keys.py --strict` 가 잡아낸
  «표에 없는 키 1개» 가 이것이다.
  → 지금은 폴백(«언어 / Language»)이 뜨므로 <b>화면이 깨지지는 않는다</b>. 다만
    어느 언어로 켜든 그 한 줄만 «언어 / Language» 로 고정된다.

■ ⚠ 목록의 «언어 이름» 은 번역하지 않는다
  <c>LanguageSetting.NameOf</c> 의 주석 그대로다 — 한국어를 못 읽는 사람이 지금 무엇이
  켜져 있는지 알 수 없게 된다. <b>번역하는 것은 창 제목뿐</b>이다.

■ 제목을 지은 방식
  이 창은 «지금 언어를 못 읽는 사람» 도 열게 된다. 그래서 각 언어 이름 뒤에
  <b>영어를 함께</b> 적는 한국어 원문의 방식(«언어 / Language»)을 그대로 물려받는다 —
  영어권 언어는 «Language» 하나로 충분하므로 겹쳐 쓰지 않는다.

■ 다음
    py -3 Tools/gen_string_table.py   →   py -3 Tools/link_string_keys.py
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET, DATA_ROW0 = 'string', 4
SOURCE = '창 코드 문구(2026-09-02)'

ROWS = {
    # ⚠ kr 은 LanguagePickerPopup 의 폴백과 <b>한 글자도 다르면 안 된다</b>.
    'ui_settings_language': dict(
        kr='언어 / Language',
        en='Language',
        es='Idioma / Language',
        fr='Langue / Language',
        de='Sprache / Language',
        ja='言語 / Language',
        ru='Язык / Language',
        pt='Idioma / Language',
        pl='Język / Language',
        note='언어 고르기 창의 제목. ⚠ 목록의 언어 이름은 번역하지 않는다'),
}


def main():
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]
    fields = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    C = {n: i + 1 for i, n in enumerate(fields)}
    LANGS = [f for f in ['kr', 'en', 'es', 'fr', 'de', 'ja', 'ru', 'pt', 'pl'] if f in C]

    where, last_row = set(), DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k is None or str(k).strip() == '':
            continue
        where.add(str(k).strip())
        last_row = r

    added = []
    for key, vals in ROWS.items():
        if key in where:
            print('  · 이미 있어 건너뜁니다: %s' % key)
            continue
        last_row += 1
        ws.cell(last_row, 1).value = key
        for L in LANGS:
            ws.cell(last_row, C[L]).value = vals[L]
        if 'source' in C:
            ws.cell(last_row, C['source']).value = SOURCE
        if 'note' in C:
            ws.cell(last_row, C['note']).value = vals.get('note', '')
        added.append(key)

    if not added:
        print('바뀐 것이 없어 저장하지 않았습니다.')
        return 0

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.20260902e.bak')
    wb.save(STRING_XLSX)
    print('저장: %s (백업 .20260902e.bak) · %d키 추가 — %s'
          % (os.path.basename(STRING_XLSX), len(added), ' · '.join(added)))
    return 0


if __name__ == '__main__':
    sys.exit(main())
