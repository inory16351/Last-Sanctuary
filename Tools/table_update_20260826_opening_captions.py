# -*- coding: utf-8 -*-
"""오프닝 내레이션 자막 열여섯 줄을 스트링 키 테이블로 옮긴다 (2026-08-26 · 17차).

유저 리포트: *"영어로 번역 안된 것들 … 등등"*. 오프닝 자막은 <b>화면에 크게 뜨는 글자</b>인데
179절의 훑기(직렬화 <b>필드</b>)에 안 걸렸다 — 자막은 <c>Slide[] slides</c> 안의
<b>중첩 배열</b> 원소라서다. 그래서 언어를 영어로 두어도 오프닝만 한국어로 흘렀다.

■ 키를 <b>음성 경로에서 만든다</b> — 표에도 코드에도 번호를 적지 않는다
      Opening/VO_01_1  →  ui_opening_vo_01_1
  ⚠ 왜 «키» 칸을 새로 만들지 않았나 — <b>자막은 씬에 저장돼 있다</b>
    (`Opening.unity` 의 OpeningDirector · 코드의 기본값보다 씬이 이긴다). 칸을 새로 만들면
    씬의 옛 값에는 그 칸이 <b>비어</b> 있어 영원히 한국어로 남는다. 음성 경로는 문장마다
    고유하고(열여섯 조각 · 141절) <b>이미 씬에 저장돼 있다</b> → 씬을 한 곳도 안 고친다.

■ kr 은 <b>씬에서 읽는다</b> (`table_update_20260826_scene_labels.py` 와 같은 규약)
  씬이 정본이므로 사람이 옮겨 적으면 어긋난다. `\\uXXXX` 디코딩 · YAML 접힘 펴기를 한다.

■ ⚠ 영어는 내가 지은 것이다 — <b>유저 검수 필요</b> (179-4절 미결 160번과 같은 성격).
  음성은 한국어 그대로다(영어 더빙이 아니다) — 자막만 영어가 된다. 타자 속도는
  «글자 수 ÷ 음성 길이» 로 재므로 문장 길이가 달라도 <b>말하는 동안</b> 다 쳐진다.

■ 다음
    py -3 Tools/gen_string_table.py   →  py -3 Tools/link_string_keys.py
"""
import io
import os
import re
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

_PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SCENE = os.path.join(_PROJECT, 'Assets', 'Scenes', 'Opening.unity')
SHEET = 'string'
DATA_ROW0 = 4
SOURCE = '오프닝 자막'
NOTE = '⚠ 음성은 한국어 그대로다 — 자막만 번역한다'

BS = chr(92)
_u = re.compile(BS + BS + 'u([0-9A-Fa-f]{4})')
_x = re.compile(BS + BS + 'x([0-9A-Fa-f]{2})')

#: 음성 파일 이름 → 영어 자막. kr 은 씬에서 읽으므로 여기 적지 않는다.
EN = {
    'VO_01_1': 'I remember — when this sanctuary shone pure white.',
    'VO_01_2': 'The song of the angels rang out from every spire,',
    'VO_01_3': 'and no darkness ever crossed this threshold.',

    'VO_02_1': 'That light went out.',
    'VO_02_2': 'In an instant the sky was stained the color of blood,',
    'VO_02_3': 'and darkness came like a flood and swallowed them.',
    'VO_02_4': 'In the defense of the place they called home,',
    'VO_02_5': 'the reason hardly mattered — knowing nothing of why,',
    'VO_02_6': 'they only fastened their armor and walked into the deepening dark.',

    'VO_03_1': 'The gates have fallen, and beasts are howling in the sky.',
    'VO_03_2': 'The flames know no mercy, and the dark spreads like roots.',
    'VO_03_3': 'All that is left is ash, and oaths that went unkept.',

    'VO_04_1': 'I cannot remember every name of the fallen.',
    'VO_04_2': 'But what they meant to protect — that I have not forgotten.',
    'VO_04_3': 'You there — before the last sanctuary sets for good,',
    'VO_04_4': 'go out to meet it.',
}


def dec(s):
    s = _u.sub(lambda m: chr(int(m.group(1), 16)), s)
    return _x.sub(lambda m: chr(int(m.group(1), 16)), s)


def read_captions():
    """씬의 `slides` 에서 (음성 이름, 한국어 자막) 을 순서대로 읽는다."""
    raw = io.open(SCENE, encoding='utf-8', errors='replace').read()
    m = re.search(r'\n  slides:\n(.*?)\n  [A-Za-z_]', raw, re.S)
    if not m:
        sys.exit('Opening.unity 에서 slides 를 못 찾았습니다.')
    lines = m.group(1).split('\n')

    out = []
    i = 0
    while i < len(lines):
        if lines[i].strip().startswith('- text:'):
            text = lines[i].split(':', 1)[1].strip()
            j = i + 1
            # ⚠ YAML 이 긴 문장을 여러 줄로 접는다 — 다음 필드까지 이어 붙인다.
            while j < len(lines) and not re.match(r'^\s+(voice|atMusicTime):', lines[j]):
                text += ' ' + lines[j].strip()
                j += 1
            voice = ''
            if j < len(lines) and lines[j].strip().startswith('voice:'):
                voice = lines[j].split(':', 1)[1].strip()
            text = dec(text).strip()
            if text.startswith('"') and text.endswith('"'):
                text = text[1:-1]
            out.append((voice, text))
            i = j
        i += 1
    return out


def main():
    caps = read_captions()
    print('[오프닝 자막 → 스트링] 씬에서 읽은 문장 %d개' % len(caps))

    novoice = [t for v, t in caps if not v]
    if novoice:
        sys.exit('✗ 음성이 없어 키를 만들 수 없는 문장 %d개: %s'
                 % (len(novoice), ' / '.join(x[:20] for x in novoice)))

    wanted = []
    for voice, text in caps:
        leaf = voice.rsplit('/', 1)[-1]
        if leaf not in EN:
            sys.exit('✗ 영어가 EN 표에 없는 문장: %s (%s)' % (leaf, text[:30]))
        wanted.append(('ui_opening_' + leaf.lower(), text, EN[leaf]))

    if len(set(k for k, _, _ in wanted)) != len(wanted):
        sys.exit('✗ 음성 이름이 겹쳐 키가 겹칩니다 — 씬을 볼 것.')

    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    rows_by_key = {}
    last_row = DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        rows_by_key[str(k).strip()] = r
        last_row = r

    added, kept = [], []
    for key, kr, en in wanted:
        if key in rows_by_key:
            kept.append(key)
            continue
        last_row += 1
        ws.cell(row=last_row, column=1).value = key
        ws.cell(row=last_row, column=2).value = kr
        ws.cell(row=last_row, column=3).value = en
        ws.cell(row=last_row, column=4).value = SOURCE
        ws.cell(row=last_row, column=5).value = NOTE
        rows_by_key[key] = last_row
        added.append((key, kr, en))

    print('  덧붙인 키 %d개 · 이미 있던 키 %d개' % (len(added), len(kept)))
    for k, kr, en in added:
        print('    +', k, '|', kr[:34], '|', en[:44])
    if kept:
        print('    (이미 있던 키:', ', '.join(kept), ')')

    if not added:
        print('  바뀐 것이 없습니다 — 저장하지 않았습니다.')
        return

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    wb.save(STRING_XLSX)
    print('  저장:', os.path.basename(STRING_XLSX), '(백업 .bak)')
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
