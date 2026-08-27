# -*- coding: utf-8 -*-
"""캐릭터 테이블에 <b>성별</b> 칸을 만들고 인물화 14장을 보고 채운다 (2026-08-27).

유저 지시
---------
  *"캐릭터 시트에 남녀 표기 칼럼 하나 추가하고 일러스트 바탕으로 파악해서 남녀 기입 해줘
    enum 으로 해서 하면 될듯"*

    Character 시트 맨 뒤에 한 칸
      gender   string   male / female        ← C# 은 CharacterGender enum 으로 읽는다

★ <b>왜 표에는 숫자가 아니라 글자를 적나</b>
──────────────────────────────────────────────────────────────────────
enum 은 <b>코드 쪽 표현</b>이다. 표에 `1`/`2` 를 적으면 사람이 표만 보고는 어느 쪽이
남자인지 알 수 없고, 언젠가 enum 에 값을 끼워 넣으면 <b>표 전체가 조용히 밀린다</b>.
그래서 표는 `male`/`female` 이라고 적고, `gen_character_assets.py` 가 enum 정수로 옮긴다
(`CharacterGenderText.Parse` 가 런타임 쪽의 같은 변환이다).
★ 빈 칸은 <b>Unknown(0)</b> 이고 «성별을 안 가린다» 는 뜻이다 — 값이 없는 것과
  «남자» 가 같은 숫자가 되지 않게 한 것이다(183-2절의 «C# 기본값» 함정과 같은 자리).

★★ <b>판정 근거는 일러스트다</b> — 이름이 아니라 그림을 봤다
──────────────────────────────────────────────────────────────────────
`Resources/Illust/illust_<이름>.png` 14장을 붙여 놓고 골랐다.
아래 주석의 «근거» 가 그 판단이다. ⚠ 얼굴이 <b>안 보이는</b> 둘(비기오르 · 세라피엘)은
갑옷·의상의 실루엣으로 판단한 것이라 원작자의 설정과 다를 수 있다 — 다르면 이 파일의
표를 고치고 다시 돌리면 된다.

⚠ <b>Excel COM 으로 쓴다</b>(136-4절 · 이 PC 에 LibreOffice 가 없다).

다음:  python Tools/gen_character_assets.py     (정의 에셋에 반영)
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "캐릭터 테이블.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")
SHEET = "Character"

NEW_COLUMNS = [
    ("성별", "gender", "string"),
]

# character_id → (gender, 근거)   ★ 인물화 14장을 보고 판단 (2026-08-27)
GENDER = {
    9001: ("female", "엘린 — 눈을 가린 흰 머리 여성"),
    9002: ("female", "비기오르 — ⚠ 투구로 얼굴이 안 보인다. 황금 갑옷의 흉갑이 여성 실루엣"),
    9003: ("female", "프레이야 — 드레스 차림의 여성"),
    9004: ("male",   "피올로 — 실크햇·부리 가면의 역병 의사. 어깨가 넓은 남성 체형"),
    9005: ("male",   "히스톤 — 검은 갑옷의 남성, 짧은 흑발"),
    9006: ("female", "시그리드 — 웃는 은발 여성"),
    9007: ("female", "시카리아 — 면사포를 쓴 여성 궁수"),
    9008: ("male",   "아루 — 등불을 든 흑발 소년(남성)"),
    9009: ("male",   "카이론 — 상반신을 드러낸 근육질 남성"),
    9010: ("female", "아르세니아 — 후드를 쓴 여성 연금술사"),
    9011: ("male",   "불칸 — 수염을 기른 노년 남성(대마법사)"),
    9012: ("female", "엘리시아 — 방패와 대검을 든 여성 기사"),
    9013: ("female", "세라피엘 — ⚠ 가면으로 얼굴이 안 보인다. 검은 슈트의 실루엣이 여성"),
    9014: ("female", "시안 — 푸른 드레스의 은발 여성"),
}

VALID = {"male", "female"}


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_성별")
    os.makedirs(folder, exist_ok=True)
    shutil.copy2(XLSX, os.path.join(folder, os.path.basename(XLSX)))
    print("백업: " + folder)


def read_layout():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    names = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None and str(v).strip():
            names[str(v).strip()] = c
    return names, ws.max_column


def main():
    bad = {cid: g for cid, (g, _) in GENDER.items() if g not in VALID}
    if bad:
        sys.exit("✗ male/female 가 아닌 값이 있습니다: %s" % bad)

    names, last = read_layout()
    print(f"「{SHEET}」 지금 칸 {last}개")
    backup()

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(XLSX))
        ws = wb.Worksheets(SHEET)

        col = {}
        nxt = last + 1
        for kr, field, typ in NEW_COLUMNS:
            if field in names:
                col[field] = names[field]
                print(f"  칸이 이미 있습니다 — {names[field]}열  {field}")
                continue
            ws.Cells(1, nxt).Value = kr
            ws.Cells(2, nxt).Value = field
            ws.Cells(3, nxt).Value = typ
            col[field] = nxt
            print(f"  칸 신설 — {nxt}열  {kr} / {field} / {typ}")
            nxt += 1

        touched = 0
        seen = set()
        row = 4
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            cid = int(v)
            seen.add(cid)
            if cid in GENDER:
                g, why = GENDER[cid]
                ws.Cells(row, col["gender"]).Value = g
                print(f"  {cid}: {g:<6} — {why}")
                touched += 1
            row += 1

        wb.Save()
        wb.Close()
        print(f"캐릭터 테이블.xlsx — {touched}명 갱신")
    finally:
        excel.Quit()

    missing = sorted(c for c in seen if c not in GENDER)
    if missing:
        print("⚠ 성별을 안 적은 인물이 있습니다(빈 칸 = Unknown = 안 가림):", missing)
    if touched != len(GENDER):
        print("⚠ 적으려던 수(%d)와 실제로 쓴 수(%d)가 다릅니다 — character_id 를 확인할 것"
              % (len(GENDER), touched))
    print("\n다음: python Tools/gen_character_assets.py")


if __name__ == "__main__":
    main()
