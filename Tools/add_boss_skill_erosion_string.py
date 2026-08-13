# -*- coding: utf-8 -*-
"""스트링 키 테이블.xlsx — 보스 스킬 2종의 기계적 설명문에 <b>침식 효과 문장</b>을 덧붙인다
(2026-08-13, `add_boss_skill_erosion_column.py` 와 세트).

`skill_type_desc_Fallen_tomb` / `skill_type_desc_Void_laser` 는 새로 만든
`value_04`(침식) 컬럼을 아직 설명하지 않는다. 기존 문장 끝에 "맞은 적은 침식이
{value_04} 만큼 오른다." 를 그대로 이어붙인다 — 다른 칸의 {value_01}~{value_03}
자리표시와 같은 규칙이다.

⚠ 이 파일은 <b>`Tools/gen_string_table.py` 가 매번 통째로 다시 쓰는 산출물</b>이라
  openpyxl 로 직접 두 칸만 고쳐도 안전하다(그 스크립트 자체가 이 방식으로 쓴다) —
  하이퍼링크는 **다른** 표(웨이브 몬스터 테이블 등)에서 이 파일 쪽으로 걸려 있을 뿐,
  이 파일 안에는 없다. 그래도 되돌릴 수 있게 백업은 남긴다.

사용법:  python Tools/add_boss_skill_erosion_string.py
        이어서 python Tools/gen_string_table.py 를 돌려 StringTable.txt 로 내보낼 것.
"""
import os
import sys
import shutil
import datetime
import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

TABLE_DIR = r'C:\Project\Last-Sanctuary-Vault\데이터 테이블'
XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
BACKUP_ROOT = os.path.join(TABLE_DIR, '_백업')
SHEET = 'string'
DATA_ROW0 = 4

# key -> 덧붙일 문장 (기존 문장 끝에 그대로 이어붙인다).
# ⚠ 원문이 마침표 없이 끝나므로(다른 여러 desc 항목과 같은 구어체 산문) 마침표를
#   직접 넣어 문장을 가른다 — 그대로 이어붙이면 "…공격한다 맞은 적은…" 처럼 붙어버린다.
APPEND = {
    'skill_type_desc_Fallen_tomb': '. 맞은 적은 침식이 {value_04} 만큼 오른다.',
    'skill_type_desc_Void_laser': '. 맞은 적은 침식이 {value_04} 만큼 오른다.',
}


def backup():
    stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    folder = os.path.join(BACKUP_ROOT, stamp + '_보스스킬침식문구')
    os.makedirs(folder, exist_ok=True)
    shutil.copy2(XLSX, os.path.join(folder, os.path.basename(XLSX)))
    return folder


def main():
    if not os.path.exists(XLSX):
        sys.exit('파일을 찾지 못했습니다: ' + XLSX)

    folder = backup()
    print('백업:', folder)

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    updated = []
    for r in range(DATA_ROW0, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if key not in APPEND:
            continue

        kr = ws.cell(r, 2).value or ''
        clause = APPEND[key]
        if clause.strip() in kr:
            print(f'  {key}: 이미 반영됨 — 건너뜀 (멱등)')
            continue

        # 마지막 문장이 "…한다" 로 끝나면 그 뒤에 마침표 없이 바로 이어붙인다 —
        # 기존 문구들이 원래 마침표가 없는 구어체 산문이다.
        ws.cell(r, 2).value = kr + clause
        updated.append(key)
        print(f'  {key}: 갱신')
        print(f'    -> {kr}{clause}')

    if not updated:
        print('\n변경 없음(이미 반영됨).')
        return

    wb.save(XLSX)
    print(f'\n{len(updated)}개 키 갱신 완료.')


if __name__ == '__main__':
    main()
