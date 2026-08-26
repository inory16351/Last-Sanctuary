# -*- coding: utf-8 -*-
"""인물화 <b>얼굴 초점</b>을 표에 넣는다 — 15장 실측 (2026-08-26).

유저 지시
---------
  *"캐릭터 로스터 초상화 일러스트 다시 측정해서 자연스럽게 바꾸기 지금 엄청 이상함
    캐릭터의 얼굴이 보이는 상체 일러스트 부분만 남기기"*

★★★ 왜 «앵커» 로는 안 됐나 — <b>얼굴 높이가 캐릭터마다 다르다</b>
──────────────────────────────────────────────────────────────────────
15장을 격자에 얹어 재 보니(스크래치패드의 `face_measure.png`) 얼굴 중심이

    세로 0.19 (비기오르 투구) ~ 0.38 (시카리아)   ← 두 배 차이
    가로 0.43 (아르세니아)    ~ 0.57 (비기오르)

로 흩어져 있다. 투구·왕관·후드·후광이 저마다 다른 높이를 먹기 때문이다.
그래서 «맨 위를 남긴다» 든 «가운데» 든 <b>한 규칙으로는 못 맞춘다</b> —
어떤 캐릭터는 얼굴이 잘리고, 어떤 캐릭터는 얼굴이 액자 밖으로 나간다.

★ 그래서 <b>캐릭터마다 얼굴 좌표</b>를 표에 적고, `PortraitFit` 이 그 점을
  <b>액자 위에서 35% 자리</b>에 놓는다 — 얼굴 위에 머리, 아래에 어깨가 남아
  «얼굴이 보이는 상체» 가 된다.

    Character 시트 맨 뒤에 두 칸
      face_x   float   얼굴 중심 가로 (0~1)
      face_y   float   얼굴 중심 세로 (0~1 · 0 이 맨 위)

⚠ 값은 <b>사람이 눈으로 잰 것</b>이다(±0.03). 원화를 다시 뽑으면 다시 재야 한다 —
  그때는 `face_measure.png` 를 만든 방식(원화 위 60% 에 0.05 격자)을 그대로 쓸 것.
⚠ <b>Excel COM 으로 쓴다</b>(136-4절).

다음:  python Tools/gen_character_assets.py     (에셋에 반영)
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "캐릭터 테이블.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")
SHEET = "Character"

NEW_COLUMNS = [
    ("얼굴 초점 X", "face_x", "float"),
    ("얼굴 초점 Y", "face_y", "float"),
]

# character_id → (face_x, face_y)   ★ 15장 실측 (2026-08-26)
FOCUS = {
    9001: (0.50, 0.30),   # 엘린      — 가시관 아래, 눈이 0.28
    9002: (0.57, 0.19),   # 비기오르  — 투구 면갑이 높다
    9003: (0.47, 0.28),   # 프레이야
    9004: (0.50, 0.29),   # 피올로    — 실크햇 아래 부리 가면
    9005: (0.44, 0.27),   # 히스톤    — 얼굴이 왼쪽으로 치우쳤다
    9006: (0.53, 0.30),   # 시그리드
    9007: (0.50, 0.38),   # 시카리아  — 면사포가 길어 얼굴이 가장 낮다
    9008: (0.48, 0.25),   # 아루      — 후드 그림자 안
    9009: (0.55, 0.23),   # 카이론
    9010: (0.43, 0.21),   # 아르세니아 — 얼굴이 가장 왼쪽·가장 높다
    9011: (0.55, 0.32),   # 불칸      — 왕관이 0.1~0.25 를 먹는다
    9012: (0.52, 0.27),   # 엘리시아
    9013: (0.47, 0.23),   # 세라피엘  — 가면
    9014: (0.55, 0.30),   # 시안
}


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_얼굴초점")
    os.makedirs(folder, exist_ok=True)
    shutil.copy2(XLSX, os.path.join(folder, os.path.basename(XLSX)))
    print("백업: " + folder)


def read_layout():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    names = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None and str(v).strip():
            names[str(v).strip()] = c
    return names, ws.max_column


def main():
    names, last = read_layout()
    print(f"「{SHEET}」 지금 칸 {last}개")
    backup()

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(XLSX))
        ws = wb.Worksheets(SHEET)

        col = {}
        nxt = last + 1
        for kr, field, typ in NEW_COLUMNS:
            if field in names:
                col[field] = names[field]
                continue
            ws.Cells(1, nxt).Value = kr
            ws.Cells(2, nxt).Value = field
            ws.Cells(3, nxt).Value = typ
            col[field] = nxt
            print(f"  칸 신설 — {nxt}열  {kr} / {field} / {typ}")
            nxt += 1

        touched = 0
        row = 4
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            cid = int(v)
            if cid in FOCUS:
                fx, fy = FOCUS[cid]
                ws.Cells(row, col["face_x"]).Value = fx
                ws.Cells(row, col["face_y"]).Value = fy
                print(f"  {cid}: 얼굴 ({fx}, {fy})")
                touched += 1
            row += 1

        wb.Save()
        wb.Close()
        print(f"캐릭터 테이블.xlsx — {touched}명 갱신")
    finally:
        excel.Quit()

    missing = [c for c in FOCUS if c not in ()]
    print(f"\n실측 {len(FOCUS)}명 · 표에 쓴 {touched}명")
    if touched != len(FOCUS):
        print("⚠ 수가 다르다 — 표의 character_id 를 확인할 것")


if __name__ == "__main__":
    main()
