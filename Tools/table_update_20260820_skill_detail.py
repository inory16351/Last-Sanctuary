# -*- coding: utf-8 -*-
"""스킬 **상세 설명 컬럼** 신설 + 카이론 플레이버 3칸 채우기 (2026-08-20, 3차).

유저 지시 둘을 한 번에 처리한다 — <b>둘 다 `Skill` 시트 한 장을 고치는 일</b>이라
엑셀을 두 번 열 이유가 없다.

★ ① 「새로 추가한 캐릭터들 설명 칼럼 이상하게 들어간거 고쳐줘」
--------------------------------------------------------------
유저 추측은 *"value_06 칼럼 추가한 원인으로 보임"* 이었지만 **밀림은 없었다.**
`gen_character_assets.py` 는 컬럼을 <b>이름으로</b> 찾으므로(`find_col`) 칸이 늘어도
뒤가 밀리지 않는다 — 그 사고는 `value_04` 때 한 번 나고 이미 고쳐졌다.

실제 원인은 <b>훨씬 단순했다</b>: 카이론의 세 줄(80025·80026·80027)만
`skill_explain` 칸이 <b>처음부터 비어 있었다</b>(커밋 `5298e5e` 「카이론 캐릭터 시트 추가」
시점부터 그렇다). 그래서:

* 표에 문구가 없다 → `gen_string_table.py` 가 `skill_explain_80025~27` 키를 <b>안 만든다</b>
* 그런데 `gen_character_assets.py` 는 `flavorKey` 를 <b>id 로 조립</b>한다
  (`"skill_explain_%d" % sid`) → 에셋에는 <b>있지도 않은 키</b>가 적힌다
* `StringTable.Get(없는 키, 빈 리터럴)` → 성장 창의 카이론 스킬 카드가 <b>빈칸</b>이 된다

⚠ 즉 «이상하게 들어간» 것이 아니라 <b>아예 안 들어간</b> 것이고, 조용히 빈칸이 되므로
  에러도 안 났다. 나머지 32개 스킬은 세 키(`name`·`explain`·`type_desc`)가 모두 성립한다
  (아래 `--audit` 로 매번 다시 검사할 수 있게 해 뒀다).

★★ ② 「스킬 상세 설명 컬럼」 — `skill_detail` 신설
--------------------------------------------------
유저 지시: *"캐릭터 스킬 시트에 칼럼 하나 추가해서 스킬 상세 설명 보여줄 칸을 만들어줘 …
밸류 타입보단 덜 상세하게"*.

<b>왜 칸이 하나 더 필요한가</b> — 지금 문구가 <b>둘</b>인데 그 사이가 비어 있다:

| 칸 | 성격 | 예 (「천벌」) |
|---|---|---|
| `skill_explain` | <b>플레이버</b> — 인물을 말한다. 수치가 없다 | *"천상에서 쫓겨난 자가 내리는…"* |
| **`skill_detail`** ← 신설 | <b>무슨 일이 일어나는가</b>. 수치를 <b>말하지 않는다</b> | *"정신을 집중한 뒤 앞쪽 직사각형 범위의 적을…"* |
| `Skill_Type.desc` | <b>정의문</b> — 수치가 전부 채워진다 | *"{value_01}초 동안 정신집중… {value_04}% 데미지…"* |

★ 규칙 하나로 갈랐다 — <b>`skill_detail` 에는 숫자를 적지 않는다.</b> 「잠시」·「크게」·
  「일정 확률로」 로 쓴다. 숫자를 적으면 <b>표의 값을 고칠 때 이 칸이 거짓말을 시작한다</b>
  (정의문은 자리표시라 저절로 따라가지만 이 칸은 사람이 적은 글이라 안 따라간다).
  그것이 곧 «밸류 타입보단 덜 상세하게» 의 뜻이라고 봤다.

⚠ 문장은 <b>정의문을 옮겨 적은 것</b>이다 — 새 규칙을 지어내지 않았다. 표현은 다듬을
  여지가 있으니 유저가 표에서 직접 고치면 된다(그러라고 표에 넣는 것이다).

실행 순서 (⚠ 이 순서대로)
-------------------------
    py -3 Tools/table_update_20260820_skill_detail.py        ← 표에 한글을 적는다
    py -3 Tools/gen_string_table.py                          ← 키를 수집한다
    py -3 Tools/convert_tables_to_string_keys.py             ← 표의 한글을 키로 바꾼다
    py -3 Tools/link_string_keys.py                          ← 하이퍼링크
    py -3 Tools/gen_character_assets.py                      ← 에셋에 반영

    py -3 Tools/table_update_20260820_skill_detail.py --audit   ← 검사만 (표를 안 고친다)
"""

import datetime
import os
import shutil
import sys

from vault_path import TABLE_DIR as TABLES

BACKUP_ROOT = os.path.join(TABLES, "_백업")
CHAR_XLSX = "캐릭터 테이블.xlsx"
FIRST_DATA_ROW = 4

#: 신설 컬럼 — (필드명, 1행 한글 머리, 3행 타입).
NEW_COLUMN = ("skill_detail", "스킬 상세 설명", "string")

#: ① 카이론의 <b>플레이버</b>. 비어 있는 칸만 채운다(값이 있으면 손대지 않는다).
#:   ⚠ 다른 스킬의 문장과 <b>같은 목소리</b>로 맞췄다 — 3인칭 서술 · 한 문장 · 수치 없음.
FLAVOR = {
    80025: "성치 않은 몸이지만 카이론은 단 한 번도 물러선 적이 없습니다.",
    80026: "그가 방패를 들면 적들의 눈에는 그 외의 것이 보이지 않습니다.",
    80027: "천상에서 쫓겨난 자가 내리는 천벌이야말로 가장 무겁습니다.",
}

#: ② 상세 설명 33종. <b>숫자를 적지 않는다</b>(맨 위 ★ 규칙).
DETAIL = {
    80001: "시야가 몸 주변으로 좁아지는 대신, 동료가 본 적을 공격할 때 사거리가 늘어납니다.",
    80002: "크게 다친 동료가 곁에 있으면 자기 체력을 깎아 그 동료를 회복시킵니다.",
    80003: "공격할 때마다 스스로 피를 흘리고, 흘린 만큼 공격력에 더해집니다.",
    80004: "침식으로 정신 이상에 걸릴 때 이로운 쪽이 나올 확률이 높아집니다.",
    80005: "주변의 적이 매초 불타는 피해를 입습니다. 피해는 자신의 현재 체력에 비례합니다.",
    80006: "주변 동료의 방어력을 올려 줍니다. 능력치 상한을 넘어설 수 있습니다.",
    80007: "자신이 직접 잡은 적마다 체력을 회복합니다.",
    80008: "자신이 직접 잡은 적마다 공격 속도와 이동 속도가 오릅니다. 겹쳐 쌓이며 상한을 넘어섭니다.",
    80009: "체력이 절반 밑으로 떨어지면 공격력이 크게 오르지만 공격할 때마다 체력이 깎입니다. "
           "받는 정신 이상은 이기심과 광분으로 고정됩니다.",
    80010: "공격당한 적의 방어력이 일정 시간 낮아집니다. 겹쳐 쌓이지는 않습니다.",
    80011: "같은 집결지의 동료가 나쁜 정신 이상에 걸리면 즉시 풀어 주고 침식도 낮춥니다.",
    80012: "체력이 크게 떨어지면 발동합니다. 공격한 적에게 「정화」를 남기고, "
           "그 적을 때린 동료가 체력을 회복합니다.",
    80013: "포지션이 전방, 공격 유형이 근거리로 고정됩니다. 근거리 공격으로도 크리티컬이 터집니다.",
    80014: "공격할 때마다 「분노」가 쌓이고 쉬는 동안 줄어듭니다. "
           "분노가 가득 찬 상태에서 죽으면 잠시 뒤 되살아납니다.",
    80015: "되살아날 때 주변의 적에게 피해를 주고 주변 동료를 회복시킵니다.",
    80016: "자신이 공격한 적이 곧 죽으면 일정 확률로 자기 체력을 나눠 주변 동료를 회복시킵니다. "
           "후퇴 기준도 고정됩니다.",
    80017: "「가학증」이 터질 때마다 공격 속도가 잠시 오릅니다. 겹치지 않고 지속시간만 늘어납니다.",
    80018: "체력이 크게 떨어지면 잠시 어떤 피해도 받지 않습니다. 회복은 그대로 받습니다.",
    80019: "사거리가 다른 캐릭터보다 깁니다. 공격 유형이 원거리면 사거리가 더 늘어납니다.",
    80020: "원거리 공격이 사거리 안의 적 여러 마리를 한 번에 맞힙니다.",
    80021: "대상 자리에 화살 비를 쏟아 원형 범위의 적을 함께 공격합니다.",
    80022: "주변에 침식이 심하거나 후퇴 중인 동료가 있으면 자기 곁으로 끌어옵니다.",
    80023: "「도움의 손길」로 옮겨진 동료는 즉시 체력 재생이 가능한 상태가 됩니다.",
    80024: "자기 공격 능력치의 일부를 모든 능력치로 쓰는 골렘을 부릅니다. "
           "골렘은 근접·전방으로 고정되고 침식·전술 변경·강화·후퇴가 없습니다. "
           "쿨타임은 골렘이 죽은 뒤부터 돕니다.",
    80025: "잠시 최대 체력에 비례하는 보호막을 두릅니다.",
    80026: "정신을 집중한 뒤 주변의 적을 도발하고, 도발이 끝나면 그 적들을 한꺼번에 때립니다.",
    80027: "정신을 집중한 뒤 직사각형 범위의 적을 내리칩니다. 맞은 적은 잠시 방어력이 낮아집니다.",
    80028: "근거리 공격 유형을 고를 수 없습니다. "
           "대신 회복과 마법에도 명중률과 크리티컬이 적용됩니다.",
    80029: "가장 전방의 동료 자리에 물약을 던져 성스러운 공간을 만듭니다. "
           "그 안의 적은 매초 피해를 입고, 동료는 받는 회복이 커집니다.",
    80030: "주변에 적이 충분히 많을 때만 터집니다. 넓은 범위에 큰 마법 피해를 주고, "
           "그 대가로 잠시 아무 행동도 할 수 없습니다.",
    80031: "공격할 때 일정 확률로 적을 불태워 일정 시간 지속 피해를 줍니다.",
    80032: "마법과 공격 속도가 영구히 오릅니다. 능력치 상한을 넘어설 수 있습니다.",
    80033: "공격 중인 대상에게 거대 화염구를 던져 떨어진 자리의 적을 함께 태웁니다.",
}


def check_locks(files):
    locked = [f for f in files if os.path.isfile(os.path.join(TABLES, "~$" + f))]
    if locked:
        raise SystemExit("⚠ 엑셀에서 열려 있는 파일이 있습니다 — 닫고 다시 실행하세요:\n   "
                         + "\n   ".join(locked))


def backup(files, tag):
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_ROOT, stamp + "_" + tag)
    os.makedirs(dst, exist_ok=True)
    for f in files:
        src = os.path.join(TABLES, f)
        if os.path.isfile(src):
            shutil.copy2(src, os.path.join(dst, f))
    print("백업:", dst)


def find_col(ws, field, max_col=40):
    """2행(필드명)으로 컬럼을 찾는다. 없으면 0."""
    for c in range(1, max_col + 1):
        v = ws.Cells(2, c).Value
        if v is not None and str(v).strip() == field:
            return c
    return 0


def last_col(ws, max_col=40):
    n = 0
    for c in range(1, max_col + 1):
        v = ws.Cells(2, c).Value
        if v is not None and str(v).strip():
            n = c
    return n


def audit():
    """표를 <b>고치지 않고</b> 세 키가 다 성립하는지만 본다. ①번 사고의 재발 감지용."""
    import openpyxl
    ch = openpyxl.load_workbook(os.path.join(TABLES, CHAR_XLSX), data_only=True)
    st = openpyxl.load_workbook(os.path.join(TABLES, "스트링 키 테이블.xlsx"),
                                data_only=True)["string"]
    have = set()
    for r in st.iter_rows(values_only=True):
        if r and r[0] and len(r) > 1 and r[1] and str(r[1]).strip():
            have.add(str(r[0]).strip())

    bad = 0
    for r in ch["Skill"].iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        if not r or r[0] is None:
            continue
        sid = int(float(r[0]))
        stype = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        want = [("이름", "skill_name_%d" % sid),
                ("플레이버", "skill_explain_%d" % sid),
                ("상세", "skill_detail_%d" % sid),
                ("정의문", "skill_type_desc_%s" % stype)]
        miss = [n for n, k in want if k not in have]
        if miss:
            print("  ⚠ %d %-24s 스트링 없음: %s" % (sid, stype, ", ".join(miss)))
            bad += 1
    print("  검사 끝 — 문제 있는 스킬 %d개" % bad)
    return bad


def fix_skill_sheet(excel):
    wb = excel.Workbooks.Open(os.path.join(TABLES, CHAR_XLSX))
    changed = 0
    try:
        ws = wb.Worksheets("Skill")
        last_row = ws.UsedRange.Rows.Count
        c_id = find_col(ws, "skill_id")
        c_exp = find_col(ws, "skill_explain")
        if not (c_id and c_exp):
            raise SystemExit("⚠ Skill 시트에서 skill_id / skill_explain 컬럼을 못 찾았습니다.")

        # ── ② 컬럼 신설 : 마지막 칸 다음에 붙인다 ────────────────────
        #    ⚠ <b>가운데에 끼워 넣지 않는다</b> — 이 표를 읽는 스크립트는 모두 이름으로
        #      찾지만(그래서 순서는 안전하다), 사람이 보는 눈에는 «맨 뒤에 새로 생긴 칸»
        #      이 더 읽기 쉽고 조건부 서식·인쇄 범위도 안 흔들린다.
        field, head, typ = NEW_COLUMN
        c_det = find_col(ws, field)
        if c_det:
            print("\n  [컬럼] %s — 이미 %d열에 있습니다" % (field, c_det))
        else:
            c_det = last_col(ws) + 1
            ws.Cells(1, c_det).Value = head
            ws.Cells(2, c_det).Value = field
            ws.Cells(3, c_det).Value = typ
            print("\n  [컬럼] %s 신설 → %d열 (머리 「%s」 · 타입 %s)"
                  % (field, c_det, head, typ))
            changed += 1

        # ── ① 플레이버 : 빈 칸만 ─────────────────────────────────────
        print("\n  [플레이버 · skill_explain] 빈 칸만 채운다")
        filled = 0
        for r in range(FIRST_DATA_ROW, last_row + 1):
            v = ws.Cells(r, c_id).Value
            if v is None:
                continue
            sid = int(float(v))
            if sid not in FLAVOR:
                continue
            cur = ws.Cells(r, c_exp).Value
            cur = str(cur).strip() if cur is not None else ""
            if cur:
                print("    · %d 이미 「%s…」 — 건드리지 않습니다" % (sid, cur[:24]))
                continue
            ws.Cells(r, c_exp).Value = FLAVOR[sid]
            print("    %d ← %s" % (sid, FLAVOR[sid]))
            filled += 1
            changed += 1
        if filled == 0:
            print("    (채울 빈 칸이 없었다)")

        # ── ② 상세 설명 : 빈 칸만 ────────────────────────────────────
        #    ⚠ 이미 값이 있으면 <b>덮지 않는다</b> — 유저가 다듬은 문장을 지우면 안 된다
        #      (`merge_string_table.py` 와 같은 규칙 · 120-1절).
        print("\n  [상세 설명 · skill_detail] 빈 칸만 채운다")
        wrote, kept, unknown = 0, 0, []
        for r in range(FIRST_DATA_ROW, last_row + 1):
            v = ws.Cells(r, c_id).Value
            if v is None:
                continue
            sid = int(float(v))
            cur = ws.Cells(r, c_det).Value
            cur = str(cur).strip() if cur is not None else ""
            if sid not in DETAIL:
                if not cur:
                    unknown.append(sid)
                continue
            if cur:
                kept += 1
                continue
            ws.Cells(r, c_det).Value = DETAIL[sid]
            wrote += 1
            changed += 1
        print("    적음 %d칸 · 그대로 둠 %d칸" % (wrote, kept))
        if unknown:
            print("    ⚠ 문장을 준비하지 못한 스킬: %s — 표에서 직접 적으세요"
                  % ", ".join(str(s) for s in unknown))

        # ★ 준비한 문장 중 표에 없는 id 가 있으면 알린다 — 스킬이 지워졌거나 오타다.
        seen = set()
        for r in range(FIRST_DATA_ROW, last_row + 1):
            v = ws.Cells(r, c_id).Value
            if v is not None:
                seen.add(int(float(v)))
        stray = sorted(set(DETAIL) - seen)
        if stray:
            print("    ⚠ 표에 없는 id 에 문장을 준비해 뒀습니다(무시됨): %s"
                  % ", ".join(str(s) for s in stray))

        wb.Save()
    finally:
        wb.Close()
    return changed


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    if "--audit" in sys.argv:
        print("[스킬 문구 검사]")
        raise SystemExit(0 if audit() == 0 else 1)

    check_locks([CHAR_XLSX])
    import win32com.client as win32
    backup([CHAR_XLSX], "스킬상세설명")

    # ⚠ DispatchEx — EnsureDispatch 는 유저가 엑셀을 열어 두면 죽는다(119-2절).
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        n = fix_skill_sheet(excel)
    finally:
        excel.Quit()

    print("\n고친 칸 %d개" % n)
    print("다음: gen_string_table.py → convert_tables_to_string_keys.py → "
          "link_string_keys.py → gen_character_assets.py")


if __name__ == "__main__":
    main()
