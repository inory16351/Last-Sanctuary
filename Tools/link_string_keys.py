# -*- coding: utf-8 -*-
"""스트링 키가 든 칸을 <b>눌러서 스트링 키 테이블로 바로 넘어가게</b> 한다.

유저 요청(2026-08-12): "스트링 컬럼 있을 때 그냥 딸깍 하면 바로 스트링 키 테이블로
넘어가는 매크로 못 만듬?"

■ ★ 왜 VBA 매크로가 아니라 하이퍼링크인가
  매크로로 만들면 모든 테이블을 <b>.xlsm 으로 바꿔야 하고</b>(파일명·확장자가 변한다)
  열 때마다 "매크로를 허용하시겠습니까" 를 통과해야 한다. 게다가 매크로가 없는 환경
  (뷰어·구글 시트·git diff)에서는 아무 기능도 안 한다.
  하이퍼링크는 <b>xlsx 표준 기능</b>이라 파일 형식을 그대로 두고, 셀 값도 키 그대로 남는다
  (`Hyperlinks.Add` 에 `TextToDisplay` 를 주지 않으면 값을 안 건드린다).
  덤으로 Excel 이 파란 밑줄을 입혀 <b>"이 칸은 누를 수 있다"</b>가 눈에 보인다.

■ ★ 이름(정의된 이름)으로 건다 — 행 번호를 박지 않는다
  `string!A5` 처럼 <b>행 번호</b>로 걸면 스트링 테이블에 행이 끼어들 때 전부 어긋난다.
  그래서 스트링 테이블에 키마다 <b>정의된 이름</b>(`key_<키>`)을 만들고 그걸 가리킨다 —
  Excel 이 행을 옮겨도 이름이 따라가므로 링크가 안 깨진다.

■ 어느 칸에 거는가
  `convert_tables_to_string_keys.py` 의 `TARGETS` 를 <b>그대로 가져다 쓴다</b> —
  "어떤 컬럼이 키를 담는가" 를 두 파일에 따로 적어두면 반드시 어긋난다.

■ 실행 (파이프라인 마지막에 한 번)
  python Tools/link_string_keys.py
  python Tools/link_string_keys.py --clear   # 링크를 전부 지운다(원상복구)
"""
import os
import sys
import shutil
import datetime
import openpyxl

from convert_tables_to_string_keys import (
    TABLE_DIR, STRING_XLSX, BACKUP_ROOT, DATA_ROW0, TARGETS,
    norm, looks_like_key,
)

STRING_SHEET = 'string'

# ★ 정의된 이름은 `gen_string_table.py` 가 만든다 — 그쪽이 스트링 키 테이블을 매번
#   새로 쓰기 때문에, 여기서만 만들면 다음 실행에서 전부 날아가고 링크가 깨진다.
#   아래 define_names 는 <b>보수용</b>이다(누락·위치 어긋남을 메운다).
NAME_PREFIX = 'key_'


def load_string_rows():
    """key → 스트링 테이블에서의 행 번호."""
    if not os.path.exists(STRING_XLSX):
        sys.exit('스트링 키 테이블이 없습니다. 먼저 python Tools/gen_string_table.py 를 돌리세요.')

    wb = openpyxl.load_workbook(STRING_XLSX, data_only=True)
    ws = wb[STRING_SHEET]
    out = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        key = norm(ws.cell(row=r, column=1).value)
        if key:
            out[key] = r
    return out


def plan_links(rows):
    """어느 파일·시트·칸에 링크를 걸지 미리 계산한다 (Excel 을 열기 전에)."""
    plans = {}
    warnings = []

    for filename, sheet, id_field, specs in TARGETS:
        path = os.path.join(TABLE_DIR, filename)
        if not os.path.exists(path):
            continue

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        # 2행(필드명) → 열 번호
        idx = {}
        for c in range(1, ws.max_column + 1):
            name = norm(ws.cell(row=2, column=c).value)
            if name:
                idx[name] = c

        for value_field, _prefix in specs:
            if value_field not in idx:
                continue
            col = idx[value_field]

            for r in range(DATA_ROW0, ws.max_row + 1):
                text = norm(ws.cell(row=r, column=col).value)
                if not text or not looks_like_key(text):
                    continue

                if text not in rows:
                    warnings.append(
                        f'{filename}/{sheet} {r}행 {value_field}: 스트링 테이블에 없는 키 "{text}" '
                        f'— 링크를 걸지 않았다(끊어진 참조다)')
                    continue

                plans.setdefault(filename, []).append((sheet, r, col, text))

    return plans, warnings


def backup(filenames):
    stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    folder = os.path.join(BACKUP_ROOT, stamp + '_링크')
    os.makedirs(folder, exist_ok=True)
    for fn in filenames:
        src = os.path.join(TABLE_DIR, fn)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(folder, fn))
    return folder


def define_names(app, rows):
    """스트링 테이블에 키마다 정의된 이름을 만든다. 이미 있으면 위치만 맞춘다."""
    wb = app.Workbooks.Open(os.path.abspath(STRING_XLSX))
    made = 0
    try:
        existing = {}
        for nm in wb.Names:
            existing[nm.Name] = nm

        for key, row in rows.items():
            name = NAME_PREFIX + key
            refers = f"='{STRING_SHEET}'!$A${row}"
            if name in existing:
                if existing[name].RefersTo != refers:
                    existing[name].RefersTo = refers
                    made += 1
                continue
            wb.Names.Add(Name=name, RefersTo=refers)
            made += 1

        # 스트링 테이블 안에서도 A열을 눌러 그 행으로 갈 이유는 없으므로 링크는 안 건다.
        wb.Save()
    finally:
        wb.Close(SaveChanges=False)
    return made


def apply_links(app, plans, clear_only):
    """대상 칸에 하이퍼링크를 건다. 항상 기존 링크를 먼저 지운다(멱등)."""
    written = 0
    cleared = 0

    for filename in sorted(plans):
        path = os.path.join(TABLE_DIR, filename)
        wb = app.Workbooks.Open(os.path.abspath(path))
        try:
            # 이 파일에서 손댈 시트·열의 기존 링크를 먼저 지운다.
            touched_cols = {}
            for sheet, r, col, key in plans[filename]:
                touched_cols.setdefault(sheet, set()).add(col)

            for sheet, cols in touched_cols.items():
                ws = wb.Worksheets(sheet)
                for col in cols:
                    rng = ws.Range(ws.Cells(DATA_ROW0, col), ws.Cells(ws.UsedRange.Rows.Count + DATA_ROW0, col))
                    if rng.Hyperlinks.Count:
                        cleared += rng.Hyperlinks.Count
                        rng.Hyperlinks.Delete()
                        # 하이퍼링크 스타일(파란 밑줄)도 같이 되돌린다.
                        rng.Font.Underline = -4142      # xlUnderlineStyleNone
                        rng.Font.ColorIndex = -4105     # xlColorIndexAutomatic

            if not clear_only:
                for sheet, r, col, key in plans[filename]:
                    ws = wb.Worksheets(sheet)
                    cell = ws.Cells(r, col)
                    # ★ TextToDisplay 를 주지 않는다 — 셀 값(키)을 건드리지 않으려는 것.
                    ws.Hyperlinks.Add(
                        Anchor=cell,
                        Address=os.path.basename(STRING_XLSX),   # 같은 폴더 → 파일명만
                        SubAddress=NAME_PREFIX + key,
                        ScreenTip=f'스트링 키 테이블로 이동: {key}')
                    written += 1

            wb.Save()
        finally:
            wb.Close(SaveChanges=False)

    return written, cleared


def verify_names(rows):
    """정의된 이름이 실제로 저장됐는지 다시 읽어 확인한다."""
    wb = openpyxl.load_workbook(STRING_XLSX)
    names = set(wb.defined_names.keys()) if hasattr(wb.defined_names, 'keys') else set()
    want = {NAME_PREFIX + k for k in rows}
    missing = want - names
    return len(want) - len(missing), sorted(missing)[:5]


def main():
    clear_only = '--clear' in sys.argv
    dry = '--dry' in sys.argv

    rows = load_string_rows()
    plans, warnings = plan_links(rows)
    total = sum(len(v) for v in plans.values())

    print(f'== 링크 대상 {total}칸 / 스트링 키 {len(rows)}개')
    for filename in sorted(plans):
        print(f'  {filename}: {len(plans[filename])}칸')
    if warnings:
        print('\n== 경고')
        for w in warnings:
            print('  !', w)

    if dry:
        print('\n--dry 이므로 파일을 고치지 않았습니다.')
        return
    if total == 0:
        print('\n링크를 걸 칸이 없습니다.')
        return

    folder = backup(sorted(list(plans.keys()) + [os.path.basename(STRING_XLSX)]))
    print(f'\n백업: {folder}')

    import win32com.client as win32
    app = win32.gencache.EnsureDispatch('Excel.Application')
    app.Visible = False
    app.DisplayAlerts = False
    try:
        if not clear_only:
            made = define_names(app, rows)
            print(f'정의된 이름: {made}개 새로/갱신')

        written, cleared = apply_links(app, plans, clear_only)
        print(f'링크 {"지움" if clear_only else "생성"}: 생성 {written}칸 · 기존 제거 {cleared}칸')
    finally:
        app.DisplayAlerts = True
        app.Quit()

    if not clear_only:
        ok, missing = verify_names(rows)
        print(f'검증: 이름 {ok}/{len(rows)}개 확인' + (f' · 누락 {missing}' if missing else ''))


if __name__ == '__main__':
    main()
