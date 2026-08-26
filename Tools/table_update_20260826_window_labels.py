# -*- coding: utf-8 -*-
"""창 «안» 의 코드 문구를 스트링 키 테이블로 옮긴다 (2026-08-26 · 13차 · 178-5절의 남은 일).

유저 지시: 178-5절의 «창 안의 코드 문구 78개» 를 이어서 처리 (*"ㅇㅇ ㄱㄱ"*).

■ 규약 — kr 은 <b>코드의 지금 기본값과 한 글자도 다르지 않게</b> 적는다
  코드는 `HudTheme.T(키, 기존값)` 으로 읽는다. 표에 키가 없으면 <b>기존값</b>이 그대로
  쓰이므로, 둘이 어긋나면 «표를 고쳤는데 화면이 안 바뀐다» 가 아니라 <b>«표를 지웠더니
  글자가 달라졌다»</b> 가 된다. 그래서 여기 적는 kr 은 코드에서 그대로 떠 온 값이다.

■ 다시 쓰는 키
  같은 문구가 여러 창에 있으면 <b>키 하나</b>를 쓴다 — «선택된 캐릭터 없음» 은 성장 창과
  전술 창이 같이 쓰고, «닫기»·«확인»·«건너뛰기» 는 창 넷이 같이 쓴다. 문구를 다듬을 때
  한 곳만 고치면 되도록.

■ ⚠ 자리표와 태그를 그대로 옮긴다
  `{0}` · `{0:0.#}` · `<size=80%>` · `<b>` 는 <b>표에도 그대로</b> 들어간다. 번역하다
  중괄호나 태그를 지우면 숫자가 사라지거나 서식이 깨진다(비고 칸에 경고를 적어 둔다).

■ ⚠ 앞뒤 공백이 뜻을 가진 칸이 다섯 있다
  «　· 처치 3/5» 처럼 <b>앞에 공백</b>이 붙은 문구는 다른 글자 뒤에 <b>이어 붙이는</b> 조각이다.
  공백이 사라지면 «각성 가능» 이 앞말에 달라붙는다. 내보낸 TSV 를 되읽어 확인한다.

■ 다음
    py -3 Tools/gen_string_table.py   →  py -3 Tools/link_string_keys.py
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET = 'string'
DATA_ROW0 = 4
SOURCE = '창 코드 문구'
FMT = '⚠ {0} 같은 자리표를 지우지 말 것'
TAG = '⚠ <b>·<size> 태그를 지우지 말 것'
PAD = '⚠ 앞의 공백이 뜻을 가진다(앞말에 이어 붙는 조각)'

#: (키, 한국어, 영어, 비고) — 한국어는 코드의 지금 기본값 그대로다.
ROWS = [
    # ── 성장 창 · 전술 창이 함께 쓰는 «선택» 문구 ──────────────────────────
    ('ui_sel_none_name', '선택된 캐릭터 없음', 'No character selected', ''),
    ('ui_sel_none_hint', '로스터에서 캐릭터를 선택하세요.', 'Pick a character in the roster.', ''),
    ('ui_sel_switch_hint', '로스터에서 다른 캐릭터를 고르면 즉시 전환됩니다.',
     'Picking another character in the roster switches instantly.', ''),

    # ── 캐릭터 성장 창 ────────────────────────────────────────────────────
    ('ui_growth_enhance_format', '강화하기 ({0})', 'Upgrade ({0})', FMT),
    ('ui_growth_enhance', '강화하기', 'Upgrade', ''),
    ('ui_growth_stat_capped', '능력치 상한', 'Stats maxed', ''),
    ('ui_growth_note_no_energy', '에너지가 부족합니다.', 'Not enough energy.', ''),
    ('ui_growth_note_summoned', '소환수는 강화할 수 없습니다.', 'Summons cannot be upgraded.', ''),
    ('ui_growth_enhance_blocked', '강화 불가', 'Cannot upgrade', ''),
    ('ui_growth_level_maxed', '만렙', 'Max level', ''),
    ('ui_growth_note_level_max', 'Lv.{0} — 더는 강화할 수 없습니다.',
     'Lv.{0} - no further upgrades.', FMT),
    ('ui_growth_pick_focus', '성장 유형 결정', 'Choose growth type', ''),
    ('ui_growth_note_pick_focus', '성장 유형을 고르면 그 계열 능력치가 더 잘 오릅니다.',
     'Choosing a growth type makes that line of stats rise more easily.', ''),
    ('ui_growth_note_focus_format', '성장 유형 : {0} — 강조된 능력치가 더 잘 오릅니다.',
     'Growth type: {0} - highlighted stats rise more easily.', FMT),
    ('ui_growth_locked_format', 'Lv.{0} 에 해금', 'Unlocks at Lv.{0}', FMT),
    ('ui_growth_unlocked', '해금됨', 'Unlocked', ''),

    # ── 패시브 카드 ──────────────────────────────────────────────────────
    ('ui_passive_click_hint', '클릭 → 상세', 'Click for details', ''),
    ('ui_passive_none', '이 캐릭터에는 지정된 스킬이 없습니다.',
     'This character has no skills assigned.', ''),
    ('ui_passive_no_selection', '캐릭터를 선택하세요.', 'Select a character.', ''),

    # ── 영웅 각성 진행 조각 (앞 공백이 뜻을 가진다) ───────────────────────
    ('ui_hero_kill_progress', ' · 처치 {0}/{1}', ' · Kills {0}/{1}', PAD),
    ('ui_hero_heal_progress', ' · 회복 {0}/{1}', ' · Healing {0}/{1}', PAD),
    ('ui_hero_ready', ' · 각성 가능', ' · Awakening ready', PAD),
    ('ui_hero_level_lock', ' · Lv.{0} 부터 각성', ' · Awakens from Lv.{0}', PAD),
    ('ui_hero_awakened', '★영웅 ', '★Hero ', PAD),

    # ── 사건 창 ──────────────────────────────────────────────────────────
    ('ui_event_title_fallback', '이름 없는 사건', 'Unnamed Event', ''),
    ('ui_event_body_fallback', '(대사 준비 중)', '(dialogue pending)', ''),
    ('ui_event_accept', '수락', 'Accept', ''),
    ('ui_event_decline', '거절', 'Decline', ''),

    # ── 발굴 창 ──────────────────────────────────────────────────────────
    ('ui_dig_title', '발굴 가능한 자리', 'Excavation Site', ''),
    ('ui_dig_title_result', '발굴 결과', 'Excavation Result', ''),
    ('ui_dig_title_boss', '빼앗은 것', 'Spoils', ''),
    ('ui_dig_choice_accept', '가까이 가서 살펴본다.', 'Go closer and take a look.', ''),
    ('ui_dig_choice_decline', '방심은 금물이다. 그냥 두자.', 'Better not risk it. Leave it.', ''),

    # ── 환경 설정 (게임 중 · 로비가 언어 형식을 함께 쓴다) ─────────────────
    ('ui_settings_language_format', '언어 : {0}', 'Language: {0}', FMT),
    ('ui_settings_saved_format', '저장했습니다 ({0})', 'Saved ({0})', FMT),
    ('ui_settings_save_failed', '저장하지 못했습니다.', 'Could not save.', ''),
    ('ui_settings_help_reset_done', '도움말을 처음 상태로 되돌렸습니다 — 각 버튼을 다시 누르면 설명이 나옵니다.',
     'Help has been reset - press each button again to see its explanation.', ''),
    ('ui_settings_help_reset_failed', '도움말 서비스를 찾지 못했습니다.',
     'Could not find the help service.', ''),
    ('ui_settings_restart_confirm', '정말 처음부터 다시 시작할까요? 한 번 더 누르면 실행됩니다 (저장이 지워집니다)',
     'Really restart from the beginning? Press again to confirm (your save will be erased).', ''),
    ('ui_settings_quit_confirm',
     '저장하지 않고 로비로 나갈까요? 한 번 더 누르면 실행됩니다 (마지막 저장 이후 진행이 사라집니다)',
     'Leave to the lobby without saving? Press again to confirm '
     '(progress since the last save will be lost).', ''),

    # ── 단축키 설정 ──────────────────────────────────────────────────────
    ('ui_hotkey_hint', '키 칸을 누르고 원하는 키를 누르세요.  Esc 취소 · Del 해제',
     'Click a key slot, then press the key you want.  Esc to cancel · Del to clear', ''),
    ('ui_hotkey_capturing', '키를 누르세요…', 'Press a key...', ''),
    ('ui_hotkey_reset', '기본값으로', 'Restore defaults', ''),

    # ── 도움말 ───────────────────────────────────────────────────────────
    ('ui_help_hint_pick', '왼쪽에서 항목을 고르면 설명이 나옵니다.',
     'Pick an item on the left to see its explanation.', ''),
    ('ui_help_progress_format', '읽은 조언 {0} / {1}', 'Advice read {0} / {1}', FMT),
    ('ui_help_see_also_format', '함께 볼 것 — {0}', 'See also - {0}', FMT),
    ('ui_help_tour', '화면에서 짚어 보기', 'Point it out on screen', ''),
    ('ui_helpcard_badge_format', '도움말 · {0}', 'Help · {0}', FMT),
    ('ui_tour_last', '다 봤습니다', 'Done', ''),

    # ── 스킬 상세 ────────────────────────────────────────────────────────
    ('ui_skill_owner_format', '{0} · 패시브 {1}', '{0} · Passive {1}', FMT),
    ('ui_skill_values_format', '수치  {0}', 'Values  {0}', FMT),
    ('ui_skill_no_effect', '효과 정의문이 비어 있습니다. 캐릭터 테이블의 Skill_Type 시트를 확인하세요.',
     'The effect definition is empty. Check the Skill_Type sheet in the character table.', ''),

    # ── 로비 ────────────────────────────────────────────────────────────
    ('ui_lobby_saved_at', '마지막 저장: {0}', 'Last save: {0}', FMT),
    ('ui_lobby_no_save', '저장된 게임이 없습니다', 'No saved game', ''),
    ('ui_lobby_help_reset_done', '도움말을 처음 상태로 되돌렸습니다 — 판을 시작하면 설명이 다시 나옵니다.',
     'Help has been reset - explanations will appear again when you start a run.', ''),
    ('ui_lobby_hotkey_missing', '단축키 설정 창을 찾지 못했습니다.',
     'Could not find the key bindings window.', ''),

    # ── 연출 ────────────────────────────────────────────────────────────
    ('ui_fx_hero_awaken', '영웅 각성!', 'Hero Awakening!', ''),
    ('ui_fx_hero_awaken_stage', '영웅 각성! {0}단계', 'Hero Awakening! Stage {0}', FMT),
    ('ui_ending_fallen_head', '돌아오지 못한 이들', 'Those who did not return', ''),
    ('ui_ending_survivor_head', '끝까지 남은 이들', 'Those who remained to the end', ''),
    ('ui_ending_roll_format', '{0}  <size=80%>{1} · Lv.{2} · {3}웨이브</size>',
     '{0}  <size=80%>{1} · Lv.{2} · Wave {3}</size>', TAG),
    ('ui_btn_skip', '건너뛰기', 'Skip', ''),
]

#: 코드가 <b>다시 쓰는</b> 키 — 이미 표에 있어야 한다(없으면 폴백 한국어가 남는다).
MUST_EXIST = ['ui_btn_close', 'ui_btn_confirm', 'ui_settings_hotkeys',
              'ui_tour_next', 'ui_tour_prev', 'ui_tour_quit',
              'ui_helpcard_more', 'ui_helpcard_ok',
              'ui_portrait_rage_format', 'ui_portrait_soul_format']

#: 앞뒤 공백이 뜻을 가지는 키 — 내보낸 뒤 되읽어 확인한다.
PADDED = ['ui_hero_kill_progress', 'ui_hero_heal_progress', 'ui_hero_ready',
          'ui_hero_level_lock', 'ui_hero_awakened']


def main():
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    rows_by_key = {}
    last_row = DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        rows_by_key[str(k).strip()] = r
        last_row = r

    added, kept = [], []
    for key, kr, en, note in ROWS:
        if key in rows_by_key:
            kept.append(key)
            continue
        last_row += 1
        ws.cell(row=last_row, column=1).value = key
        ws.cell(row=last_row, column=2).value = kr
        ws.cell(row=last_row, column=3).value = en
        ws.cell(row=last_row, column=4).value = SOURCE
        ws.cell(row=last_row, column=5).value = note
        rows_by_key[key] = last_row
        added.append((key, kr, en))

    missing = [k for k in MUST_EXIST if k not in rows_by_key]

    print('[창 코드 문구 → 스트링] 대상 %d개 · 덧붙인 키 %d개 · 이미 있던 키 %d개'
          % (len(ROWS), len(added), len(kept)))
    for k, kr, en in added:
        print('    +', k, '|', kr[:40].replace('\n', ' '), '|', en[:40])
    if kept:
        print('    (이미 있던 키:', ', '.join(kept), ')')
    if missing:
        wb.close()
        sys.exit('\n✗ 코드가 다시 쓰는데 표에 없는 키: %s' % ', '.join(missing))

    if not added:
        print('  바뀐 것이 없습니다 — 저장하지 않았습니다.')
        return

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    wb.save(STRING_XLSX)
    print('  저장:', os.path.basename(STRING_XLSX), '(백업 .bak)')
    print('  ⚠ 앞뒤 공백 확인 대상:', ', '.join(PADDED))
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
