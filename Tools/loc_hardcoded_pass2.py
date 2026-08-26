# -*- coding: utf-8 -*-
"""하드코딩 한글 이관 2차 — 허드 액션 · 부대 · 토벌 · 승리/패배 · 배속 · 웨이브 표시.

★★★ 이 차수의 방법 — <b>«문구 칸을 표 값으로 갈아 끼우는 한 함수»</b>
──────────────────────────────────────────────────────────────────────
이 창들은 문구를 <c>[SerializeField] string</c> 칸에 두고 여러 자리에서 쓴다. 쓰는 자리를
전부 고치면 편집이 수십 곳이 되고, 그때마다 «어느 자리를 빠뜨렸나» 가 생긴다.
그래서 <b>칸을 한 번에 갈아 끼우는 함수</b>를 창마다 하나 둔다:

    void LocalizeLabels()
    {
        createFormat = HudTheme.T("ui_action_create", createFormat);
        …
    }

  ★ 쓰는 자리는 <b>한 줄도 안 고친다</b>.
  ★ <see cref="Data.StringTable"/> 에 키가 없으면 칸의 값이 그대로 남는다(폴백).
  ★ 언어를 바꾸면 다시 부른다 — 그래서 <b>창을 닫았다 열지 않아도</b> 바뀐다.
  ⚠ 칸을 덮어쓰므로 «원래 한글» 은 표가 정본이 된다. 표에 키가 없는 채로 언어를
    바꾸면 그 칸만 이전 언어의 글로 남는다 — 그래서 <b>이 차수의 키는 전부 표에 넣는다</b>.

다음: python Tools/gen_string_table.py
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SCRIPTS = os.path.join("Assets", "_Project", "Scripts")
STRING_XLSX = os.path.join(TABLE_DIR, "스트링 키 테이블.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")
T = "HudTheme.T"
G = "Data.StringTable.Get"


def localize_block(pairs, indent=12):
    """(칸, 키) 목록 → LocalizeLabels 본문."""
    pad = " " * indent
    return "\n".join(f"{pad}{field} = {T}(\"{key}\", {field});" for field, key in pairs)


ACTION = [
    ("createFormat", "ui_action_create"),
    ("createAtLimit", "ui_action_create_cap"),
    ("createOutOfCandidates", "ui_action_create_none"),
    ("tacticsIdle", "ui_action_tactics"),
    ("tacticsOpen", "ui_action_tactics_close"),
    ("squadIdle", "ui_action_squad"),
    ("squadOpen", "ui_action_squad_close"),
    ("squadPicking", "ui_action_squad_picking"),
    ("subjugateOpen", "ui_action_subjugate_close"),
    ("subjugateFound", "ui_action_subjugate_found"),
    ("subjugateNone", "ui_action_subjugate"),
    ("relicFound", "ui_action_relic_found"),
    ("relicIdle", "ui_action_relic"),
    ("relicOpen", "ui_action_relic_close"),
    ("settingsIdle", "ui_action_settings"),
    ("settingsOpen", "ui_action_settings_close"),
    ("helpUnread", "ui_action_help_unread"),
    ("helpIdle", "ui_action_help"),
    ("helpOpen", "ui_action_help_close"),
]

SQUAD = [
    ("title", "ui_squad_title"),
    ("hint", "ui_squad_hint"),
    ("hintNoSquad", "ui_squad_hint_none"),
    ("memberFormat", "ui_squad_members"),
    ("rallySetIdle", "ui_squad_rally_set"),
    ("rallySetPicking", "ui_squad_rally_picking"),
    ("rallySetMove", "ui_squad_rally_move"),
    ("rallyClear", "ui_squad_rally_clear"),
    ("coopExpeditionOn", "ui_squad_coop_on"),
    ("coopExpeditionOff", "ui_squad_coop_off"),
]

SUBJ = [
    ("hintPickSquad", "ui_subj_hint_squad"),
    ("hintPickTarget", "ui_subj_hint_target"),
    ("hintNoSquad", "ui_subj_hint_no_squad"),
    ("hintNoTarget", "ui_subj_hint_no_target"),
    ("memberFormat", "ui_squad_members"),
    ("orderNone", "ui_subj_order_none"),
    ("orderBusy", "ui_subj_order_busy"),
    ("levelFormat", "ui_subj_level"),
    ("squadCountFormat", "ui_subj_squad_count"),
    ("hintTargetFull", "ui_subj_hint_full"),
]

VICTORY = [
    ("titleText", "ui_victory_title"),
    ("reasonFormat", "ui_victory_reason"),
    ("summaryFormat", "ui_victory_summary"),
    ("restartLabel", "ui_restart"),
]

DEFEAT = [
    ("summaryFormat", "ui_defeat_summary"),
    ("restartLabel", "ui_restart"),
]

SPEED = [
    ("pauseLabel", "ui_speed_pause"),
    ("resumeLabel", "ui_speed_resume"),
]

EDITS = {
    # ── 허드 액션 버튼 ──
    "UI/ActionPanel.cs": [
        ("        [SerializeField] string helpOpen = \"도움말 닫기\";",
         "        [SerializeField] string helpOpen = \"도움말 닫기\";\n"
         "\n"
         "        /// <summary>\n"
         "        /// ★★★ <b>문구 칸을 표의 값으로 갈아 끼운다</b> (2026-08-26 · 유저 지시:\n"
         "        /// *\"하드 코딩으로 들어가 있는 텍스트들 … 스트링 키 테이블에도 옮기고 영어로도\"*).\n"
         "        /// 표에 키가 없으면 칸의 값이 그대로 남는다 — 그래서 <b>실패해도 화면은 멀쩡하다</b>.\n"
         "        /// </summary>\n"
         "        void LocalizeLabels()\n"
         "        {\n"
         + localize_block(ACTION) + "\n"
         "        }"),
    ],
    "UI/SquadPanel.cs": [
        ("        [SerializeField] string coopExpeditionOff = \"협동 탐험 OFF\";",
         "        [SerializeField] string coopExpeditionOff = \"협동 탐험 OFF\";\n"
         "\n"
         "        /// <summary>문구 칸을 표의 값으로 갈아 끼운다(2026-08-26 · 하드코딩 이관).</summary>\n"
         "        void LocalizeLabels()\n"
         "        {\n"
         + localize_block(SQUAD) + "\n"
         "        }"),
    ],
    "UI/SubjugationPanel.cs": [
        ("        [SerializeField] string hintTargetFull = \"이 대상에는 이미 {0}개 부대가 가 있습니다.\";",
         "        [SerializeField] string hintTargetFull = \"이 대상에는 이미 {0}개 부대가 가 있습니다.\";\n"
         "\n"
         "        /// <summary>문구 칸을 표의 값으로 갈아 끼운다(2026-08-26 · 하드코딩 이관).</summary>\n"
         "        void LocalizeLabels()\n"
         "        {\n"
         + localize_block(SUBJ) + "\n"
         "        }"),
    ],
    "UI/VictoryPanel.cs": [
        ("        [SerializeField] string restartLabel = \"다시 시작\";",
         "        [SerializeField] string restartLabel = \"다시 시작\";\n"
         "\n"
         "        /// <summary>문구 칸을 표의 값으로 갈아 끼운다(2026-08-26 · 하드코딩 이관).</summary>\n"
         "        void LocalizeLabels()\n"
         "        {\n"
         + localize_block(VICTORY) + "\n"
         "        }"),
    ],
    "UI/DefeatPanel.cs": [
        ("        [SerializeField] string restartLabel = \"다시 시작\";",
         "        [SerializeField] string restartLabel = \"다시 시작\";\n"
         "\n"
         "        /// <summary>\n"
         "        /// 문구 칸을 표의 값으로 갈아 끼운다(2026-08-26 · 하드코딩 이관).\n"
         "        /// ★ 제목·패배 사유는 이미 <c>…Key</c> 칸으로 표를 보고 있다 — 여기는 나머지다.\n"
         "        /// </summary>\n"
         "        void LocalizeLabels()\n"
         "        {\n"
         + localize_block(DEFEAT) + "\n"
         "        }"),
    ],
    "UI/GameSpeedPanel.cs": [
        ("        [SerializeField] string resumeLabel = \"재개\";",
         "        [SerializeField] string resumeLabel = \"재개\";\n"
         "\n"
         "        /// <summary>문구 칸을 표의 값으로 갈아 끼운다(2026-08-26 · 하드코딩 이관).</summary>\n"
         "        void LocalizeLabels()\n"
         "        {\n"
         + localize_block(SPEED) + "\n"
         "        }"),
    ],
}


def apply_code_edits():
    total = 0
    for rel, pairs in EDITS.items():
        path = os.path.join(SCRIPTS, rel.replace("/", os.sep))
        with open(path, encoding="utf-8-sig", newline="") as f:
            src = f.read()
        crlf = "\r\n" in src
        flat = src.replace("\r\n", "\n")

        hit = 0
        for old, new in pairs:
            if new in flat:
                continue
            if old not in flat:
                sys.exit(f"! 못 찾음 — {rel}\n  찾던 것: {old[:100]}")
            flat = flat.replace(old, new, 1)
            hit += 1

        if hit:
            out = flat.replace("\n", "\r\n") if crlf else flat
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                f.write(out)
        print(f"  {rel} — {hit}곳")
        total += hit
    return total


# ══════════════════════════════════════════════════════════════════════
#  스트링 키 — (키, 한국어, 영어)
# ══════════════════════════════════════════════════════════════════════
KEYS = [
    # 허드 액션
    ("ui_action_create", "캐릭터 생성 {0}", "Create character {0}"),
    ("ui_action_create_cap", "인원 상한", "Roster full"),
    ("ui_action_create_none", "등장할 인물 없음", "No one left to arrive"),
    ("ui_action_tactics", "전술 지침", "Tactical orders"),
    ("ui_action_tactics_close", "전술 지침 닫기", "Close orders"),
    ("ui_action_squad", "부대 설정", "Squads"),
    ("ui_action_squad_close", "부대 설정 닫기", "Close squads"),
    ("ui_action_squad_picking", "집결지 지정 중", "Placing rally point"),
    ("ui_action_subjugate", "토벌 지시", "Subjugation"),
    ("ui_action_subjugate_found", "토벌 지시 ({0})", "Subjugation ({0})"),
    ("ui_action_subjugate_close", "토벌 지시 닫기", "Close subjugation"),
    ("ui_action_relic", "유물 관리", "Relics"),
    ("ui_action_relic_found", "유물 관리 (발굴 {0})", "Relics (dug {0})"),
    ("ui_action_relic_close", "유물 관리 닫기", "Close relics"),
    ("ui_action_settings", "환경 설정", "Settings"),
    ("ui_action_settings_close", "환경 설정 닫기", "Close settings"),
    ("ui_action_help", "도움말 (F1)", "Help (F1)"),
    ("ui_action_help_unread", "도움말 (새 {0})", "Help ({0} new)"),
    ("ui_action_help_close", "도움말 닫기", "Close help"),
    # 부대 설정
    ("ui_squad_title", "부대 설정", "Squads"),
    ("ui_squad_hint",
     "부대를 고른 뒤 로스터에서 캐릭터를 클릭하면 배정됩니다(다시 누르면 해제). 부대 이름은 직접 고칠 수 있습니다.",
     "Choose a squad, then click a character in the roster to assign them (click again to "
     "remove). Squad names can be edited."),
    ("ui_squad_hint_none", "부대가 없습니다. '부대 추가'로 만드세요.",
     "No squads yet. Use 'Add squad' to make one."),
    ("ui_squad_members", "{0}명", "{0}"),
    ("ui_squad_rally_set", "집결지 설정", "Set rally point"),
    ("ui_squad_rally_picking", "맵을 클릭", "Click the map"),
    ("ui_squad_rally_move", "집결지 이동", "Move rally point"),
    ("ui_squad_rally_clear", "집결지 해제", "Clear rally point"),
    ("ui_squad_coop_on", "협동 탐험 ON", "Joint expedition ON"),
    ("ui_squad_coop_off", "협동 탐험 OFF", "Joint expedition OFF"),
    # 토벌 지시
    ("ui_subj_hint_squad", "토벌을 맡길 부대를 고르세요.", "Choose the squad to send."),
    ("ui_subj_hint_target", "잡을 대상을 고르세요. 같은 대상을 다시 누르면 명령이 해제됩니다.",
     "Choose the target. Press the same target again to cancel the order."),
    ("ui_subj_hint_no_squad", "부대가 없습니다. '부대 설정'에서 먼저 만드세요.",
     "No squads yet. Make one in 'Squads' first."),
    ("ui_subj_hint_no_target", "아직 발견한 에픽 몬스터가 없습니다. 부대를 외곽까지 탐험 보내세요.",
     "No epic monsters found yet. Send a squad exploring to the outer ring."),
    ("ui_subj_order_none", "명령 없음", "No order"),
    ("ui_subj_order_busy", "토벌 중", "On the hunt"),
    ("ui_subj_level", "적정 Lv.{0}", "Suggested Lv.{0}"),
    ("ui_subj_squad_count", "부대 {0}/{1}", "Squads {0}/{1}"),
    ("ui_subj_hint_full", "이 대상에는 이미 {0}개 부대가 가 있습니다.",
     "{0} squads are already on this target."),
    # 승리 · 패배 · 배속
    ("ui_victory_title", "승리", "Victory"),
    ("ui_victory_reason", "웨이브 {0}까지 방어에 성공했습니다.", "You held through wave {0}."),
    ("ui_victory_summary", "웨이브 {0} 클리어 · 생존 {1} · 남은 인원 {2}명",
     "Wave {0} cleared - survived {1} - {2} left"),
    ("ui_defeat_summary", "웨이브 {0} 도달 · 생존 {1} · 남은 인원 {2}명",
     "Reached wave {0} - survived {1} - {2} left"),
    ("ui_restart", "다시 시작", "Restart"),
    ("ui_speed_pause", "정지", "Pause"),
    ("ui_speed_resume", "재개", "Resume"),
    # 웨이브 단계
    ("ui_phase_idle", "대기 전", "Before the wave"),
    ("ui_phase_prep", "정비", "Preparation"),
    ("ui_phase_advance", "진군", "Advance"),
    ("ui_phase_combat", "전투", "Battle"),
    ("ui_phase_enraged", "광폭화", "Enraged"),
    ("ui_phase_defeat", "패배", "Defeat"),
    ("ui_phase_victory", "승리", "Victory"),
    ("ui_defeat_party", "캐릭터가 전멸했습니다", "Every character has fallen"),
    ("ui_defeat_nexus", "성역이 파괴되었습니다", "The sanctuary has been destroyed"),
    # 1차에서 빠진 것
    ("ui_relic_row_wearer", "{0}", "{0}"),
]


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_스트링키2차")
    os.makedirs(folder, exist_ok=True)
    shutil.copy2(STRING_XLSX, os.path.join(folder, os.path.basename(STRING_XLSX)))
    print("백업: " + folder)


def apply_string_rows():
    wb = openpyxl.load_workbook(STRING_XLSX, data_only=True)
    ws = wb["string"]
    where = {}
    last = 3
    for r in range(4, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k is None or not str(k).strip():
            continue
        where[str(k).strip()] = r
        last = r

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    added = filled = 0
    try:
        wbc = excel.Workbooks.Open(os.path.abspath(STRING_XLSX))
        sh = wbc.Worksheets("string")
        row = last + 1
        for key, kr, en in KEYS:
            if key in where:
                r = where[key]
                if not str(sh.Cells(r, 2).Value or "").strip():
                    sh.Cells(r, 2).Value = kr
                    filled += 1
                if not str(sh.Cells(r, 3).Value or "").strip():
                    sh.Cells(r, 3).Value = en
                    filled += 1
                continue
            sh.Cells(row, 1).Value = key
            sh.Cells(row, 2).Value = kr
            sh.Cells(row, 3).Value = en
            sh.Cells(row, 4).Value = "code(하드코딩)"
            sh.Cells(row, 5).Value = "2026-08-26 하드코딩 이관 2차"
            row += 1
            added += 1
        wbc.Save()
        wbc.Close()
    finally:
        excel.Quit()
    return added, filled


def main():
    print("① 코드 교체")
    n = apply_code_edits()
    print(f"   → {n}곳\n")
    backup()
    print("② 스트링 키 테이블")
    added, filled = apply_string_rows()
    print(f"   → 새 키 {added}개 · 빈 칸 채움 {filled}개 (전체 {len(KEYS)}개 중)")
    print("\n⚠ LocalizeLabels() 를 <b>부르는 자리</b>는 손으로 잇는다(창마다 배선 함수가 다르다).")


if __name__ == "__main__":
    main()
