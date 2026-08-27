# -*- coding: utf-8 -*-
"""HUD 로그·창 문구의 하드코딩 한글을 스트링 키 테이블로 옮긴다
(2026-08-27 · 유저 리포트: *"현재 언어에서 영어로 변경해도 영어로 변경되지 않는 UI들이
있어(ex 로그, 단축키 설정 등) 해당 내역들도 전부 데이터 시트(특히 스트링키 테이블)을
참조해서 번역하도록 수정해"*).

■ 무엇을 옮겼나 — 182-10절이 «다음 차례» 로 적어 둔 그 일이다
  ① <b>HUD 로그</b>(`HudLog.Add` · `BattleLogPanel.Append`) — 182-10절의 실측은 64줄이었는데
     이번에 다시 세니 <b>84줄</b>이었다(그 절은 `Scripts/UI` 밖의 몇 파일을 덜 셌다).
     ★ 그중 최대 밀집 구간은 `HudLog.SkillLine(caster, skill, <b>detail</b>)` 의 detail 인자다 —
       이름 둘은 이미 표를 거치는데 <b>덧붙이는 말만</b> 한글로 박혀 있었다.
  ② <b>단축키 설정 창</b> — `HotkeyService.Label` 열여섯 줄. 창의 <b>왼쪽 열 전체</b>가
     한글 리터럴이라 영어로 바꿔도 그대로였다.
  ③ <b>씬 라벨 구멍</b> — 「토벌 지시」·「유물 관리」 제목과 초상화 자리의 「캐릭터 선택」.
     `UiLocalizer` 지도에도 없고 코드도 안 건드려 <b>영영 한국어</b>였다.
  ④ 사건 보상 요약 · 웨이브 상태 · 유물 장착 실패 사유 · 저장 알림 · 발굴 결과.

■ ★★★ <b>kr 을 손으로 적지 않는다</b> — 코드에서 읽어 온다
  규약은 «kr 은 <b>코드의 지금 폴백과 한 글자도 다르지 않게</b>» 인데, 그것을 사람이 옮겨
  적으면 반드시 한 글자가 어긋난다. 그래서 이 스크립트는
  <c>Tools/extract_string_fallbacks.py</c> 로 <b>코드에서 (키, 폴백) 을 읽어</b> kr 로 쓴다.
  `table_update_20260826_scene_labels.py` 가 <b>씬에서</b> kr 을 읽는 것과 같은 원리다.
  → 여기서 사람이 정하는 것은 <b>영어뿐</b>이다(아래 `EN`).

■ 규약
  ★ <b>같은 문구는 키 하나로 묶는다</b>(179-1절). 묶은 근거는 `EN` 의 주석에 적는다.
  ⚠ 자리표가 있는 키는 <b>비고에 «{0} 을 지우지 말 것» 을 적는다</b>(178-4절의 그 경고).
    이 스크립트가 kr 을 보고 <b>자동으로</b> 붙인다 — 사람이 빠뜨릴 수 없게.
  ⚠ <b>이미 있는 키는 건드리지 않는다</b> — 178-6절에서 «표 전체를 훑는 치환» 이 유저 번역
    21칸을 덮은 사고가 있었다. 이 스크립트는 <b>없는 키만 덧붙인다</b>.
  ⚠ <b>영어를 안 지은 키가 하나라도 있으면 실패로 끝낸다</b> — 조용히 빈 칸을 만들면
    «표에는 있는데 화면은 한국어» 가 되어 182-5절과 같은 종류의 구멍이 생긴다.

■ 다음
    py -3 Tools/gen_string_table.py   →   py -3 Tools/link_string_keys.py
    py -3 Tools/check_string_keys.py --strict      (빠진 키 0 인지 검산)
"""
import os
import re
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR
from extract_string_fallbacks import scan as scan_fallbacks

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET = 'string'
DATA_ROW0 = 4
SOURCE = '로그·단축키 이관(2026-08-27)'

PLACEHOLDER = re.compile(r'\{\d+[^}]*\}')


# ══════════════════════════════════════════════════════════════════════
#  영어 — <b>사람이 정하는 유일한 칸</b>
# ══════════════════════════════════════════════════════════════════════
#  ⚠ 자리표는 <b>개수와 번호를 한국어와 똑같이</b> 유지할 것. 어순은 달라도 된다
#    (그러라고 조각을 안 이어 붙이고 «형식 하나» 로 만든 것이다 · 173-6절).
#  ⚠ 서식 자리표(`{1:0.#}`)는 영어에서도 <b>서식까지 그대로</b> 적는다 —
#    `{1}` 로 줄이면 소수점이 그대로 쏟아진다.
EN = {}

#: ⚠ <b>폴백이 «변수» 라 코드에서 읽어낼 수 없는 키</b> — kr 을 손으로 적는 <b>유일한</b> 자리다.
#
#  세 가지 모양이 여기 걸린다:
#    ① `HudTheme.T("키", someField)`  — 폴백이 인스펙터 필드다(값은 씬/프리팹에 있다)
#    ② `UiLocalizer` 의 지도 `E("경로", "키")` — 애초에 T() 호출이 아니다(kr 은 <b>씬</b>에 있다)
#  ★ 그래서 <b>여기 세 줄이 늘어나면 의심할 것</b> — 대부분의 새 문구는 리터럴 폴백을 갖는다.
#  ⚠ ②의 kr 은 씬에서 읽은 값과 같아야 한다. `Tools/table_update_20260826_scene_labels.py`
#    가 씬에서 다시 읽어 검산하므로, 틀리면 그쪽이 잡아 준다.
KR_EXPLICIT = {
    # ① EndingDirector.rollSurvivorFormat — 「끝까지 남은 이들」 줄의 형식
    'ui_ending_roll_survivor_format': '{0}  <size=80%>{1} · Lv.{2}</size>',
    # ① SquadPanel.squadRemove — 부대 카드의 「부대 해제」 버튼
    'ui_squad_remove': '부대 해제',
    # ② UiLocalizer 지도 — 성장 창·전술 지침 창의 초상화 자리(씬의 글자)
    'ui_portrait_pick_hint': '캐릭터 선택',
}


def en(group, mapping):
    """영어 한 묶음. group 은 사람이 읽을 이름일 뿐이다."""
    for k, v in mapping.items():
        if k in EN and EN[k] != v:
            sys.exit('✗ 같은 키에 영어를 두 번 다르게 적었습니다: %s' % k)
        EN[k] = v


# ── 단축키 설정 창 ────────────────────────────────────────────────────
#  ★ 열여섯 줄 중 <b>일곱은 HUD 액션 버튼과 키를 나눠 쓴다</b>(ui_action_settings 등) —
#    이미 표에 있으므로 여기 없다. 나눠 쓸 수 없던 셋은 HUD 쪽에 덧말·자리표가 붙어 있다:
#    ui_action_help «도움말 (F1)» · ui_action_create «캐릭터 생성 {0}» · ui_speed_pause «정지».
en('단축키', {
    'ui_hotkey_action_help': 'Help',
    'ui_hotkey_action_create': 'Create Character',
    'ui_hotkey_action_pause': 'Pause',
    # 배속 1~4단계를 형식 하나로 묶었다(배속 값은 x1·x2·x4·x8 이라 «N배» 가 아니라 «N단계» 다).
    'ui_hotkey_action_speed': 'Speed step {0}',
    'ui_hotkey_action_recenter': 'Center view on the Sanctuary',
    'ui_hotkey_key_none': 'None',
    'log_hotkey_reset_all': 'Key bindings restored to defaults',
    'log_hotkey_stolen': '"{0}" lost its key — {1} was taken by "{2}"',
    'log_hotkey_assigned': '"{0}" key → {1}',
})

# ── 씬 라벨 · 창 라벨 ─────────────────────────────────────────────────
en('창 라벨', {
    'ui_portrait_pick_hint': 'Pick a character',
    'ui_squad_remove': 'Disband',
    'ui_ending_roll_survivor_format': '{0}  <size=80%>{1} · Lv.{2}</size>',
    'ui_relic_none': 'None',
    'ui_growth_cost_format': 'Energy {0}',
    'ui_squad_all': 'All',
    'ui_squad_numbered': 'Squad #{0}',
})

# ── HUD 로그: 집결지 ──────────────────────────────────────────────────
en('집결지', {
    'log_rally_drag_begin': '{0} rally move — release where you want it',
    'log_rally_move_cancel': 'Rally move canceled',
    'log_rally_pick_begin': '{0} rally point — click the map (Esc to cancel)',
    'log_rally_pick_cancel': 'Rally placement canceled',
    'log_rally_moved': '{0} rally point moved',
    'log_rally_set': '{0} rally point set',
    'log_rally_assign': 'Rally #{0} → {1}',
    'log_rally_removed': 'Rally #{0} cleared',
    'log_rally_removed_squad': '{0} rally point cleared',
    'log_rally_removed_all': 'All rally points cleared',
})

# ── HUD 로그: 전투 로그 · 생성 · 배속 · 승패 ──────────────────────────
en('전투 로그', {
    'log_battlelog_ready': 'Battle log ready',
    'log_nexus_destroyed': 'Sanctuary destroyed',
    'log_ally_died': '{0} died',
    # ⚠ 「중립 몬스터/몬스터」를 조각으로 이어 붙이던 것을 종류마다 «형식 하나» 로 갈랐다.
    'log_kill_neutral': 'Neutral monster slain — {0}',
    'log_kill_monster': 'Monster slain — {0}',
    'log_energy_gain': 'Energy +{0} (total {1})',
    'log_upgraded': '{0} Lv.{1} (−{2})',
    'log_char_joined': '{0} joined',
})

en('캐릭터 생성', {
    'log_char_joined_free': '{0} joined ({1})',
    # 한국어는 «{0}명» 처럼 단위가 붙는다 — 영어는 괄호로 밀어 어순 문제를 흡수한다.
    'log_create_at_limit': 'Roster limit reached ({0})',
    'log_create_no_candidates': 'No characters left to recruit',
    'log_create_no_energy': 'Not enough energy — {0} needed',
    'log_char_created': '{0} created (−{1})',
})

en('배속·승패·웨이브', {
    'log_paused': 'Paused',
    'log_resumed': 'Resumed ({0})',
    'log_game_speed': 'Game speed {0}',
    # ⚠ log_victory 와 log_wave_cleared 는 승리 시 <b>두 줄이 나란히</b> 뜬다 — 묶지 말 것.
    'log_victory': 'Victory — wave {0} cleared',
    'log_wave_spawned': 'Wave {0} monsters spawned',
    'log_wave_cleared': 'Wave {0} cleared — victory!',
    'log_defeat': 'Defeat — {0}',
    'log_minimap_alert': 'Wave {0} spawn — {1} minimap alerts',
    'log_hotkey_panel_missing': 'Hotkey settings window not found',
})

# ── 건설 ──────────────────────────────────────────────────────────────
en('건설', {
    'log_build_cancelled': 'Build canceled',
    'log_build_cannot_place_here': "You can't build there",
    'log_build_completed': '{0} built',
    'log_build_limit_reached': 'No more can be built',
    'log_build_not_enough_energy': 'Not enough energy (build cost {0})',
    'log_build_pick_site': 'Build site — click a {0}x{1} area (cost {2}, Esc to cancel)',
    'log_build_site_refunded': 'Build order canceled (energy +{0})',
    'log_build_site_reserved': '{0} queued (energy {1})',
})

# ── 저장·불러오기 ─────────────────────────────────────────────────────
#  ⚠ ui_save_* 는 <b>로비 화면</b>에 그대로 뜬다(LobbyPanel 이 SaveService.LastMessage 를 찍는다).
en('저장', {
    'log_autosave': 'Autosaved — {0}',
    'log_autosave_died': '{0} died',
    'log_autosave_upgraded': '{0} upgraded',
    'log_autosave_wave_cleared': 'Wave {0} cleared',
    'log_load_done': 'Loaded — wave {0} ({1})',
    'log_load_skipped_no_characters': 'No characters in the save — nothing was loaded',
    'log_save_skipped_no_characters': 'No characters — nothing was saved',
    'ui_save_load_failed': "Couldn't read the save file.",
    'ui_save_version_mismatch': 'Save format differs (file {0} · now {1}).',
    'ui_save_write_failed': "Couldn't save.",
    'ui_save_written': 'Saved ({0})',
    'ui_unit_generic': 'Character',
})

# ── 패시브·보스 스킬의 «덧붙이는 말»(HudLog.SkillLine 의 detail) ────────
#  ⚠⚠ 서식 자리표(`{1:0.#}`)는 <b>서식까지 그대로</b> 옮긴다 — `{1}` 로 줄이면
#    소수점이 그대로 쏟아진다(«2.7초» 가 «2.7000000000초» 가 된다).
en('패시브 덧말', {
    'log_detail_at_ally_seconds': "at {0}'s side · {1:0.#}s",
    'log_detail_box_hits': '{0:0.#}x{1:0.#} tiles · {2} hit',
    'log_detail_box_seconds': '{0:0.#}x{1:0.#} tiles · {2:0.#}s',
    'log_detail_channeling': 'channeling {0:0.#}s',
    'log_detail_ferried': '{0} ferried',
    'log_detail_ferried_cleansed': '{0} ferried · regen cleansed',
    'log_detail_golem_summoned': 'golem summoned',
    'log_detail_healed': '{0} healed',
    'log_detail_hits': '{0} hit',
    'log_detail_hits_and_heal': '{0} hit · {1} allies +{2}',
    'log_detail_hits_exhaust': '{0} hit · exhausted {1:0.#}s',
    'log_detail_hits_souls': '{0} hit · souls {1}',
    'log_detail_hp_gain': 'HP +{0}',
    'log_detail_hp_gain_over': 'HP +{0} (over {1:0.#}s)',
    'log_detail_invincible': 'invincible {0:0.#}s',
    'log_detail_leap_shots': 'leap {0:0.#} tiles · {1} shots',
    'log_detail_radius': 'radius {0:0.#} tiles',
    'log_detail_radius_hits': 'radius {0:0.#} tiles · {1} hit',
    'log_detail_revive_in': 'revives in {0:0.#}s',
    'log_detail_riposte_hits': 'riposte {0}',
    # ⚠ {1} 이 <b>두 번</b> 쓰인다(주는 양과 자기가 잃는 양이 같다) — 영어도 그대로 둘 것.
    'log_detail_share_hp': '{0} allies +{1} · self −{1}',
    'log_detail_shield': 'shield {0} · {1:0.#}s',
    'log_detail_taunt': 'taunt {0} · {1:0.#}s',
    'log_revived': '{0} revived',
    'log_hero_awaken': '{0} hero awakening!',
    'log_hero_awaken_stage': '{0} hero awakening! (stage {1})',
    # ⚠ 조사(이/가)는 코드가 <b>한국어일 때만</b> 붙인다 — 영어에는 주격 조사가 없다.
    'log_mental_error_onset': '{0} falls into {1}.',
    'log_status_applied': '{0} {1}!',
})

# ── 상태 이름 — 로그와 로스터 「임무」 칸이 함께 쓴다 ──────────────────
en('상태 이름', {
    'ui_status_bind': 'Bound',
    'ui_status_burn': 'Burning',
    'ui_status_exhausted': 'Exhausted',
    'ui_status_poison': 'Poisoned',
})

# ── 발굴·유물 ─────────────────────────────────────────────────────────
en('발굴·유물', {
    'log_dig_epic_promotion': "You've collected every common relic — epics can be found there now",
    'log_dig_finished_empty': 'The dig turned up nothing',
    'log_dig_ordered': 'Dig ordered — the nearest character is on the way',
    'log_dig_result': '{0} — dug up: {1}',
    'log_dig_site_found': 'Found a spot worth digging',
    # ⚠ 서식 태그(<color=#…>)는 <b>그대로</b> 둔다 — 지우면 등급 색이 사라진다.
    'log_relic_dropped': '{0} left behind <color=#{1}>"{2}" ({3})</color>',
    'log_relic_revived': '{0} — "{1}" brought them back.',
    'ui_dig_dialogue_accept': 'One of the angels heads for the spot.',
    'ui_dig_dialogue_boss_drop': 'Something was taken from the fallen.',
    'ui_dig_dialogue_decline': 'The spot is left alone.',
    'ui_dig_dialogue_discover': 'Something seems to be buried here.',
    'ui_dig_dialogue_result': 'The dug-up spot is examined.',
    'ui_dig_digger_unknown': 'Someone',
    'ui_dig_outcome_energy': 'Energy +{0}',
    'ui_dig_outcome_erosion_down': 'Corruption −{0}',
    'ui_dig_outcome_erosion_up': 'Corruption +{0}',
    'ui_dig_outcome_heal': 'HP +{0}%',
    'ui_dig_outcome_hurt': 'HP −{0}%',
    'ui_dig_relic_all_collected': "You've already collected every {0} relic",
    'ui_dig_relic_granted': 'Relic <color=#{0}>"{1}" ({2})</color>',
    'ui_relic_boss_drop_title': 'Relic "{0}" ({1})',
    'ui_relic_equip_no_definition': 'This character has no definition.',
    'ui_relic_equip_no_target': 'No target.',
    'ui_relic_equip_none_left': 'None left (another character has it equipped).',
    'ui_relic_equip_slots_full': 'Relic slots are full ({0}). Unequip one first.',
    'ui_relic_equip_summon_denied': 'Summons cannot equip relics.',
})

# ── 토벌 · 사건 ───────────────────────────────────────────────────────
en('토벌·사건', {
    'log_epic_monster_found': 'Epic monster found — {0}',
    'log_subjugation_order_cleared': '{0} slay order cleared',
    'log_subjugation_ordered': '{0} → slay {1}',
    'log_subjugation_target_full': '{0} already has {1} squads assigned',
    'ui_squad_default_name': 'Squad {0}',
    'log_event_begin': '<b>[Event]</b> {0}',
    'log_event_reward': '[Event] {0} — {1}',
    'log_help_reset': 'Help has been reset to its initial state',
})

# ── 사건 보상 요약 — 사건 창의 결과 문장 ──────────────────────────────
#  ⚠ 한국어의 세는 단위(«{0}명» · «{0}마리»)는 영어에서 «allies» · «enemies» 로 옮긴다 —
#    숫자만 두면 «누구를 셌는지» 가 사라진다(사건 창은 자리가 넉넉하다).
en('사건 보상', {
    'ui_reward_char_die': '{0} died',
    'ui_reward_char_join': '{0} joined',
    'ui_reward_enemy_atk_down': 'Enemy ATK −{0}% for {1}s ({2} enemies)',
    'ui_reward_enemy_atk_spd': 'Enemy attack speed {0}% for {1}s ({2} enemies)',
    'ui_reward_enemy_bind': 'Enemies bound {0}s ({1} enemies)',
    'ui_reward_enemy_burn': 'Enemies burning {0}%/s for {1}s ({2} enemies)',
    'ui_reward_enemy_def_down': 'Enemy DEF −{0}% for {1}s ({2} enemies)',
    'ui_reward_enemy_hp_loss': 'Enemy HP −{0}% ({1} enemies)',
    'ui_reward_enemy_move_spd': 'Enemy move speed {0}% for {1}s ({2} enemies)',
    'ui_reward_energy': 'Energy {0}',
    'ui_reward_erosion': 'Corruption {0} ({1} allies)',
    'ui_reward_heal_received': 'Healing received +{0}% ({1} allies)',
    'ui_reward_heal_received_timed': 'Healing received +{0}% for {1}s ({2} allies)',
    'ui_reward_hp_percent': 'HP {0}% ({1} allies)',
    'ui_reward_kill_grant': '{0} kill count +{1}',
    'ui_reward_mental_cure': 'Mental disorders cleared ({0} allies)',
    'ui_reward_nexus_damage': 'Sanctuary damaged −{0}',
    'ui_reward_nexus_heal': 'Sanctuary repaired +{0}',
    'ui_reward_permanent_stat': '{0} {1} +{2} (permanent)',
    'ui_reward_range': 'Attack range {0}% ({1} allies)',
    'ui_reward_range_timed': 'Attack range {0}% for {1}s ({2} allies)',
    'ui_reward_relic_gain': '{0} obtained ({1})',
    'ui_reward_shield': 'Shield {0}% of max HP ({1} allies)',
    'ui_reward_source_event': 'Event',
    'ui_reward_stat': '{0} {1}% ({2} allies)',
    'ui_reward_stat_timed': '{0} {1}% for {2}s ({3} allies)',
    'ui_reward_summon_enemy': '{0} more enemies summoned',
    'ui_reward_vision': 'Vision {0}% ({1} allies)',
    'ui_reward_vision_timed': 'Vision {0}% for {1}s ({2} allies)',
})


# ══════════════════════════════════════════════════════════════════════

def note_for(kr):
    """비고 — 자리표가 있으면 «지우지 말 것» 을 <b>자동으로</b> 붙인다."""
    holes = PLACEHOLDER.findall(kr)
    if not holes:
        return ''
    return '⚠ 자리표 %s 를 지우지 말 것 (개수·번호를 한국어와 같게 유지)' % ' '.join(sorted(set(holes)))


def main():
    apply_ = '--apply' in sys.argv

    found = scan_fallbacks()          # {키: [(폴백, 파일:줄), …]}

    # 폴백이 변수라 코드에서 못 읽는 키를 보탠다(위 KR_EXPLICIT 의 ⚠).
    for k, kr in KR_EXPLICIT.items():
        found.setdefault(k, [(kr, '(KR_EXPLICIT)')])

    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    rows_by_key = {}
    last_row = DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        rows_by_key[str(k).strip()] = r
        last_row = r

    missing = sorted(k for k in found if k not in rows_by_key)

    no_en = [k for k in missing if k not in EN]
    if no_en:
        wb.close()
        print('✗ 영어를 안 지은 키가 %d개 있습니다 — 아래를 EN 에 적으세요:' % len(no_en))
        for k in no_en:
            fb, where = found[k][0]
            print("    %-42s %-50r %s" % (k, fb[:48], where))
        sys.exit(1)

    # 영어를 적어 놨는데 코드에 없는 키 — 오타이거나 지운 자리다
    stale = sorted(k for k in EN if k not in found and k not in rows_by_key)
    if stale:
        wb.close()
        sys.exit('✗ EN 에 적었는데 코드가 부르지 않는 키: %s' % ', '.join(stale))

    # ⚠ 자리표 개수가 한국어와 다르면 실행 중에 터진다(string.Format) — 미리 막는다
    bad = []
    for k in missing:
        kr = found[k][0][0]
        kn = {re.match(r'\{(\d+)', h).group(1) for h in PLACEHOLDER.findall(kr)}
        e = {re.match(r'\{(\d+)', h).group(1) for h in PLACEHOLDER.findall(EN[k])}
        if kn != e:
            bad.append((k, sorted(kn), sorted(e)))
    if bad:
        wb.close()
        print('✗ 자리표 번호가 한국어와 영어에서 다릅니다 — string.Format 이 터집니다:')
        for k, kn, e in bad:
            print('    %-42s kr=%s  en=%s' % (k, kn, e))
        sys.exit(1)

    print('코드가 부르는 키 %d개 · 표에 없는 키 %d개' % (len(found), len(missing)))
    for k in missing:
        kr = found[k][0][0]
        print('    + %-40s %r' % (k, kr[:46]))

    if not apply_:
        print('\n(미리보기입니다 — 실제로 넣으려면 --apply 를 붙이세요)')
        return

    for k in missing:
        kr, where = found[k][0]
        last_row += 1
        ws.cell(row=last_row, column=1).value = k
        ws.cell(row=last_row, column=2).value = kr
        ws.cell(row=last_row, column=3).value = EN[k]
        ws.cell(row=last_row, column=4).value = SOURCE
        ws.cell(row=last_row, column=5).value = note_for(kr)

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    wb.save(STRING_XLSX)
    print('\n  저장: %s (백업 .bak) · %d키 추가' % (os.path.basename(STRING_XLSX), len(missing)))
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
