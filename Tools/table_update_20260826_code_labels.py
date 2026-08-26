# -*- coding: utf-8 -*-
"""항상 보이는 HUD 의 <b>코드 문구</b>를 스트링 키 테이블로 옮긴다 (2026-08-26 · 12차).

유저 지시: *"씬에 구운 것들도 테이블로 스트링 키 써서 영어로 바뀌게 해줘"*.

■ 왜 «씬» 이 아니라 «코드» 인가
  이 자리들은 <b>씬에도 한국어가 구워져 있지만</b>, 실행하면 코드가 직렬화 문자열이나
  하드코딩 리터럴로 <b>덮어쓴다</b>. 그래서 `UiLocalizer`(씬 정적 라벨용)로는 못 고친다 —
  덮어쓰는 쪽을 표로 돌려야 한다. 대상은 <b>항상 보이는 HUD</b> 로 한정했다:

    HUD_Actions  건물 건설 · 캐릭터 성장 버튼      (BuildButtonUI · UpgradeButtonUI)
    HUD_Roster   제목 머리말 «캐릭터»              (CharacterRosterPanel)
    로스터·전술·성장의 침식 라벨 «침식 42»         (ErosionGaugeView)
    HUD_Energy   «에너지 30»                       (EnergyLabel)
    HUD_Wave     «웨이브 3 · 정비» · 진군/광폭화   (WaveStatusPanel)

■ ⚠ 서식 지정자를 그대로 옮긴다
  `{0}` · `{0:00}` 은 <b>표에도 그대로</b> 들어간다(`string.Format` 이 그대로 받는다).
  번역할 때 <b>중괄호를 지우면 숫자가 사라진다</b> — 표의 비고 칸에 그 경고를 적어 둔다.

■ 다음
    py -3 Tools/gen_string_table.py   →  py -3 Tools/link_string_keys.py
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET = 'string'
DATA_ROW0 = 4
SOURCE = 'HUD 코드 문구'
FMT_NOTE = '⚠ {0} 같은 자리표를 지우지 말 것'

#: (키, 한국어, 영어, 비고)
ROWS = [
    ('ui_action_build',         '건물 건설 {0}',   'Build {0}', FMT_NOTE),
    ('ui_action_build_picking', '자리 지정 중 (Esc 취소)',
                                'Choosing a spot (Esc to cancel)', ''),
    ('ui_action_build_cap',     '건설 상한',       'Build limit', ''),
    ('ui_action_upgrade',       '캐릭터 성장',     'Character Growth', ''),
    ('ui_action_upgrade_close', '캐릭터 성장 닫기', 'Close Character Growth', ''),

    ('ui_roster_title',         '캐릭터',          'Characters', ''),

    # ★ 영어 낱말은 유저가 쓰는 «Corruption» 을 따른다(도움말 en 이 그 낱말이다).
    ('ui_erosion_value',        '침식 {0}',        'Corruption {0}', FMT_NOTE),
    ('ui_erosion_value_state',  '침식 {0} · {1}',  'Corruption {0} · {1}', FMT_NOTE),
    ('ui_erosion_none',         '침식 -',          'Corruption -', ''),

    ('ui_energy_format',        '에너지 {0:00}',   'Energy {0:00}', FMT_NOTE),

    ('ui_wave_unknown',         '웨이브 정보 없음', 'No wave data', ''),
    ('ui_wave_phase_format',    '웨이브 {0} · {1}', 'Wave {0} · {1}', FMT_NOTE),
    ('ui_timer_marching',       '진군 중',         'Marching', ''),
    ('ui_timer_enraged',        '광폭화!',         'Enraged!', ''),
]

#: 코드가 이미 쓰고 있어 <b>반드시 있어야</b> 하는 키 — 없으면 폴백 한국어가 화면에 남는다.
MUST_EXIST = ['ui_phase_defeat', 'ui_phase_victory', 'ui_phase_idle', 'ui_phase_prep',
              'ui_phase_advance', 'ui_phase_combat', 'ui_phase_enraged']


def main():
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    rows_by_key = {}
    last_row = DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        rows_by_key[str(k).strip()] = r
        last_row = r

    added, kept = [], []
    for key, kr, en, note in ROWS:
        if key in rows_by_key:
            kept.append(key)
            continue
        last_row += 1
        ws.cell(row=last_row, column=1).value = key
        ws.cell(row=last_row, column=2).value = kr
        ws.cell(row=last_row, column=3).value = en
        ws.cell(row=last_row, column=4).value = SOURCE
        ws.cell(row=last_row, column=5).value = note
        rows_by_key[key] = last_row
        added.append((key, kr, en))

    missing = [k for k in MUST_EXIST if k not in rows_by_key]

    print('[HUD 코드 문구 → 스트링] 대상 %d개 · 덧붙인 키 %d개 · 이미 있던 키 %d개'
          % (len(ROWS), len(added), len(kept)))
    for k, kr, en in added:
        print('    +', k, '|', kr, '|', en)
    if kept:
        print('    (그대로 둔 키:', ', '.join(kept), ')')
    if missing:
        wb.close()
        sys.exit('\n✗ 코드가 쓰는데 표에 없는 키: %s' % ', '.join(missing))

    if not added:
        print('  바뀐 것이 없습니다 — 저장하지 않았습니다.')
        return

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    wb.save(STRING_XLSX)
    print('  저장:', os.path.basename(STRING_XLSX), '(백업 .bak)')
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
