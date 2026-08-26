# -*- coding: utf-8 -*-
"""179절 뒤에 남아 있던 «표를 안 거치는 문구» 를 스트링 키 테이블로 옮긴다
(2026-08-26 · 17차 · 유저 리포트 *"영어로 번역 안된 것들 … 로비버튼들 / 알겠습니다 /
캐릭터 성장 능력치들 등등"*).

■ 무엇이 남아 있었나 — 179절의 훑기(`Scripts/UI` 의 <b>직렬화 필드</b>)가 못 본 세 갈래
  ① 서비스 클래스의 직렬화 필드 — `SquadService.squadNameFormat`(«{0}부대»).
     `Scripts/UI` 밖이라 179절의 대상이 아니었다(미결 158번의 그 자리).
  ② 코드 안에 <b>바로 박힌</b> 보간 문자열 — 필드가 아니라 `$""` 라서 «필드를 훑는»
     검사에 안 걸린다. `SkillDetailPanel` 의 «재사용 대기시간 N초» ·
     `TacticalOrderPanel` 의 «사용 안 함».
  ★ 로비 버튼·성장 창 능력치는 <b>여기서 다루지 않는다</b> — 전자는 씬 라벨이라
    `table_update_20260826_scene_labels.py`(지도가 정본), 후자는 이미 있는
    `ui_stat_*` 키를 <b>다시 쓰기만</b> 한다(키를 새로 만들지 않았다).

■ 규약 — kr 은 <b>코드의 지금 폴백과 한 글자도 다르지 않게</b> 적는다
  (`table_update_20260826_window_labels.py` 의 머리글과 같은 이유.)

■ 다음
    py -3 Tools/gen_string_table.py   →  py -3 Tools/link_string_keys.py
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET = 'string'
DATA_ROW0 = 4
SOURCE = '남은 코드 문구'
FMT = '⚠ {0} 같은 자리표를 지우지 말 것'

#: (키, 한국어, 영어, 비고)
ROWS = [
    ('ui_squad_name_format', '{0}부대', 'Squad {0}',
     FMT + ' · 새로 만드는 부대의 기본 이름(이미 만든 부대의 이름은 바뀌지 않는다)'),
    ('ui_skill_cooltime_format', '재사용 대기시간 {0}초', 'Cooldown {0}s', FMT),
    ('ui_tactics_retreat_off', '사용 안 함', 'Off', ''),
]

#: 이미 표에 있지만 <b>영어 칸이 비어 있던</b> 키 — 채운다.
#  177-1절이 «en 빈칸 14개» 로 센 그 자리다. 그중 `event_dialogue_10001` ·
#  `skill_explain_2001~2006` 은 <b>한국어도 비어</b> 있어(내용 자체가 없는 자리표) 여기서
#  다루지 않는다 — 번역 문제가 아니라 데이터 구멍이다.
#  ⚠⚠ <b>비어 있을 때만 쓴다</b> — 178-6절에서 «표 전체를 훑는 치환» 이 유저 번역 21칸을
#    건드린 사고가 있었다. 값이 있는 칸은 무슨 일이 있어도 건드리지 않는다.
FILL_EN = {
    'dig_outcome_desc_dig_erosion_down':
        "The character who dug here loses {value_01} Corruption.",
    'dig_outcome_desc_dig_erosion_up':
        "The character who dug here gains {value_01} Corruption.",
    'dig_outcome_script_dig_erosion_down':
        'A cool grain brushed against them, and what had been seething settled for a while.',
    'dig_outcome_script_dig_erosion_up':
        'Something tainted, pooled below, rose up on their breath.',
    'help_erosion_title': 'A Mind Being Eaten Away',
    'help_erosion_summary':
        "Your companions' minds are <b>corrupted</b> while they fight."
        '\nWhen Corruption reaches 100, the mind breaks.',
    'help_erosion_body':
        'Corruption rises while they fight, and falls on its own a little after the fighting stops.'
        '\nHow fast it rises and how fast it falls are decided by <b>Resistance</b>.'
        '\nYou can watch Corruption on the purple gauge in the roster and the growth window.'
        '\nPulling a companion at risk back during the lull lets them recover that much.',
}
#: ⚠ 줄바꿈은 <b>진짜 줄바꿈</b>이다 — xlsx 칸에는 그렇게 들어 있고
#  (`gen_string_table.py` 가 내보낼 때 «\n» 두 글자로 바꾼다) 한국어 칸과 같은 모양이어야 한다.

#: 이번에 <b>다시 쓰기만</b> 하는 키 — 표에 이미 있어야 한다(없으면 폴백 한국어가 남는다).
#  · ui_stat_*        : 성장 창의 능력치 열두 칸이 StatBlock.DisplayName 을 거치게 됐다
#  · ui_action_settings·ui_settings_* : 로비 창이 게임 창과 같은 키를 본다
MUST_EXIST = [
    'ui_stat_hp', 'ui_stat_melee_atk', 'ui_stat_ranged_atk', 'ui_stat_magic',
    'ui_stat_cure', 'ui_stat_def', 'ui_stat_regen', 'ui_stat_accuracy',
    'ui_stat_critical', 'ui_stat_atk_speed', 'ui_stat_move_speed', 'ui_stat_resistance',
    'ui_action_settings', 'ui_settings_volume', 'ui_settings_hotkeys',
    'ui_settings_help_reset',
    'ui_help_title', 'ui_help_see_also', 'ui_helpcard_more', 'ui_helpcard_ok',
    'ui_tour_next', 'ui_tour_prev', 'ui_tour_quit',
]


def main():
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    rows_by_key = {}
    last_row = DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        rows_by_key[str(k).strip()] = r
        last_row = r

    added, kept = [], []
    for key, kr, en, note in ROWS:
        if key in rows_by_key:
            kept.append(key)
            continue
        last_row += 1
        ws.cell(row=last_row, column=1).value = key
        ws.cell(row=last_row, column=2).value = kr
        ws.cell(row=last_row, column=3).value = en
        ws.cell(row=last_row, column=4).value = SOURCE
        ws.cell(row=last_row, column=5).value = note
        rows_by_key[key] = last_row
        added.append((key, kr, en))

    # ── 영어 칸이 비어 있던 키 채우기 (있는 값은 절대 건드리지 않는다) ────────
    filled, occupied, absent = [], [], []
    for key, en in FILL_EN.items():
        row = rows_by_key.get(key)
        if row is None:
            absent.append(key)
            continue
        cell = ws.cell(row=row, column=3)
        if cell.value is not None and str(cell.value).strip() != '':
            occupied.append(key)
            continue
        cell.value = en
        filled.append(key)

    missing = [k for k in MUST_EXIST if k not in rows_by_key]

    print('[남은 코드 문구 → 스트링] 대상 %d개 · 덧붙인 키 %d개 · 이미 있던 키 %d개'
          % (len(ROWS), len(added), len(kept)))
    for k, kr, en in added:
        print('    +', k, '|', kr, '|', en)
    if kept:
        print('    (이미 있던 키:', ', '.join(kept), ')')

    print('  영어 빈칸 채움 %d개 / 대상 %d개' % (len(filled), len(FILL_EN)))
    for k in filled:
        print('    en+', k)
    if occupied:
        print('    (이미 영어가 있어 건드리지 않은 키:', ', '.join(occupied), ')')
    if absent:
        wb.close()
        sys.exit('\n✗ 영어를 채우려는 키가 표에 없습니다: %s' % ', '.join(absent))

    if missing:
        wb.close()
        sys.exit('\n✗ 코드가 다시 쓰는데 표에 없는 키: %s' % ', '.join(missing))
    print('  다시 쓰는 키 %d개 전부 표에 있습니다.' % len(MUST_EXIST))

    if not added and not filled:
        print('  바뀐 것이 없습니다 — 저장하지 않았습니다.')
        return

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.bak')
    wb.save(STRING_XLSX)
    print('  저장:', os.path.basename(STRING_XLSX), '(백업 .bak)')
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
