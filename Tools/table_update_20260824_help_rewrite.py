# -*- coding: utf-8 -*-
"""도움말 표 — <b>문구를 처음 하는 사람 기준으로 다시 쓰고</b>, «화면에서 짚어 주기» 단계를
새로 만든다 (2026-08-24).

유저 지시: *"말 좀 더 다듬어줘 지금 너무 ai 티 나니까 문어체로 깔끔하게 다듬어서 이해하기
쉽게 설명해줘 게임을 아예 처음 하는 사람도 이해할 수 있도록. 자세히 보기에서 실제 ui로
연결하고 <b>빨간 테두리 선으로 하나하나 설명</b>해주는 기능 넣어주고"*.

★★ <b>«AI 티» 가 무엇이었는지 먼저 찾았다</b>
------------------------------------------
고칠 데를 느낌으로 찾으면 결국 취향이 된다. 그래서 앞 초안에서 <b>무엇이 되풀이되는지</b>
세어 보고 그 버릇을 규칙으로 적었다.

| 버릇 | 앞 초안 | 왜 «AI 티» 로 읽히는가 |
|---|---|---|
| 문장 끝의 격언 | 27개 중 <b>25개</b>가 ★ 로 시작하는 한 줄로 끝났다 | 설명이 아니라 <b>논평</b>이다 |
| «…» 기호 | <b>31군데</b> | 사람이 쓰는 글에는 이렇게 자주 나오지 않는다 |
| 대구 | 「…은 …이고, …은 …입니다」 | 뜻은 안 늘고 길이만 는다 |
| 추상 요약 | 「모든 판단의 기준은 …입니다」 | 처음 하는 사람에게 <b>아무것도 알려주지 않는다</b> |
| 화면 이름 없음 | 「액션 패널의」 | 초보자는 «액션 패널» 이 어디인지 모른다 |

<b>그래서 세운 규칙 다섯</b>
  ① <b>한 문장에 한 가지</b>만 담는다. 문장을 짧게 끊는다.
  ② <b>기호를 쓰지 않는다.</b> ★ · ⚠ · «» 를 전부 걷어내고 <b>강조만</b> 남긴다.
  ③ <b>화면의 어디인지 말한다.</b> 「오른쪽 버튼 묶음의 강화」처럼 자리를 함께 적는다.
  ④ <b>논평하지 않는다.</b> 「…가 이 게임의 리듬입니다」 같은 줄을 지우고, 대신
     «그래서 무엇을 하면 되는가» 를 적는다.
  ⑤ <b>말투는 문어체 ~입니다</b>를 그대로 유지한다(앞선 확정사항이고, 이 프로젝트의
     플레이어용 글이 전부 그 말투다).

⚠ <b>뜻은 하나도 바꾸지 않았다.</b> 규칙·수치·조건은 앞 초안 그대로다 —
  «다시 쓴다» 가 «다시 기획한다» 가 되면 도움말이 게임과 어긋난다.

★★ <b>HelpStep 시트 신설 — 빨간 테두리로 짚어 줄 단계</b>
-----------------------------------------------------
`help_id · step_order · target_path · step_text` 네 칸이다.
`target_path` 는 <b>씬의 경로</b>이고, 비우면 짚지 않고 글만 보여준다.

⚠ <b>창 안의 칸을 가리키지 않았다</b> — 창은 평소 닫혀 있어 짚어도 안 보인다.
  대신 <b>그 창을 여는 버튼</b>을 가리킨다. 그것이 유저가 실제로 눌러야 하는 곳이다.
⚠ <b>월드 오브젝트도 가리키지 않았다</b>(넥서스·몬스터) — 짚는 쪽이 UI 좌표로 계산하므로
  RectTransform 이 아닌 것은 못 짚는다. 그런 단계는 `target_path` 를 비웠다.

사용법:  py -3 Tools/table_update_20260824_help_rewrite.py
다음:    py -3 Tools/help_string_merge.py
         py -3 Tools/gen_string_table.py
         py -3 Tools/link_string_keys.py
         py -3 Tools/gen_help_assets.py
"""

import io
import os
import shutil
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

XLSX = os.path.join(TABLE_DIR, "Last_Sanctuary_도움말테이블_Ver01.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="1F3B4D")
HEAD_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
BODY_FONT = Font(name="맑은 고딕", size=10)
THIN = Side(style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ──────────────────────────────────────────────────────────────────────────
# 문구 — help_id → (제목, 요약, 본문)
#
# 요약 = 조언 카드에 뜨는 두 줄. «무엇인가» 와 «그래서 무엇을 하면 되는가».
# 본문 = 백과에 뜨는 세~다섯 줄. 한 줄에 한 가지.
# ──────────────────────────────────────────────────────────────────────────
TEXTS = {
    # ══════════════ 기본 ══════════════
    "help_nexus": (
        "성역과 넥서스",
        "화면 가운데의 큰 건물이 넥서스입니다.\n"
        "넥서스가 부서지면 그 자리에서 게임이 끝납니다.",
        "넥서스는 이 성역의 심장입니다. 체력이 0이 되면 즉시 패배합니다.\n"
        "넥서스 둘레의 밝은 땅이 여러분이 지켜야 할 범위입니다.\n"
        "캐릭터는 따로 지시하지 않으면 스스로 넥서스 쪽으로 돌아와 지킵니다.\n"
        "넥서스를 마우스로 누르면 남은 체력과 일러스트를 볼 수 있습니다."),

    "help_wave": (
        "웨이브와 정비 시간",
        "지금은 <b>정비 시간</b>입니다. 몬스터가 아직 없습니다.\n"
        "시간이 다 되면 몬스터가 몰려오니, 그 전에 준비를 마치십시오.",
        "한 웨이브는 정비, 진군, 전투의 세 단계로 흘러갑니다.\n"
        "<b>정비</b>는 몬스터가 없는 시간입니다. 캐릭터를 만들고 강화하는 시간입니다.\n"
        "<b>진군</b>은 몬스터가 걸어오는 동안입니다. 아직 전투 시간은 흐르지 않습니다.\n"
        "<b>전투</b>는 제한 시간이 흐릅니다. 시간 안에 다 잡으면 웨이브가 끝납니다.\n"
        "웨이브를 일찍 끝내면 다음 정비 시간이 그만큼 길어집니다."),

    "help_energy": (
        "에너지",
        "에너지는 이 게임의 하나뿐인 자원입니다.\n"
        "캐릭터를 만들고 강화하는 데 모두 이 에너지를 씁니다.",
        "에너지는 화면 왼쪽 위에 숫자로 표시됩니다.\n"
        "<b>쓰는 곳</b>은 캐릭터 생성과 강화, 그리고 포탑 건설입니다.\n"
        "<b>버는 곳</b>은 맵에 사는 중립 몬스터 사냥과 웨이브 보상입니다.\n"
        "캐릭터를 만들 때마다 다음 생성 비용이 오르고, 강화 비용도 레벨마다 오릅니다.\n"
        "그래서 한 명을 더 만들지, 있는 사람을 키울지 매번 고민하게 됩니다."),

    "help_create": (
        "캐릭터 생성",
        "오른쪽 버튼 묶음의 <b>캐릭터 생성</b>을 누르면 아군이 한 명 늘어납니다.\n"
        "누가 나올지는 정해져 있지 않습니다.",
        "버튼에 적힌 숫자가 지금 필요한 에너지입니다.\n"
        "등장하는 인물과 능력치는 무작위로 정해집니다. 한 판에 같은 인물은 한 번만 나옵니다.\n"
        "인물마다 근접, 원거리, 마법, 회복 중 하나의 역할을 맡습니다.\n"
        "역할에 따라 적과 유지하는 거리와 싸우는 방식이 다릅니다.\n"
        "새로 만든 캐릭터는 지시가 없으면 주변을 돌아보다가 넥서스로 돌아옵니다."),

    "help_camera": (
        "화면 보기와 선택",
        "빈 땅을 누른 채 끌면 화면이 움직입니다.\n"
        "마우스 바퀴를 굴리면 확대와 축소가 됩니다.",
        "<b>화면 옮기기</b>는 빈 땅을 누른 채 끌면 됩니다.\n"
        "<b>확대와 축소</b>는 마우스 바퀴로 합니다.\n"
        "캐릭터나 몬스터를 누르면 초상화 창이 열려 상태를 볼 수 있습니다.\n"
        "오른쪽 아래 미니맵에는 지금까지 밝혀 둔 곳만 나타납니다.\n"
        "초상화 창과 미니맵은 자리가 고정입니다. 다른 창은 배경을 잡아 끌어 옮길 수 있습니다."),

    # ══════════════ 전투 ══════════════
    "help_battle": (
        "스스로 벌어지는 전투",
        "공격 명령은 따로 없습니다. 캐릭터가 알아서 싸웁니다.\n"
        "여러분이 정하는 것은 <b>누구를 어디에 세울지</b>입니다.",
        "캐릭터는 적을 찾고, 다가가고, 공격하고, 물러나는 일을 스스로 합니다.\n"
        "그래서 이 게임에서 하는 일은 전투 조작이 아니라 배치와 지침입니다.\n"
        "배치는 집결지와 부대로, 싸우는 방식은 전술 지침으로 정합니다.\n"
        "무슨 일이 있었는지는 왼쪽 아래 기록창에 한 줄씩 남습니다.\n"
        "안개에 가려진 적은 캐릭터도 보지 못합니다. 먼저 살펴봐야 싸울 수 있습니다."),

    "help_enrage": (
        "광폭화",
        "전투 시간이 다 됐는데 몬스터가 남아 <b>광폭화</b>가 시작됐습니다.\n"
        "남은 적이 시간이 갈수록 강해지니 서둘러 정리하십시오.",
        "광폭화는 시간을 끌면 반드시 불리해지도록 만든 단계입니다.\n"
        "남은 적의 능력치가 시간이 지날수록 계속 올라갑니다.\n"
        "그래서 버티는 것으로는 끝나지 않습니다. 남은 적을 지금 잡아야 합니다.\n"
        "흩어진 부대를 한곳으로 모아 남은 무리에 붙이는 것이 가장 확실합니다.\n"
        "웨이브가 끝나면 올라간 능력치는 원래대로 돌아갑니다."),

    "help_death": (
        "쓰러진 캐릭터",
        "쓰러진 캐릭터는 판에서 사라집니다. 대부분 되살릴 수 없습니다.\n"
        "죽기 전에 물러나게 하는 것이 유일한 대비입니다.",
        "체력이 0이 되면 그 캐릭터는 사라집니다.\n"
        "그동안 올린 강화와 끼운 유물도 함께 사라집니다.\n"
        "되살리는 일은 일부 인물이 가진 특별한 능력으로만 가능합니다.\n"
        "전술 지침의 <b>후퇴 기준</b>을 올려 두면 위험해지기 전에 스스로 물러납니다.\n"
        "회복 역할의 캐릭터를 뒤에 두면 물러난 동료를 치료해 줍니다."),

    "help_hit": (
        "명중과 크리티컬",
        "공격이 빗나가는 것은 고장이 아닙니다.\n"
        "맞을 확률과 크게 맞을 확률은 능력치로 정해집니다.",
        "공격할 때마다 먼저 맞았는지 판정하고, 맞았으면 크리티컬인지 다시 판정합니다.\n"
        "원거리 역할은 처음 명중률이 낮게 시작하며, 강화로 올릴 수 있습니다.\n"
        "빗나감과 크리티컬은 캐릭터 머리 위 숫자와 기록창에 표시됩니다.\n"
        "방어력은 받는 피해를 줄여 줍니다. 다만 일부 보스 기술은 방어력을 무시합니다."),

    "help_boss": (
        "보스",
        "보스가 나타났습니다. 화면 위쪽에 보스의 체력바가 보입니다.\n"
        "보스는 일반 몬스터와 같은 방식으로는 버티기 어렵습니다.",
        "정해진 웨이브마다 보스가 한 마리 등장합니다.\n"
        "보스에게는 고유한 기술이 있습니다. 묶거나 약하게 만들거나 광선을 쏘는 것들입니다.\n"
        "어떤 기술을 썼는지는 기록창에 이름으로 찍힙니다.\n"
        "일반 몬스터와 보스를 나눠서 상대하면 둘 다 놓칩니다. 한쪽부터 정리하십시오.\n"
        "보스를 누르면 일러스트와 지금 상태를 볼 수 있습니다."),

    # ══════════════ 성장 ══════════════
    "help_upgrade": (
        "강화",
        "에너지를 써서 캐릭터의 레벨을 올렸습니다.\n"
        "레벨이 오를 때마다 능력치 몇 칸이 함께 올라갑니다.",
        "오른쪽 버튼 묶음의 <b>강화</b>를 누르면 성장 창이 열립니다.\n"
        "레벨이 오를 때마다 어떤 능력치가 오를지는 그때그때 정해집니다.\n"
        "인물이 주로 쓰는 능력치가 더 자주, 더 많이 오릅니다.\n"
        "능력치는 <b>100</b>이 상한입니다. 영웅 각성만 그 위로 올라갑니다.\n"
        "강화 비용은 레벨마다 오릅니다. 캐릭터가 죽으면 올린 것도 함께 사라집니다."),

    "help_stats": (
        "능력치 읽는 법",
        "능력치는 열두 칸입니다.\n"
        "성장 창에 보이는 숫자는 지금 실제로 적용되는 값입니다.",
        "체력, 공격력, 방어력, 재생, 명중, 크리티컬, 공격 속도, 이동 속도, 사거리, 저항력 등입니다.\n"
        "성장 창의 숫자에는 유물과 특별한 능력, 각성 보너스가 이미 더해져 있습니다.\n"
        "<b>저항력</b>은 정신이 침식되는 속도와 회복되는 속도를 정합니다.\n"
        "능력치의 상한은 100입니다. 인물마다 주로 쓰는 칸이 다릅니다."),

    "help_awaken": (
        "영웅 각성",
        "많이 싸운 캐릭터가 드물게 <b>영웅</b>으로 각성합니다.\n"
        "각성하면 능력치 상한 100을 넘어섭니다.",
        "적을 많이 잡고 레벨 조건까지 채운 캐릭터에게 낮은 확률로 일어납니다.\n"
        "각성하면 능력치에 보너스가 붙고, 상한 100을 넘어설 수 있습니다.\n"
        "회복 역할은 적을 잡지 않아도 동료를 많이 치료하면 각성할 수 있습니다.\n"
        "각성한 순간 화면에 금색 글자가 뜨고, 그 뒤로 이름이 금색으로 남습니다."),

    "help_relic_dig": (
        "유물 발굴",
        "맵에 노란 <b>느낌표</b>가 떴습니다. 무언가 묻혀 있는 자리입니다.\n"
        "느낌표를 누르면 캐릭터를 보내 파낼 수 있습니다.",
        "발굴할 수 있는 자리는 캐릭터의 시야에 들어오면 느낌표로 나타납니다.\n"
        "느낌표를 누르면 창이 열리고, 파낼지 그냥 둘지 고를 수 있습니다.\n"
        "파러 보낸 캐릭터는 그동안 자리를 비웁니다. 정비 시간에 해 두는 편이 안전합니다.\n"
        "무엇이 나오는지는 파낸 뒤에 정해집니다. 에너지나 유물이 나오고, 손해를 볼 때도 있습니다.\n"
        "다 파내면 느낌표는 사라집니다. 보스와 강한 중립 몬스터도 유물을 남깁니다."),

    "help_relic_equip": (
        "유물 장착",
        "유물을 얻었습니다. 캐릭터에게 끼워야 효과가 생깁니다.\n"
        "오른쪽 버튼 묶음의 <b>유물 관리</b>에서 끼울 수 있습니다.",
        "유물 관리 창에서 유물을 고른 뒤, 끼울 캐릭터를 로스터에서 선택합니다.\n"
        "성장 창 아래쪽에서도 지금 낀 유물을 보고 바꿀 수 있습니다.\n"
        "등급은 일반, 레어, 에픽 순으로 좋아지고 그만큼 드물게 나옵니다.\n"
        "일부 유물이 주는 능력치는 상한 100을 넘어서 붙습니다.\n"
        "유물을 낀 캐릭터가 죽으면 그 유물도 함께 잃습니다."),

    # ══════════════ 지휘 ══════════════
    "help_squad": (
        "부대",
        "여러 캐릭터를 한 <b>부대</b>로 묶으면 함께 움직입니다.\n"
        "혼자 떨어져 죽는 일을 막는 가장 쉬운 방법입니다.",
        "오른쪽 버튼 묶음의 <b>부대 설정</b>에서 부대를 만듭니다.\n"
        "부대를 만든 뒤 왼쪽 로스터에서 캐릭터를 눌러 배정합니다.\n"
        "같은 부대에 든 캐릭터는 함께 돌아다니고 함께 사냥합니다.\n"
        "부대마다 색이 정해지고, 로스터의 캐릭터 테두리가 그 색으로 묶입니다.\n"
        "부대 이름은 원하는 대로 고칠 수 있습니다."),

    "help_rally": (
        "집결지",
        "<b>집결지</b>를 찍으면 그 부대가 그 자리를 맡아 지킵니다.\n"
        "깃발을 잠깐 누른 채 끌면 자리를 옮길 수 있습니다.",
        "부대 설정 창의 부대 칸에서 집결지 만들기를 누르고, 맵에서 자리를 정합니다.\n"
        "깃발을 누르면 그 부대가 맡은 범위가 테두리로 보입니다.\n"
        "깃발을 1초쯤 누른 채 끌면 집결지를 다른 곳으로 옮길 수 있습니다.\n"
        "집결지가 있는 부대는 그 자리를 중심으로 싸웁니다. 없으면 넥서스로 돌아옵니다.\n"
        "집결지는 언제든 해제할 수 있고, 저장할 때 함께 저장됩니다."),

    "help_tactics": (
        "전술 지침",
        "캐릭터마다 <b>전방, 중위, 후방</b> 중 어디에 설지 정할 수 있습니다.\n"
        "오른쪽 버튼 묶음의 <b>전술 지침</b>에서 바꿉니다.",
        "<b>전방</b>은 앞에 서서 적을 직접 받아냅니다.\n"
        "<b>중위</b>는 조금 뒤에서 앞줄을 돕습니다.\n"
        "<b>후방</b>은 사거리 끝에서 싸웁니다. 원거리와 회복 역할에 맞습니다.\n"
        "중위와 후방도 자기 구역에서 전투가 벌어지면 사거리 안까지는 다가갑니다.\n"
        "한 번 붙은 적과는 교전을 계속합니다. 제자리로 돌아가 빙빙 돌지 않습니다."),

    "help_retreat": (
        "후퇴 기준",
        "체력이 얼마나 남았을 때 물러날지 정할 수 있습니다.\n"
        "전술 지침 창에서 캐릭터마다 따로 정합니다.",
        "후퇴 기준을 60퍼센트로 두면 체력이 그만큼 남았을 때 스스로 물러납니다.\n"
        "기준을 너무 높이면 조금만 맞아도 물러나 전선이 자꾸 무너집니다.\n"
        "기준을 너무 낮추면 물러나기 전에 죽습니다.\n"
        "앞줄은 낮게, 뒷줄은 높게 잡는 것이 무난한 시작점입니다.\n"
        "물러난 캐릭터는 회복 역할이 가까이 있으면 치료를 받습니다."),

    # ⚠ "help_build"(포탑 건설)는 <b>지웠다</b> (2026-08-25 · 유저: *"도움말에서 포탑
    #   건설 관련 설명 삭제해 해당 기능 없어졌어"*). 여기 되살리면 다시 돌릴 때
    #   표에 <b>도로 살아난다</b> — 씬의 BuildButton 도 꺼져 있다.

    # ══════════════ 위험 ══════════════
    "help_erosion": (
        "침식",
        "싸우는 동안 캐릭터의 정신이 <b>침식</b>됩니다.\n"
        "침식이 100에 닿으면 정신 이상이 나타납니다.",
        "전투 중에는 침식이 오르고, 전투가 끝나고 잠시 뒤부터 저절로 회복됩니다.\n"
        "오르는 속도와 회복되는 속도는 <b>저항력</b> 능력치가 정합니다.\n"
        "침식은 로스터와 성장 창의 보라색 게이지로 확인할 수 있습니다.\n"
        "100에 닿으면 정신 이상이 하나 나타납니다.\n"
        "정비 시간에 위험한 캐릭터를 뒤로 빼 두면 그만큼 회복됩니다."),

    "help_mental_error": (
        "정신 이상",
        "침식이 100에 닿아 <b>정신 이상</b>이 나타났습니다.\n"
        "나쁜 것만 있는 것은 아닙니다.",
        "정신 이상은 여러 종류 중에서 하나가 무작위로 걸립니다.\n"
        "능력치를 깎는 것도 있고, 행동을 바꾸는 것도 있습니다.\n"
        "드물게는 오히려 도움이 되는 것도 섞여 있습니다.\n"
        "나타나면 화면에 이름이 뜨고, 로스터에 상태로 표시됩니다.\n"
        "침식을 아예 쌓지 않는 것은 불가능합니다. 누구에게 언제 쌓게 할지를 고르는 일입니다."),

    "help_neutral": (
        "중립 몬스터",
        "맵에 사는 중립 몬스터를 잡았습니다. 이것이 주된 수입입니다.\n"
        "먼저 건드리지 않으면 이쪽으로 덤비지는 않습니다.",
        "중립 몬스터는 자기 서식지 주변을 돌아다닙니다. 서식지는 바닥 색이 다릅니다.\n"
        "캐릭터는 주변을 돌아다니거나 이동하다가 가까운 중립을 스스로 사냥합니다.\n"
        "종류마다 주는 에너지가 다릅니다.\n"
        "같은 종을 많이 잡으면 그 종의 다음 개체가 조금 더 강해집니다.\n"
        "서식지의 주인이 죽으면 그 땅은 바깥부터 서서히 사라집니다."),

    "help_epic": (
        "에픽 중립과 토벌 지시",
        "아주 강한 <b>에픽</b> 중립 몬스터를 발견했습니다.\n"
        "준비 없이 붙으면 부대가 통째로 사라지니 조심하십시오.",
        "에픽을 발견하면 오른쪽 버튼 묶음의 <b>토벌 지시</b> 목록에 올라옵니다.\n"
        "토벌 지시를 내린 부대만 잡으러 갑니다. 지시가 없으면 지나쳐 갑니다.\n"
        "목록에는 그 에픽을 잡는 데 필요한 수준이 함께 나옵니다. 그것을 보고 판단하십시오.\n"
        "에픽은 크게 보상합니다. 에너지도 많고 고유한 유물을 남깁니다.\n"
        "잡아도 시간이 지나면 다시 나타납니다. 다시 나타날 때마다 더 강해집니다."),

    "help_event": (
        "사건",
        "<b>사건</b>이 일어나 창이 열렸습니다.\n"
        "고른 선택지에 따라 결과가 달라지고, 되돌릴 수 없습니다.",
        "사건은 웨이브가 끝날 때나 시간이 지날 때 조건이 맞으면 일어납니다.\n"
        "상황 설명을 읽고 선택지 중 하나를 고릅니다.\n"
        "좋은 결과도 나쁜 결과도 고른 그 자리에서 바로 적용됩니다.\n"
        "선택에 따라 유물을 얻는 사건도 있습니다.\n"
        "사건 창이 열려 있는 동안에도 시간은 흐릅니다. 전투 중이라면 상황을 먼저 살피십시오."),

    # ══════════════ 운영 ══════════════
    "help_save": (
        "저장과 이어하기",
        "게임이 자동으로 저장되었습니다.\n"
        "저장 칸은 하나뿐이라, 새로 저장하면 앞의 것을 덮어씁니다.",
        "강화할 때, 캐릭터가 죽을 때, 웨이브를 넘길 때 자동으로 저장됩니다.\n"
        "환경 설정 창에서 직접 저장하거나, 저장하고 로비로 나갈 수 있습니다.\n"
        "로비의 이어하기를 누르면 저장한 판을 그대로 잇습니다.\n"
        "로비의 새로하기는 저장을 <b>지웁니다</b>. 이어할 판이 있으면 주의하십시오.\n"
        "맵, 밝혀 둔 곳, 중립 몬스터, 유물, 부대까지 함께 저장됩니다."),

    "help_speed": (
        "배속과 일시정지",
        "화면 아래 버튼으로 게임 속도를 바꿀 수 있습니다.\n"
        "P 키를 누르면 멈추고, 다시 누르면 이어집니다.",
        "속도는 1배, 2배, 4배, 8배 중에서 고릅니다. 키보드 1부터 4로도 바꿉니다.\n"
        "멈춰 있는 동안에도 창을 열어 보고 지침을 바꿀 수 있습니다.\n"
        "몬스터가 없는 정비 시간은 빠르게 넘기고, 전투에서는 속도를 낮추면 편합니다.\n"
        "이기거나 지면 시간은 저절로 멈춥니다."),
}

# ──────────────────────────────────────────────────────────────────────────
# 「자세히 보기」가 <b>열어야 하는 창</b> — help_id → 씬 경로
#
# ★★ 2026-08-24 유저 지시로 <b>구조를 통째로 바꿨다</b>:
#   *"듀토리얼 이벤트의 배선이 어수선해 … 자세히 보기를 누르면 다음 기능을 설명하는 것이
#   아니라 <b>해당 ui를 직접 띄워서</b> 설명하는 방식으로 만들어야 하고"* ·
#   *"전술 지침을 누르면 … 거기서 자세히 보기를 누르면 <b>실제 전술 지침 ui를 띄워놓고
#   각 영역에 대해</b> 빨간색 테두리로 설명해 주어야 함"*.
#
#   앞 초안은 단계마다 <b>다른 HUD</b>를 짚었다 — 「전술 지침 버튼」 → 「로스터」 → 「기록창」.
#   눈이 화면을 세 번 건너뛰는데 정작 <b>그 창 안에 무엇이 있는지</b>는 알려주지 않았다.
#   그것이 «어수선하다» 의 정체였다.
#
# ⚠ 여기 적은 창은 <b>바깥에서 열 수 있어야</b> 한다 —
#   `HudExclusive.TryOpen` 에 그 창의 가지가 있어야 한다. 없으면 안내가 경고를 찍고
#   창 없이 글만 보여준다(사건 창·발굴 창은 «상황이 만들어» 뜨는 창이라 넣지 않았다).
# ──────────────────────────────────────────────────────────────────────────
OPEN_PANEL = {
    "help_upgrade": "UI_Root/HUD_Growth",
    "help_stats": "UI_Root/HUD_Growth",
    "help_awaken": "UI_Root/HUD_Growth",
    "help_erosion": "UI_Root/HUD_Growth",
    "help_mental_error": "UI_Root/HUD_Growth",
    "help_tactics": "UI_Root/HUD_Tactics",
    "help_retreat": "UI_Root/HUD_Tactics",
    "help_squad": "UI_Root/HUD_Squad",
    "help_rally": "UI_Root/HUD_Squad",
    "help_relic_equip": "UI_Root/HUD_Relics",
    "help_epic": "UI_Root/HUD_Subjugate",
    "help_save": "UI_Root/HUD_Settings",
}

# ──────────────────────────────────────────────────────────────────────────
# 화면에서 짚어 주기 — (help_id, step_order, target_path, step_text)
#
# ★★ <b>규칙 하나 — 한 항목의 단계는 «한 UI 안» 에서만 머문다.</b>
#   창을 여는 항목은 <b>그 창 안</b>만, 창이 없는 항목은 <b>늘 보이는 HUD 하나</b> 안만 짚는다.
#   `check()` 가 이 규칙을 검산한다.
#
# ★★ <b>단계를 아예 두지 않은 항목이 열한 개다</b> (유저 지시:
#   *"단순히 넥서스가 파괴되면 게임이 종료된다는 간단한 규칙 같은거
#   (다른 ui와 연결되지 않아도 되는 기능)은 그냥 자세히 보기 없어도 됨"*).
#   그 항목들은 「자세히 보기」 버튼이 <b>뜨지 않는다</b> — 넥서스·전투·광폭화·사망·명중·보스·
#   중립·발굴·사건·조작·건설. 짚을 칸이 없는 <b>규칙과 개념</b>이라 글로 끝내는 것이 맞다.
#
# ⚠ target_path 를 비우면 짚지 않고 글만 보여준다(조작법처럼 칸이 아닌 것).
# ──────────────────────────────────────────────────────────────────────────
STEPS = [
    # ══════════ 창을 열지 않고 «늘 보이는 HUD» 를 짚는 것 (4항목) ══════════
    ("help_wave", 1, "UI_Root/HUD_Wave/Phase",
     "지금이 어느 단계인지 알려 주는 칸입니다. 정비, 진군, 전투, 광폭화 중 하나가 적힙니다."),
    ("help_wave", 2, "UI_Root/HUD_Wave/Timer",
     "남은 시간입니다. 정비 단계에서 이 시간이 0이 되면 몬스터가 몰려옵니다."),

    ("help_energy", 1, "UI_Root/HUD_Energy/Energy",
     "지금 가진 에너지입니다. 캐릭터를 만들거나 강화할 때 이 숫자에서 빠져나갑니다."),

    ("help_create", 1, "UI_Root/HUD_Actions/Buttons/CreateButton",
     "이 버튼을 누르면 아군이 한 명 늘어납니다. 버튼에 적힌 숫자가 지금 드는 에너지입니다."),

    ("help_speed", 1, "UI_Root/HUD_Speed/x1",
     "게임 속도를 고르는 버튼입니다. 지금 걸린 속도는 색이 밝게 표시됩니다."),
    ("help_speed", 2, "UI_Root/HUD_Speed/x8",
     "가장 빠른 8배입니다. 몬스터가 없는 정비 시간을 넘길 때 쓰면 편합니다."),
    ("help_speed", 3, "UI_Root/HUD_Speed/Pause",
     "이 버튼이나 P 키로 시간을 멈춥니다. 멈춘 동안에도 창을 열어 지침을 바꿀 수 있습니다."),

    # ══════════ 성장 창(HUD_Growth) 을 띄워 놓고 짚는 것 (5항목) ══════════
    ("help_upgrade", 1, "UI_Root/HUD_Growth/Info/EnhanceButton",
     "이 버튼이 강화입니다. 한 번 누를 때마다 고른 캐릭터의 레벨이 1 올라갑니다."),
    ("help_upgrade", 2, "UI_Root/HUD_Growth/Info/CostValue",
     "한 번 강화하는 데 드는 에너지입니다. 레벨이 오를수록 이 값도 함께 커집니다."),
    ("help_upgrade", 3, "UI_Root/HUD_Growth/Info/Level",
     "지금 레벨입니다. 강화할 때마다 여기가 올라갑니다."),
    ("help_upgrade", 4, "UI_Root/HUD_Growth/Stats/Grid",
     "레벨이 오르면 이 능력치 칸들 중 몇 개가 함께 올라갑니다. 어느 칸이 오를지는 그때 정해집니다."),

    ("help_stats", 1, "UI_Root/HUD_Growth/Stats/Grid",
     "능력치 열두 칸입니다. 여기 보이는 숫자에는 유물과 각성 보너스가 이미 더해져 있습니다."),
    ("help_stats", 2, "UI_Root/HUD_Growth/Stats/GrowthTypes",
     "이 캐릭터가 주로 쓰는 능력치입니다. 강화할 때 이쪽이 더 자주, 더 많이 오릅니다."),
    ("help_stats", 3, "UI_Root/HUD_Growth/Stats/PassiveGrid",
     "이 캐릭터가 타고난 특별한 능력입니다. 강화로는 바뀌지 않습니다."),
    ("help_stats", 4, "UI_Root/HUD_Growth/RelicBar",
     "지금 끼고 있는 유물입니다. 유물이 주는 능력치는 상한 100을 넘어서 붙습니다."),

    ("help_awaken", 1, "UI_Root/HUD_Growth/Info/Name",
     "각성한 캐릭터는 이름이 금색으로 바뀝니다. 로스터에서도 같은 색으로 보입니다."),
    ("help_awaken", 2, "UI_Root/HUD_Growth/Info/Note",
     "각성 여부와 남은 조건이 여기에 적힙니다."),
    ("help_awaken", 3, "UI_Root/HUD_Growth/Stats/Grid",
     "각성하면 이 칸들에 보너스가 얹히고, 상한 100을 넘어설 수 있습니다."),

    ("help_erosion", 1, "UI_Root/HUD_Growth/Info/ErosionBack",
     "보라색 게이지가 침식입니다. 싸우는 동안 차오르고, 전투가 끝나면 저절로 줄어듭니다."),
    ("help_erosion", 2, "UI_Root/HUD_Growth/Stats/Grid",
     "능력치 중 저항력이 침식이 차는 속도와 줄어드는 속도를 정합니다."),

    ("help_mental_error", 1, "UI_Root/HUD_Growth/Info/ErosionBack",
     "이 게이지가 끝까지 차면 정신 이상이 하나 나타납니다."),
    ("help_mental_error", 2, "UI_Root/HUD_Growth/Info/Note",
     "지금 걸린 정신 이상의 이름이 여기에 적힙니다. 로스터에도 함께 표시됩니다."),

    # ══════════ 전술 지침 창(HUD_Tactics) 을 띄워 놓고 짚는 것 (2항목) ══════════
    ("help_tactics", 1, "UI_Root/HUD_Tactics/Col1/Pos",
     "전방, 중위, 후방 중 어디에 설지 고릅니다. 앞줄은 적을 직접 받아내고 뒷줄은 사거리 끝에서 싸웁니다."),
    ("help_tactics", 2, "UI_Root/HUD_Tactics/Col1/Type",
     "어떤 방식으로 공격할지 고릅니다."),
    ("help_tactics", 3, "UI_Root/HUD_Tactics/Col2/Target",
     "여러 적이 있을 때 누구를 먼저 칠지 고릅니다."),
    ("help_tactics", 4, "UI_Root/HUD_Tactics/Col3/Wave",
     "몬스터가 몰려올 때 어떻게 움직일지 고릅니다."),
    ("help_tactics", 5, "UI_Root/HUD_Tactics/Col3/Summary",
     "고른 지침을 한눈에 정리해 보여 줍니다. 여기만 봐도 지금 설정이 확인됩니다."),

    ("help_retreat", 1, "UI_Root/HUD_Tactics/Col2/RetreatSlider",
     "이 막대를 끌어 후퇴 기준을 정합니다. 체력이 그만큼 남으면 스스로 물러납니다."),
    ("help_retreat", 2, "UI_Root/HUD_Tactics/Col2/RetreatValue",
     "지금 정한 기준입니다. 앞줄은 낮게, 뒷줄은 높게 잡는 것이 무난합니다."),
    ("help_retreat", 3, "UI_Root/HUD_Tactics/Col2/RetreatAction",
     "물러난 뒤 무엇을 할지 고릅니다."),

    # ══════════ 부대 설정 창(HUD_Squad) 을 띄워 놓고 짚는 것 (2항목) ══════════
    ("help_squad", 1, "UI_Root/HUD_Squad/Header/AddButton",
     "이 버튼으로 새 부대를 만듭니다. 부대는 여섯 개까지 만들 수 있습니다."),
    ("help_squad", 2, "UI_Root/HUD_Squad/Body/Grid",
     "만든 부대가 여기에 칸으로 늘어섭니다. 왼쪽 로스터에서 캐릭터를 눌러 배정합니다."),

    ("help_rally", 1, "UI_Root/HUD_Squad/Body/Grid",
     "부대 칸마다 집결지를 만드는 버튼과 해제하는 버튼이 있습니다."),
    ("help_rally", 2, "",
     "만들기를 누른 뒤 맵에서 자리를 찍으십시오. 깃발을 1초쯤 누른 채 끌면 자리를 옮길 수 있습니다."),

    # ══════════ 유물 관리 창(HUD_Relics) 을 띄워 놓고 짚는 것 (1항목) ══════════
    ("help_relic_equip", 1, "UI_Root/HUD_Relics/List",
     "지금까지 얻은 유물이 여기에 쌓입니다. 등급이 높은 것이 위로 옵니다."),
    ("help_relic_equip", 2, "UI_Root/HUD_Relics/Detail/Effect",
     "고른 유물이 무엇을 해 주는지 여기에 적힙니다."),
    ("help_relic_equip", 3, "UI_Root/HUD_Relics/Detail/EquipButton",
     "이 버튼으로 끼웁니다. 끼울 대상은 왼쪽 로스터에서 고른 캐릭터입니다."),
    ("help_relic_equip", 4, "UI_Root/HUD_Relics/Detail/Wearer",
     "그 유물을 이미 누가 끼고 있으면 여기에 이름이 나옵니다."),

    # ══════════ 토벌 지시 창(HUD_Subjugate) 을 띄워 놓고 짚는 것 (1항목) ══════════
    ("help_epic", 1, "UI_Root/HUD_Subjugate/Targets/List",
     "발견한 에픽 몬스터가 여기에 오릅니다. 잡는 데 필요한 수준도 함께 나옵니다."),
    ("help_epic", 2, "UI_Root/HUD_Subjugate/Squads/List",
     "어느 부대를 보낼지 여기서 고릅니다. 지시를 내린 부대만 잡으러 갑니다."),
    ("help_epic", 3, "UI_Root/HUD_Subjugate/Hint",
     "지금 무엇을 골라야 하는지 여기에 안내가 나옵니다."),

    # ══════════ 환경 설정 창(HUD_Settings) 을 띄워 놓고 짚는 것 (1항목) ══════════
    ("help_save", 1, "UI_Root/HUD_Settings/Body/SaveButton",
     "이 버튼으로 지금 상태를 저장합니다. 저장 칸은 하나뿐입니다."),
    ("help_save", 2, "UI_Root/HUD_Settings/Body/LobbyButton",
     "저장하고 로비로 나갑니다. 로비의 이어하기로 이 판을 다시 잇습니다."),
    ("help_save", 3, "UI_Root/HUD_Settings/Body/Status",
     "마지막으로 저장한 때가 여기에 적힙니다."),
]

#: ⚠ 머리글은 <b>필드명 한 줄</b>이다 — 이 표의 다른 시트(`Help`·`StringKeys`)와 같은 규약이고,
#   `gen_help_assets.py` 가 1행에서 필드명을 읽는다. 한글 라벨을 1행에 넣으면 <b>전부 못 읽는다</b>.
#   (다른 데이터 테이블은 «1행 한글 · 2행 필드 · 3행 타입» 세 줄이다 — 이 표만 한 줄이다.)
STEP_KEYS = ["help_id", "step_order", "target_path", "step_text"]
STEP_WIDTH = [20, 10, 46, 84]


def check():
    """쓰기 <b>전에</b> 스스로 검산한다 — 틀린 표를 만드는 것이 안 만드는 것보다 나쁘다."""
    bad = []

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ids = []
    for r in wb["Help"].iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            ids.append(str(r[0]).strip())

    for hid in TEXTS:
        if hid not in ids:
            bad.append("문구를 쓴 %s 가 Help 시트에 없습니다" % hid)
    for hid in ids:
        if hid not in TEXTS:
            bad.append("Help 시트의 %s 에 새 문구가 없습니다" % hid)

    seen = set()
    for hid, order, path, text in STEPS:
        if hid not in ids:
            bad.append("단계의 %s 가 Help 시트에 없습니다" % hid)
        if (hid, order) in seen:
            bad.append("%s 의 단계 %d 가 겹칩니다" % (hid, order))
        seen.add((hid, order))
        if not str(text).strip():
            bad.append("%s 단계 %d 의 설명이 비었습니다" % (hid, order))
        if path and not path.startswith("UI_Root/"):
            bad.append("%s 단계 %d 의 경로가 UI_Root 로 시작하지 않습니다: %s" % (hid, order, path))

    # ★★ <b>한 항목의 단계는 «한 UI 안» 에서만 머문다</b> (맨 위 ★★ 의 규칙).
    #   창을 여는 항목은 그 창 안만, 창이 없는 항목은 늘 보이는 HUD 하나 안만 짚어야 한다.
    #   ⚠ 이 검산이 «어수선함» 을 막는 유일한 장치다 — 사람이 표를 늘릴 때 반드시 다시 흩어진다.
    by_entry = {}
    for hid, order, path, text in STEPS:
        if path:
            by_entry.setdefault(hid, []).append(path)

    for hid, paths in by_entry.items():
        panel = OPEN_PANEL.get(hid)
        if panel:
            for p in paths:
                if not p.startswith(panel + "/"):
                    bad.append("%s 는 %s 를 여는데 단계가 그 창 밖을 짚습니다: %s"
                               % (hid, panel, p))
        else:
            # 창이 없으면 «UI_Root/HUD_xxx» 두 토막이 전부 같아야 한다.
            roots = {"/".join(p.split("/")[:2]) for p in paths}
            if len(roots) > 1:
                bad.append("%s 의 단계가 여러 UI 를 건너뜁니다(어수선함): %s"
                           % (hid, " · ".join(sorted(roots))))

    for hid in OPEN_PANEL:
        if hid not in ids:
            bad.append("여는 창을 적은 %s 가 Help 시트에 없습니다" % hid)
        if hid not in by_entry:
            bad.append("%s 는 여는 창만 있고 짚을 단계가 없습니다 — 빈 창만 뜹니다" % hid)

    # ★ 말투 검사 — 존댓말은 예외 없이 «니다» 로 끝난다(140-7절이 오탐으로 배운 규칙).
    #   ⚠ «다.» 로 잡으면 «~입니다» 도 걸려 전부 오탐이 된다.
    for hid, (title, summary, body) in TEXTS.items():
        for label, text in (("요약", summary), ("본문", body)):
            for line in text.split("\n"):
                line = line.strip().rstrip("</b>").strip()
                if not line:
                    continue
                if line.endswith("다.") and "니다." not in line:
                    bad.append("%s %s: 반말로 끝나는 줄 — %s" % (hid, label, line[:40]))
        # ★ 앞 초안의 버릇이 되살아나지 않게 <b>기호</b>를 검사한다(맨 위 규칙 ②).
        for mark in ("★", "⚠", "«", "»"):
            if mark in summary or mark in body or mark in title:
                bad.append("%s: 걷어내기로 한 기호가 남아 있습니다 — %s" % (hid, mark))
    return bad


def write_steps(wb):
    """`HelpStep` 시트를 <b>통째로 다시</b> 쓴다 — 단계는 이 스크립트가 정본이다."""
    if "HelpStep" in wb.sheetnames:
        del wb["HelpStep"]
    ws = wb.create_sheet("HelpStep")

    for c, v in enumerate(STEP_KEYS, start=1):
        cell = ws.cell(1, c, v)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.border = BORDER

    row = 2
    for hid, order, path, text in sorted(STEPS, key=lambda s: (s[0], s[1])):
        ws.cell(row, 1, hid).font = BODY_FONT
        ws.cell(row, 2, order).font = BODY_FONT
        ws.cell(row, 3, path).font = BODY_FONT
        c = ws.cell(row, 4, text)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 5):
            ws.cell(row, col).border = BORDER
        row += 1

    for c, w in enumerate(STEP_WIDTH, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    return row - 2


def write_open_panel(wb):
    """
    `Help` 시트에 <b>`open_panel` 열</b>을 만들고 채운다.

    ★ 이미 있으면 그 열을 쓴다(멱등) — 없으면 <b>맨 오른쪽에</b> 새로 만든다.
    ⚠ 비고 열(`비고(제목)`) 뒤에 붙이므로 사람이 보던 열 순서가 흐트러지지 않는다.
    """
    ws = wb["Help"]

    col = 0
    last = ws.max_column
    for c in range(1, last + 1):
        if str(ws.cell(1, c).value or "").strip() == "open_panel":
            col = c
            break
    if col == 0:
        col = last + 1
        cell = ws.cell(1, col, "open_panel")
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 30

    filled = 0
    for r in range(2, ws.max_row + 1):
        hid = str(ws.cell(r, 1).value or "").strip()
        if not hid:
            continue
        want = OPEN_PANEL.get(hid, "")
        cell = ws.cell(r, col, want)
        cell.font = BODY_FONT
        cell.border = BORDER
        if want:
            filled += 1
    return filled


def write_strings(wb):
    """`StringKeys` 시트의 kr 칸을 새 문구로 <b>덮어쓴다</b>. 키는 그대로 둔다."""
    ws = wb["StringKeys"]
    where = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k:
            where[str(k).strip()] = r

    changed, missing = 0, []
    for hid, (title, summary, body) in TEXTS.items():
        for suffix, value in (("_title", title), ("_summary", summary), ("_body", body)):
            key = hid + suffix
            r = where.get(key)
            if r is None:
                missing.append(key)
                continue
            before = ws.cell(r, 2).value
            if (before or "") == value:
                continue
            cell = ws.cell(r, 2)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            changed += 1
    return changed, missing


def main():
    if not os.path.isfile(XLSX):
        print("파일이 없습니다:", XLSX)
        return 1

    bad = check()
    if bad:
        print("[도움말 문구] ⚠ 검산 실패 — 표를 고치지 않았습니다")
        for b in bad:
            print("   " + b)
        return 1

    shutil.copy2(XLSX, XLSX + ".bak")
    wb = openpyxl.load_workbook(XLSX)

    changed, missing = write_strings(wb)
    steps = write_steps(wb)
    panels = write_open_panel(wb)

    wb.save(XLSX)

    with_tour = sorted({s[0] for s in STEPS})
    no_tour = sorted(set(TEXTS) - set(with_tour))

    print("[도움말 문구 다시 쓰기]")
    print("  ✓ 검산 통과 — 반말 0건 · 기호 0건 · 단계가 «한 UI 안» 규칙을 지킴 · 항목 %d개"
          % len(TEXTS))
    print("  문구 %d칸을 덮어썼습니다 (제목·요약·본문 %d칸 중)" % (changed, len(TEXTS) * 3))
    if missing:
        print("  ⚠ StringKeys 에 없어서 못 쓴 키 %d개: %s" % (len(missing), missing))
    print("  HelpStep 시트 %d줄 — 짚을 곳이 있는 단계 %d개 · 글만 보여주는 단계 %d개"
          % (steps,
             sum(1 for s in STEPS if s[2]),
             sum(1 for s in STEPS if not s[2])))
    print("  open_panel 열 — 창을 여는 항목 %d개" % panels)
    print("  「자세히 보기」가 뜨는 항목 %d개 : %s" % (len(with_tour), " · ".join(with_tour)))
    print("  「자세히 보기」가 <b>안 뜨는</b> 항목 %d개 (규칙·개념) : %s"
          % (len(no_tour), " · ".join(no_tour)))
    print("  백업: %s" % os.path.basename(XLSX + ".bak"))
    print("  다음: py -3 Tools/help_string_merge.py  →  gen_string_table.py  →  "
          "link_string_keys.py  →  gen_help_assets.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
