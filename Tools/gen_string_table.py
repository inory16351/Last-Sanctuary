# -*- coding: utf-8 -*-
"""모든 데이터 테이블의 게임 표시 문자열을 '스트링 키 테이블.xlsx' 한 곳으로 모으고,
그것을 Unity 가 읽는 TSV(`Resources/Data/StringTable.txt`)로 내보낸다.

유저 지시(2026-08-12): "웨이브 몬스터 테이블.xlsx 의 string 시트처럼 string key 테이블을
따로 빼서 모든 테이블 스트링을 따로 관리하고 싶다."

■ 왜 스크립트인가
  손으로 옮기면 테이블이 바뀔 때마다 두 곳을 고쳐야 하고 반드시 어긋난다.
  엑셀이 원본이고 이 스크립트가 그걸 그대로 옮긴다(다른 Tools/*.py 와 같은 규칙).

■ ★ 재실행해도 사람이 고친 번역을 덮지 않는다 (이 스크립트의 핵심 규칙)
  `스트링 키 테이블.xlsx` 에 이미 있는 키의 kr/en 은 **그대로 둔다** — 새로 발견된
  키만 뒤에 덧붙인다(merge). 안 그러면 유저가 스트링 테이블에서 문구를 다듬어도
  다음 실행에서 원본 테이블의 옛 리터럴로 되돌아간다.
  `--rebuild` 를 주면 수집값으로 강제로 덮어쓴다(원본 테이블을 정본으로 되돌릴 때만).

■ 키 규칙 — 유저가 이미 쓴 형식을 그대로 따른다
  `wave_mid_boss` 시트에 `monster_name_110001` 이 손으로 들어가 있었다. 그래서
      <필드명>_<id>                     예: monster_name_110001 · skill_explain_80002
      <필드명>_<enum 값>                예: skill_type_desc_Innate_delicacy
  UI·대사처럼 id 가 없는 것은 사람이 직접 짓는다(예: ui_btn_open) — 수집 대상이 아니라
  스트링 테이블에만 존재하며, 이 스크립트는 그런 행을 건드리지 않는다.

■ 수집하지 않는 것
  `능력치 및 공식 정리.xlsx` 와 건물 시트의 Information/Docs/DEF 시트는 **기획 문서**다
  (공식 설명·주석 산문). 게임에 뜨는 문자열이 아니므로 키로 만들지 않는다.
"""
import os
import sys
import openpyxl

# ⚠ 이 스크립트만 이 한 줄이 <b>빠져 있었다</b> (2026-08-25 에 겪고 넣었다). 윈도우 콘솔은
#   기본이 cp949 라, 표에 «−»(U+2212 · 진짜 빼기표) 같은 글자가 있으면 <b>보고를 찍다가</b>
#   `UnicodeEncodeError` 로 죽는다. 표는 이미 저장된 뒤라 결과물은 멀쩡하지만, 무엇이
#   추가됐는지 <b>보지 못한 채</b> 실패로 보인다 — 다른 Tools 스크립트는 다 갖고 있는 방어다.
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from vault_path import TABLE_DIR   # ★ PC 마다 다른 볼트 위치를 찾아준다(2026-08-15)
OUT_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')

# 스크립트 위치에서 프로젝트 루트를 역산한다 (gen_character_assets.py 의 교훈 —
# 경로를 박아두면 엉뚱한 폴더를 새로 만들고 "성공"만 찍힌다).
_PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR_UNITY = os.path.join(_PROJECT, 'Assets', '_Project', 'Resources', 'Data')
OUT_TSV = os.path.join(OUT_DIR_UNITY, 'StringTable.txt')

SHEET = 'string'
INFO_SHEET = 'Info'

# 하이퍼링크 표적이 되는 정의된 이름의 접두사. link_string_keys.py 와 같아야 한다.
NAME_PREFIX = 'key_'

# 3행 헤더 — 웨이브 몬스터 테이블의 string 시트와 완전히 같은 규약이다.
#   1행 한글 라벨 / 2행 필드명 / 3행 자료형 / 4행부터 데이터
HEADER_KR = ['스트링키', '한국어', '영어', '출처 테이블', '비고']
HEADER_FIELD = ['string_key', 'kr', 'en', 'source', 'note']
HEADER_TYPE = ['string', 'string', 'string', '-', '-']
DATA_ROW0 = 4

FONT = 'Arial'


# ---------------------------------------------------------------------------
# 수집 규칙 — (파일, 시트, id 컬럼, [(값 컬럼, 키 접두사, 언어)])
#   언어: 'kr' 이면 한국어 칸으로, 'en' 이면 영어 칸으로 들어간다.
#   ★ 같은 키에 kr 과 en 을 각각 넣는 것이 이 표의 요점이다 —
#     예전에는 `character_name` 과 `character_name_EG` 두 컬럼으로 갈라져 있었다.
# ---------------------------------------------------------------------------
# ★ 2026-08-13 — 원본 표의 <b>영어 이름 컬럼(`*_EG`)을 전부 지웠다</b>(유저 지시:
#   "스트링 테이블로 영어 이름 배정해야 하는데 지금 영어 이름 칼럼이 삭제되지 않고
#   남아있는 것들 확인해서 없애줘"). 같은 값을 두 곳에 적어두면 반드시 어긋나기 때문이다.
#   <b>이제 영어는 이 파일(스트링 키 테이블)의 `en` 칸이 유일한 정본</b>이고, 아래 규칙에서도
#   수집 대상이 아니다. 이미 들어간 en 값은 merge 규칙(기존 우선)상 그대로 남는다 —
#   `--rebuild` 를 주면 수집값으로 덮으므로, 지금 상태에서 `--rebuild` 를 돌리면
#   ⚠ <b>영어 이름이 전부 비워진다.</b> 절대 그렇게 돌리지 말 것.
RULES = [
    # ── 웨이브 몬스터 테이블 ────────────────────────────────────────────────
    ('웨이브 몬스터 테이블.xlsx', 'wave_nom', 'monster_id', [
        ('monster_name', 'monster_name', 'kr'),
    ]),
    # ⚠ `wave_mid_boss` 시트는 <b>더 이상 없다</b> — 2026-08-18 의 보스 개편이 중간보스를
    #   없애면서 시트째 지웠다(`table_update_20260818_boss_rework.py`). 그런데 이 규칙만
    #   남아 있어 돌릴 때마다 «시트 없음 wave_mid_boss» 경고가 떴다. 경고가 늘 하나 떠 있으면
    #   <b>진짜 경고가 그 소음에 묻힌다</b> — 그래서 규칙을 지운다(2026-08-27).
    #   중간보스의 칭호 칸(H열) 이야기도 함께 사라진다. 최종보스의 `boss_title` 은 아래에 남는다.
    ('웨이브 몬스터 테이블.xlsx', 'wave_top_boss', 'monster_id', [
        ('monster_name', 'monster_name', 'kr'),
        ('boss_title', 'boss_title', 'kr'),
    ]),
    ('웨이브 몬스터 테이블.xlsx', 'Skill', 'skill_id', [
        ('skill_name', 'skill_name', 'kr'),
        ('skill_explain', 'skill_explain', 'kr'),
        # 2026-08-19 신설 — 이 스킬이 거는 「구속」의 화면 표시 이름(예: 기절). 비어 있는
        # 스킬은 코드 기본값("구속")을 그대로 쓴다. `BossSkillSO.statusNameKey` 가 읽는다.
        ('status_name', 'status_name', 'kr'),
    ]),
    ('웨이브 몬스터 테이블.xlsx', 'Skill_Type', 'skill_type', [
        ('desc', 'skill_type_desc', 'kr'),
    ]),

    # ── 캐릭터 테이블 ──────────────────────────────────────────────────────
    ('캐릭터 테이블.xlsx', 'Character', 'character_id', [
        ('character_name', 'character_name', 'kr'),
        # `character_name_EG` 는 2026-08-13 에 삭제됐다 — 위 ★ 주석 참조.
        # ⚠ 이 en 값은 `gen_character_assets.py` 가 <b>에셋 파일 이름</b>으로 쓴다
        #   (`Character_9001_Elin`). 스트링 키 테이블에서 지우면 guid 가 바뀌어 참조가 끊긴다.
        #
        # 2026-08-19 신설 — <b>칭호</b>. `wave_top_boss` 의 boss_title/boss_title_EG 와 같은 짜임이다
        # (상세 카드 112절의 칭호 칸이 캐릭터에서만 비어 있었다).
        ('character_title', 'character_title', 'kr'),
        ('character_title_EG', 'character_title', 'en'),
    ]),
    ('캐릭터 테이블.xlsx', 'Skill', 'skill_id', [
        ('skill_name', 'skill_name', 'kr'),
        ('skill_explain', 'skill_explain', 'kr'),
        # 2026-08-20 신설 — 「상세 설명」. <b>수치를 적지 않는</b> 중간 문구
        # (플레이버보다 구체적이고 정의문보다 헐렁하다 · 유저 지시
        #  *"밸류 타입보단 덜 상세하게"*).
        ('skill_detail', 'skill_detail', 'kr'),
    ]),
    ('캐릭터 테이블.xlsx', 'Skill_Type', 'skill_type', [
        ('desc', 'skill_type_desc', 'kr'),
    ]),

    # ── 중립 몬스터 ────────────────────────────────────────────────────────
    ('임시용 중립 몬스터.xlsx', 'neutrality_mon', 'mon_id', [
        ('mon_name', 'mon_name', 'kr'),
        # 칭호 — 2026-08-15 에 표에 생긴 칸. 웨이브 보스의 boss_title 과 같은 역할이다
        # (지금은 비어 있어 키가 만들어지지 않는다 — 값을 적으면 그때 생긴다).
        ('mon_title', 'mon_title', 'kr'),
    ]),
    # 2026-08-19 신설 — 에픽 중립 보스(카르시노스·아니사킬 등)의 스킬이 거는 「구속」의
    # 화면 표시 이름. 위 웨이브 쪽과 같은 컬럼·같은 뜻이다(아니사킬의 「거대한 위협
    # 포효」가 "기절"을 쓴다 — 정의문 자체가 그렇게 부른다).
    ('임시용 중립 몬스터.xlsx', 'Skill', 'skill_id', [
        # ★★★ 2026-08-25 신설 — <b>이 두 줄이 없어서 로그에 스킬 «번호» 가 나왔다.</b>
        #   유저 리포트: *"보스 스킬 쓸때 스킬 이름이 아니라 번호 나오는거 고쳐줘 로그에"*.
        #   웨이브·캐릭터 표에는 skill_name 이 걷히는데 <b>중립 표만 빠져 있었다</b>.
        #   그래서 skill_name_2003~2009 가 StringTable 에 아예 없었고,
        #   `BossSkillSO.DisplayName` 의 폴백이 <b>에셋 이름</b>("BossSkill_2003")으로
        #   떨어졌다 — 그것이 화면에 «번호» 로 보인 것이다.
        #   ⚠ 2001·2002 만 멀쩡했던 이유는 그 둘이 스트링 키 표에 <b>손으로</b> 들어가
        #     있었기 때문이다(출처 칸이 비어 있다). 손으로 넣은 두 줄이 구멍을 가리고 있었다.
        ('skill_name', 'skill_name', 'kr'),
        ('skill_explain', 'skill_explain', 'kr'),
        ('status_name', 'status_name', 'kr'),
    ]),

    # ── 정신 이상 ──────────────────────────────────────────────────────────
    ('정신 이상 테이블.xlsx', 'mental_error', 'mental_error_id', [
        ('Korean_explain', 'mental_error_name', 'kr'),
    ]),
    ('정신 이상 테이블.xlsx', 'mental_error_type', 'mental_error_type', [
        ('desc', 'mental_error_type_desc', 'kr'),
    ]),

    # ── 건물 ───────────────────────────────────────────────────────────────
    # Construction 시트만 본다. Information·Docs·DEF 는 기획 산문이다(위 doc 참조).
    ('Last_Sanctuary_건물데이터시트_Ver05.xlsx', 'Construction', 'Const_id', [
        ('Const_name', 'const_name', 'kr'),
    ]),

    # ══════════════════════════════════════════════════════════════════════
    #  ★★★ 유물 · 이벤트 (2026-08-25 신설 — 유저 지시:
    #      *"이벤트랑 유물 테이블도 스트링 키 테이블 연동"*)
    # ══════════════════════════════════════════════════════════════════════
    # 이 둘은 131·134·124절에서 <b>나중에</b> 생긴 표라 이 규칙에 들어오지 못했다.
    # 그래서 <b>플레이어에게 보이는 글의 절반 이상</b>(유물 이름·설명·서사 · 사건 대사·
    # 선택지·결과)이 스트링 키 테이블 <b>밖</b>에 있었다 — 51절이 «모든 테이블 문구를 한
    # 파일로» 라고 세운 방향에서 두 표만 빠져 있던 셈이다.
    #
    # ⚠ <b>기획 산문 시트는 넣지 않는다</b> — 유물의 `EffectType.effect_desc`·`Drop.drop_desc`,
    #   사건의 `Condition.desc`·`RewardType.reward_type_desc` 는 <b>표를 읽는 사람</b>을 위한
    #   글이지 화면에 나오는 글이 아니다(건물 표의 Information·Docs 를 뺀 것과 같은 기준).

    # ── 유물 ───────────────────────────────────────────────────────────────
    ('Last_Sanctuary_유물테이블_Ver02.xlsx', 'Relic', 'relic_id', [
        ('relic_name', 'relic_name', 'kr'),
        ('relic_desc', 'relic_desc', 'kr'),
        ('relic_flavor', 'relic_flavor', 'kr'),
    ]),
    # 등급 이름(일반·레어·에픽) — ⚠ id 가 <b>정수가 아니라 enum</b>이다(`common`).
    #   키는 `relic_grade_common` 이 된다. 수집 코드는 id 를 문자열로 다루므로 그대로 된다.
    ('Last_Sanctuary_유물테이블_Ver02.xlsx', 'Grade', 'grade', [
        ('grade_name', 'relic_grade', 'kr'),
    ]),
    # 발굴 결과 — ⚠ `outcome_desc` 에는 <c>{value_01}</c> 같은 <b>자리표</b>가 들어 있다.
    #   번역할 때 그 표시를 <b>지우면 안 된다</b>(런타임에 숫자가 들어갈 자리다).
    ('Last_Sanctuary_유물테이블_Ver02.xlsx', 'DigOutcome', 'outcome_type', [
        ('outcome_desc', 'dig_outcome_desc', 'kr'),
        ('outcome_script', 'dig_outcome_script', 'kr'),
    ]),
    ('Last_Sanctuary_유물테이블_Ver02.xlsx', 'DigChoice', 'choice_id', [
        ('choice_text', 'dig_choice_text', 'kr'),
    ]),
    ('Last_Sanctuary_유물테이블_Ver02.xlsx', 'Dialogue', 'dialogue_id', [
        ('script', 'relic_dialogue', 'kr'),
    ]),

    # ── 사건(이벤트) ───────────────────────────────────────────────────────
    ('Last_Sanctuary_이벤트테이블_Ver013.xlsx', 'Event', 'event_id', [
        ('event_name', 'event_name', 'kr'),
        ('event_script', 'event_script', 'kr'),
    ]),
    # ⚠ 선택지의 id 는 `choice_group_id` 가 아니라 <b>`choice_id`</b> 다 — 그룹 하나에
    #   선택지가 여럿이라 그룹으로 키를 만들면 <b>서로 덮어쓴다</b>.
    ('Last_Sanctuary_이벤트테이블_Ver013.xlsx', 'ChoiceGroup', 'choice_id', [
        ('choice_text', 'event_choice_text', 'kr'),
        ('result_script', 'event_result_script', 'kr'),
        ('result_effect', 'event_result_effect', 'kr'),
    ]),

    # ── 대체 이름 (2026-08-26 신설) ────────────────────────────────────────
    # 같은 인물이 이번 판에 <b>두 번째로</b> 등장할 때 받는 «다른 이름» 주머니.
    # 표를 만들고 키를 처음 붙이는 것은 `Tools/gen_alt_name_table.py` 가 한다 —
    # 여기 규칙은 그 뒤로 <b>한국어를 표와 같게 유지</b>하기 위한 것이다(영어는 이 표의 en 이 정본).
    # ⚠ 키가 `character_altname_1` 부터 <b>구멍 없이</b> 이어져야 한다 — 코드가 빈 번호에서 멈춘다.
    ('대체 이름 테이블.xlsx', 'AltName', 'alt_name_id', [
        ('kr', 'character_altname', 'kr'),
    ]),
]


def norm(v):
    """셀 값을 문자열로 정규화한다. 앞뒤 공백을 없애 키가 어긋나지 않게 한다.

    ⚠️ 원본 테이블에는 ' 공허의 속삭임' 처럼 **앞에 공백이 붙은 값**이 실제로 있었다.
    그대로 쓰면 같은 이름이 두 행으로 갈라진다.
    """
    if v is None:
        return ''
    return str(v).strip()


def field_index(ws):
    """2행(필드명) → 컬럼 인덱스. 필드명도 공백을 없애 맞춘다
    (' resistance' 처럼 앞 공백이 붙은 헤더가 실제로 있다)."""
    out = {}
    for c in range(1, ws.max_column + 1):
        name = norm(ws.cell(row=2, column=c).value)
        if name:
            out[name] = c
    return out


def looks_like_key(text):
    """
    이미 스트링 키가 들어가 있는 칸인가 (키로 변환된 칸).

    판정: <b>ASCII 이고 · 공백이 없고 · 밑줄이 있다.</b>
    한국어 문구는 `isascii()` 에서 걸리고, 영어 문구('The Lord Of Endless Forms')는
    공백에서 걸리고, 영어 단어 하나('SoulArcher')는 밑줄이 없어서 걸린다.

    ⚠️ 예전에는 "전부 소문자" 조건을 넣었는데 `skill_type_desc_Innate_delicacy` 처럼
    <b>대문자가 섞인 키를 놓쳤다</b>(enum 값이 대문자로 시작한다). 그러면 그 키의
    kr 칸에 키 문자열이 그대로 들어간다.
    """
    if not text:
        return False
    return text.isascii() and ('_' in text) and not any(c.isspace() for c in text)


# 이미 존재하는 string 시트 — 유저가 손으로 적어둔 번역이 들어있다.
# ★ 이걸 먼저 읽어야 한다. `wave_mid_boss` 의 monster_name 칸은 이미 키
#   (`monster_name_110001`)로 바뀌어 있고, 그 키의 실제 문구('혈인')는 여기에만 있다.
#   이 시트를 안 읽으면 그 문구가 조용히 사라진다.
SEED_SHEETS = [('웨이브 몬스터 테이블.xlsx', 'string')]


def collect_seed_sheets():
    """기존 string 시트를 (key, kr, en) 그대로 읽어온다."""
    rows = {}
    warnings = []

    for filename, sheet in SEED_SHEETS:
        path = os.path.join(TABLE_DIR, filename)
        if not os.path.exists(path):
            continue

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]
        idx = field_index(ws)
        for r in range(DATA_ROW0, ws.max_row + 1):
            key = norm(ws.cell(row=r, column=idx.get('string_key', 1)).value)
            kr = norm(ws.cell(row=r, column=idx.get('kr', 2)).value)
            en = norm(ws.cell(row=r, column=idx.get('en', 3)).value)

            if not key:
                # ⚠️ 실제로 이런 행이 있었다 — 키 없이 영어만 적힌 줄. 어느 키의 것인지
                #    알 수 없으므로 버리지 않고 경고로 남긴다(유저가 직접 붙여야 한다).
                if kr or en:
                    warnings.append(
                        f'{filename}/{sheet} {r}행: 키가 비어 있어 옮기지 못했다 — kr="{kr}" en="{en}"')
                continue

            rows[key] = {'kr': kr, 'en': en, 'source': f'{sheet}(기존)'}

    return rows, warnings


def collect():
    """규칙대로 훑어 {key: {'kr':…, 'en':…, 'source':…}} 를 만든다."""
    rows, warnings = collect_seed_sheets()

    for filename, sheet, id_field, specs in RULES:
        path = os.path.join(TABLE_DIR, filename)
        if not os.path.exists(path):
            warnings.append(f'파일 없음: {filename}')
            continue

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            warnings.append(f'{filename}: 시트 없음 {sheet}')
            continue

        ws = wb[sheet]
        idx = field_index(ws)
        if id_field not in idx:
            warnings.append(f'{filename}/{sheet}: id 컬럼 없음 {id_field}')
            continue

        for r in range(DATA_ROW0, ws.max_row + 1):
            row_id = norm(ws.cell(row=r, column=idx[id_field]).value)
            if not row_id:
                continue
            if row_id.endswith('.0'):        # 엑셀이 int 를 float 로 읽는 경우
                row_id = row_id[:-2]

            for value_field, prefix, lang in specs:
                if value_field not in idx:
                    continue
                text = norm(ws.cell(row=r, column=idx[value_field]).value)
                if not text:
                    continue

                key = f'{prefix}_{row_id}'

                # 이미 키로 바꿔둔 칸이면 그 값 자체가 키다 — 리터럴이 아니므로
                # 번역으로 넣지 않는다(그러면 kr 칸에 키 문자열이 들어간다).
                #
                # ★ 2026-08-15 수정 — <b>칸에 적힌 키를 그대로 쓴다</b>.
                #   예전에는 여기서도 `f'{prefix}_{row_id}'` 로 <b>규칙에서 유추한</b> 키를
                #   등록했다. 둘이 같을 때는 티가 안 났지만, 중립 표의 1004 는
                #   `mon_title` 칸에 <b>`epic_boss_title_1004`</b> 라고 적혀 있어
                #   실재하지 않는 `mon_title_1004` 가 <b>빈 값으로 새로 생겼다</b>
                #   (그리고 진짜 칭호는 어디에도 안 연결됐다).
                #   칸이 정본이므로 유추한 이름이 아니라 적힌 이름을 등록해야 한다.
                if looks_like_key(text):
                    key = text
                    rows.setdefault(key, {'kr': '', 'en': '', 'source': ''})
                    rows[key]['source'] = rows[key]['source'] or f'{sheet}.{value_field}'
                    continue

                entry = rows.setdefault(key, {'kr': '', 'en': '', 'source': ''})

                # 기존 string 시트에서 이미 읽어온 값이 있으면 그쪽이 정본이다 —
                # 사람이 그 시트에서 문구를 다듬었을 수 있다.
                if entry[lang]:
                    if entry[lang] != text:
                        warnings.append(
                            f'{key}[{lang}]: string 시트 "{entry[lang]}" 를 유지하고 '
                            f'{sheet}.{value_field} 의 "{text}" 는 무시했다')
                    continue

                entry[lang] = text
                entry['source'] = f'{sheet}.{value_field}'

    return rows, warnings


def read_existing():
    """이미 있는 스트링 키 테이블을 읽는다. 사람이 고친 번역을 지키기 위한 것."""
    if not os.path.exists(OUT_XLSX):
        return {}, []

    wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
    if SHEET not in wb.sheetnames:
        return {}, []

    ws = wb[SHEET]
    idx = field_index(ws)
    order = []
    out = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        key = norm(ws.cell(row=r, column=idx.get('string_key', 1)).value)
        if not key:
            continue
        out[key] = {
            'kr': norm(ws.cell(row=r, column=idx.get('kr', 2)).value),
            'en': norm(ws.cell(row=r, column=idx.get('en', 3)).value),
            'source': norm(ws.cell(row=r, column=idx.get('source', 4)).value),
            'note': norm(ws.cell(row=r, column=idx.get('note', 5)).value),
        }
        order.append(key)
    return out, order


# 스트링 테이블에만 존재하는(테이블에서 수집되지 않는) 키의 초기 목록.
# 웨이브 몬스터 테이블의 string 시트에 유저가 이미 적어둔 것들을 그대로 옮긴다.
SEED_MANUAL = [
    ('event_dialogue_10001', '', '', '(수동)', '이벤트 대사 — id 는 이벤트 테이블 기준'),
    ('ui_btn_open', '열기', 'Open', '(수동)', 'UI 공용 버튼'),
    ('ui_btn_close', '닫기', 'Close', '(수동)', 'UI 공용 버튼'),
]

# 더 이상 쓰지 않는 키 — 최종 표에서 지운다.
# `monster_name_melee/ranged/boss` 는 "게임에 보이는 이름(근거리 암세포)을 그대로 두자"고
# 잠깐 만들었던 임시 키다. 유저가 <b>웨이브 몬스터 테이블을 정본으로 확정</b>했으므로
# (2026-08-12) 이제 세 에셋이 monster_name_100001/100002/120001 을 가리킨다.
RETIRED_KEYS = {'monster_name_melee', 'monster_name_ranged', 'monster_name_boss'}


def merge(collected, existing, existing_order, rebuild):
    """수집값과 기존 파일을 합친다. 기본은 '기존 우선'(사람이 고친 번역을 지킨다)."""
    result = {}
    order = []

    for key in existing_order:
        if key in RETIRED_KEYS:
            continue
        result[key] = dict(existing[key])
        order.append(key)

    for key, seed_kr, seed_en, src, note in SEED_MANUAL:
        if key in result:
            continue
        result[key] = {'kr': seed_kr, 'en': seed_en, 'source': src, 'note': note}
        order.append(key)

    added, updated = [], []
    for key in sorted(collected.keys()):
        got = collected[key]
        if key not in result:
            result[key] = {'kr': got['kr'], 'en': got['en'],
                           'source': got['source'], 'note': ''}
            order.append(key)
            added.append(key)
            continue

        cur = result[key]
        cur['source'] = got['source'] or cur.get('source', '')
        for lang in ('kr', 'en'):
            if not got[lang]:
                continue
            if rebuild or not cur.get(lang):
                if cur.get(lang) != got[lang]:
                    updated.append(f'{key}[{lang}]')
                cur[lang] = got[lang]

    return result, order, added, updated


def write_xlsx(rows, order):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = SHEET

    for c, (kr, field, typ) in enumerate(zip(HEADER_KR, HEADER_FIELD, HEADER_TYPE), 1):
        ws.cell(row=1, column=c, value=kr)
        ws.cell(row=2, column=c, value=field)
        ws.cell(row=3, column=c, value=typ)

    head_fill = PatternFill('solid', fgColor='DDE6F0')
    for r in (1, 2, 3):
        for c in range(1, len(HEADER_KR) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, bold=(r == 1))
            cell.fill = head_fill
            cell.alignment = Alignment(vertical='center')

    for i, key in enumerate(order):
        v = rows[key]
        r = DATA_ROW0 + i
        ws.cell(row=r, column=1, value=key).font = Font(name=FONT)
        ws.cell(row=r, column=2, value=v.get('kr', '')).font = Font(name=FONT)
        ws.cell(row=r, column=3, value=v.get('en', '')).font = Font(name=FONT)
        ws.cell(row=r, column=4, value=v.get('source', '')).font = Font(name=FONT, color='808080')
        ws.cell(row=r, column=5, value=v.get('note', '')).font = Font(name=FONT, color='808080')
        for c in (2, 3, 5):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

    for c, width in zip(range(1, 6), (32, 52, 52, 22, 34)):
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = 'B4'

    add_key_names(wb, order)
    write_info(wb.create_sheet(INFO_SHEET), rows)
    wb.save(OUT_XLSX)


def add_key_names(wb, order):
    """
    키마다 <b>정의된 이름</b>(`key_<키>`)을 만든다 — 원본 테이블의 하이퍼링크가 가리키는 표적이다
    (`Tools/link_string_keys.py`, 유저 요청 "딸깍 하면 스트링 키 테이블로 넘어가게").

    ★ <b>여기서 만들어야 한다.</b> 이 함수가 스트링 키 테이블을 <b>매번 새 통합문서로 새로 쓰기</b>
      때문에, 이름을 Excel 쪽에서만 만들어 두면 <b>다음 실행에서 전부 사라지고 링크가 통째로
      깨진다.</b> 실제로 그 순서로 만들었다가 이 구멍을 발견해 여기로 옮겼다.

    행 번호가 아니라 이름으로 거는 이유는 link_string_keys.py 의 doc 참조 —
    스트링 테이블에 행이 끼어들어도 링크가 안 깨진다.
    """
    from openpyxl.workbook.defined_name import DefinedName

    for i, key in enumerate(order):
        row = DATA_ROW0 + i
        # 시트 이름에 공백이 없으므로 따옴표 없이 쓴다.
        wb.defined_names.add(DefinedName(NAME_PREFIX + key,
                                         attr_text=f'{SHEET}!$A${row}'))


INFO_LINES = [
    ('Last Sanctuary — 스트링 키 테이블', True),
    ('', False),
    ('■ 이 파일이 게임에 뜨는 모든 문자열의 정본이다.', True),
    ('다른 테이블(캐릭터·웨이브 몬스터·중립 몬스터·정신 이상·건물)의 이름/설명 컬럼은', False),
    ('여기의 string_key 를 가리키기만 한다. 문구를 고칠 때는 이 파일만 고치면 된다.', False),
    ('', False),
    ('■ 시트 규약 — 웨이브 몬스터 테이블의 string 시트와 완전히 같다', True),
    ('1행 한글 라벨 / 2행 필드명 / 3행 자료형 / 4행부터 데이터.', False),
    ('앞 3열(string_key · kr · en)의 위치와 뜻은 그 시트와 동일하다.', False),
    ('source · note 두 열은 관리용으로 뒤에 덧붙인 것이다(게임은 읽지 않는다).', False),
    ('', False),
    ('■ 키 규칙', True),
    ('<필드명>_<id>          예: monster_name_110001 · skill_explain_80002', False),
    ('<필드명>_<enum 값>     예: skill_type_desc_Innate_delicacy', False),
    ('id 가 없는 UI·대사는 사람이 직접 짓는다. 예: ui_btn_open · event_dialogue_10001', False),
    ('', False),
    ('■ 새 문구를 넣는 방법', True),
    ('1) 원본 테이블에 한국어를 그대로 적고 Tools/gen_string_table.py 를 돌린다', False),
    ('   → 키가 자동으로 생기고 kr 칸이 채워진다. 그다음 원본 테이블의 그 칸을 키로 바꾼다.', False),
    ('2) 또는 이 시트에 직접 한 줄 추가하고 원본 테이블에 그 키를 적는다.', False),
    ('', False),
    ('■ ★ 다시 돌려도 여기서 다듬은 번역은 안 지워진다', True),
    ('이미 있는 키의 kr/en 은 그대로 두고 새 키만 덧붙인다(merge).', False),
    ('원본 테이블 값으로 강제로 되돌리려면 --rebuild 를 준다.', False),
    ('', False),
    ('■ 게임에 반영하는 방법', True),
    ('python Tools/gen_string_table.py', False),
    ('  → 이 파일을 갱신하고 Assets/_Project/Resources/Data/StringTable.txt 로 내보낸다.', False),
    ('  → Unity 의 LastSanctuary.Data.StringTable 이 그 파일을 읽는다.', False),
    ('', False),
    ('■ ★ 원본 테이블의 키 칸을 누르면 여기로 바로 온다', True),
    ('각 테이블의 키 칸에 하이퍼링크가 걸려 있다(파란 밑줄). 누르면 이 시트의 그 키 행으로 이동한다.', False),
    ('링크는 행 번호가 아니라 "정의된 이름"(key_<키>)을 가리키므로 행이 밀려도 안 깨진다.', False),
    ('키를 새로 추가했으면 python Tools/link_string_keys.py 를 한 번 돌려 링크를 붙인다.', False),
    ('  (--clear 를 주면 링크를 전부 지운다)', False),
    ('', False),
    ('■ 수집하지 않는 파일', True),
    ('능력치 및 공식 정리.xlsx — 공식 설명 문서(게임에 뜨는 문자열이 아니다)', False),
    ('건물데이터시트의 Information · Docs · DEF 시트 — 기획 산문', False),
]


def write_info(ws, rows):
    ws.column_dimensions['A'].width = 104
    r = 1
    for text, bold in INFO_LINES:
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(name=FONT, bold=bold, size=13 if r == 1 else 11)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=f'■ 현재 키 개수: {len(rows)}').font = Font(name=FONT, bold=True)
    r += 1
    missing_kr = [k for k, v in rows.items() if not v.get('kr')]
    missing_en = [k for k, v in rows.items() if not v.get('en')]
    ws.cell(row=r, column=1,
            value=f'한국어 미기입 {len(missing_kr)}개 · 영어 미기입 {len(missing_en)}개'
                  ' (영어는 아직 대부분 비어 있다 — 게임은 한국어로 폴백한다)'
            ).font = Font(name=FONT)


def write_tsv(rows, order):
    """Unity 로 내보낸다. TSV 인 이유 — 한국어 문구에 쉼표가 흔해서 CSV 는 인용부호
    처리가 필요하다. 탭은 문구에 나올 일이 없다.
    확장자를 .txt 로 두는 이유 — Unity 는 .tsv 를 TextAsset 으로 임포트하지 않는다."""
    os.makedirs(OUT_DIR_UNITY, exist_ok=True)

    lines = ['# Last Sanctuary 스트링 테이블 — Tools/gen_string_table.py 가 생성한다. 직접 고치지 말 것.',
             '# 원본: 데이터 테이블/스트링 키 테이블.xlsx',
             'string_key\tkr\ten']
    for key in order:
        v = rows[key]
        kr = v.get('kr', '').replace('\t', ' ').replace('\r', '')
        en = v.get('en', '').replace('\t', ' ').replace('\r', '')
        # 줄바꿈은 \n 리터럴로 접어 한 줄에 담는다(런타임에서 되돌린다).
        kr = kr.replace('\n', '\\n')
        en = en.replace('\n', '\\n')
        lines.append(f'{key}\t{kr}\t{en}')

    with open(OUT_TSV, 'w', encoding='utf-8', newline='\n') as f:
        f.write('\n'.join(lines) + '\n')

    write_meta(OUT_TSV)


def write_meta(path):
    """.meta 가 없으면 만든다. guid 는 경로에서 결정적으로 뽑아 재실행에도 같게 한다
    (gen_character_assets.py 와 같은 방식). 이미 있으면 건드리지 않는다."""
    import hashlib
    meta = path + '.meta'
    if os.path.exists(meta):
        return
    rel = os.path.relpath(path, _PROJECT).replace('\\', '/')
    guid = hashlib.md5(rel.encode('utf-8')).hexdigest()
    with open(meta, 'w', encoding='utf-8', newline='\n') as f:
        f.write('fileFormatVersion: 2\n')
        f.write(f'guid: {guid}\n')
        f.write('TextScriptImporter:\n')
        f.write('  externalObjects: {}\n')
        f.write('  userData: \n')
        f.write('  assetBundleName: \n')
        f.write('  assetBundleVariant: \n')


def main():
    rebuild = '--rebuild' in sys.argv

    collected, warnings = collect()
    existing, existing_order = read_existing()
    rows, order, added, updated = merge(collected, existing, existing_order, rebuild)

    write_xlsx(rows, order)
    write_tsv(rows, order)

    print(f'수집 {len(collected)}개 / 기존 {len(existing)}개 → 최종 {len(rows)}개')
    if added:
        print(f'추가된 키 {len(added)}개:')
        for k in added:
            print('  +', k, '|', rows[k].get('kr', '')[:40])
    if updated:
        print(f'값이 채워진 칸 {len(updated)}개: ' + ', '.join(updated[:20]))
    if warnings:
        print('경고:')
        for w in warnings:
            print('  !', w)
    print('->', OUT_XLSX)
    print('->', OUT_TSV)


if __name__ == '__main__':
    main()
