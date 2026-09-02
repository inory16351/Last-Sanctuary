# -*- coding: utf-8 -*-
"""번역 다듬기 — <b>같은 말을 다르게 부르는 자리</b>를 아홉 언어에서 한꺼번에 잡는다 (2026-09-02).

■ 어떻게 찾았나
  «영어가 비었나» 는 이미 0이다. 그래서 그물을 뒤집어 «<b>같은 한국어인데 번역이 다른 칸</b>»
  을 셌다. 16쌍이 나왔고 그중 아홉이 진짜였다.
  ⚠⚠ 그리고 <b>일곱 언어가 영어의 흔들림을 그대로 물려받았다</b> — 번역자가 영어를 보고
    옮겼으므로 당연한 결과다. 즉 <b>영어만 고치면 나머지 일곱은 어긋난 채 남는다</b>.

■ ★★★ 그래서 «정본 행에서 언어별로 복사» 한다
  여덟 언어 × 아홉 쌍 = 72칸을 손으로 적으면 반드시 하나를 틀린다.
  <b>어느 행이 정본인지</b>만 정하고, 값은 그 행에서 그대로 가져온다.

    event_name_207001·2·3   ← epic_boss_title_1101·2·3   짐승 이름이 정본
                                (플레이어가 <b>먼저 보는</b> 것이 짐승 쪽이다)
    epic_boss_title_1104    ← event_name_207004          이쪽만 «숙적 = Nemesis» 가 낫다
    ui_dig_choice_accept    ← dig_choice_text_340001     표의 문장이 정본(창 코드는 폴백)
    ui_dig_choice_decline   ← dig_choice_text_340002
    ui_save_write_failed    ← ui_settings_save_failed    같은 사고를 두 창이 다르게 알렸다
    ui_tactics_title        ← ui_action_tactics          ★★★ 아래 참조
    help_tactics_title      ← ui_action_tactics

  ★★★ <b>「전술 지침」을 한 창인데 세 이름으로 부르고 있었다</b> —
     버튼 «Tactical orders» · 그 버튼이 여는 창 «Tactical Directives» · 백과 «Tactics».
     es·fr·de·ru·pt·pl 도 같은 방식으로 갈라져 있었다(ja 만 셋 다 「戦術指針」이었다).
     → <b>버튼의 낱말</b>로 통일한다. 코드 이름(`TacticalOrderPanel`)과도 같아진다.
  ⚠ `help_tactics_title` 의 en 을 여기서 고쳐도 된다 — «help_ 는 도움말 표가 정본» 은
    <b>kr</b> 에 대한 규칙이고, 그 표의 en 열은 «자리만 비워 둔다» 가 확정사항 ④ 라
    `help_string_merge.py` 는 <b>비어 있을 때만</b> 넣는다.

■ 영어만 손대는 두 칸 (대소문자·짝맞춤)
  ui_action_tactics        Tactical orders → Tactical Orders
  ui_action_tactics_close  Close orders    → Close Tactical Orders

■ 미결 179 — 죽은 쌍둥이 키 둘을 지운다
  `ui_portrait_rage` · `ui_portrait_souls` 는 `_format` 쪽과 값이 같은데 <b>아무도 안 부른다</b>
  (초상화·성장 창 둘 다 `_format` 을 쓴다 — 코드 전수로 확인했다).
  ★ 실제로 이미 영어가 어긋나 있었다(Souls gathered ↔ Souls collected). 게다가 이번
    다국어 작업으로 <b>쓰이지 않는 칸에 번역이 아홉 벌</b> 붙었다 — 지금 지우는 편이 싸다.
  ⚠ 지워도 되살아나지 않는다 — `ui_` 키는 원본 테이블에서 수집하는 대상이 아니다.

■ 그물을 빠져나간 문구 하나 — `ui_growth_relic_hint`
  성장 창 유물 칸의 <b>효과 줄</b>이 호출문에 박힌 리터럴이라 178·179절 그물을 통과했다
  (184-4절 「빗나감」과 똑같은 종류). 코드는 인스펙터 칸(`relicSlotHint`)으로 빼 두었다.
  ⚠ 「」 는 영어권 언어에서 <b>큰따옴표</b>로 옮긴다(이 표의 기존 21칸 규약).
  ⚠⚠ `<Open Relics>` 처럼 홑화살괄호로 감싸면 TMP 가 <b>서식 태그로 먹어</b> 글자가 사라진다.

■ 손대지 않은 것 (일부러) — 다음에 또 세면 같은 것이 나오므로 적어 둔다
  광폭화 Enrage/Enraged(제목↔상태) · 사냥 Hunt/Hunting(버튼↔하는 일) ·
  캐릭터 Characters/Character(수) · 일시정지 Pause/Paused(이름↔결과) ·
  전투 Battle/Combat(단계↔분류) · 공포 Fear/Dread · 곪은 자리 사건↔유물
  (뒤의 둘은 <b>서로 다른 물건</b>이 한국어 이름만 같은 경우다).

■ 다음
    py -3 Tools/gen_string_table.py   →   py -3 Tools/link_string_keys.py
    py -3 Tools/check_string_keys.py --strict
"""
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET, DATA_ROW0 = 'string', 4
SOURCE = '번역 다듬기(2026-09-02)'

#: 고칠 키 → (정본 키, 왜). 값은 <b>언어마다 정본 행에서 그대로 복사</b>한다.
ALIGN = {
    'event_name_207001':    ('epic_boss_title_1101', '짐승 이름과 맞춤'),
    'event_name_207002':    ('epic_boss_title_1102', '짐승 이름과 맞춤'),
    'event_name_207003':    ('epic_boss_title_1103', '짐승 이름과 맞춤'),
    'epic_boss_title_1104': ('event_name_207004',    '사건 이름과 맞춤(숙적 = Nemesis)'),
    'ui_dig_choice_accept':  ('dig_choice_text_340001', '표의 선택지 문장과 맞춤'),
    'ui_dig_choice_decline': ('dig_choice_text_340002', '표의 선택지 문장과 맞춤'),
    'ui_save_write_failed': ('ui_settings_save_failed', '같은 사고를 알리는 다른 창과 맞춤'),
    'ui_tactics_title':     ('ui_action_tactics', '창을 여는 버튼과 맞춤'),
    'help_tactics_title':   ('ui_action_tactics', '창을 여는 버튼과 맞춤'),
}

#: 영어만 손보는 칸 → (기대하는 지금 값, 새 값, 왜)
EN_ONLY = {
    'ui_action_tactics':       ('Tactical orders', 'Tactical Orders',
                                '다른 버튼과 같은 대소문자'),
    'ui_action_tactics_close': ('Close orders', 'Close Tactical Orders',
                                '여는 버튼과 짝을 맞춘다'),
}

#: 새 키 → 언어별 문구
ADD = {
    'ui_growth_relic_hint': dict(
        kr='「유물 관리 열기」로 이 캐릭터에게 유물을 끼웁니다.',
        en='Use "Open Relics" to equip a relic on this character.',
        es='Usa «Abrir reliquias» para equipar una reliquia a este personaje.',
        fr='Utilisez « Ouvrir les reliques » pour équiper une relique à ce personnage.',
        de='Über „Relikte öffnen“ rüstest du diesem Charakter ein Relikt aus.',
        ja='「遺物管理を開く」からこのキャラクターに遺物を装着します。',
        ru='Откройте «Реликвии», чтобы надеть реликвию на этого персонажа.',
        pt='Use "Abrir relíquias" para equipar uma relíquia neste personagem.',
        pl='Użyj „Otwórz relikty”, aby założyć relikt temu bohaterowi.',
        note='성장 창 · 유물 칸의 효과 줄(하나도 안 꼈을 때)'),
}

#: 지울 죽은 키 → (기대하는 지금 kr, 왜)
REMOVE = {
    'ui_portrait_rage':  ('분노 {0:0.#} / {1:0}',   '_format 쌍둥이 · 아무도 안 부른다(미결 179)'),
    'ui_portrait_souls': ('획득한 영혼 수: {0:0}개', '_format 쌍둥이 · 아무도 안 부른다(미결 179)'),
}


def norm(v):
    return '' if v is None else str(v).strip()


def main():
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]
    fields = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    C = {n: i + 1 for i, n in enumerate(fields)}
    LANGS = [f for f in ['kr', 'en', 'es', 'fr', 'de', 'ja', 'ru', 'pt', 'pl'] if f in C]
    print('언어 열: ' + ' '.join(LANGS))

    where, last_row = {}, DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = norm(ws.cell(r, 1).value)
        if not k:
            continue
        where[k] = r
        last_row = r

    changed, skipped, added, removed = [], [], [], []

    # ── ① 영어만 손보는 칸을 <b>먼저</b> ─────────────────────────────
    # ⚠⚠ 순서가 중요하다. ②의 복사가 먼저 돌면 «아직 안 고친 정본» 을 퍼가서
    #    `ui_tactics_title` 이 소문자 «Tactical orders» 로 굳는다 — 실제로 한 번 겪었다.
    for key, (expect, new_en, why) in EN_ONLY.items():
        r = where.get(key)
        if r is None:
            skipped.append('%s (키 없음)' % key)
            continue
        now = norm(ws.cell(r, C['en']).value)
        if now == new_en:
            print('  · 이미 그렇습니다: %s' % key)
            continue
        if now != expect:
            skipped.append('%s (지금 «%s» — 기대한 «%s» 가 아니라 안 건드렸다)' % (key, now, expect))
            continue
        ws.cell(r, C['en']).value = new_en
        changed.append('%s en: %s → %s' % (key, expect, new_en))

    # ── ② 정본에서 언어별로 복사 ──────────────────────────────────────
    for key, (src, why) in ALIGN.items():
        r, rs = where.get(key), where.get(src)
        if r is None or rs is None:
            skipped.append('%s (키 없음)' % key)
            continue
        # ⚠ 한국어가 서로 다르면 «같은 말» 이 아니다 — 손대지 않는다.
        if norm(ws.cell(r, C['kr']).value) != norm(ws.cell(rs, C['kr']).value):
            skipped.append('%s (한국어가 %s 와 달라 안 건드렸다)' % (key, src))
            continue
        hits = []
        for L in LANGS:
            if L == 'kr':
                continue
            now, want = norm(ws.cell(r, C[L]).value), norm(ws.cell(rs, C[L]).value)
            if not want or now == want:
                continue
            ws.cell(r, C[L]).value = want
            hits.append(L)
        if hits:
            note = norm(ws.cell(r, C['note']).value) if 'note' in C else ''
            if 'note' in C:
                ws.cell(r, C['note']).value = (note + ' · ' + why) if note else why
            changed.append('%s ← %s (%s)' % (key, src, ','.join(hits)))

    # ── ③ 새 키 ───────────────────────────────────────────────────
    for key, vals in ADD.items():
        if key in where:
            print('  · 이미 있어 건너뜁니다: %s' % key)
            continue
        last_row += 1
        ws.cell(last_row, 1).value = key
        for L in LANGS:
            ws.cell(last_row, C[L]).value = vals[L]
        if 'source' in C:
            ws.cell(last_row, C['source']).value = SOURCE
        if 'note' in C:
            ws.cell(last_row, C['note']).value = vals.get('note', '')
        where[key] = last_row
        added.append(key)

    # ── ④ 죽은 키 지우기 (아래에서 위로) ───────────────────────────────
    kill = []
    for key, (expect_kr, why) in REMOVE.items():
        r = where.get(key)
        if r is None:
            print('  · 이미 없습니다: %s' % key)
            continue
        if norm(ws.cell(r, C['kr']).value) != expect_kr:
            skipped.append('%s (kr 이 바뀌어 있어 안 지웠다)' % key)
            continue
        kill.append((r, key, why))
    for r, key, why in sorted(kill, reverse=True):
        ws.delete_rows(r, 1)
        removed.append('%s (%s)' % (key, why))

    if not (changed or added or removed):
        print('바뀐 것이 없어 저장하지 않았습니다.')
        return 0

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.20260902d.bak')
    wb.save(STRING_XLSX)
    print('저장: %s (백업 .20260902d.bak)' % os.path.basename(STRING_XLSX))
    print('  다듬은 행 %d' % len(changed))
    for c in changed:
        print('     ~ %s' % c)
    print('  새 키 %d: %s' % (len(added), ' · '.join(added) or '-'))
    print('  지운 죽은 키 %d: %s' % (len(removed), ' · '.join(removed) or '-'))
    if skipped:
        print('  ⚠ 건너뜀 %d:' % len(skipped))
        for s in skipped:
            print('     ! %s' % s)
    return 0


if __name__ == '__main__':
    sys.exit(main())
