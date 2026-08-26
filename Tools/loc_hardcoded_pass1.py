# -*- coding: utf-8 -*-
"""하드코딩 한글을 <b>스트링 키</b>로 옮긴다 — 1차: 능력치·성장유형·전술지침·로스터·유물·초상화.

유저 지시
---------
  *"지금 하드 코딩으로 들어가 있는 텍스트들 단위별로 전부 스트링 키 테이블에도 옮기고
    영어로도 번역해줘 지금 언어를 바꿔도 번역 안되는 한글들이 많음"*

★★★ 어떻게 옮기나 — <b>리터럴을 «폴백» 으로 남긴다</b>
──────────────────────────────────────────────────────────────────────
    "체력"   →   Data.StringTable.Get("ui_stat_hp", "체력")

이 프로젝트가 이미 쓰는 모양이다(`BossSkillSO.DisplayName` 등). 이렇게 하면 —
  ★ 키가 표에 없거나 값이 비어도 <b>지금과 똑같이</b> 동작한다(폴백이 이긴다).
  ★ 코드를 읽는 사람이 <b>무슨 글이 나오는지</b> 그 자리에서 안다.
  ⚠ 그래서 <b>표를 고치면 그때부터</b> 표가 이긴다 — 폴백은 «표가 없을 때» 의 값이다.

자리표(`{0}`)가 있는 글은 <c>string.Format(Get(...), …)</c> 으로 바꾼다 —
<b>영어는 어순이 다르다</b>. 그것이 «$"" 보간을 그대로 못 두는» 이유다.

★ 이 스크립트가 하는 일
  ① C# 파일의 <b>정확한 문자열</b>을 찾아 위 모양으로 바꾼다(못 찾으면 <b>멈춘다</b>).
  ② `스트링 키 테이블.xlsx` 의 `string` 시트에 <b>키 · 한국어 · 영어</b>를 넣는다
     (이미 있는 키는 <b>영어가 비어 있을 때만</b> 채운다 — 사람이 고친 번역을 안 덮는다).
  ③ 멱등이다 — 두 번 돌려도 결과가 같다(이미 바뀐 줄은 «찾을 문자열» 이 없으므로 건너뛴다).

다음: python Tools/gen_string_table.py   (StringTable.txt 다시 내보내기)
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SCRIPTS = os.path.join("Assets", "_Project", "Scripts")
STRING_XLSX = os.path.join(TABLE_DIR, "스트링 키 테이블.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")

G = 'Data.StringTable.Get'

# ══════════════════════════════════════════════════════════════════════
#  ① 코드 교체 — {파일: [(찾을 것, 바꿀 것), …]}
# ══════════════════════════════════════════════════════════════════════
EDITS = {
    "Combat/StatBlock.cs": [
        ('StatType.Hp => "체력",', f'StatType.Hp => {G}("ui_stat_hp", "체력"),'),
        ('StatType.Attack => "근거리 공격력",',
         f'StatType.Attack => {G}("ui_stat_melee_atk", "근거리 공격력"),'),
        ('StatType.Defense => "방어력",', f'StatType.Defense => {G}("ui_stat_def", "방어력"),'),
        ('StatType.Regen => "체력 재생",', f'StatType.Regen => {G}("ui_stat_regen", "체력 재생"),'),
        ('StatType.RangedAttack => "원거리 공격력",',
         f'StatType.RangedAttack => {G}("ui_stat_ranged_atk", "원거리 공격력"),'),
        ('StatType.Magic => "마법",', f'StatType.Magic => {G}("ui_stat_magic", "마법"),'),
        ('StatType.Cure => "회복력",', f'StatType.Cure => {G}("ui_stat_cure", "회복력"),'),
        ('StatType.Accuracy => "명중률",', f'StatType.Accuracy => {G}("ui_stat_accuracy", "명중률"),'),
        ('StatType.Critical => "크리티컬 확률",',
         f'StatType.Critical => {G}("ui_stat_critical", "크리티컬 확률"),'),
        ('StatType.AttackSpeed => "공격 속도",',
         f'StatType.AttackSpeed => {G}("ui_stat_atk_speed", "공격 속도"),'),
        ('StatType.MoveSpeed => "이동속도",',
         f'StatType.MoveSpeed => {G}("ui_stat_move_speed", "이동속도"),'),
        ('StatType.Resistance => "저항력",',
         f'StatType.Resistance => {G}("ui_stat_resistance", "저항력"),'),
    ],
    "Combat/StatGrowthFocus.cs": [
        ('StatGrowthFocus.Tank      => "탱커",',
         f'StatGrowthFocus.Tank      => {G}("ui_focus_tank", "탱커"),'),
        ('StatGrowthFocus.MeleeDps  => "근거리 딜러",',
         f'StatGrowthFocus.MeleeDps  => {G}("ui_focus_melee_dps", "근거리 딜러"),'),
        ('StatGrowthFocus.RangedDps => "원거리 딜러",',
         f'StatGrowthFocus.RangedDps => {G}("ui_focus_ranged_dps", "원거리 딜러"),'),
        ('StatGrowthFocus.MagicDps  => "마법 딜러",',
         f'StatGrowthFocus.MagicDps  => {G}("ui_focus_magic_dps", "마법 딜러"),'),
        ('StatGrowthFocus.Support   => "지원가",',
         f'StatGrowthFocus.Support   => {G}("ui_focus_support", "지원가"),'),
        ('_                         => "미선택",',
         f'_                         => {G}("ui_focus_none", "미선택"),'),
    ],
    "Combat/TacticalOrder.cs": [
        # ── 요약 한 문장 — 어순이 언어마다 달라 <b>자리표 형식</b>으로 옮긴다 ──
        ('            string retreat = retreatHpPercent > 0 ? $"체력 {retreatHpPercent}% 이하에서 후퇴" : "후퇴하지 않음";\n'
         '            return $"{Label(position)} 에서 {Label(attackType)} 공격으로 {Label(targetPriority)}을(를) 노리고, " +\n'
         '                   $"{Label(attackReaction)}. 탐험 유형은 {Label(expeditionType)}({Label(roamRange)}), " +\n'
         '                   $"웨이브에는 {Label(waveReaction)}. {retreat}, 앞이 빠지면 {Label(retreatAction)}.";',
         '            string retreat = retreatHpPercent > 0\n'
         f'                ? string.Format({G}("ui_order_retreat_at", "체력 {{0}}% 이하에서 후퇴"),\n'
         '                                retreatHpPercent)\n'
         f'                : {G}("ui_order_retreat_never", "후퇴하지 않음");\n'
         '\n'
         '            // ⚠ <b>자리표 순서를 바꾸지 말 것</b> — 영어 문장은 어순이 다르므로\n'
         '            //   표의 영어 칸이 같은 번호를 다른 자리에 놓는다.\n'
         f'            return string.Format({G}("ui_order_summary",\n'
         '                    "{0} 에서 {1} 공격으로 {2}을(를) 노리고, {3}. "\n'
         '                    + "탐험 유형은 {4}({5}), 웨이브에는 {6}. {7}, 앞이 빠지면 {8}."),\n'
         '                Label(position), Label(attackType), Label(targetPriority),\n'
         '                Label(attackReaction), Label(expeditionType), Label(roamRange),\n'
         '                Label(waveReaction), retreat, Label(retreatAction));'),

        ('TacticalAttackType.Ranged => "원거리",',
         f'TacticalAttackType.Ranged => {G}("ui_atk_ranged", "원거리"),'),
        ('TacticalAttackType.Magic  => "마법",',
         f'TacticalAttackType.Magic  => {G}("ui_atk_magic", "마법"),'),
        ('TacticalAttackType.Heal   => "치유",',
         f'TacticalAttackType.Heal   => {G}("ui_atk_heal", "치유"),'),
        ('_                         => "근거리",',
         f'_                         => {G}("ui_atk_melee", "근거리"),'),

        ('TacticalRetreatAction.FallBackWithAlly => "동료와 함께 후퇴",',
         f'TacticalRetreatAction.FallBackWithAlly => {G}("ui_retreat_with_ally", "동료와 함께 후퇴"),'),
        ('_                                      => "공격 유지",',
         f'_                                      => {G}("ui_retreat_hold", "공격 유지"),'),

        ('TacticalPosition.Front => "전방",',
         f'TacticalPosition.Front => {G}("ui_pos_front", "전방"),'),
        ('TacticalPosition.Back  => "후방",',
         f'TacticalPosition.Back  => {G}("ui_pos_back", "후방"),'),
        ('_                      => "중위",',
         f'_                      => {G}("ui_pos_mid", "중위"),'),

        ('TacticalTargetPriority.Strongest => "가장 강력한 적",',
         f'TacticalTargetPriority.Strongest => {G}("ui_target_strongest", "가장 강력한 적"),'),
        ('TacticalTargetPriority.Farthest  => "가장 먼 적",',
         f'TacticalTargetPriority.Farthest  => {G}("ui_target_farthest", "가장 먼 적"),'),
        ('TacticalTargetPriority.Weakest   => "가장 체력이 적은 적",',
         f'TacticalTargetPriority.Weakest   => {G}("ui_target_weakest", "가장 체력이 적은 적"),'),
        ('_                                => "가장 가까운 적",',
         f'_                                => {G}("ui_target_nearest", "가장 가까운 적"),'),

        ('            v == TacticalAttackReaction.HoldGround ? "사거리에 들어올 때까지 대기" : "시야 내의 적을 쫓아가 공격";',
         '            v == TacticalAttackReaction.HoldGround\n'
         f'                ? {G}("ui_reaction_hold", "사거리에 들어올 때까지 대기")\n'
         f'                : {G}("ui_reaction_chase", "시야 내의 적을 쫓아가 공격");'),

        ('TacticalExpeditionType.Explore => "탐색",',
         f'TacticalExpeditionType.Explore => {G}("ui_scout_explore", "탐색"),'),
        ('_                              => "사냥",',
         f'_                              => {G}("ui_scout_hunt", "사냥"),'),

        ('            TacticalExpeditionType.Explore =>\n'
         '                "탐색 — 탐험 중 중립 몬스터를 절대 공격하지 않고 안개만 밝히며 돌아다닙니다. " +\n'
         '                "중립 몬스터에게 공격당해도 반격 없이 그 자리를 벗어나 도망칩니다(웨이브 " +\n'
         '                "몬스터에게 맞을 때는 평소처럼 맞서 싸웁니다). 전투를 최대한 피해 안전하게 " +\n'
         '                "시야를 넓히고 싶을 때 적합합니다.",',
         '            TacticalExpeditionType.Explore =>\n'
         f'                {G}("ui_scout_explore_desc",\n'
         '                    "탐색 — 탐험 중 중립 몬스터를 절대 공격하지 않고 안개만 밝히며 돌아다닙니다. " +\n'
         '                    "중립 몬스터에게 공격당해도 반격 없이 그 자리를 벗어나 도망칩니다(웨이브 " +\n'
         '                    "몬스터에게 맞을 때는 평소처럼 맞서 싸웁니다). 전투를 최대한 피해 안전하게 " +\n'
         '                    "시야를 넓히고 싶을 때 적합합니다."),'),

        ('            _ =>\n'
         '                "사냥 — 탐험 중 중립 몬스터를 발견하면 먼저 다가가 공격해 사냥합니다. 처치하면 " +\n'
         '                "에너지를 얻지만, 사냥에 나서는 동안 부대에서 떨어지거나 강한 개체와 마주쳐 " +\n'
         '                "위험해질 수 있습니다.",',
         '            _ =>\n'
         f'                {G}("ui_scout_hunt_desc",\n'
         '                    "사냥 — 탐험 중 중립 몬스터를 발견하면 먼저 다가가 공격해 사냥합니다. 처치하면 " +\n'
         '                    "에너지를 얻지만, 사냥에 나서는 동안 부대에서 떨어지거나 강한 개체와 마주쳐 " +\n'
         '                    "위험해질 수 있습니다."),'),

        ('TacticalRoamRange.Mid => "외곽",',
         f'TacticalRoamRange.Mid => {G}("ui_roam_mid", "외곽"),'),
        ('TacticalRoamRange.Far => "전역",',
         f'TacticalRoamRange.Far => {G}("ui_roam_far", "전역"),'),
        ('_                     => "근방",',
         f'_                     => {G}("ui_roam_near", "근방"),'),

        ('            TacticalRoamRange.Mid =>\n'
         '                "성역 바깥 지대까지 나가 탐험합니다. 밝힐 곳이 더 많지만 그만큼 위험합니다.",',
         '            TacticalRoamRange.Mid =>\n'
         f'                {G}("ui_roam_mid_desc",\n'
         '                    "성역 바깥 지대까지 나가 탐험합니다. 밝힐 곳이 더 많지만 그만큼 위험합니다."),'),
        ('            TacticalRoamRange.Far =>\n'
         '                "제한 없이 맵 끝까지 나갑니다. 미지의 존재와 마주칠 각오가 필요합니다.",',
         '            TacticalRoamRange.Far =>\n'
         f'                {G}("ui_roam_far_desc",\n'
         '                    "제한 없이 맵 끝까지 나갑니다. 미지의 존재와 마주칠 각오가 필요합니다."),'),
        ('            _ =>\n'
         '                "성역 가까운 곳만 돕니다. 위험은 적지만 넓은 지역을 밝히지 못합니다.",',
         '            _ =>\n'
         f'                {G}("ui_roam_near_desc",\n'
         '                    "성역 가까운 곳만 돕니다. 위험은 적지만 넓은 지역을 밝히지 못합니다."),'),

        ('            v == TacticalWaveReaction.KeepExploring ? "탐험 우선" : "즉시 방어";',
         '            v == TacticalWaveReaction.KeepExploring\n'
         f'                ? {G}("ui_wave_keep_exploring", "탐험 우선")\n'
         f'                : {G}("ui_wave_defend_now", "즉시 방어");'),
    ],
    "UI/CharacterRosterPanel.cs": [
        ('if (row.Duty != null) { row.Duty.text = "사망"; row.Duty.color = deadTextColor; }',
         'if (row.Duty != null)\n'
         '            {\n'
         f'                row.Duty.text = {G}("ui_duty_dead", "사망");\n'
         '                row.Duty.color = deadTextColor;\n'
         '            }'),
        ('if (behavior != null && behavior.IsRetreating) return "후퇴";',
         f'if (behavior != null && behavior.IsRetreating) return {G}("ui_duty_retreat", "후퇴");'),
        ('if (behavior != null && behavior.IsFleeing) return "도망";',
         f'if (behavior != null && behavior.IsFleeing) return {G}("ui_duty_flee", "도망");'),
        ('if (combat.AttackType == TacticalAttackType.Heal) return "치유";',
         f'if (combat.AttackType == TacticalAttackType.Heal) return {G}("ui_duty_heal", "치유");'),
        ('return combat.IsHunting ? "사냥" : "교전";',
         f'return combat.IsHunting ? {G}("ui_duty_hunt", "사냥")\n'
         f'                                        : {G}("ui_duty_fight", "교전");'),
        ('CharacterDuty.Expedition   => "탐험",',
         f'CharacterDuty.Expedition   => {G}("ui_duty_expedition", "탐험"),'),
        ('CharacterDuty.Rally   => "집결",',
         f'CharacterDuty.Rally   => {G}("ui_duty_rally", "집결"),'),
        ('CharacterDuty.Retreat => "후퇴",',
         f'CharacterDuty.Retreat => {G}("ui_duty_retreat", "후퇴"),'),
        ('CharacterDuty.Flee    => "도망",',
         f'CharacterDuty.Flee    => {G}("ui_duty_flee", "도망"),'),
        ('CharacterDuty.Build   => "건설",',
         f'CharacterDuty.Build   => {G}("ui_duty_build", "건설"),'),
        ('_                     => "방어",',
         f'_                     => {G}("ui_duty_guard", "방어"),'),
    ],
    "UI/RelicPanel.cs": [
        ('[SerializeField] string hintEmpty = "아직 얻은 유물이 없습니다. 발굴하거나 사냥해 보세요.";',
         '[SerializeField] string hintEmpty = "아직 얻은 유물이 없습니다. 발굴하거나 사냥해 보세요.";\n'
         '        [SerializeField] string hintEmptyKey = "ui_relic_hint_empty";'),
        ('[SerializeField] string hintPick = "유물을 고르면 설명이 나옵니다.";',
         '[SerializeField] string hintPick = "유물을 고르면 설명이 나옵니다.";\n'
         '        [SerializeField] string hintPickKey = "ui_relic_hint_pick";'),
        ('[SerializeField] string hintNoCharacter = "장착하려면 로스터에서 캐릭터를 먼저 고르세요.";',
         '[SerializeField] string hintNoCharacter = "장착하려면 로스터에서 캐릭터를 먼저 고르세요.";\n'
         '        [SerializeField] string hintNoCharacterKey = "ui_relic_hint_no_character";'),
        ('            RelicSource.Boss => "보스 처치로만 얻습니다.",',
         f'            RelicSource.Boss => {G}("ui_relic_src_boss", "보스 처치로만 얻습니다."),'),
        ('            RelicSource.DigOnly => "발굴로만 얻습니다.",',
         f'            RelicSource.DigOnly => {G}("ui_relic_src_dig", "발굴로만 얻습니다."),'),
        ('            RelicSource.Event => "사건에서만 얻습니다.",',
         f'            RelicSource.Event => {G}("ui_relic_src_event", "사건에서만 얻습니다."),'),
        ('            _ => "발굴하거나 일반 몬스터를 사냥해 얻습니다.",',
         f'            _ => {G}("ui_relic_src_common", "발굴하거나 일반 몬스터를 사냥해 얻습니다."),'),
        ('            return c != null ? c.DisplayName : "다른 캐릭터";',
         f'            return c != null ? c.DisplayName : {G}("ui_relic_other_wearer", "다른 캐릭터");'),
    ],
}

# ══════════════════════════════════════════════════════════════════════
#  ② 스트링 키 — (키, 한국어, 영어)
# ══════════════════════════════════════════════════════════════════════
KEYS = [
    # ── 능력치 열두 칸 ──
    ("ui_stat_hp", "체력", "Health"),
    ("ui_stat_melee_atk", "근거리 공격력", "Melee Attack"),
    ("ui_stat_def", "방어력", "Defense"),
    ("ui_stat_regen", "체력 재생", "Health Regeneration"),
    ("ui_stat_ranged_atk", "원거리 공격력", "Ranged Attack"),
    ("ui_stat_magic", "마법", "Magic"),
    ("ui_stat_cure", "회복력", "Healing"),
    ("ui_stat_accuracy", "명중률", "Accuracy"),
    ("ui_stat_critical", "크리티컬 확률", "Critical Chance"),
    ("ui_stat_atk_speed", "공격 속도", "Attack Speed"),
    ("ui_stat_move_speed", "이동속도", "Movement Speed"),
    ("ui_stat_resistance", "저항력", "Resistance"),
    # ── 성장 유형 ──
    ("ui_focus_tank", "탱커", "Tank"),
    ("ui_focus_melee_dps", "근거리 딜러", "Melee Damage"),
    ("ui_focus_ranged_dps", "원거리 딜러", "Ranged Damage"),
    ("ui_focus_magic_dps", "마법 딜러", "Magic Damage"),
    ("ui_focus_support", "지원가", "Support"),
    ("ui_focus_none", "미선택", "Not chosen"),
    # ── 전술 지침 ──
    ("ui_atk_melee", "근거리", "Melee"),
    ("ui_atk_ranged", "원거리", "Ranged"),
    ("ui_atk_magic", "마법", "Magic"),
    ("ui_atk_heal", "치유", "Healing"),
    ("ui_pos_front", "전방", "Front"),
    ("ui_pos_mid", "중위", "Middle"),
    ("ui_pos_back", "후방", "Rear"),
    ("ui_target_nearest", "가장 가까운 적", "Nearest enemy"),
    ("ui_target_strongest", "가장 강력한 적", "Strongest enemy"),
    ("ui_target_farthest", "가장 먼 적", "Farthest enemy"),
    ("ui_target_weakest", "가장 체력이 적은 적", "Enemy with the least health"),
    ("ui_reaction_hold", "사거리에 들어올 때까지 대기", "Hold until they come into range"),
    ("ui_reaction_chase", "시야 내의 적을 쫓아가 공격", "Chase and attack enemies in sight"),
    ("ui_retreat_with_ally", "동료와 함께 후퇴", "Fall back with allies"),
    ("ui_retreat_hold", "공격 유지", "Keep attacking"),
    ("ui_retreat_at", "체력 {0}% 이하에서 후퇴", "Retreat below {0}% health"),
    ("ui_order_retreat_at", "체력 {0}% 이하에서 후퇴", "Retreat below {0}% health"),
    ("ui_order_retreat_never", "후퇴하지 않음", "Never retreats"),
    ("ui_order_summary",
     "{0} 에서 {1} 공격으로 {2}을(를) 노리고, {3}. 탐험 유형은 {4}({5}), 웨이브에는 {6}. {7}, 앞이 빠지면 {8}.",
     "Holds the {0} line, attacks at {1} range, targets the {2}, and will {3}. "
     "Expedition: {4} ({5}). On a wave: {6}. {7}; if the front falls back, {8}."),
    ("ui_scout_explore", "탐색", "Scout"),
    ("ui_scout_hunt", "사냥", "Hunt"),
    ("ui_scout_explore_desc",
     "탐색 — 탐험 중 중립 몬스터를 절대 공격하지 않고 안개만 밝히며 돌아다닙니다. "
     "중립 몬스터에게 공격당해도 반격 없이 그 자리를 벗어나 도망칩니다(웨이브 "
     "몬스터에게 맞을 때는 평소처럼 맞서 싸웁니다). 전투를 최대한 피해 안전하게 "
     "시야를 넓히고 싶을 때 적합합니다.",
     "Scout - never attacks neutral monsters while exploring; only lifts the fog. "
     "If a neutral monster strikes them, they leave without striking back (they still "
     "fight wave monsters as usual). Suited to widening your sight while avoiding battle."),
    ("ui_scout_hunt_desc",
     "사냥 — 탐험 중 중립 몬스터를 발견하면 먼저 다가가 공격해 사냥합니다. 처치하면 "
     "에너지를 얻지만, 사냥에 나서는 동안 부대에서 떨어지거나 강한 개체와 마주쳐 "
     "위험해질 수 있습니다.",
     "Hunt - approaches and attacks any neutral monster found while exploring. Kills "
     "bring energy, but a hunter can drift away from the squad or run into something "
     "far stronger."),
    ("ui_roam_near", "근방", "Near"),
    ("ui_roam_mid", "외곽", "Outer"),
    ("ui_roam_far", "전역", "Whole map"),
    ("ui_roam_near_desc", "성역 가까운 곳만 돕니다. 위험은 적지만 넓은 지역을 밝히지 못합니다.",
     "Stays close to the sanctuary. Safer, but little of the map gets lifted."),
    ("ui_roam_mid_desc", "성역 바깥 지대까지 나가 탐험합니다. 밝힐 곳이 더 많지만 그만큼 위험합니다.",
     "Goes out past the sanctuary's edge. More to uncover, and more danger with it."),
    ("ui_roam_far_desc", "제한 없이 맵 끝까지 나갑니다. 미지의 존재와 마주칠 각오가 필요합니다.",
     "Roams to the edge of the map without limit. Be ready to meet the unknown."),
    ("ui_wave_keep_exploring", "탐험 우선", "Keep exploring"),
    ("ui_wave_defend_now", "즉시 방어", "Defend at once"),
    # ── 로스터의 «지금 무엇을 하는가» ──
    ("ui_duty_dead", "사망", "Dead"),
    ("ui_duty_retreat", "후퇴", "Retreating"),
    ("ui_duty_flee", "도망", "Fleeing"),
    ("ui_duty_heal", "치유", "Healing"),
    ("ui_duty_hunt", "사냥", "Hunting"),
    ("ui_duty_fight", "교전", "Fighting"),
    ("ui_duty_expedition", "탐험", "Exploring"),
    ("ui_duty_rally", "집결", "Rallying"),
    ("ui_duty_build", "건설", "Building"),
    ("ui_duty_guard", "방어", "Guarding"),
    # ── 유물 관리 ──
    ("ui_relic_hint_empty", "아직 얻은 유물이 없습니다. 발굴하거나 사냥해 보세요.",
     "You have no relics yet. Try digging or hunting."),
    ("ui_relic_hint_pick", "유물을 고르면 설명이 나옵니다.", "Choose a relic to see what it does."),
    ("ui_relic_hint_no_character", "장착하려면 로스터에서 캐릭터를 먼저 고르세요.",
     "Pick a character in the roster first to equip it."),
    ("ui_relic_slot_format", "(유물 칸 {0}/{1})", "(relic slots {0}/{1})"),
    ("ui_relic_equip", "장착", "Equip"),
    ("ui_relic_unequip", "해제", "Unequip"),
    ("ui_relic_wearer", "{0} 착용 중", "Worn by {0}"),
    ("ui_relic_wearer_more", "{0} 외 {1}", "{0} and {1} more"),
    ("ui_relic_count", "x{0}", "x{0}"),
    ("ui_relic_src_boss", "보스 처치로만 얻습니다.", "Only dropped by bosses."),
    ("ui_relic_src_dig", "발굴로만 얻습니다.", "Only found by digging."),
    ("ui_relic_src_event", "사건에서만 얻습니다.", "Only given by events."),
    ("ui_relic_src_common", "발굴하거나 일반 몬스터를 사냥해 얻습니다.",
     "Found by digging or by hunting ordinary monsters."),
    ("ui_relic_other_wearer", "다른 캐릭터", "another character"),
    ("ui_relic_empty_slot", "빈 칸", "Empty"),
    ("ui_relic_empty_hint", "눌러서 유물을 끼웁니다.", "Press to equip a relic."),
    ("ui_relic_pick_character", "캐릭터를 선택하세요", "Select a character"),
    # ── 초상화 창 ──
    ("ui_portrait_no_art", "일러스트 없음", "No illustration"),
    ("ui_portrait_rage", "분노 {0:0.#} / {1:0}", "Rage {0:0.#} / {1:0}"),
    ("ui_portrait_souls", "획득한 영혼 수: {0:0}개", "Souls gathered: {0:0}"),
]


def apply_code_edits():
    changed = 0
    for rel, pairs in EDITS.items():
        path = os.path.join(SCRIPTS, rel.replace("/", os.sep))
        with open(path, encoding="utf-8-sig", newline="") as f:
            src = f.read()
        crlf = "\r\n" in src
        flat = src.replace("\r\n", "\n")

        hit = 0
        for old, new in pairs:
            if new in flat:
                continue                        # 이미 바꿨다(멱등)
            if old not in flat:
                sys.exit(f"! 못 찾음 — {rel}\n  찾던 것: {old[:90]}")
            flat = flat.replace(old, new, 1)
            hit += 1

        if hit == 0:
            print(f"  {rel} — 바꿀 것이 없다(이미 반영)")
            continue

        out = flat.replace("\n", "\r\n") if crlf else flat
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(out)
        print(f"  {rel} — {hit}곳")
        changed += hit
    return changed


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_스트링키1차")
    os.makedirs(folder, exist_ok=True)
    shutil.copy2(STRING_XLSX, os.path.join(folder, os.path.basename(STRING_XLSX)))
    print("백업: " + folder)


def apply_string_rows():
    """`string` 시트에 키를 넣는다. 이미 있으면 <b>빈 칸만</b> 채운다."""
    wb = openpyxl.load_workbook(STRING_XLSX, data_only=True)
    ws = wb["string"]
    where = {}
    last = 3
    for r in range(4, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k is None or not str(k).strip():
            continue
        where[str(k).strip()] = r
        last = r

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    added = filled = 0
    try:
        wbc = excel.Workbooks.Open(os.path.abspath(STRING_XLSX))
        sh = wbc.Worksheets("string")
        row = last + 1
        for key, kr, en in KEYS:
            if key in where:
                r = where[key]
                if not str(sh.Cells(r, 2).Value or "").strip():
                    sh.Cells(r, 2).Value = kr
                    filled += 1
                if not str(sh.Cells(r, 3).Value or "").strip():
                    sh.Cells(r, 3).Value = en
                    filled += 1
                continue
            sh.Cells(row, 1).Value = key
            sh.Cells(row, 2).Value = kr
            sh.Cells(row, 3).Value = en
            sh.Cells(row, 4).Value = "code(하드코딩)"
            sh.Cells(row, 5).Value = "2026-08-26 하드코딩 이관 1차"
            row += 1
            added += 1
        wbc.Save()
        wbc.Close()
    finally:
        excel.Quit()
    return added, filled


def main():
    print("① 코드 교체")
    n = apply_code_edits()
    print(f"   → {n}곳 바꿨다\n")

    backup()
    print("② 스트링 키 테이블")
    added, filled = apply_string_rows()
    print(f"   → 새 키 {added}개 · 빈 칸 채움 {filled}개 (전체 {len(KEYS)}개 중)")
    print("\n다음: python Tools/gen_string_table.py")


if __name__ == "__main__":
    main()
