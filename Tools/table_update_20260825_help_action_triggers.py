# -*- coding: utf-8 -*-
"""도움말 표 — 계기를 «허드 액션 버튼의 첫 클릭» 으로 옮기고 포탑 건설을 지운다 (2026-08-25).

유저 지시 셋 + 하나:

  1. *"게임 최초 시작 시 중앙건물->정비시간 타이머에 대한 설명 -> 자원에 대한 설명으로
     기본 듀토리얼 진행"*
  2. *"이후엔 허드 액션의 각 버튼을 최초로 눌렀을때 해당 기능에 대한 도움말이 등장"*
  3. *"허드 액션에 있는 버튼을 눌러서 뜬 도움말을 자세히 보기를 눌렀을 때 실제 해당 ui가
     켜지고 각 기능에 대한 설명 시작"*
  4. *"도움말에서 포탑 건설 관련 설명 삭제해 해당 기능 없어졌어"*

★★ <b>이 표는 사람이 엑셀에서 고쳐 나가는 문서다</b> — `gen_help_table.py` 처럼 통째로
   다시 굽지 않고, <b>바꿀 칸만</b> 고친다(144·145절이 쓴 방식 그대로다).

★★★ <b>왜 «결과» 에서 «버튼» 으로 되돌리는가</b>
────────────────────────────────────────────
144-2 절은 계기를 «버튼» 이 아니라 <b>«기능이 실제로 일어난 자리»</b> 에 걸었다. 이유는
«누르는 통로가 여럿이라 버튼마다 세면 하나를 빠뜨린다» 였고, 그 판단은 <b>지금도 맞다</b>.

그런데 그 설계에는 유저가 원한 흐름이 <b>담기지 않는다</b> — 「강화가 <b>끝난 뒤</b>에
강화를 설명하는」 순서가 되기 때문이다. 튜토리얼은 <b>하기 전에</b> 알려주는 것이다.
그래서 <b>허드 액션의 일곱 버튼만</b> 버튼으로 옮긴다. 통로를 빠뜨릴 걱정은 없다 —
이 일곱은 <b>그 창을 여는 유일한 통로</b>이고, 창이 열리는 순간이 곧 «처음 쓴다» 다.

  ⚠ 나머지 계기(전투·죽음·침식·사건 …)는 <b>그대로 둔다</b>. 그것들은 버튼이 아니라
    <b>상황</b>이고, 144-2 의 이유가 그쪽에서는 여전히 유효하다.

사용법:  py -3 Tools/table_update_20260825_help_action_triggers.py
다음:    py -3 Tools/help_string_merge.py
         py -3 Tools/gen_string_table.py
         py -3 Tools/link_string_keys.py
         py -3 Tools/gen_help_assets.py
"""

import io
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HELP_XLSX = os.path.join(TABLE_DIR, "Last_Sanctuary_도움말테이블_Ver01.xlsx")
STRING_XLSX = os.path.join(TABLE_DIR, "스트링 키 테이블.xlsx")

# ══════════════════════════════════════════════════════════════════════
#  ① 기본 튜토리얼 세 장 — 판을 켜면 «넥서스 → 정비 타이머 → 자원» 순서로 뜬다
# ══════════════════════════════════════════════════════════════════════
# 대기줄은 priority → order 로 정렬된다(HelpService.Fire). 셋을 같은 계기에 걸고
# priority 로 순서를 못박는다. ★ 에너지는 예전에 «에너지가 처음 늘었을 때» 였는데,
#   그것은 <b>몬스터를 잡은 뒤</b>라 «기본 튜토리얼» 안에 들어오지 못했다.
BASIC_TUTORIAL = {
    "help_nexus":  ("NewRunFirstPreparation", 1),   # 중앙건물
    "help_wave":   ("NewRunFirstPreparation", 2),   # 정비 시간 타이머
    "help_energy": ("NewRunFirstPreparation", 3),   # 자원
}

# ══════════════════════════════════════════════════════════════════════
#  ② 허드 액션 버튼 일곱 — «처음 눌렀을 때» 그 기능의 도움말이 뜬다
# ══════════════════════════════════════════════════════════════════════
# ⚠ 여기 적는 이름은 <b>C# 의 HelpTrigger enum 과 글자 그대로 같아야 한다</b>
#   (gen_help_assets.py 의 TRIGGER 표가 그 다리다).
ACTION_TRIGGER = {
    "help_create":      "ActionCreate",      # CreateButton   — 캐릭터 생성
    "help_upgrade":     "ActionUpgrade",     # Upgrade        — 캐릭터 성장(강화)
    "help_squad":       "ActionSquad",       # SquadButton    — 부대 설정
    "help_tactics":     "ActionTactics",     # TacticsButton  — 전술 지침
    "help_epic":        "ActionSubjugate",   # SubjugateButton— 토벌 지시
    "help_relic_equip": "ActionRelic",       # RelicButton    — 유물 관리
    "help_save":        "ActionSettings",    # SettingsButton — 환경 설정
}

# ══════════════════════════════════════════════════════════════════════
#  ③ 지운다 — 포탑 건설 (유저: *"해당 기능 없어졌어"*)
# ══════════════════════════════════════════════════════════════════════
# ★ 씬에서도 이미 꺼져 있다(`HUD_Actions/Buttons/BuildButton` 의 activeSelf = false).
#   도움말만 남으면 <b>있지도 않은 버튼을 가리키는 설명</b>이 된다.
DROP_IDS = {"help_build"}


def col_index(ws, header_row, name):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, c).value or "").strip() == name:
            return c
    raise SystemExit(f"[실패] '{ws.title}' 시트에 '{name}' 열이 없습니다")


def main():
    if not os.path.exists(HELP_XLSX):
        raise SystemExit(f"[실패] 표를 찾지 못했습니다: {HELP_XLSX}")

    shutil.copy2(HELP_XLSX, HELP_XLSX + ".bak")
    wb = openpyxl.load_workbook(HELP_XLSX)

    # ── Help 시트 ────────────────────────────────────────────────────
    ws = wb["Help"]
    c_id = col_index(ws, 1, "help_id")
    c_trigger = col_index(ws, 1, "trigger")
    c_priority = col_index(ws, 1, "priority")
    c_see = col_index(ws, 1, "see_also")

    changed, dropped_rows = [], []
    for r in range(2, ws.max_row + 1):
        hid = str(ws.cell(r, c_id).value or "").strip()
        if not hid:
            continue

        if hid in DROP_IDS:
            dropped_rows.append(r)
            continue

        old = str(ws.cell(r, c_trigger).value or "").strip()

        if hid in BASIC_TUTORIAL:
            trig, prio = BASIC_TUTORIAL[hid]
            ws.cell(r, c_trigger).value = trig
            ws.cell(r, c_priority).value = prio
            if old != trig:
                changed.append((hid, old, trig, prio))
        elif hid in ACTION_TRIGGER:
            trig = ACTION_TRIGGER[hid]
            ws.cell(r, c_trigger).value = trig
            if old != trig:
                changed.append((hid, old, trig, ws.cell(r, c_priority).value))

        # ⚠ 지워진 항목을 가리키는 see_also 가 남으면 <b>없는 버튼</b>이 백과에 뜬다.
        if str(ws.cell(r, c_see).value or "").strip() in DROP_IDS:
            ws.cell(r, c_see).value = None
            changed.append((hid, "see_also→help_build", "(지움)", ""))

    # ⚠ 아래에서 위로 지운다 — 위에서 지우면 뒤 행 번호가 밀린다.
    for r in reversed(dropped_rows):
        ws.delete_rows(r)

    # ── StringKeys 시트 ──────────────────────────────────────────────
    wsk = wb["StringKeys"]
    c_key = col_index(wsk, 1, "string_key")
    kill_prefix = tuple(f"{i}_" for i in DROP_IDS)
    kill_rows = [r for r in range(2, wsk.max_row + 1)
                 if str(wsk.cell(r, c_key).value or "").strip().startswith(kill_prefix)]
    killed_keys = [str(wsk.cell(r, c_key).value).strip() for r in kill_rows]
    for r in reversed(kill_rows):
        wsk.delete_rows(r)

    # ── HelpStep 시트 — 지운 항목의 단계도 함께 ──────────────────────
    wss = wb["HelpStep"]
    c_sid = col_index(wss, 1, "help_id")
    step_rows = [r for r in range(2, wss.max_row + 1)
                 if str(wss.cell(r, c_sid).value or "").strip() in DROP_IDS]
    for r in reversed(step_rows):
        wss.delete_rows(r)

    try:
        wb.save(HELP_XLSX)
    except PermissionError:
        raise SystemExit(f"[실패] 엑셀에서 열려 있습니다 — 닫고 다시 돌리십시오:\n  {HELP_XLSX}")

    print(f"[도움말 표] {os.path.basename(HELP_XLSX)}")
    for hid, old, new, prio in changed:
        print(f"   {hid:<18} {old or '(없음)':<24} → {new}   prio {prio}")
    print(f"   지운 항목 {len(dropped_rows)}개 · StringKeys {len(kill_rows)}줄 · "
          f"HelpStep {len(step_rows)}줄")

    # ── 스트링 키 테이블에서도 지운다 ────────────────────────────────
    # ⚠ help_string_merge.py 는 «덮어쓰기» 만 한다 — <b>지우지는 않는다</b>.
    #   여기서 안 지우면 「포탑 건설」 문구가 스트링 테이블에 <b>영영 남는다</b>.
    if killed_keys and os.path.exists(STRING_XLSX):
        shutil.copy2(STRING_XLSX, STRING_XLSX + ".bak")
        swb = openpyxl.load_workbook(STRING_XLSX)
        sws = swb["string"]
        want = set(killed_keys)
        hit = [r for r in range(4, sws.max_row + 1)
               if str(sws.cell(r, 1).value or "").strip() in want]
        for r in reversed(hit):
            sws.delete_rows(r)
        try:
            swb.save(STRING_XLSX)
        except PermissionError:
            raise SystemExit(f"[실패] 엑셀에서 열려 있습니다 — 닫고 다시 돌리십시오:\n  {STRING_XLSX}")
        print(f"[스트링 키 테이블] help_build 문구 {len(hit)}줄을 지웠습니다")


if __name__ == "__main__":
    main()
