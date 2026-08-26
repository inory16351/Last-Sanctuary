# -*- coding: utf-8 -*-
"""경제 재밸런싱 — 「30웨이브에 12명 평균 Lv35」로 착지점을 옮긴다 (2026-08-26).

유저 지시: *"경제 밸런스 12명 35lv 기준으로 재 밸런싱(데이터 변경 허용하나 중대한 변경점은
질문으로 처리 / 변경한 데이터 값은 테이블에 반드시 반영)"*.

유저 확정 넷 (질문으로 받은 답):

  ① 강화 비용의 «벽» 은 **Lv30 에 그대로 둔다** (35 로 옮기지 않는다)
  ② 몬스터 수치는 **그대로 둔다** — 기획서가 이미 «Lv35 3부대면 무난하게 승리» 로 정의한 그 상태가 된다
  ③ 모자란 수입은 **중립·에픽 보상만** 올려서 채운다
  ④ 유물 칸이 셋이 되어도 유물 수치는 그대로 (효과가 그대로 3배)

★★★ 왜 «중립·에픽만» 으로 곡선이 만들어지는가
────────────────────────────────────────────
중립 몬스터는 표에 **성역으로부터의 등장 거리**가 박혀 있다:

    1001 종양 거미   15~99 타일    ← 초반부터 닿는다
    1002 종양충      100~199       ← 부대를 꾸리면 닿는다 (중반)
    1004 고르도네    100~199
    1003 종양 두더지 200~320       ← 원정을 보내야 닿는다 (후반)
    1101~1104 에픽   200~320       ← 요구 레벨이 10/15/20/25 (후반)

그래서 **먼 종의 보상만 올리면 «후반에만» 수입이 오른다.** 구간마다 배율을 따로 넣는
장치를 새로 만들 필요가 없다 — 거리와 요구 레벨이 이미 그 일을 하고 있다.
밸런스 기획서의 «초반은 처치만으로 충분 / 중반은 적극적인 중립 사냥 / 후반은 에픽 보상을
적극 활용» 이 정확히 이 모양이다.

    1001 유지 (초반을 안 건드린다) · 1002·1004 ×1.3 · 1003 ×3.0 · 에픽 4종 ×3.0

경제 모델로 역산한 «중립 수입/웨이브» 배율은 구간별로
1~10 ×1.00 · 11~15 ×1.15 · 16~20 ×1.30 · 21~25 ×1.55 · 26~30 ×2.33 이고,
위 종별 배수를 구간별 사냥 구성에 얹으면 그 곡선에 들어맞는다(오차 ±10%).

⚠ 이 표들은 **사람이 엑셀에서 고쳐 나가는 문서**다 — `gen_*.py` 처럼 통째로 다시 굽지 않고
  **바꿀 칸만** 고친다(136·144·145절이 쓴 방식 그대로).
⚠ **Excel COM 으로 쓴다** — openpyxl 로 저장하면 하이퍼링크·주석·수식 캐시가 날아간다(136-4절).

사용법:  python Tools/table_update_20260826_econ_lv35.py
다음:    python Tools/sync_tables_to_assets.py     (중립 에셋에 반영)
         Unity 에서 Assets/Refresh
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

NEUTRAL_XLSX = os.path.join(TABLE_DIR, "임시용 중립 몬스터.xlsx")
FORMULA_XLSX = os.path.join(TABLE_DIR, "능력치 및 공식 정리.xlsx")

# ══════════════════════════════════════════════════════════════════════
#  ① 중립·에픽 보상 — mon_id → (min_energy, max_energy)
# ══════════════════════════════════════════════════════════════════════
# ★ 1001 은 <b>일부러 뺐다</b> — 초반 «처치만으로도 충분» 구간을 건드리지 않는다.
NEW_ENERGY = {
    1002: (34, 57),          # ×1.3  (26~44)
    1004: (34, 57),          # ×1.3  (26~44)
    1003: (600, 990),        # ×3.0  (200~330)
    1101: (1800, 2700),      # ×3.0  (600~900)
    1102: (1800, 2700),      # ×3.0
    1103: (1800, 2700),      # ×3.0
    1104: (2700, 3600),      # ×3.0  (900~1200)
}

# ══════════════════════════════════════════════════════════════════════
#  ② 「웨이브 부하」 시트 — 경제 모델을 다시 돌린 결과
# ══════════════════════════════════════════════════════════════════════
# 구간별 중립 수입 배율 (위 ★★★).
BANDS = [(10, 1.00), (15, 1.15), (20, 1.30), (25, 1.55), (30, 2.329)]

# 성장 곡선 — 136-0절 확정값에서 그대로 온다(실측으로 재확인: 시트의 옛 값과 한 칸도 안 어긋난다).
FOCUS_PER_LEVEL = 2.62     # 주력 스탯 (성장 유형에 묶인 것)
PLAIN_PER_LEVEL = 1.62     # 일반 스탯
STAT_BASE_FOCUS = 10
STAT_BASE_PLAIN = 5
STAT_MAX = 100             # BalanceConfig.statMax — Lv35 에 주력이 여기 닿는다

HP_PER_STAT = 6.1832       # 캐릭터 HP = 90 + 6.1832 × (주력 − 10)  (시트 실측 역산)
HP_BASE = 90

# 파티 DPS 1인분 = a·s² + b·s + c   (s = 주력 스탯)
# ★ 시트의 옛 30행에서 최소제곱으로 뽑은 <b>맞춘 곡선</b>이다 — 원래 계산 스크립트가
#   남아 있지 않아(135-1절의 스크래치패드) 값에서 되짚었다. 옛 행과의 오차는 ±5% 다.
DPS_A, DPS_B, DPS_C = 0.01961, 3.329, -9.50

# 잡몹 부하 = 총 잡몹 체력 ÷ (파티 DPS × EFFECTIVE × 120초)
# ★ 시트의 옛 행 전부에서 같은 값(0.647)이 나왔다 — 27-7절의 «실효 화력 60%» 가정이다.
EFFECTIVE = 0.647
BATTLE_SECONDS = 120

# 유료 생성 비용 = BASE + STEP × (지금까지 만든 수)
CREATE_BASE, CREATE_STEP = 200, 150
FREE_CHARACTERS = 4        # 판 시작에 그냥 서 있는 인원

# 강화 비용 = BASE + STEP × Lv, Lv30 부터 등비 (유저 확정 ① — 벽은 30 에 그대로)
UP_BASE, UP_STEP = 40, 10
UP_STEEP_LEVEL, UP_STEEP_RATE, UP_ROUND = 30, 1.35, 10

WAVE_KILL_ENERGY = 10      # 웨이브 잡몹 한 마리당 (변경 없음)

SHEET = "웨이브 부하"
FIRST_ROW = 7              # 표 머리 다음 줄
COL = dict(wave=1, lv=2, people=3, focus=4, plain=5, hp=6,
           hp_scale=7, atk_scale=8, count=9, mob_hp=10, mob_hit=11,
           hits=12, hit_ratio=13, dps=14, load=15,
           boss=16, boss_raw=17, boss_hp=18, boss_goal=19, neutral=20)


def band_multiplier(wave):
    for upto, mult in BANDS:
        if wave <= upto:
            return mult
    return BANDS[-1][1]


def upgrade_cost(level):
    linear = UP_BASE + UP_STEP * level
    if level < UP_STEEP_LEVEL:
        return linear
    at_start = UP_BASE + UP_STEP * UP_STEEP_LEVEL
    cost = at_start * (UP_STEEP_RATE ** (level - UP_STEEP_LEVEL))
    return int(round(cost / UP_ROUND) * UP_ROUND)


def run_model(rows):
    """
    각 웨이브의 «그 웨이브를 시작할 때» 평균 Lv 를 구한다.

    ⚠ 순서가 곧 모델이다 — ① 충원 ② 남는 에너지를 <b>최저 레벨</b> 캐릭터에게
    ③ 그 상태를 기록 ④ 그 웨이브를 치르며 번다. 옛 시트를 이 순서로 재현했을 때
    30행이 전부 ±0.3 Lv 안에 들어왔다(그래서 이 모델이 옛 시트와 같은 것이라고 본다).
    """
    energy = 0.0
    levels = []
    created = 0
    out = []

    for r in rows:
        while len(levels) < r["people"]:
            if len(levels) < FREE_CHARACTERS:
                levels.append(0)
                continue
            cost = CREATE_BASE + CREATE_STEP * created
            if energy < cost:
                break
            energy -= cost
            created += 1
            levels.append(0)

        guard = 0
        while levels and guard < 500000:
            guard += 1
            i = min(range(len(levels)), key=lambda j: levels[j])
            c = upgrade_cost(levels[i])
            if energy < c:
                break
            energy -= c
            levels[i] += 1

        out.append(sum(levels) / len(levels) if levels else 0.0)

        energy += r["count"] * 2 * WAVE_KILL_ENERGY
        energy += r["neutral"] * band_multiplier(r["wave"])

    return out


def read_rows():
    wb = openpyxl.load_workbook(FORMULA_XLSX, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=FIRST_ROW, max_row=ws.max_row, values_only=True):
        if r[0] is None:
            continue
        rows.append(dict(
            wave=int(r[0]), lv=float(r[1]), people=int(r[2]),
            focus=float(r[3]), plain=float(r[4]), hp=float(r[5]),
            count=int(r[8]), mob_hp=float(r[9]), mob_hit=float(r[10]),
            hits=float(r[11]), hit_ratio=float(r[12]), dps=float(r[13]),
            load=float(r[14]), neutral=float(r[19]),
        ))
    return rows


def main():
    rows = read_rows()
    print(f"「{SHEET}」 {len(rows)}행 읽음")

    new_lv = run_model(rows)

    # ── 파생 칸 다시 계산 ────────────────────────────────────────────
    plan = []
    for r, lv in zip(rows, new_lv):
        n = round(lv)
        focus = min(STAT_MAX, STAT_BASE_FOCUS + FOCUS_PER_LEVEL * n)
        plain = min(STAT_MAX, STAT_BASE_PLAIN + PLAIN_PER_LEVEL * n)
        hp = round(HP_BASE + HP_PER_STAT * (focus - STAT_BASE_FOCUS))

        per = DPS_A * focus * focus + DPS_B * focus + DPS_C
        dps = round(r["people"] * per)

        total_mob_hp = r["count"] * 2 * r["mob_hp"]
        load = total_mob_hp / (dps * EFFECTIVE * BATTLE_SECONDS) if dps > 0 else 0.0

        # ⚠ «한 대» 는 잡몹 타격이 안 바뀌었으므로 <b>체력이 늘어난 만큼만</b> 줄어든다.
        #   실제로는 방어력도 같이 올라 더 줄지만, 그 감쇠식이 시트에 남아 있지 않아
        #   <b>보수적으로</b>(피해를 더 크게) 잡았다.
        ratio = r["hit_ratio"] * (r["hp"] / hp) if hp > 0 else 0.0
        hits = r["hits"] * (hp / r["hp"]) if r["hp"] > 0 else 0.0

        plan.append(dict(
            wave=r["wave"], lv=round(lv, 1), focus=round(focus, 1), plain=round(plain, 1),
            hp=hp, dps=dps, load=round(load, 2),
            ratio=round(ratio, 3), hits=round(hits, 1),
            neutral=round(r["neutral"] * band_multiplier(r["wave"])),
        ))

    print("\n웨이브 | 평균Lv (전) | 주력 | 캐릭터HP | 파티DPS | 잡몹부하 | 중립수입 (전)")
    for p, r in zip(plan, rows):
        if p["wave"] % 5 == 0 or p["wave"] <= 2:
            print(f"{p['wave']:6d} | {p['lv']:5.1f} ({r['lv']:4.1f}) | {p['focus']:5.1f} | "
                  f"{p['hp']:8d} | {p['dps']:7d} | {p['load']:8.2f} | "
                  f"{p['neutral']:6d} ({int(r['neutral'])})")

    old_total = sum(r["neutral"] for r in rows)
    new_total = sum(p["neutral"] for p in plan)
    print(f"\n중립 수입 총합 {old_total:,.0f} → {new_total:,.0f}  (×{new_total/old_total:.3f})")

    write_excel(plan)


def write_excel(plan):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        # ── ① 중립 보상 ────────────────────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(NEUTRAL_XLSX))
        ws = wb.Worksheets("neutrality_mon")
        # 3행이 자료형 줄, 4행부터 자료. id 열 A · min_energy J · max_energy K
        touched = 0
        row = 4
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            mon_id = int(v)
            if mon_id in NEW_ENERGY:
                lo, hi = NEW_ENERGY[mon_id]
                before = (ws.Cells(row, 10).Value, ws.Cells(row, 11).Value)
                ws.Cells(row, 10).Value = lo
                ws.Cells(row, 11).Value = hi
                print(f"  중립 {mon_id}: {int(before[0])}~{int(before[1])} → {lo}~{hi}")
                touched += 1
            row += 1
        wb.Save()
        wb.Close()
        print(f"임시용 중립 몬스터.xlsx — {touched}종 갱신")

        # ── ② 「웨이브 부하」 + 「계수」 ─────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(FORMULA_XLSX))
        ws = wb.Worksheets(SHEET)

        for i, p in enumerate(plan):
            r = FIRST_ROW + i
            assert int(ws.Cells(r, COL["wave"]).Value) == p["wave"], \
                f"{r}행이 웨이브 {p['wave']} 가 아니다"
            ws.Cells(r, COL["lv"]).Value = p["lv"]
            ws.Cells(r, COL["focus"]).Value = p["focus"]
            ws.Cells(r, COL["plain"]).Value = p["plain"]
            ws.Cells(r, COL["hp"]).Value = p["hp"]
            ws.Cells(r, COL["hits"]).Value = p["hits"]
            ws.Cells(r, COL["hit_ratio"]).Value = p["ratio"]
            ws.Cells(r, COL["dps"]).Value = p["dps"]
            ws.Cells(r, COL["load"]).Value = p["load"]
            ws.Cells(r, COL["neutral"]).Value = p["neutral"]

        # 머리말에 이번 개정을 한 줄 남긴다 — 이 시트가 «재밸런스의 근거» 이므로
        # 언제 무엇이 왜 바뀌었는지가 시트 안에 있어야 한다(136-1절이 세운 규칙).
        note = ("※ 2026-08-26 개정: 착지점을 «w30 = 12명 Lv30» 에서 «12명 Lv35» 로 옮겼다. "
                "유저 확정 — 강화 벽은 Lv30 유지 · 몬스터 수치 그대로 · 수입은 중립/에픽 보상만 상향. "
                "중립 수입 곡선 ×1.00/1.15/1.30/1.55/2.33 (1~10 / 11~15 / 16~20 / 21~25 / 26~30). "
                "⚠ 파티 DPS 는 옛 30행에서 되짚은 «맞춘 곡선»(2차식)이고, "
                "«한 대÷캐릭터 HP» 는 방어력 상승을 빼고 보수적으로 잡았다.")
        ws.Cells(5, 1).Value = note

        # 계수 시트 — 이번에 바뀐 상수 넷을 맨 아래에 덧붙인다.
        # ⚠ 행을 끼워넣지 않는다 — 다른 시트가 절대 행 번호로 참조한다(38-2절).
        coef = wb.Worksheets("계수")
        r = coef.UsedRange.Rows.Count + 1
        for label, value, where, why in [
            ("■ 2026-08-26 — 경제 착지점 Lv35 · 만렙 · 유물 칸", "", "", ""),
            ("만렙(강화 횟수 상한)", 40, "CharacterUpgradeService.maxLevel",
             "★ 유저 지시 «만렙 40 LV». 경제 모델의 착지점은 Lv35 이고, 그 위 다섯 칸은 유물·에픽 보상으로만 닿는 구간으로 남겼다. 0 이면 만렙이 없다"),
            ("유물 장착 칸", 3, "RelicInventory.equipSlots",
             "★ 유저 지시 «유물 장착 인벤토리 3칸». 유물 수치는 그대로 두었으므로 캐릭터당 유물 효과가 최대 3배가 된다(유저 확정)"),
            ("영웅 각성 최소 레벨", 15, "HeroAwakeningService.awakenMinLevel",
             "값은 136절부터 15 였다. 2026-08-26 에 «각성 가능» 표시가 이 조건을 안 보던 것을 고쳤다(판정은 원래부터 막고 있었다)"),
            ("중립 수입 배율(11~15 / 16~20 / 21~25 / 26~30)", "1.15 / 1.30 / 1.55 / 2.33",
             "임시용 중립 몬스터.min_energy·max_energy",
             "★ 구간 배율을 넣는 장치는 없다 — 종별 등장 «거리»(1001 근 / 1002·1004 중 / 1003·에픽 원)가 그 곡선을 만든다. 1001 은 안 건드렸다"),
        ]:
            coef.Cells(r, 1).Value = label
            if value != "":
                coef.Cells(r, 2).Value = value
            if where:
                coef.Cells(r, 3).Value = where
            if why:
                coef.Cells(r, 4).Value = why
            r += 1

        wb.Save()
        wb.Close()
        print("능력치 및 공식 정리.xlsx — 「웨이브 부하」 9열 × 30행 · 「계수」 5행 추가")
    finally:
        excel.Quit()


if __name__ == "__main__":
    main()
