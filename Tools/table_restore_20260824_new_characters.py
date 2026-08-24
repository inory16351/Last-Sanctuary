# -*- coding: utf-8 -*-
"""★★ 신규 3인(9012·9013·9014)이 표에서 <b>사라진 것</b>을 되돌린다 (2026-08-24).

유저 지시: *"캐릭터 시트에 마지막 캐릭터 2인이 삭제되었는데 테이블에 추가하고
캐릭터 테이블 스트링 키 연동 해줘"* / *"세라피엘이랑 시안 스킬 아이콘도 추가해줘"*
/ *"캐릭터 시트 복구할때는 진행상황.md 확인해보고 복구"*.

무엇이 사라졌나 — <b>실측</b>
------------------------------
백업 ``_백업/20260821_194541_명사수_밸류/`` 와 지금 파일을 한 칸씩 비교했다.
<b>덧붙임만 있고 충돌은 하나도 없다</b>(값이 다른 칸 0개) — 즉 «누가 고쳤다» 가 아니라
<b>되돌아갔다</b>:

    캐릭터 테이블.xlsx
      Character   9013·9014 <b>행 삭제</b> · 9012 는 스트링 키 → <b>한글 리터럴로 되돌아감</b>
      first_Stat  9013·9014 <b>행 삭제</b>
      Skill       80037~80042 <b>행 삭제</b> · 80034~80036 리터럴로 되돌아감
      Skill_Type  6줄 <b>삭제</b> · 3줄 리터럴로 되돌아감
    스트링 키 테이블.xlsx
      42행 <b>통째로 삭제</b> (300행 → 258행)

★ 128절이 «인게임 구현» 을 끝냈으므로 <b>유니티 에셋에는 세 인물이 다 살아 있다</b>
  (`Resources/Characters/Character_9013_Seraphiel.asset` 등). 그래서 게임은 지금도 돌고,
  <b>표만</b> 어긋나 있었다 — 그 상태로 `gen_character_assets.py` 를 돌리면 두 인물이
  «표에 없다» 며 <b>에셋에서도 사라진다</b>. 되돌리는 것이 먼저다.

무엇을 쓰나
-----------
백업의 값을 그대로 쓴다(진행상황 125절이 그 값의 출처를 적어 두었다 — 유저가 넣은
`Character`/`first_Stat`/`Skill` 9행과 125절이 채운 설명 15칸).
백업 이후에 결정된 두 가지만 <b>덧입힌다</b>:

  ① <b>「명사수」(80038) 밸류</b> — `table_update_20260821_sharpshooter_value.py` 가
     `value_01 = 20` 으로 고치고 정의문의 «20의» 를 «{value_01}의» 로 바꿨다.
     그 백업은 <b>고치기 전</b>에 뜬 것이라 되돌리면 구멍이 되살아난다(125-4절 1번).
  ② <b>스킬 아이콘 9칸</b> — 125-4절 2번·129-5절이 «남은 일» 로 적어둔 것.
     비어 있으면 성장 창·상세 카드의 패시브 칸이 <b>빈 사각</b>으로 뜬다.

⚠ `캐릭터 테이블.xlsx` 는 <b>Excel COM</b> 으로 쓴다 — 하이퍼링크가 154칸 있고
  (51-11절 «키 칸을 눌러 스트링 테이블로 이동») openpyxl 로 저장하면 전부 날아간다
  (64-2·69-10·125-3절이 겪은 그 사고).
⚠ `DispatchEx` — 유저가 엑셀을 켜 두었을 때 그 창에 붙지 않게 한다(112-7절 함정 2번).
⚠ `스트링 키 테이블.xlsx` 는 openpyxl 로 쓴다 — `gen_string_table.py` 가 매번 통째로
  다시 쓰는 파일이라 지켜야 할 서식이 없다.

⚠ <b>멱등하다</b> — 이미 값이 맞으면 건드리지 않는다. 두 번 돌려도 행이 겹쳐 생기지 않는다.

실행:
    python Tools/table_restore_20260824_new_characters.py
다음:
    python Tools/gen_string_table.py          # StringTable.txt 내보내기
    python Tools/link_string_keys.py          # 되살린 칸에 하이퍼링크 다시 걸기
    python Tools/gen_character_assets.py      # 유니티 에셋 재생성
"""

import datetime
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR as TABLES

CHAR_XLSX = os.path.join(TABLES, "캐릭터 테이블.xlsx")
STRING_XLSX = os.path.join(TABLES, "스트링 키 테이블.xlsx")
BACKUP_ROOT = os.path.join(TABLES, "_백업")

#: 값의 출처. 이 백업이 «사라지기 직전» 의 마지막 온전한 상태다.
SOURCE = os.path.join(BACKUP_ROOT, "20260821_194541_명사수_밸류")

DATA_ROW0 = 4          # 1행 한글 라벨 · 2행 필드명 · 3행 자료형 · 4행부터 데이터

#: 되살릴 대상 — 이 셋만 본다. 나머지 행은 읽지도 않는다.
CHAR_IDS = [9012, 9013, 9014]
SKILL_IDS = list(range(80034, 80043))
SKILL_TYPES = [
    "Strong_mind", "The_Legion’s_Shield", "Blessing_of_Four_Wings",
    "Evasive_maneuver", "Sharpshooter", "Declaration_of_the_End",
    "Soul_Absorption", "The_Reaper’s_Scythe", "Breaking_through_limits",
]

# ── ① 「명사수」 밸류 구멍 (125-4절 1번 · table_update_20260821_sharpshooter_value.py) ──
SHARPSHOOTER_VALUE01 = 20
SHARPSHOOTER_DESC_FIX = ("영구적으로 세라피엘이 20의", "영구적으로 세라피엘이 {value_01}의")

# ── ② 스킬 아이콘 (129-5절이 남긴 일) ─────────────────────────────────────
#
# ★ <b>이미 있는 아이콘에서 고른다</b> — `Resources/SkillIcons` 에 90장이 있고
#   그중 33장만 쓰이고 있다. 새로 자를 것이 없다(119절이 시트를 이미 다 갈라 뒀다).
# ★ 되도록 <b>새 세트</b>(픽셀아트 타일 · `skill_icon_build.py` 의 66장)에서 골랐다 —
#   옛 24장은 테두리 없는 납작한 그림이라 한 화면에 섞이면 «빠진 것처럼» 보인다.
#   ⚠ 다만 «방패 돔»·«회복 십자» 는 새 세트에 대체할 그림이 없어 옛 것을 썼다.
# ⚠ 이미 쓰인 33장과 <b>겹치지 않게</b> 골랐다 — 같은 그림이 두 스킬에 붙으면
#   상세 카드에서 어느 쪽인지 구분이 안 된다.
SKILL_ICONS = {
    # 엘리시아 9012
    80034: "icon_heal_cross",      # 강인한 정신   — 초록 회복 십자 (스스로 되찾는다)
    80035: "icon_barrier_dome",    # 군단의 방패   — 사람을 감싸는 방패 돔
    80036: "icon_meteor_circle",   # 네 날개의 가호 — 원형 범위로 내리꽂히는 빛
    # 세라피엘 9013
    80037: "icon_sprint_dash",     # 회피 기동     — 달려 빠지는 형상
    80038: "icon_snipe_mark",      # 명사수        — 조준선과 화살
    80039: "icon_cannon_battery",  # 종말의 선언   — 집중 포격
    # 시안 9014
    80040: "icon_ghost_wail",      # 영혼 흡수     — 거둬들이는 망령
    80041: "icon_hooded_reaper",   # 사신의 낫     — 두건 쓴 사신
    80042: "icon_red_slash",       # 한계 돌파     — 붉은 참격 (근접 공격력)
}


# ---------------------------------------------------------------------------

def backup(paths, tag):
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_ROOT, stamp + "_" + tag)
    os.makedirs(dst, exist_ok=True)
    for p in paths:
        shutil.copy2(p, os.path.join(dst, os.path.basename(p)))
    print("백업:", dst)


def read_source():
    """백업에서 되살릴 값을 읽어 (시트 → {키: 행 튜플}) 로 돌려준다."""
    wb = openpyxl.load_workbook(os.path.join(SOURCE, "캐릭터 테이블.xlsx"), data_only=True)
    out = {}
    for sheet, keys in (("Character", CHAR_IDS), ("first_Stat", CHAR_IDS),
                        ("Skill", SKILL_IDS), ("Skill_Type", SKILL_TYPES)):
        rows = {}
        for r in wb[sheet].iter_rows(values_only=True):
            if r[0] in keys:
                rows[r[0]] = list(r)
        missing = [k for k in keys if k not in rows]
        if missing:
            raise SystemExit("⚠ 백업에 없는 키: %s / %s" % (sheet, missing))
        out[sheet] = rows

    ws = openpyxl.load_workbook(os.path.join(SOURCE, "스트링 키 테이블.xlsx"),
                                data_only=True)["string"]
    out["string"] = {r[0]: list(r) for r in ws.iter_rows(values_only=True) if r[0]}
    return out


# ── 캐릭터 테이블 (Excel COM) ──────────────────────────────────────────────

def rows_of(ws, last_row):
    """1열 값 → 행 번호."""
    found = {}
    for r in range(DATA_ROW0, last_row + 1):
        v = ws.Cells(r, 1).Value
        if v is None:
            continue
        found[int(v) if isinstance(v, float) and v.is_integer() else v] = r
    return found


def put_sheet(wb, sheet, wanted, width, log):
    """`wanted` 의 행을 시트에 반영한다 — 있으면 고치고 없으면 <b>맨 아래에 붙인다</b>."""
    ws = wb.Worksheets(sheet)
    last = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
    where = rows_of(ws, last)
    tail = max([r for r in where.values()] + [DATA_ROW0 - 1])

    for key, values in wanted.items():
        row = where.get(key)
        if row is None:
            tail += 1
            row = tail
            log.append("  %s %s — 행 신설 (%d행)" % (sheet, key, row))

        for c in range(1, width + 1):
            v = values[c - 1] if c - 1 < len(values) else None
            cur = ws.Cells(row, c).Value
            same = cur == v or (cur is None and v is None) or \
                   (isinstance(cur, float) and isinstance(v, (int, float)) and cur == v)
            if same:
                continue
            ws.Cells(row, c).Value = v
            log.append("  %s %s 열%d: %r → %r" % (sheet, key, c, cur, v))


def update_character_table(excel, src):
    wb = excel.Workbooks.Open(CHAR_XLSX)
    log = []
    try:
        # ── Skill: 아이콘·명사수 밸류를 덧입힌 뒤 쓴다 ──
        skills = {k: list(v) for k, v in src["Skill"].items()}
        for sid, icon in SKILL_ICONS.items():
            skills[sid][10] = icon                       # 11열 = skill_icon
        skills[80038][3] = SHARPSHOOTER_VALUE01          # 4열 = value_01

        put_sheet(wb, "Character", src["Character"], 9, log)
        put_sheet(wb, "first_Stat", src["first_Stat"], 13, log)
        put_sheet(wb, "Skill", skills, 13, log)
        put_sheet(wb, "Skill_Type", src["Skill_Type"], 2, log)

        links = sum(ws.Hyperlinks.Count for ws in wb.Worksheets)
        if log:
            wb.Save()
        print("\n".join(log) if log else "  (바뀐 칸 없음)")
        print("  하이퍼링크 %d칸 (저장 후에도 남아 있어야 한다)" % links)
    finally:
        wb.Close(False)
    return len(log)


# ── 스트링 키 테이블 (openpyxl) ────────────────────────────────────────────

def update_string_table(src):
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET_NAME]
    have = {ws.cell(r, 1).value: r for r in range(DATA_ROW0, ws.max_row + 1)
            if ws.cell(r, 1).value}

    # 되살릴 키 = 백업에만 있고 지금 없는 것 + 이번에 고칠 「명사수」 정의문
    added = 0
    for key, row in src["string"].items():
        if key in have:
            continue
        # 신규 3인과 그 스킬에 관한 것만 되살린다 — 백업의 다른 행까지 끌고 오지 않는다.
        if not any(t in str(key) for t in
                   ("_9012", "_9013", "_9014", "_8003", "_8004") + tuple(SKILL_TYPES)):
            continue
        r = ws.max_row + 1
        for c, v in enumerate(row[:5], start=1):
            ws.cell(r, c).value = v
        added += 1
        print("  + %s" % key)

    # ① 「명사수」 정의문의 하드코딩 숫자 → 자리표시 (멱등)
    fixed = 0
    for r in range(DATA_ROW0, ws.max_row + 1):
        if ws.cell(r, 1).value != "skill_type_desc_Sharpshooter":
            continue
        kr = ws.cell(r, 2).value or ""
        if SHARPSHOOTER_DESC_FIX[0] in kr:
            ws.cell(r, 2).value = kr.replace(*SHARPSHOOTER_DESC_FIX)
            fixed = 1
            print("  ~ skill_type_desc_Sharpshooter — «20의» → «{value_01}의»")

    if added or fixed:
        wb.save(STRING_XLSX)
    return added + fixed


SHEET_NAME = "string"


def main():
    for p in (CHAR_XLSX, STRING_XLSX):
        if not os.path.isfile(p):
            print("⚠ 파일 없음:", p)
            return 1
    if not os.path.isdir(SOURCE):
        print("⚠ 값의 출처 백업이 없습니다:", SOURCE)
        return 1

    src = read_source()
    backup([CHAR_XLSX, STRING_XLSX], "신규3인_복구전")

    print("\n[스트링 키 테이블]")
    n2 = update_string_table(src)
    print("  %d칸" % n2)

    import win32com.client as win32
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        print("\n[캐릭터 테이블]")
        n1 = update_character_table(excel, src)
    finally:
        excel.Quit()

    print("\n캐릭터 테이블 %d칸 · 스트링 테이블 %d칸" % (n1, n2))
    print("다음: gen_string_table.py → link_string_keys.py → gen_character_assets.py")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.exit(main())
