# -*- coding: utf-8 -*-
"""중반 구간이 <b>너무 쉬워진 것</b>을 되돌린다 — 후반 몫을 «에픽» 으로 옮긴다 (2026-08-26, 3차).

유저 지시
---------
  *"중앙 구간이 너무 쉬워진 것을 수정해야돼 두더지 밸류를 낮추고 고르게 분포 시켰기 때문이다.
    해당 작업을 진행하면서 변경된 데이터는 반드시 테이블에 연동"*

★★★ 무엇이 잘못돼 있었나 — <b>«고르게» 와 «후반에 몰리게» 는 같이 못 간다</b>
──────────────────────────────────────────────────────────────────────
UI-67(169-1절)은 착지점을 «w30 = 12명 Lv35» 로 잡고 <b>먼 종만</b> 올렸다.
2차(`table_update_20260826_neutral_spread.py`)는 «두더지 한 종에 다 실렸다» 는 지시를 받아
1003 을 795→340 으로 내리고 그만큼 <b>1001·1002·1004(가까운 종)를 올렸다</b>.

그 결과가 이번 지시의 증상이다 — <b>수입이 앞으로 당겨졌다</b>:

    구간 몫    1~5   6~10  11~15  16~20  21~25  26~30
      UI-67    0.8%  1.6%   4.7%  10.6%  24.3%  58.0%
      2차       1.3%  2.5%   7.9%  18.2%  25.9%  44.1%   ← 중반이 두 배 가까이

    파티 평균 Lv        w15    w20    w25    w30
      기획서 요구선      13     18     23    (30)
      UI-67            11.9   20.4   24.4   35.0
      2차               15.0   25.9   28.5   35.2   ← w20 이 요구선보다 <b>8 Lv</b> 위
      이번(3차)         12.1   20.6   25.2   35.0

2차의 주석은 이 대가를 <b>미리 적어 뒀다</b>(«⚠⚠ 대가 — 중반이 세진다 … 너무 싱거우면
그때 그 구간 몬스터를 잡을 것»). 유저가 «너무 쉬워졌다» 고 한 것이 그 자리다.

★ <b>왜 몬스터를 안 잡고 경제를 되돌리나</b>
──────────────────────────────────────────────────────────────────────
「밸런스 기획서」는 <b>보스마다 요구 파티 Lv 를 못박았다</b>(w20 라린길 = 20Lv 1부대 +
15Lv 1부대면 «적당»). 파티가 Lv25.9 로 오면 그 보스를 Lv25.9 에 맞춰 올려야 하는데,
그러면 <b>기획서가 정한 요구선 자체를 옮기는</b> 일이 된다. 경제를 요구선으로 되돌리면
잡몹 부하도 저절로 제자리로 온다(아래 표) — 몬스터 수치를 한 칸도 안 고치고 끝난다.
UI-67 의 유저 확정 ②(«몬스터는 그대로»)도 그대로 지킨다.

★ <b>어디를 고쳤나 — 두더지는 손대지 않았다</b>
──────────────────────────────────────────────────────────────────────
구간별 사냥 구성(2차의 MIX)을 보면 11~20 웨이브의 수입은 <b>1002·1004 하나</b>로 만들어진다.
그러니 중반을 내리려면 그 둘을 내리는 수밖에 없고, 그만큼을 <b>21~30 에만 있는 종</b>으로
옮겨야 한다. 그 자리에 <b>두더지 대신 에픽</b>을 썼다 —

    1001 종양 거미     11~17  →   8~12    (평균 14 →   10)   2차 인상분 되돌림
    1002 종양귀       62~98  →  37~55    (평균 80 →   46)   ← 2차 전 값
    1004 고르도네     62~98  →  37~55    (평균 80 →   46)   ← 2차 전 값
    1003 종양 두더지 260~420 →  <b>그대로</b>  (평균 340)        ★ 2차 지시를 그대로 둔다
    1101~1103 에픽  1800~2700 → 4400~6600 (평균 2250 → 5500)
    1104 폴리르     2700~3600 → 6600~8800 (평균 3150 → 7700)

  ⚠ 에픽은 <b>지속 수급을 흔들지 않는다</b> — 동시 1마리 · 리스폰 1200~1600초라
    초당 수급이 4.6 E/s 다(1001 48 · 1002+1004 234 · 1003 258). 즉 «밭» 이 아니라 <b>사건</b>이고,
    「밸런스 기획서」의 «후반부 … 에픽 몬스터의 보상을 적극적으로 활용할 것을 요구하는 구간»
    이 정확히 이 모양이다. 레벨 제한(10/15/20/25)이 초반 남용도 이미 막는다.
  ★ 두더지의 «마리당 값이 중반 종의 17배» 였던 문제(2차)는 <b>되살아나지 않는다</b> —
    340 ÷ 46 = 7.4배, 지속 수급은 258 : 234 로 거의 같다.

★ 결과 — <b>총 수입과 착지점은 그대로, 실리는 구간만 되돌린다</b>
──────────────────────────────────────────────────────────────────────
    총 중립 수입 108,555 → 108,920 (+0.3%) — «w30 12명 Lv35» 착지점 유지
    구간 몫 1.3/2.5/7.9/18.2/25.9/44.1% → <b>1.0/1.9/5.1/11.5/26.0/55.3%</b>

    잡몹 부하   w11   w16   w20   w21   w25   w26   w30
      2차       0.40  0.47  0.27  0.50  0.31  0.60  0.41
      3차       0.36  0.58  0.34  0.71  0.36  0.71  0.42   ← 136절 곡선으로 복귀

⚠ w20 20.6 · w25 25.2 는 기획서 요구선(18 · 23)보다 여전히 2 Lv 위다 — 이것은 UI-67 이
  받아들인 상태(20.4 · 24.4)와 <b>같은 수준</b>이다. 더 내리려면 1002 를 2차 전 값 아래로
  깎아야 하는데, 그러면 «중반은 적극적인 중립 사냥» 이라는 기획서 문장과 어긋난다.

⚠ 이 표들은 <b>사람이 엑셀에서 고쳐 나가는 문서</b>다 — 통째로 다시 굽지 않고 바꿀 칸만 고친다.
⚠ <b>Excel COM 으로 쓴다</b> — openpyxl 로 저장하면 하이퍼링크·주석·수식 캐시가 날아간다(136-4절).

사용법:  python Tools/table_update_20260826_mid_band.py
다음:    python Tools/sync_tables_to_assets.py     (중립 에셋에 반영)
         Unity 에서 Assets/Refresh
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

NEUTRAL_XLSX = os.path.join(TABLE_DIR, "임시용 중립 몬스터.xlsx")
FORMULA_XLSX = os.path.join(TABLE_DIR, "능력치 및 공식 정리.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")

# ══════════════════════════════════════════════════════════════════════
#  ① 중립 보상 — mon_id → (min_energy, max_energy)
#     ★ 1003(종양 두더지)은 <b>일부러 빠져 있다</b> — 2차 지시를 그대로 둔다.
# ══════════════════════════════════════════════════════════════════════
NEW_ENERGY = {
    1001: (8, 12),          # 평균 14 →   10
    1002: (37, 55),         # 평균 80 →   46   (2차 전 값)
    1004: (37, 55),         # 평균 80 →   46   (2차 전 값)
    1101: (4400, 6600),     # 평균 2250 → 5500
    1102: (4400, 6600),
    1103: (4400, 6600),
    1104: (6600, 8800),     # 평균 3150 → 7700
}

MOLE_AVG = 340.0            # 1003 — 안 고치지만 모델에는 들어간다

# ══════════════════════════════════════════════════════════════════════
#  ② 구간별 사냥 구성 — 2차와 <b>같은 표</b>를 쓴다 (두 벌이 되면 안 된다)
#     (구간 상한, (1001, 1002+1004, 1003, 에픽))
# ══════════════════════════════════════════════════════════════════════
MIX = [
    (5,  (20,  0,  0, 0.0)),
    (10, (39,  0,  0, 0.0)),
    (15, (20, 18,  0, 0.0)),
    (20, (20, 46,  0, 0.0)),
    (25, (20, 40,  3, 0.5)),
    (30, (20, 40, 10, 1.2)),
]


def avg(mon_id):
    lo, hi = NEW_ENERGY[mon_id]
    return (lo + hi) / 2.0


EPIC_AVG = None             # main() 에서 1101 의 평균으로 채운다


def mix_for(wave):
    for upto, m in MIX:
        if wave <= upto:
            return m
    return MIX[-1][1]


def neutral_income(wave):
    n1, n2, n3, ne = mix_for(wave)
    return n1 * avg(1001) + n2 * avg(1002) + n3 * MOLE_AVG + ne * EPIC_AVG


# ══════════════════════════════════════════════════════════════════════
#  ③ 성장 모델 — UI-67·2차의 계수를 <b>그대로</b> 쓴다
# ══════════════════════════════════════════════════════════════════════
FOCUS_PER_LEVEL = 2.62
PLAIN_PER_LEVEL = 1.62
STAT_BASE_FOCUS = 10
STAT_BASE_PLAIN = 5
STAT_MAX = 100

HP_PER_STAT = 6.1832
HP_BASE = 90

DPS_A, DPS_B, DPS_C = 0.01961, 3.329, -9.50
EFFECTIVE = 0.647
BATTLE_SECONDS = 120

CREATE_BASE, CREATE_STEP = 200, 150
FREE_CHARACTERS = 4

UP_BASE, UP_STEP = 40, 10
UP_STEEP_LEVEL, UP_STEEP_RATE, UP_ROUND = 30, 1.35, 10

WAVE_KILL_ENERGY = 10

SHEET = "웨이브 부하"
FIRST_ROW = 7
COL = dict(wave=1, lv=2, people=3, focus=4, plain=5, hp=6,
           hp_scale=7, atk_scale=8, count=9, mob_hp=10, mob_hit=11,
           hits=12, hit_ratio=13, dps=14, load=15,
           boss=16, boss_raw=17, boss_hp=18, boss_goal=19, neutral=20)


def upgrade_cost(level):
    linear = UP_BASE + UP_STEP * level
    if level < UP_STEEP_LEVEL:
        return linear
    at_start = UP_BASE + UP_STEP * UP_STEEP_LEVEL
    cost = at_start * (UP_STEEP_RATE ** (level - UP_STEEP_LEVEL))
    return int(round(cost / UP_ROUND) * UP_ROUND)


def run_model(rows, neutral):
    """UI-67 의 run_model 과 <b>같은 순서</b>다 — ① 충원 ② 최저 레벨부터 강화 ③ 기록 ④ 번다."""
    energy = 0.0
    levels = []
    created = 0
    out = []

    for r, nt in zip(rows, neutral):
        while len(levels) < r["people"]:
            if len(levels) < FREE_CHARACTERS:
                levels.append(0)
                continue
            cost = CREATE_BASE + CREATE_STEP * created
            if energy < cost:
                break
            energy -= cost
            created += 1
            levels.append(0)

        guard = 0
        while levels and guard < 500000:
            guard += 1
            i = min(range(len(levels)), key=lambda j: levels[j])
            c = upgrade_cost(levels[i])
            if energy < c:
                break
            energy -= c
            levels[i] += 1

        out.append(sum(levels) / len(levels) if levels else 0.0)

        energy += r["count"] * 2 * WAVE_KILL_ENERGY
        energy += nt

    return out


def read_rows():
    wb = openpyxl.load_workbook(FORMULA_XLSX, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=FIRST_ROW, max_row=ws.max_row, values_only=True):
        if r[0] is None:
            continue
        rows.append(dict(
            wave=int(r[0]), lv=float(r[1]), people=int(r[2]),
            focus=float(r[3]), plain=float(r[4]), hp=float(r[5]),
            count=int(r[8]), mob_hp=float(r[9]), mob_hit=float(r[10]),
            hits=float(r[11]), hit_ratio=float(r[12]), dps=float(r[13]),
            load=float(r[14]), neutral=float(r[19]),
        ))
    return rows


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_중반재조정")
    os.makedirs(folder, exist_ok=True)
    for f in (NEUTRAL_XLSX, FORMULA_XLSX):
        shutil.copy2(f, os.path.join(folder, os.path.basename(f)))
    print("백업: " + folder)
    return folder


def main():
    global EPIC_AVG
    EPIC_AVG = avg(1101)

    rows = read_rows()
    print(f"「{SHEET}」 {len(rows)}행 읽음 · 에픽 평균 {EPIC_AVG:,.0f}")

    neutral = [neutral_income(r["wave"]) for r in rows]
    old_total = sum(r["neutral"] for r in rows)
    new_total = sum(neutral)
    print(f"\n중립 총 수입 {old_total:,.0f} → {new_total:,.0f}  "
          f"({(new_total / old_total - 1) * 100:+.2f}%)")

    print("\n구간 몫")
    seen = set()
    for r, nt in zip(rows, neutral):
        band = next(u for u, _ in MIX if r["wave"] <= u)
        if band in seen:
            continue
        seen.add(band)
        print(f"  ~{band:2d}웨이브  {r['neutral']:7.0f} → {nt:7.0f} /웨이브   "
              f"몫 {r['neutral'] * 5 / old_total * 100:5.1f}% → {nt * 5 / new_total * 100:5.1f}%")

    new_lv = run_model(rows, neutral)

    plan = []
    for r, lv, nt in zip(rows, new_lv, neutral):
        n = round(lv)
        focus = min(STAT_MAX, STAT_BASE_FOCUS + FOCUS_PER_LEVEL * n)
        plain = min(STAT_MAX, STAT_BASE_PLAIN + PLAIN_PER_LEVEL * n)
        hp = round(HP_BASE + HP_PER_STAT * (focus - STAT_BASE_FOCUS))

        per = DPS_A * focus * focus + DPS_B * focus + DPS_C
        dps = round(r["people"] * per)

        total_mob_hp = r["count"] * 2 * r["mob_hp"]
        load = total_mob_hp / (dps * EFFECTIVE * BATTLE_SECONDS) if dps > 0 else 0.0

        ratio = r["hit_ratio"] * (r["hp"] / hp) if hp > 0 else 0.0
        hits = r["hits"] * (hp / r["hp"]) if r["hp"] > 0 else 0.0

        plan.append(dict(
            wave=r["wave"], lv=round(lv, 1), focus=round(focus, 1), plain=round(plain, 1),
            hp=hp, dps=dps, load=round(load, 2),
            ratio=round(ratio, 3), hits=round(hits, 1), neutral=round(nt),
        ))

    print("\n웨이브 | 평균Lv (전) | 캐릭터HP | 파티DPS | 잡몹부하 (전) | 중립수입 (전)")
    for p, r in zip(plan, rows):
        if p["wave"] % 5 == 0 or p["wave"] in (11, 16, 21, 26):
            print(f"{p['wave']:6d} | {p['lv']:5.1f} ({r['lv']:4.1f}) | {p['hp']:8d} | "
                  f"{p['dps']:7d} | {p['load']:5.2f} ({r['load']:4.2f}) | "
                  f"{p['neutral']:6d} ({int(r['neutral'])})")

    backup()
    write_excel(plan)


def write_excel(plan):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        # ── ① 중립 보상 ────────────────────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(NEUTRAL_XLSX))
        ws = wb.Worksheets("neutrality_mon")
        touched = 0
        row = 4                      # 3행이 자료형 줄, 4행부터 자료. id A · min_energy J · max_energy K
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            mon_id = int(v)
            if mon_id in NEW_ENERGY:
                lo, hi = NEW_ENERGY[mon_id]
                before = (ws.Cells(row, 10).Value, ws.Cells(row, 11).Value)
                ws.Cells(row, 10).Value = lo
                ws.Cells(row, 11).Value = hi
                print(f"  중립 {mon_id}: {int(before[0])}~{int(before[1])} → {lo}~{hi}")
                touched += 1
            row += 1
        wb.Save()
        wb.Close()
        print(f"임시용 중립 몬스터.xlsx — {touched}종 갱신 (1003 은 손대지 않았다)")

        # ── ② 「웨이브 부하」 + 「계수」 ─────────────────────────────
        wb = excel.Workbooks.Open(os.path.abspath(FORMULA_XLSX))
        ws = wb.Worksheets(SHEET)

        for i, p in enumerate(plan):
            r = FIRST_ROW + i
            assert int(ws.Cells(r, COL["wave"]).Value) == p["wave"], \
                f"{r}행이 웨이브 {p['wave']} 가 아니다"
            ws.Cells(r, COL["lv"]).Value = p["lv"]
            ws.Cells(r, COL["focus"]).Value = p["focus"]
            ws.Cells(r, COL["plain"]).Value = p["plain"]
            ws.Cells(r, COL["hp"]).Value = p["hp"]
            ws.Cells(r, COL["hits"]).Value = p["hits"]
            ws.Cells(r, COL["hit_ratio"]).Value = p["ratio"]
            ws.Cells(r, COL["dps"]).Value = p["dps"]
            ws.Cells(r, COL["load"]).Value = p["load"]
            ws.Cells(r, COL["neutral"]).Value = p["neutral"]

        note = ("※ 2026-08-26 개정 3차(유저 지시 — 중앙 구간이 너무 쉬워졌다): 2차가 «전 구간에 나누면서» "
                "수입을 앞으로 당겨 w20 파티가 요구선(18)보다 8Lv 위(25.9)까지 올라갔다. "
                "1002·1004 를 2차 전 값(평균 46)으로 되돌리고 1001 도 10 으로 내린 뒤, 그만큼을 에픽 보상으로 옮겼다"
                "(1101~1103 평균 2250→5500 · 1104 3150→7700). ★ 1003 종양 두더지는 2차 값(340) 그대로 — 두더지 쏠림은 되살리지 않았다. "
                "에픽은 동시 1마리·리스폰 1200초라 지속 수급이 4.6E/s 뿐이어서 «밭» 이 아니라 사건이다(기획서: 후반은 에픽 보상 활용 구간). "
                "구간 몫 1.3/2.5/7.9/18.2/25.9/44.1% → 1.0/1.9/5.1/11.5/26.0/55.3%. 총 중립 수입은 그대로(108,9xx)라 착지점 «w30 12명 Lv35» 도 그대로다. "
                "⚠ 몬스터 수치는 한 칸도 고치지 않았다 — 파티 Lv 가 요구선으로 돌아오면 잡몹 부하도 136절 곡선으로 저절로 돌아온다.")
        ws.Cells(5, 1).Value = note

        coef = wb.Worksheets("계수")
        r = coef.UsedRange.Rows.Count + 1
        for label, value, where, why in [
            ("■ 2026-08-26 (3차) — 중반 구간 되돌리기 · 후반 몫을 에픽으로", "", "", ""),
            ("종양 거미 1001 평균 보상", 10, "임시용 중립 몬스터.min/max_energy 8~12",
             "2차가 14 로 올려 둔 것을 되돌렸다 — 이 값은 모든 구간에 똑같이 얹혀 중반까지 밀어올린다"),
            ("종양귀 1002 · 고르도네 1004 평균 보상", 46, "임시용 중립 몬스터 37~55",
             "11~20 웨이브 수입은 이 둘 하나로 만들어진다 — 2차 전 값으로 되돌려 중반을 요구선에 맞췄다"),
            ("종양 두더지 1003 평균 보상", 340, "임시용 중립 몬스터 260~420 (고치지 않음)",
             "★ 2차 지시를 그대로 둔다. 340÷46 = 7.4배 · 지속 수급 258 : 234 로 중반 종과 거의 같다"),
            ("에픽 1101~1103 평균 보상", 5500, "임시용 중립 몬스터 4400~6600",
             "후반으로 옮긴 몫을 여기에 실었다. 동시 1마리·리스폰 1200초 → 지속 수급 4.6E/s (밭이 아니라 사건)"),
            ("에픽 1104 폴리르 평균 보상", 7700, "임시용 중립 몬스터 6600~8800",
             "1101~1103 과 같은 배수(x2.44). 요구 Lv 25~30 이라 초반 남용이 막힌다"),
            ("구간 몫 (1~5 / 6~10 / 11~15 / 16~20 / 21~25 / 26~30)",
             "1.0 / 1.9 / 5.1 / 11.5 / 26.0 / 55.3 %", "위 값들에서 파생",
             "2차 1.3/2.5/7.9/18.2/25.9/44.1% — 중반 몫이 절반으로 줄고 그만큼 후반으로 돌아갔다"),
            ("파티 평균 Lv (w15 / w20 / w25 / w30)", "12.1 / 20.6 / 25.2 / 35.0",
             "성장 모델(이 스크립트)", "2차 15.0/25.9/28.5/35.2 · 기획서 요구선 13/18/23/(30) · 착지점 Lv35 유지"),
        ]:
            coef.Cells(r, 1).Value = label
            if value != "":
                coef.Cells(r, 2).Value = value
            if where:
                coef.Cells(r, 3).Value = where
            if why:
                coef.Cells(r, 4).Value = why
            r += 1

        wb.Save()
        wb.Close()
        print("능력치 및 공식 정리.xlsx — 「웨이브 부하」 9열 × 30행 · 「계수」 8행 추가")
    finally:
        excel.Quit()


if __name__ == "__main__":
    main()
