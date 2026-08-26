# -*- coding: utf-8 -*-
"""씬에 «구운» 정적 라벨을 스트링 키 테이블로 옮긴다 (2026-08-26 · 12차).

유저 지시: *"씬에 구운 것들도 테이블로 스트링 키 써서 영어로 바뀌게 해줘"*.

■ 무엇을 하나
  ① `Assets/_Project/Scripts/UI/UiLocalizer.cs` 의 <b>지도</b>(경로 → 스트링 키)를 읽는다.
     — 지도가 정본이다. 이 스크립트는 <b>키 목록을 손으로 다시 적지 않는다</b>.
  ② 씬(`SCENES` — `Proto_01.unity` · `Lobby.unity`)에서 그 경로의 <b>지금 한국어 문구</b>를
     읽어 kr 로 쓴다. — 사람이 옮겨 적으면 반드시 어긋난다. 화면에 있는 그 글자가 그대로
     표에 들어간다. ★ 2026-08-26 에 <b>로비 씬</b>이 붙었다(로비 버튼이 영어로 안 바뀌던 건) —
     지도 하나를 두 씬이 나눠 쓴다. 그래서 «이 씬에서 경로를 못 찾은 칸» 은 <b>다른 씬의
     칸일 수 있다</b>: 아래 경고는 두 씬을 <b>합친 뒤</b>에 남은 것만 나온다.
  ③ 표에 없는 키만 <b>덧붙인다</b>(en 은 아래 `EN` 표에서). 이미 있는 키는 <b>손대지 않는다</b>
     — 유저가 다듬은 번역을 덮지 않는 gen_string_table.py 의 규칙과 같다.
  ④ 지도의 <b>모든</b> 키가 표에 있는지 검산하고, 없으면 <b>실패</b>로 끝낸다.

■ 왜 씬에서 kr 을 읽나 — 씬의 한글은 `\\uXXXX` 로 직렬화된다
  그냥 grep 하면 <b>한 칸도 안 잡힌다</b>(2026-08-26 에 실제로 «0개» 를 보고 한참 헤맸다).
  게다가 긴 문장은 YAML 이 <b>여러 줄로 접는다</b> — 한 줄만 읽으면 문장이 잘린다.

■ ⚠ kr 을 «고쳐서» 넣는 칸이 둘 있다 (`KR_FIX`)
  156절이 「넥서스」를 <b>「성역」</b>으로 바꿨는데 <b>이 두 라벨은 씬에 남아 있었다</b>.
  표로 옮기는 김에 바로잡는다 — 씬 문구를 그대로 옮기면 옛 낱말이 정본으로 굳는다.

■ 다음
    py -3 Tools/gen_string_table.py      → StringTable.txt 로 내보낸다
    py -3 Tools/link_string_keys.py      → 하이퍼링크 재생성
  그리고 씬의 `UI_Root` 에 `UiLocalizer` 컴포넌트를 붙인다(MCP).
"""
import io
import os
import re
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

_PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
MAP_CS = os.path.join(_PROJECT, 'Assets', '_Project', 'Scripts', 'UI', 'UiLocalizer.cs')
#: 지도가 걸치는 씬들 — 2026-08-26 에 <b>로비 씬</b>이 붙었다(로비 버튼이 영어로 안 바뀌던 건).
#  ⚠ 경로는 두 씬 모두 «UI_Root 아래» 로 적히므로(chain[1:]) 그대로 합쳐도 겹치지 않는다.
SCENES = [
    os.path.join(_PROJECT, 'Assets', 'Scenes', 'Proto_01.unity'),
    os.path.join(_PROJECT, 'Assets', 'Scenes', 'Lobby.unity'),
]

SHEET = 'string'
DATA_ROW0 = 4          # 3행 헤더 + 4행부터 데이터 (gen_string_table.py 와 같은 규약)
SOURCE_NOTE = '씬 정적 라벨(UiLocalizer)'

BS = chr(92)

#: 새 키의 영어. kr 은 씬에서 읽으므로 여기 적지 않는다.
EN = {
    'ui_log_title': 'Log',
    'ui_minimap_title': 'Minimap',
    'ui_head_relic': 'Relic',
    'ui_head_skill': 'Skill',
    'ui_head_squads': 'Squads',
    'ui_head_hp_now': 'Current HP',

    'ui_settings_save': 'Save',
    'ui_settings_to_lobby': 'Save and Return to Lobby',
    'ui_settings_quit': 'Quit Without Saving',
    'ui_settings_restart': 'Restart Game',
    'ui_settings_hotkeys': 'Key Bindings',
    'ui_settings_help_reset': 'Show Help Again',
    'ui_settings_volume': 'Volume',

    'ui_squad_subtitle': 'Pick a squad, then click a character in the roster to assign them.',
    'ui_squad_add': 'Add Squad',
    'ui_subj_targets_head': 'Epic Monsters Found',

    'ui_help_title': 'Help',
    'ui_help_see_also': 'See Also',
    'ui_helpcard_more': 'Learn More',
    'ui_helpcard_ok': 'Got It',
    'ui_tour_next': 'Next',
    'ui_tour_prev': 'Back',
    'ui_tour_quit': 'Stop Tour',

    'ui_skill_effect_head': 'Effects',
    'ui_btn_confirm': 'Confirm',
    'ui_btn_reset': 'Reset',

    'ui_growth_title': 'Character Growth',
    'ui_growth_subtitle': ('Upgrades the selected character’s stats. '
                           'Picking another character in the roster switches instantly.'),
    'ui_growth_cost_head': 'Upgrade Cost',
    'ui_growth_stats_head': 'Stats (1–100 before upgrades)',
    'ui_growth_focus_head': 'Growth Type  —  stats in the chosen line rise more easily',
    'ui_growth_passive_head': 'Passive Skills',
    'ui_growth_relic_head': 'Equipped Relics',
    'ui_growth_relic_open': 'Open Relics',

    'ui_tactics_title': 'Tactical Directives',
    'ui_tactics_subtitle': ('Characters act on their own according to these directives. '
                            'Construction is not a directive — the nearest character takes it.'),
    'ui_tactics_col1_head': '1  Engagement',
    'ui_tactics_pos_head': 'Position',
    'ui_tactics_pos_hint': 'Measured from the rally point · the side away from the Sanctuary is the front',
    'ui_tactics_react_head': 'Attack Reaction',
    'ui_reaction_chase': 'Chase enemies within sight',
    'ui_tactics_type_head': 'Attack Type',
    'ui_tactics_col2_head': '2  Combat Behavior',
    'ui_tactics_target_head': 'Target Priority',
    'ui_tactics_retreat_head': 'Retreat Threshold',
    'ui_tactics_retreat_hint': ('Falls back to the Sanctuary to recover when HP drops below the '
                                'threshold. At 0% they never retreat.'),
    'ui_tactics_retreat_action_head': 'On Retreat',
    'ui_tactics_retreat_action_hint': (
        'What this character does when the ally in front falls back. '
        '‘Retreat with ally’ keeps maximum range and withdraws together. '
        'If their own HP is below the threshold they retreat either way. '
        'When the position is ‘Front’ this is fixed to ‘Keep attacking’.'),
    'ui_tactics_col3_head': '3  Exploration · Waves',
    'ui_tactics_scout_head': 'Exploration Type',
    'ui_scout_patrol': 'Patrol',
    'ui_tactics_roam_head': 'Roaming Range',
    'ui_tactics_wave_head': 'Wave Response',
    'ui_tactics_note': ('Characters act on their own according to these directives. '
                        'You can switch characters in the roster while this window stays open.'),

    # ── 로비 씬 (2026-08-26) ─────────────────────────────────────────────
    'ui_lobby_continue': 'Continue',
    'ui_lobby_new_game': 'New Game',
    'ui_lobby_quit': 'Quit Game',
}

#: 씬 문구를 그대로 쓰지 않는 칸 — 156절의 「넥서스 → 성역」이 안 닿은 자리다.
KR_FIX = {
    'ui_tactics_pos_hint': '집결지 기준 · 성역에서 먼 쪽이 전방',
    'ui_tactics_retreat_hint': '체력이 기준치 이하로 떨어지면 성역으로 물러나 회복합니다. 0% 면 후퇴하지 않습니다.',
}


# ── UiLocalizer.cs 의 지도 읽기 ────────────────────────────────────────────
def read_map():
    src = io.open(MAP_CS, encoding='utf-8').read()
    pairs = re.findall(r'E\(\s*"([^"]+)"\s*,\s*"([^"]+)"\s*\)', src)
    if not pairs:
        sys.exit('UiLocalizer.cs 에서 지도를 못 읽었습니다.')
    return pairs


# ── 씬의 TMP 문구 읽기 (\\uXXXX · \\xNN 디코딩 + YAML 접힘 펴기) ────────────
_u = re.compile(BS + BS + 'u([0-9A-Fa-f]{4})')
_x = re.compile(BS + BS + 'x([0-9A-Fa-f]{2})')


def read_all_scenes():
    """모든 대상 씬의 «경로 → 문구» 를 합친다 (앞 씬이 이긴다)."""
    out = {}
    for path in SCENES:
        for k, v in read_scene_text(path).items():
            out.setdefault(k, v)
    return out


def read_scene_text(scene_path):
    t = io.open(scene_path, encoding='utf-8', errors='replace').read()
    blocks = {}
    for m in re.finditer(r'--- !u!(\d+) &(\d+)\n(.*?)(?=--- !u!|\Z)', t, re.S):
        blocks[m.group(2)] = (m.group(1), m.group(3))

    tr_of_go, go_of_tr = {}, {}
    for fid, (cls, b) in blocks.items():
        if cls in ('224', '4'):
            g = re.search(r'm_GameObject: \{fileID: (\d+)\}', b)
            if g:
                tr_of_go[g.group(1)] = fid
                go_of_tr[fid] = g.group(1)

    def nm(gid):
        b = blocks.get(gid)
        m = re.search(r'm_Name: (.*)', b[1]) if b else None
        return m.group(1).strip() if m else '?'

    def chain(gid, out=None, d=0):
        out = out if out is not None else []
        out.insert(0, nm(gid))
        tr = tr_of_go.get(gid)
        if tr and d < 10:
            f = re.search(r'm_Father: \{fileID: (\d+)\}', blocks[tr][1])
            if f and f.group(1) != '0' and f.group(1) in go_of_tr:
                return chain(go_of_tr[f.group(1)], out, d + 1)
        return out

    def dec(s):
        s = _u.sub(lambda m: chr(int(m.group(1), 16)), s)
        return _x.sub(lambda m: chr(int(m.group(1), 16)), s)

    out = {}
    for fid, (cls, b) in blocks.items():
        if cls != '114':
            continue
        m = re.search(r'^[ \t]*m_text: (.*)$', b, re.M)
        if not m:
            continue
        # ⚠ YAML 은 긴 문장을 여러 줄로 접는다 — 다음 필드가 나올 때까지 이어 붙인다.
        parts = [m.group(1).rstrip()]
        for ln in b[m.end():].split('\n')[1:]:
            if re.match(r'^\s*m_[A-Za-z]\w*:', ln) or ln.strip() == '':
                break
            parts.append(ln.strip())
        raw = ' '.join(parts).strip()
        if raw.startswith('"'):
            raw = raw[1:]
            if raw.endswith('"'):
                raw = raw[:-1]
        g = re.search(r'm_GameObject: \{fileID: (\d+)\}', b)
        if not g:
            continue
        path = '/'.join(chain(g.group(1))[1:])
        out[path] = dec(raw)
    return out


def main():
    pairs = read_map()
    scene = read_all_scenes()

    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]

    existing = {}
    last_row = DATA_ROW0 - 1
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None or str(k).strip() == '':
            continue
        existing[str(k).strip()] = r
        last_row = r

    # 지도의 키마다 kr 을 정한다 (같은 키가 여러 경로에 붙어 있으면 첫 경로의 문구를 쓴다)
    want = {}
    no_scene = []
    for path, key in pairs:
        if key in want:
            continue
        if key in KR_FIX:
            want[key] = KR_FIX[key]
            continue
        text = scene.get(path)
        if text is None:
            no_scene.append((path, key))
            continue
        want[key] = text

    added, skipped = [], []
    for path, key in pairs:
        if key in existing or key not in want:
            continue
        if key in [a[0] for a in added]:
            continue
        en = EN.get(key)
        if en is None:
            skipped.append(key)
            continue
        last_row += 1
        ws.cell(row=last_row, column=1).value = key
        ws.cell(row=last_row, column=2).value = want[key]
        ws.cell(row=last_row, column=3).value = en
        ws.cell(row=last_row, column=4).value = SOURCE_NOTE
        existing[key] = last_row
        added.append((key, want[key], en))

    # ── 검산: 지도의 모든 키가 표에 있는가 ────────────────────────────────
    missing = sorted({k for _, k in pairs if k not in existing})

    print('[씬 → 스트링] 지도 %d칸 · 서로 다른 키 %d개' % (len(pairs), len({k for _, k in pairs})))
    if no_scene:
        print('  ⚠ 씬에서 경로를 못 찾은 칸 %d개 (경로가 바뀌었는지 볼 것):' % len(no_scene))
        for p, k in no_scene:
            print('     ?', p, '->', k)
    if skipped:
        print('  ⚠ 영어가 EN 표에 없어 건너뛴 키:', ', '.join(skipped))
    print('  덧붙인 키 %d개:' % len(added))
    for k, kr, en in added:
        print('    +', k, '|', kr[:44].replace('\n', ' '), '|', en[:40])
    if missing:
        wb.close()
        sys.exit('\n✗ 표에 아직 없는 키 %d개 — 저장하지 않았습니다:\n  %s'
                 % (len(missing), '\n  '.join(missing)))

    if not added:
        print('  표가 이미 지도와 같습니다 — 저장하지 않았습니다.')
        return

    bak = STRING_XLSX + '.bak'
    shutil.copyfile(STRING_XLSX, bak)
    wb.save(STRING_XLSX)
    print('  저장:', os.path.basename(STRING_XLSX), '(백업 .bak)')
    print('  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py')


if __name__ == '__main__':
    main()
