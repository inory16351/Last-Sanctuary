# -*- coding: utf-8 -*-
"""데이터 테이블의 <b>표시 문구 컬럼을 스트링 키로 바꾼다.</b>

유저 지시(2026-08-12): "웨이브 몬스터 테이블이 최신 버전이니 파일들 다 거기에 맞춰서 수정해."
그 파일의 `wave_mid_boss` 시트가 이미 `monster_name` 칸에 리터럴 대신
`monster_name_110001` 을 넣어둔 것이 목표 형태다 — 나머지 테이블도 그 모양으로 맞춘다.

■ ★ Excel COM 으로 고치는 이유 (openpyxl 을 쓰지 않는다)
  건물 데이터 시트는 <b>셀 주석(빨간 삼각형)에 세부 규칙이 적혀 있다</b>(그 파일 Docs 시트가
  그렇게 안내한다). openpyxl 로 열어 저장하면 <b>주석과 일부 서식이 사라진다.</b>
  Excel 로 열어 셀 값만 바꾸고 저장하면 그대로 남는다.

■ ★ 안전 규칙 — 값이 예상과 다르면 건드리지 않는다
  각 칸마다 "이 칸의 키" 를 계산해 스트링 테이블에서 kr 을 읽고, <b>칸의 리터럴과 같을 때만</b>
  키로 바꾼다. 다르면 경고만 남기고 그대로 둔다 — 표를 수정한 뒤 스트링 테이블을 다시
  내보내지 않은 상태에서 돌리면 문구가 사라질 수 있기 때문이다.
  이미 키가 들어 있는 칸은 조용히 넘어간다(멱등).

■ 바꾸지 않는 것
  · `*_EG`(영어 이름) 컬럼 — 에셋·폴더 이름을 맞추는 <b>식별자</b>다.
    실제로 `gen_character_assets.py` 가 이 값으로 에셋 파일명(Character_9001_Elin)을 만든다.
    화면에 뜨는 영어는 스트링 테이블의 `en` 컬럼이 정본이다.
  · `skill_type` · `mental_error_type` 같은 enum 값 — 분기용 식별자다.
  · `능력치 및 공식 정리.xlsx`, 건물 시트의 Information·Docs·DEF — 기획 산문.

■ 실행
  python Tools/gen_string_table.py            # 먼저 스트링 테이블을 최신으로
  python Tools/convert_tables_to_string_keys.py
"""
import os
import shutil
import sys
import datetime
import openpyxl

# ⚠ 콘솔이 cp949 라 경고 문구의 '—' 같은 글자에서 죽는다. 그런데 경고 출력은
#   <b>저장보다 앞</b>이라(main 참조) 여기서 죽으면 <b>변환이 통째로 안 된 채 끝난다</b> —
#   "변환 계획"만 찍히고 표는 그대로여서 성공한 것처럼 보인다(2026-08-19 실제로 겪었다).
#   출력만 UTF-8 로 바꾼다(파일 내용과 무관).
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

from vault_path import TABLE_DIR   # ★ PC 마다 다른 볼트 위치를 찾아준다(2026-08-15)
STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
BACKUP_ROOT = os.path.join(TABLE_DIR, '_백업')

DATA_ROW0 = 4

# (파일, 시트, id 컬럼, [(문구 컬럼, 키 접두사)])
# gen_string_table.py 의 RULES 와 <b>키 규칙이 반드시 같아야 한다</b>.
TARGETS = [
    ('웨이브 몬스터 테이블.xlsx', 'wave_nom', 'monster_id', [('monster_name', 'monster_name')]),
    ('웨이브 몬스터 테이블.xlsx', 'wave_mid_boss', 'monster_id', [('monster_name', 'monster_name')]),
    ('웨이브 몬스터 테이블.xlsx', 'wave_top_boss', 'monster_id', [
        ('monster_name', 'monster_name'), ('boss_title', 'boss_title')]),
    ('웨이브 몬스터 테이블.xlsx', 'Skill', 'skill_id', [
        ('skill_name', 'skill_name'), ('skill_explain', 'skill_explain')]),
    ('웨이브 몬스터 테이블.xlsx', 'Skill_Type', 'skill_type', [('desc', 'skill_type_desc')]),

    ('캐릭터 테이블.xlsx', 'Character', 'character_id', [
        ('character_name', 'character_name'),
        # 2026-08-19 신설 — 칭호. `character_title_EG` 는 영어 <b>문구</b>가 정본이라
        # 키로 바꾸지 않는다(`boss_title_EG` · `character_name_EG` 와 같은 규칙, 위 doc 참조).
        ('character_title', 'character_title')]),
    ('캐릭터 테이블.xlsx', 'Skill', 'skill_id', [
        ('skill_name', 'skill_name'), ('skill_explain', 'skill_explain')]),
    ('캐릭터 테이블.xlsx', 'Skill_Type', 'skill_type', [('desc', 'skill_type_desc')]),

    ('임시용 중립 몬스터.xlsx', 'neutrality_mon', 'mon_id', [
        ('mon_name', 'mon_name'), ('mon_title', 'mon_title')]),

    ('정신 이상 테이블.xlsx', 'mental_error', 'mental_error_id', [
        ('Korean_explain', 'mental_error_name')]),
    ('정신 이상 테이블.xlsx', 'mental_error_type', 'mental_error_type', [
        ('desc', 'mental_error_type_desc')]),

    ('Last_Sanctuary_건물데이터시트_Ver05.xlsx', 'Construction', 'Const_id', [
        ('Const_name', 'const_name')]),
]

# 스트링 키 테이블로 이관이 끝나 <b>원본에서 빼야 하는</b> 시트.
# 유저 지시의 "따로 빼서" 가 이것이다 — 두 곳에 같은 표가 남으면 어느 쪽이 정본인지 알 수 없다.
SHEETS_TO_REMOVE = [('웨이브 몬스터 테이블.xlsx', 'string')]


def norm(v):
    return '' if v is None else str(v).strip()


def looks_like_key(text):
    """gen_string_table.py 의 같은 이름 함수와 규칙이 같아야 한다."""
    if not text:
        return False
    return text.isascii() and ('_' in text) and not any(c.isspace() for c in text)


def load_string_table():
    """key → kr. 변환 전에 '이 키에 이 문구가 들어있다'를 확인하는 데 쓴다."""
    if not os.path.exists(STRING_XLSX):
        sys.exit('스트링 키 테이블이 없습니다. 먼저 python Tools/gen_string_table.py 를 돌리세요.')

    wb = openpyxl.load_workbook(STRING_XLSX, data_only=True)
    ws = wb['string']
    out = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        key = norm(ws.cell(row=r, column=1).value)
        if key:
            out[key] = norm(ws.cell(row=r, column=2).value)
    return out


def field_row_map(ws_vals):
    """2행(필드명) → 열 번호 (1-based). openpyxl 로 읽은 값 격자를 받는다."""
    out = {}
    header = ws_vals[1] if len(ws_vals) > 1 else []
    for c, name in enumerate(header, 1):
        name = norm(name)
        if name:
            out[name] = c
    return out


def plan_edits(strings):
    """무엇을 어떻게 바꿀지 먼저 전부 계산한다 (Excel 을 열기 전에).

    Excel 을 띄운 상태에서 판단까지 하면, 중간에 예외가 나면 <b>절반만 바뀐 파일</b>이
    남는다. 계획을 먼저 세우고 쓰기는 한 번에 한다.
    """
    plans = {}          # 파일 → [(시트, 행, 열, 옛값, 새값)]
    warnings = []
    skipped = 0

    for filename, sheet, id_field, specs in TARGETS:
        path = os.path.join(TABLE_DIR, filename)
        if not os.path.exists(path):
            warnings.append(f'파일 없음: {filename}')
            continue

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            warnings.append(f'{filename}: 시트 없음 {sheet}')
            continue

        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        idx = field_row_map(rows)
        if id_field not in idx:
            warnings.append(f'{filename}/{sheet}: id 컬럼 없음 {id_field}')
            continue

        for r in range(DATA_ROW0, ws.max_row + 1):
            row_id = norm(ws.cell(row=r, column=idx[id_field]).value)
            if not row_id:
                continue
            if row_id.endswith('.0'):
                row_id = row_id[:-2]

            for value_field, prefix in specs:
                if value_field not in idx:
                    continue

                col = idx[value_field]
                text = norm(ws.cell(row=r, column=col).value)
                if not text:
                    continue

                key = f'{prefix}_{row_id}'

                if looks_like_key(text):
                    if text != key:
                        warnings.append(
                            f'{filename}/{sheet} {r}행 {value_field}: 이미 다른 키가 있다 '
                            f'— "{text}" (예상 "{key}")')
                    skipped += 1
                    continue

                if key not in strings:
                    warnings.append(
                        f'{filename}/{sheet} {r}행 {value_field}: 스트링 테이블에 "{key}" 가 없어 '
                        f'건너뛴다 — gen_string_table.py 를 먼저 돌리세요')
                    continue

                if strings[key] != text:
                    warnings.append(
                        f'{filename}/{sheet} {r}행 {value_field}: 값이 스트링 테이블과 달라 '
                        f'건드리지 않았다 — 표 "{text}" vs 스트링 "{strings[key]}"')
                    continue

                plans.setdefault(filename, []).append((sheet, r, col, text, key))

    return plans, warnings, skipped


def backup(filenames):
    stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    folder = os.path.join(BACKUP_ROOT, stamp)
    os.makedirs(folder, exist_ok=True)
    for fn in filenames:
        src = os.path.join(TABLE_DIR, fn)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(folder, fn))
    return folder


def apply_with_excel(plans, remove_sheets):
    """Excel 로 열어 셀 값만 바꾸고 저장한다 (주석·서식 보존).

    ⚠ ★ <b>`DispatchEx` 를 쓴다 — `EnsureDispatch`/`Dispatch` 가 아니다</b> (2026-08-19).

      `EnsureDispatch` 는 <b>유저가 엑셀을 켜 두면 실패한다</b>:
        TypeError: This COM object can not automate the makepy process
      이미 떠 있는 인스턴스에 붙으면서 makepy 캐시를 새로 구울 수 없기 때문이다.
      (gen_py 캐시를 지워도 그대로다 — 원인이 캐시가 아니라 '떠 있는 인스턴스'다.)

      더 위험한 건 그다음이다. 그냥 `Dispatch` 로 바꾸면 <b>유저가 열어 둔 그 엑셀에 붙고</b>,
      이 함수 끝의 `app.Quit()` 이 <b>유저의 창을 닫아 버린다</b> — 저장 안 한 표가 있으면
      그대로 날아간다.

      `DispatchEx` 는 <b>항상 새 인스턴스를 띄운다.</b> 그래서 유저 창과 완전히 분리되고
      `Quit()` 도 이 스크립트가 띄운 인스턴스만 닫는다. 이 프로젝트는 표를 열어 놓고
      스크립트를 돌리는 일이 잦으므로 <b>COM 을 쓰는 모든 스크립트가 이 방식이어야 한다.</b>
    """
    import win32com.client as win32

    app = win32.DispatchEx('Excel.Application')
    app.Visible = False
    app.DisplayAlerts = False       # 시트를 지울 때 확인 창이 뜨지 않게

    written = 0
    try:
        touched = set(plans.keys()) | {fn for fn, _ in remove_sheets}
        for filename in sorted(touched):
            path = os.path.join(TABLE_DIR, filename)
            wb = app.Workbooks.Open(os.path.abspath(path))
            try:
                for sheet, r, c, old, new in plans.get(filename, []):
                    wb.Worksheets(sheet).Cells(r, c).Value = new
                    written += 1

                for fn, sheet_name in remove_sheets:
                    if fn != filename:
                        continue
                    names = [ws.Name for ws in wb.Worksheets]
                    if sheet_name in names:
                        if len(names) <= 1:
                            print(f'  ! {filename}: 시트가 하나뿐이라 {sheet_name} 을 지우지 않았다')
                        else:
                            wb.Worksheets(sheet_name).Delete()
                            print(f'  - 시트 제거: {filename} / {sheet_name}')

                wb.Save()
            finally:
                wb.Close(SaveChanges=False)
    finally:
        app.DisplayAlerts = True
        app.Quit()

    return written


def main():
    dry = '--dry' in sys.argv

    strings = load_string_table()
    plans, warnings, skipped = plan_edits(strings)

    total = sum(len(v) for v in plans.values())
    print(f'== 변환 계획: {total}칸 (이미 키인 칸 {skipped}개는 건너뜀)')
    for filename in sorted(plans):
        print(f'\n[{filename}]')
        for sheet, r, c, old, new in plans[filename]:
            print(f'  {sheet} {r}행{c}열: "{old[:34]}" -> {new}')

    if warnings:
        print('\n== 경고')
        for w in warnings:
            print('  !', w)

    if dry:
        print('\n--dry 이므로 파일을 고치지 않았습니다.')
        return

    if total == 0 and not SHEETS_TO_REMOVE:
        print('\n바꿀 것이 없습니다.')
        return

    folder = backup(sorted(set(list(plans.keys()) + [fn for fn, _ in SHEETS_TO_REMOVE])))
    print(f'\n백업: {folder}')

    written = apply_with_excel(plans, SHEETS_TO_REMOVE)
    print(f'변환 완료: {written}칸')


if __name__ == '__main__':
    main()
