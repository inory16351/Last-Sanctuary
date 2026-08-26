# -*- coding: utf-8 -*-
"""보상 재분배 <b>최종</b> — 일반을 올리고 에픽은 «자주 싸게» (2026-08-26, 5차 · <b>정본</b>).

유저 지시
---------
  *"이러면 에픽 보상이 너무 좋음 지금 수정한 값에서 한 2배 정도로 일반 몬스터를 올리고
    에픽을 좀 줄여야 할듯"*
  → 질문으로 확정: **«절충 — 일반 ×1.4 · 두더지 ×2 · 에픽 1/3»**

★★★ 왜 «일반 2배» 를 그대로 못 했나 — <b>산술이 막는다</b>
──────────────────────────────────────────────────────────────────────
11~20 웨이브의 중립 수입은 <b>종양귀(1002)·고르도네(1004) 둘로만</b> 만들어진다
(사냥 구성표에 그 구간에는 두더지·에픽이 없다). 그래서 «일반을 2배» 는 곧
«11~20 구간 수입을 2배» 이고, 파티 레벨이 이렇게 된다:

    파티 평균 Lv       w15    w20    w25    w30
      4차(직전)        11.6   19.4   24.6   34.9
      일반 ×2 라면     16.2   <b>26.5</b>   28.8   35.1   ← 이 세션 첫 지시(«중앙이 너무 쉽다»)로 되돌아간다
      <b>이번(확정)</b>  13.6   <b>22.5</b>   26.3   35.0
    기획서 요구선        13     18     23    (30)

★★ <b>에픽은 «값» 이 아니라 «빈도» 로 줄였다</b>
──────────────────────────────────────────────────────────────────────
«에픽 보상이 너무 좋다» 는 <b>마리당 스티커</b>(폴리르 한 마리 16,100)의 문제다. 그런데
후반 수입에서 에픽이 큰 몫을 지는 것은 <b>수고가 실제로 거기 있기 때문</b>이다 —
26~30 구간에서 «잡는 체력» 의 79%가 에픽이다(에픽 1.2마리 × 2,038 vs 잡몹 70마리 × 4~15).
그래서 <b>총량은 두고 스티커만 나눴다</b>:

    에픽 재생성 1200초 → <b>600초</b> (폴리르 1600 → 800)
      → 사냥 구성의 에픽 칸이 0.5·1.2 → <b>1.0·2.4마리/웨이브</b>
      → 마리당은 <b>1/2.7</b> 로 내려가는데 구간 수입은 그대로

  ⚠ 에픽 마리 수는 <b>공급이 정하는 값</b>이다(서식지마다 동시 1마리 · 넷을 합쳐도
    1200초에 4마리). 그래서 «리스폰 절반» 이 «구성표의 마리 수 2배» 로 곧바로 이어진다.
  ★ 첫 등장 지연(300초)과 요구 레벨(10/15/20/25)은 <b>안 건드렸다</b> — 초반 남용을 막는 것은 그 둘이다.

★ 확정값 — 규칙은 4차와 <b>같다</b> (마리당 = 체력 × 단가 × 거리계수)
──────────────────────────────────────────────────────────────────────
    단가       일반 <b>3.5</b> E/체력 (4차 2.5 ×1.4) · 에픽 <b>1.49</b> E/체력 (4차 4.0)
    거리 계수  15~99 ×1.0 · 100~199 ×1.3 · 200~320 <b>×2.6</b> (4차 1.8 — 두더지만 ×2 가 된다)

    id    이름          체력    4차      5차(확정)      검산
    1001  종양 거미        4     10 →    <b>14</b>    4×3.5×1.0 = 14
    1002  종양귀           8     26 →    <b>36</b>    8×3.5×1.3 = 36.4
    1004  고르도네        15     48 →    <b>68</b>   15×3.5×1.3 = 68.3
    1003  종양 두더지     14     63 →   <b>125</b>   14×3.5×2.6 = 127.4
    1101  카르키노스    1365  5,500 → <b>2,000</b> 1365×1.49 = 2,034
    1102  아니사킬      2035  8,150 → <b>3,000</b>
    1103  바리올라      2715 10,850 → <b>4,050</b>
    1104  폴리르        4024 16,100 → <b>6,000</b>

    에픽 ÷ 두더지 마리당   4차 87배  →  <b>16배</b>
    지속 수급 E/s          1001 67 · 1002 93 · 1004 174 · 1003 203 (26마리÷16초)
                           ★ «멀수록 좋다» 가 처음으로 <b>단조롭게</b> 성립한다

⚠ <b>중반이 3.1Lv 만큼 쉬워지는 것을 유저가 받아들였다</b>(w20 19.4 → 22.5). 기획서 요구선
  (18)보다 4.5 위다. 더 조이려면 일반 단가를 3.5 → 3.0 으로 내리면 w20 이 21 쯤으로 온다.

⚠ 이 표들은 <b>사람이 엑셀에서 고쳐 나가는 문서</b>다 — 통째로 다시 굽지 않고 바꿀 칸만 고친다.
⚠ <b>Excel COM 으로 쓴다</b>(136-4절).

사용법:  python Tools/table_update_20260826_reward_spread_final.py
다음:    python Tools/sync_tables_to_assets.py     ·  Unity 에서 Assets/Refresh
"""

import datetime
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import win32com.client

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

NEUTRAL_XLSX = os.path.join(TABLE_DIR, "임시용 중립 몬스터.xlsx")
FORMULA_XLSX = os.path.join(TABLE_DIR, "능력치 및 공식 정리.xlsx")
BACKUP_ROOT = os.path.join(TABLE_DIR, "_백업")

# ══════════════════════════════════════════════════════════════════════
#  ① 규칙 — 마리당 = 체력 × 단가 × 거리 계수
# ══════════════════════════════════════════════════════════════════════
UNIT_NORMAL = 3.5           # 일반 E/체력 (4차 2.5 → 유저 «2배 정도» 지시를 ×1.4 로 절충)
UNIT_EPIC = 1.49            # 에픽 E/체력 (4차 4.0 — 빈도를 2배로 올렸으니 마리당은 내려간다)

HP = {1001: 4, 1002: 8, 1003: 14, 1004: 15,
      1101: 1365, 1102: 2035, 1103: 2715, 1104: 4024}

DIST = {1001: 1.0, 1002: 1.3, 1004: 1.3, 1003: 2.6}

SPREAD = 0.2                # min/max = 평균 ±20%


def round_to(v):
    """표를 사람이 읽는다 — 값이 클수록 굵게 끊는다."""
    if v < 100:
        return int(round(v))
    if v < 1000:
        return int(round(v / 10.0) * 10)
    return int(round(v / 100.0) * 100)


def energy_of(mon_id):
    unit = UNIT_EPIC if mon_id >= 1100 else UNIT_NORMAL
    avg = HP[mon_id] * unit * DIST.get(mon_id, 1.0)
    lo = round_to(avg * (1.0 - SPREAD))
    hi = round_to(avg * (1.0 + SPREAD))
    return lo, hi, (lo + hi) / 2.0


NEW_ENERGY = {m: energy_of(m)[:2] for m in HP}

# ══════════════════════════════════════════════════════════════════════
#  ② 개체수 · 재생성 — mon_id → (max_alive, respawn_seconds)
#     ★ 에픽 리스폰 절반이 이번 지시의 본체다("에픽을 좀 줄여야")
# ══════════════════════════════════════════════════════════════════════
NEW_DENSITY = {
    1003: (26, 16),         # 4차에서 이미 넣었다 — 다시 돌려도 같다(멱등)
    1101: (1, 600),
    1102: (1, 600),
    1103: (1, 600),
    1104: (1, 800),
}

# ══════════════════════════════════════════════════════════════════════
#  ③ 구간별 사냥 구성 — ★ 에픽 칸이 <b>2배</b>다 (리스폰이 절반이므로)
#     (구간 상한, (1001, 1002+1004, 1003, 에픽))
# ══════════════════════════════════════════════════════════════════════
MIX = [
    (5,  (20,  0,  0, 0.0)),
    (10, (39,  0,  0, 0.0)),
    (15, (20, 18,  0, 0.0)),
    (20, (20, 46,  0, 0.0)),
    (25, (20, 40,  3, 1.0)),        # 4차 0.5
    (30, (20, 40, 10, 2.4)),        # 4차 1.2
]

# ══════════════════════════════════════════════════════════════════════
#  ④ 성장 모델 — UI-67 부터 계수는 <b>한 번도 안 바꿨다</b>
# ══════════════════════════════════════════════════════════════════════
FOCUS_PER_LEVEL = 2.62
PLAIN_PER_LEVEL = 1.62
STAT_BASE_FOCUS = 10
STAT_BASE_PLAIN = 5
STAT_MAX = 100

HP_PER_STAT = 6.1832
HP_BASE = 90

DPS_A, DPS_B, DPS_C = 0.01961, 3.329, -9.50
EFFECTIVE = 0.647
BATTLE_SECONDS = 120

CREATE_BASE, CREATE_STEP = 200, 150
FREE_CHARACTERS = 4

UP_BASE, UP_STEP = 40, 10
UP_STEEP_LEVEL, UP_STEEP_RATE, UP_ROUND = 30, 1.35, 10

WAVE_KILL_ENERGY = 10

SHEET = "웨이브 부하"
FIRST_ROW = 7
COL = dict(wave=1, lv=2, people=3, focus=4, plain=5, hp=6,
           hp_scale=7, atk_scale=8, count=9, mob_hp=10, mob_hit=11,
           hits=12, hit_ratio=13, dps=14, load=15,
           boss=16, boss_raw=17, boss_hp=18, boss_goal=19, neutral=20)


def upgrade_cost(level):
    if level < UP_STEEP_LEVEL:
        return UP_BASE + UP_STEP * level
    at_start = UP_BASE + UP_STEP * UP_STEEP_LEVEL
    return int(round(at_start * (UP_STEEP_RATE ** (level - UP_STEEP_LEVEL)) / UP_ROUND) * UP_ROUND)


def mix_for(wave):
    for upto, m in MIX:
        if wave <= upto:
            return m
    return MIX[-1][1]


def avg_of(mon_id):
    return energy_of(mon_id)[2]


def neutral_income(wave):
    n1, n2, n3, ne = mix_for(wave)
    mid = (avg_of(1002) + avg_of(1004)) / 2.0
    epic = (avg_of(1101) + avg_of(1102) + avg_of(1103)) / 3.0
    return n1 * avg_of(1001) + n2 * mid + n3 * avg_of(1003) + ne * epic


def run_model(rows, neutral):
    """① 충원 ② 최저 레벨부터 강화 ③ 기록 ④ 번다 — UI-67 과 같은 순서."""
    energy = 0.0
    levels = []
    created = 0
    out = []
    for r, nt in zip(rows, neutral):
        while len(levels) < r["people"]:
            if len(levels) < FREE_CHARACTERS:
                levels.append(0)
                continue
            cost = CREATE_BASE + CREATE_STEP * created
            if energy < cost:
                break
            energy -= cost
            created += 1
            levels.append(0)
        guard = 0
        while levels and guard < 500000:
            guard += 1
            i = min(range(len(levels)), key=lambda j: levels[j])
            c = upgrade_cost(levels[i])
            if energy < c:
                break
            energy -= c
            levels[i] += 1
        out.append(sum(levels) / len(levels) if levels else 0.0)
        energy += r["count"] * 2 * WAVE_KILL_ENERGY + nt
    return out


def read_rows():
    wb = openpyxl.load_workbook(FORMULA_XLSX, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=FIRST_ROW, max_row=ws.max_row, values_only=True):
        if r[0] is None:
            continue
        rows.append(dict(
            wave=int(r[0]), lv=float(r[1]), people=int(r[2]),
            focus=float(r[3]), plain=float(r[4]), hp=float(r[5]),
            count=int(r[8]), mob_hp=float(r[9]), mob_hit=float(r[10]),
            hits=float(r[11]), hit_ratio=float(r[12]), dps=float(r[13]),
            load=float(r[14]), neutral=float(r[19]),
        ))
    return rows


def backup():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(BACKUP_ROOT, stamp + "_보상재분배최종")
    os.makedirs(folder, exist_ok=True)
    for f in (NEUTRAL_XLSX, FORMULA_XLSX):
        shutil.copy2(f, os.path.join(folder, os.path.basename(f)))
    print("백업: " + folder)


def main():
    print(f"규칙 — 마리당 = 체력 × 단가 × 거리계수 (일반 {UNIT_NORMAL} · 에픽 {UNIT_EPIC} E/체력)")
    alive = {1001: (43, 9), 1002: (33, 13), 1004: (33, 13)}
    alive.update({m: v for m, v in NEW_DENSITY.items()})
    for m in sorted(HP):
        lo, hi, av = energy_of(m)
        a, rs = alive[m]
        print(f"  {m}  체력 {HP[m]:5d}  거리 ×{DIST.get(m, 1.0):.1f}  →  {lo:6d}~{hi:6d} "
              f"(평균 {av:8.0f})   지속 {a / rs * av:6.1f} E/s  ({a}마리 ÷ {rs:.0f}초)")

    rows = read_rows()
    neutral = [neutral_income(r["wave"]) for r in rows]
    old_total = sum(r["neutral"] for r in rows)
    new_total = sum(neutral)
    print(f"\n중립 총 수입 {old_total:,.0f} → {new_total:,.0f}  "
          f"({(new_total / old_total - 1) * 100:+.2f}%)")

    print("\n구간 몫")
    seen = set()
    for r, nt in zip(rows, neutral):
        band = next(u for u, _ in MIX if r["wave"] <= u)
        if band in seen:
            continue
        seen.add(band)
        print(f"  ~{band:2d}웨이브  {r['neutral']:7.0f} → {nt:7.0f} /웨이브   "
              f"몫 {r['neutral'] * 5 / old_total * 100:5.1f}% → {nt * 5 / new_total * 100:5.1f}%")

    new_lv = run_model(rows, neutral)

    plan = []
    for r, lv, nt in zip(rows, new_lv, neutral):
        n = round(lv)
        focus = min(STAT_MAX, STAT_BASE_FOCUS + FOCUS_PER_LEVEL * n)
        plain = min(STAT_MAX, STAT_BASE_PLAIN + PLAIN_PER_LEVEL * n)
        hp = round(HP_BASE + HP_PER_STAT * (focus - STAT_BASE_FOCUS))
        per = DPS_A * focus * focus + DPS_B * focus + DPS_C
        dps = round(r["people"] * per)
        load = (r["count"] * 2 * r["mob_hp"]) / (dps * EFFECTIVE * BATTLE_SECONDS) if dps > 0 else 0.0
        ratio = r["hit_ratio"] * (r["hp"] / hp) if hp > 0 else 0.0
        hits = r["hits"] * (hp / r["hp"]) if r["hp"] > 0 else 0.0
        plan.append(dict(wave=r["wave"], lv=round(lv, 1), focus=round(focus, 1),
                         plain=round(plain, 1), hp=hp, dps=dps, load=round(load, 2),
                         ratio=round(ratio, 3), hits=round(hits, 1), neutral=round(nt)))

    print("\n웨이브 | 평균Lv (전) | 캐릭터HP | 파티DPS | 잡몹부하 (전) | 중립수입 (전)")
    for p, r in zip(plan, rows):
        if p["wave"] % 5 == 0 or p["wave"] in (11, 16, 21, 26):
            print(f"{p['wave']:6d} | {p['lv']:5.1f} ({r['lv']:4.1f}) | {p['hp']:8d} | "
                  f"{p['dps']:7d} | {p['load']:5.2f} ({r['load']:4.2f}) | "
                  f"{p['neutral']:6d} ({int(r['neutral'])})")

    backup()
    write_excel(plan)


def write_excel(plan):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(NEUTRAL_XLSX))
        ws = wb.Worksheets("neutrality_mon")
        touched = 0
        row = 4
        while True:
            v = ws.Cells(row, 1).Value
            if v is None:
                break
            mon_id = int(v)
            if mon_id in NEW_ENERGY:
                lo, hi = NEW_ENERGY[mon_id]
                before = (ws.Cells(row, 10).Value, ws.Cells(row, 11).Value)
                ws.Cells(row, 10).Value = lo
                ws.Cells(row, 11).Value = hi
                print(f"  중립 {mon_id}: 에너지 {int(before[0])}~{int(before[1])} → {lo}~{hi}")
                touched += 1
            if mon_id in NEW_DENSITY:
                a, rs = NEW_DENSITY[mon_id]
                was = (ws.Cells(row, 15).Value, ws.Cells(row, 16).Value)
                ws.Cells(row, 15).Value = a
                ws.Cells(row, 16).Value = rs
                if (int(was[0]), int(was[1])) != (a, rs):
                    print(f"  중립 {mon_id}: 밀도 {int(was[0])}마리/{int(was[1])}초 → {a}마리/{rs}초")
            row += 1
        wb.Save()
        wb.Close()
        print(f"임시용 중립 몬스터.xlsx — {touched}종 갱신")

        wb = excel.Workbooks.Open(os.path.abspath(FORMULA_XLSX))
        ws = wb.Worksheets(SHEET)
        for i, p in enumerate(plan):
            r = FIRST_ROW + i
            assert int(ws.Cells(r, COL["wave"]).Value) == p["wave"], \
                f"{r}행이 웨이브 {p['wave']} 가 아니다"
            ws.Cells(r, COL["lv"]).Value = p["lv"]
            ws.Cells(r, COL["focus"]).Value = p["focus"]
            ws.Cells(r, COL["plain"]).Value = p["plain"]
            ws.Cells(r, COL["hp"]).Value = p["hp"]
            ws.Cells(r, COL["hits"]).Value = p["hits"]
            ws.Cells(r, COL["hit_ratio"]).Value = p["ratio"]
            ws.Cells(r, COL["dps"]).Value = p["dps"]
            ws.Cells(r, COL["load"]).Value = p["load"]
            ws.Cells(r, COL["neutral"]).Value = p["neutral"]

        note = ("※ 2026-08-26 개정 5차(정본 · 유저 확정 «절충»): 4차의 «체력 비례» 규칙은 그대로 두고 단가만 옮겼다 — "
                "일반 2.5→3.5 E/체력(유저 «2배 정도» 를 ×1.4 로 절충) · 두더지 거리계수 1.8→2.6(두더지만 ×2) · "
                "에픽 4.0→1.49 E/체력. ★ 에픽은 «값» 이 아니라 «빈도» 로 줄였다 — 재생성 1200→600초(폴리르 1600→800)라 "
                "구간 수입은 그대로인데 마리당 스티커가 1/2.7 이 된다(폴리르 16,100→6,000). "
                "마리당 1001 14 · 1002 36 · 1004 68 · 1003 125 · 에픽 2,000/3,000/4,050/6,000. 에픽÷두더지 87배→16배. "
                "지속 수급 E/s 67/93/174/203 으로 «멀수록 좋다» 가 처음으로 단조롭다. "
                "⚠ 대가 — 11~20 구간 수입은 종양귀·고르도네 둘로만 만들어지므로 일반을 올리면 그 구간이 쉬워진다: "
                "파티 Lv w15 13.6 · w20 22.5 · w25 26.3 · w30 35.0 (4차 11.6/19.4/24.6/34.9 · 기획서 요구선 13/18/23). "
                "유저가 이 대가를 확인하고 고른 값이다. 더 조이려면 일반 단가를 3.0 으로 내릴 것(w20 ≈ 21). "
                "⚠ 몬스터 스펙(체력·공격력)과 보스는 한 칸도 고치지 않았다.")
        ws.Cells(5, 1).Value = note

        coef = wb.Worksheets("계수")
        r = coef.UsedRange.Rows.Count + 1
        rows_out = [
            ("■ 2026-08-26 (5차 · 정본) — 일반을 올리고 에픽은 «자주 싸게»", "", "", ""),
            ("단가 — 일반 / 에픽", "3.5 / 1.49 E/체력", "Tools/table_update_20260826_reward_spread_final.py",
             "4차 2.5/4.0 → 유저 지시 «일반을 2배로, 에픽을 줄여» 를 ×1.4 로 절충(2배면 w20 이 26.5Lv 가 된다)"),
            ("거리 계수 — 15~99 / 100~199 / 200~320 타일", "1.0 / 1.3 / 2.6", "위 규칙의 계수",
             "두더지만 ×2 가 되도록 200~320 을 1.8 → 2.6 으로. 지속 수급이 처음으로 «멀수록 좋다» 로 단조로워졌다"),
            ("에픽 재생성 주기", "600초 (폴리르 800초)", "임시용 중립 몬스터.respawn_seconds",
             "★ 에픽을 «값» 이 아니라 «빈도» 로 줄였다 — 구간 수입은 그대로인데 마리당 스티커가 1/2.7 이 된다"),
        ]
        for m in (1001, 1002, 1004, 1003, 1101, 1102, 1103, 1104):
            lo, hi, av = energy_of(m)
            rows_out.append((f"{m} 마리당 평균 보상", int(av), f"임시용 중립 몬스터 {lo}~{hi}",
                             f"체력 {HP[m]} × {UNIT_EPIC if m >= 1100 else UNIT_NORMAL} × {DIST.get(m, 1.0)}"))
        rows_out += [
            ("에픽 ÷ 두더지 마리당", "16배", "위 값에서 파생",
             "2차 0.4배(두더지가 더 좋았다) · 4차 87배 · 이번 16배"),
            ("파티 평균 Lv (w15 / w20 / w25 / w30)", "13.6 / 22.5 / 26.3 / 35.0",
             "성장 모델(5차 스크립트)",
             "기획서 요구선 13/18/23/(30) · 착지점 Lv35 유지 · 4차 11.6/19.4/24.6/34.9"),
        ]
        for label, value, where, why in rows_out:
            coef.Cells(r, 1).Value = label
            if value != "":
                coef.Cells(r, 2).Value = value
            if where:
                coef.Cells(r, 3).Value = where
            if why:
                coef.Cells(r, 4).Value = why
            r += 1

        wb.Save()
        wb.Close()
        print(f"능력치 및 공식 정리.xlsx — 「웨이브 부하」 9열 × 30행 · 「계수」 {len(rows_out)}행 추가")
    finally:
        excel.Quit()


if __name__ == "__main__":
    main()
