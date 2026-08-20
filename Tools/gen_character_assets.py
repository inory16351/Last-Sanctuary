# -*- coding: utf-8 -*-
"""캐릭터 테이블.xlsx 를 읽어 Unity ScriptableObject 에셋(.asset + .meta)을 생성한다.

엑셀이 원본이고 이 스크립트가 그걸 그대로 옮긴다 — 값을 손으로 옮겨 적지 않는다.
다시 돌려도 같은 결과가 나온다(guid 를 경로에서 결정적으로 만든다).

⚠ SO 에셋을 만드는 MCP 도구가 없어서 YAML 을 직접 쓴다 — 진행상황 5절·8절에서
이미 쓰던 방식이다. .asset YAML 에 빈 줄을 넣으면 Unity 파서가 그 뒤 필드를
전부 무시하므로(8절 3번) 절대 빈 줄을 넣지 않는다.
"""
import os, hashlib
import openpyxl

from vault_path import TABLE_DIR   # ★ PC 마다 다른 볼트 위치를 찾아준다(2026-08-15)
import os as _os
XLSX = _os.path.join(TABLE_DIR, '캐릭터 테이블.xlsx')

# ⚠ 예전에는 프로젝트 경로를 'C:\Project\Last Sanctuary' 로 박아뒀는데 실제 폴더는
#   'C:\Project\Last-Sanctuary' 다(하이픈). 그대로 돌리면 엉뚱한 폴더를 새로 만들고
#   진짜 에셋은 하나도 안 바뀐 채 "생성 완료" 만 찍혔다. 스크립트 위치에서 역산한다.
_PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_SKILL = os.path.join(_PROJECT, 'Assets', '_Project', 'Resources', 'PassiveSkills')
OUT_CHAR = os.path.join(_PROJECT, 'Assets', '_Project', 'Resources', 'Characters')

SCRIPT_GUID_SKILL = '21498a0a43b4e824ea7e6db210ef2e29'   # PassiveSkillSO.cs
SCRIPT_GUID_CHAR = 'bacf4f16746e56b4da254173d578cf4e'    # CharacterDefinitionSO.cs

# 뺄 캐릭터 — 지금은 없다.
# 프레이야(9003)는 인게임 에셋 제작 중이라 빠져 있었으나 2026-08-11 임포트 완료
# (Tools/char_asset_preyja_build.py → Skin_Preyja).
EXCLUDE_CHARACTER_IDS = set()

# ── 인게임 외형 배정 (Resources/Skins/<이름>.asset) ────────────────────────
#
# ★★ 2026-08-20 — <b>손으로 적던 표를 없앴다.</b> 유저 지시: *"하드 코딩 최대한 자제하고
#   웬만한건 다 mcp로 직접 만들어줘"*.
#
# 예전에는 캐릭터 id → 스킨 이름을 여기 다섯 줄로 적어 뒀다. 그런데 다섯 줄이 전부
# <b>같은 규칙</b>이었다: `Skin_<영어 이름>` (Elin→Skin_Elin · Bigior→Skin_Bigior · …).
# 즉 표는 규칙의 <b>중복</b>이었고, 그래서 시그리드(9006)를 추가했을 때 이 줄을 안 고쳐
# `skinAssetName` 이 <b>빈 문자열</b>로 나갔다 — 캐릭터가 외형 없이 생성될 상태였다.
#
# 이제 규칙으로 만든다: 영어 이름으로 `Skin_<En>` 을 짓고, <b>그 에셋이 실제로 있을 때만</b>
# 적는다. 원화가 아직 없는 캐릭터가 없는 스킨을 가리켜 런타임 경고를 내지 않게 하려는 것이다.
#
# ⚠ `SKIN_OVERRIDE` 는 <b>예외용으로만</b> 남긴다 — 이름 규칙에서 벗어나는 인물이 생기면
#   여기 한 줄을 적는다. 지금은 비어 있는 것이 정상이다.
SKIN_OVERRIDE = {}


def skin_asset_name(cid, name_en):
    """
    이 캐릭터가 쓸 스킨 에셋 이름. 규칙은 `Skin_<영어 이름>` 하나다(위 ★★).
    에셋이 없으면 빈 문자열 — 없는 것을 가리키는 것보다 비는 편이 낫다
    (`CharacterAnimator` 가 비면 무작위 스킨으로 떨어진다).
    """
    if cid in SKIN_OVERRIDE:
        return SKIN_OVERRIDE[cid]
    if not name_en:
        return ''
    guess = 'Skin_%s' % name_en
    path = os.path.join(_PROJECT, 'Assets', '_Project', 'Resources', 'Skins', guess + '.asset')
    return guess if os.path.isfile(path) else ''

# 역할 고정 — 능력치 역산(CharacterRole)을 덮어쓸 인물만 적는다. 값은 RoleAttackPreset /
# RolePositionPreset enum 의 정수다 (0=Auto · 공격 1=Melee 2=Ranged 3=Magic 4=Heal ·
# 위치 1=Front 2=Mid 3=Back).
#
# ★ 히스톤 — 표의 패시브 <b>선봉장(Vanguard)</b> 이 "포지션은 전방 / 공격 유형은 근거리로
#   고정된다" 고 못박고 있다. 능력치 역산은 근거리까지는 맞히지만(근접 9 가 최고),
#   맷집이 체력 8 + 방어 5 = 13 으로 기준(15)에 못 미쳐 <b>중위</b>가 나온다 —
#   스킬이 규정한 값과 어긋나므로 여기서 고정한다.
#
# ★ 시그리드 — <b>유저 지시</b>(2026-08-20): *"시그리드 캐릭터 생성 시 초기 전술 포지션이
#   원거리로 설정되어있는데 생성하면 근거리로 설정되게 변경"*.
#   능력치 역산은 원거리를 고른다 — 원거리 공격력 8 이 근접 6 보다 높다(`first_Stat`).
#   그런데 스킬 셋이 <b>전부 붙어서 싸우는 전제</b>다:
#     가학증               «때린 적이 2초 안에 죽으면» — 자기가 마무리를 쳐야 터진다
#     통제할 수 없는 쾌락  체력 10% 아래에서 무적 — 맞는 자리에 서 있어야 뜻이 있다
#     (가학증이 후퇴 기준을 5% 로 못박는 것도 같은 방향 · UI-42절)
#   히스톤과 <b>같은 종류의 예외</b>다 — 스킬이 규정한 자리와 역산이 어긋난다.
#
# ⚠ 위치(positionPreset)는 <b>Auto 로 둔다</b> — 유저가 말한 것은 「원거리 → 근거리」
#   (공격 유형)뿐이다. 근거리 + 맷집 체력 6 + 방어 2 = 8 이라 역산은 <b>중위</b>가 되고,
#   그게 «무른 근접» 인 시그리드에게 맞는다(프레이야와 같은 자리). 전방으로 못박으면
#   가장 먼저 맞아 무적 구간을 쓰기도 전에 죽는다.
ROLE_OVERRIDE = {
    9005: (1, 1),   # (attackPreset=Melee, positionPreset=Front)
    9006: (1, 0),   # (attackPreset=Melee, positionPreset=Auto → 중위)
}

NARRATIVE = {
    9001: '눈을 가린 채 가장 먼저 전장에 뛰어드는 호중구. 제 몸을 그물처럼 풀어 적을 '
          '붙잡고 스러진다. 보지 못하지만 누구보다 먼저 침입자를 알아챈다.',
    9002: '전열을 지탱하는 대식세포. 삼킨 것을 태워 열을 내고, 그 열로 곁의 아군까지 지킨다. '
          '무너지지 않는 것이 그의 역할이다.',
    9003: '굶주린 자연살해세포. 표식이 사라진 것을 먼저 알아보고, 먹어치울수록 빨라진다.',
}

FOLDER_META = """fileFormatVersion: 2
guid: {guid}
folderAsset: yes
DefaultImporter:
  externalObjects: {{}}
  userData:
  assetBundleName:
  assetBundleVariant:
"""

ASSET_META = """fileFormatVersion: 2
guid: {guid}
NativeFormatImporter:
  externalObjects: {{}}
  mainObjectFileID: 11400000
  userData:
  assetBundleName:
  assetBundleVariant:
"""

HEADER = """%YAML 1.1
%TAG !u! tag:unity3d.com,2011:
--- !u!114 &11400000
MonoBehaviour:
  m_ObjectHideFlags: 0
  m_CorrespondingSourceObject: {{fileID: 0}}
  m_PrefabInstance: {{fileID: 0}}
  m_PrefabAsset: {{fileID: 0}}
  m_GameObject: {{fileID: 0}}
  m_Enabled: 1
  m_EditorHideFlags: 0
  m_Script: {{fileID: 11500000, guid: {script_guid}, type: 3}}
  m_Name: {name}
  m_EditorClassIdentifier:
"""


def guid_for(key):
    return hashlib.md5(('LastSanctuary/' + key).encode('utf-8')).hexdigest()


# ---------------------------------------------------------------------------
# 스트링 테이블 — 2026-08-12부터 캐릭터 테이블의 문구 컬럼에는 <b>키</b>가 들어 있다
# (Tools/convert_tables_to_string_keys.py 가 바꿨다). 그래서 에셋에 넣을
# <b>리터럴 폴백</b>은 여기서 되돌려 읽어야 한다 — 안 그러면 폴백 칸에 키 문자열이 들어가고,
# 스트링 테이블을 못 읽는 상황에서 화면에 'character_name_9001' 이 그대로 뜬다.
# ---------------------------------------------------------------------------
STRING_XLSX = _os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')


def load_strings(column=2):
    """스트링 키 테이블에서 `{키: 문구}` 를 읽는다. column 2 = 한국어(kr), 3 = 영어(en)."""
    if not os.path.exists(STRING_XLSX):
        print('  ! 스트링 키 테이블이 없습니다 — 리터럴 폴백이 비게 됩니다.')
        return {}
    wb = openpyxl.load_workbook(STRING_XLSX, data_only=True)
    ws = wb['string']
    out = {}
    for r in range(4, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k:
            out[str(k).strip()] = (ws.cell(r, column).value or '')
    return out


_STRINGS = load_strings()

# ★ 영어 이름은 <b>스트링 키 테이블의 en 칸</b>이 정본이다 (2026-08-13).
#   예전에는 캐릭터 테이블의 `character_name_EG` 컬럼을 읽어 에셋 파일 이름
#   (`Character_9001_Elin`)을 만들었다. 유저 지시로 그 컬럼을 지웠으므로
#   ("영어 이름 칼럼이 삭제되지 않고 남아있는 것들 확인해서 없애줘") 여기서 읽는다.
#   ⚠ 에셋 <b>파일 이름</b>에 쓰이므로 값이 비면 guid 가 바뀌어 참조가 끊긴다 —
#   비면 아래에서 에러를 내고 멈춘다.
_STRINGS_EN = load_strings(column=3)


def english_name(key, fallback=''):
    """스트링 키의 영어 문구. 공백은 지운다 — 파일 이름에 쓰이기 때문이다."""
    s = str(_STRINGS_EN.get(key, '') or '').strip()
    return s.replace(' ', '') if s else fallback


def looks_like_key(text):
    """convert_tables_to_string_keys.py 와 같은 규칙."""
    return bool(text) and text.isascii() and ('_' in text) and \
        not any(c.isspace() for c in text)


def text_of(value):
    """셀 값 → 사람이 읽는 문구. 셀이 스트링 키면 스트링 테이블에서 되돌려 읽는다."""
    s = '' if value is None else str(value).strip()
    if looks_like_key(s) and s in _STRINGS:
        return str(_STRINGS[s]).strip()
    return s


def yaml_str(s):
    """Unity YAML 의 문자열. 줄바꿈·특수문자를 안전하게 처리한다."""
    if s is None:
        return "''"
    s = str(s).strip()
    if s == '':
        return "''"
    # 여러 줄이면 접힌 블록 대신 이스케이프한 한 줄로 넣는다 (파서 사고 방지)
    s = s.replace('\\', '\\\\').replace('\n', '\\n').replace('\r', '')
    if "'" in s:
        return '"' + s.replace('"', '\\"') + '"'
    return "'" + s + "'"


def num(v):
    if v is None:
        return 0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0
    return int(f) if f == int(f) else f


# ------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---- Skill 시트 → PassiveSkillSO ----
ws = wb['Skill']
skill_types = {}
wt = wb['Skill_Type']
for r in range(4, wt.max_row + 1):
    k = wt.cell(r, 1).value
    if k:
        skill_types[str(k).strip()] = wt.cell(r, 2).value or ''

os.makedirs(OUT_SKILL, exist_ok=True)
os.makedirs(OUT_CHAR, exist_ok=True)
for folder, key in ((OUT_SKILL, 'Resources/PassiveSkills'), (OUT_CHAR, 'Resources/Characters')):
    fm = folder + '.meta'
    if not os.path.exists(fm):
        with open(fm, 'w', encoding='utf-8', newline='\n') as f:
            f.write(FOLDER_META.format(guid=guid_for(key)))

skill_guid_by_id = {}
made = 0
# ★★ Skill 시트도 <b>이름으로</b> 읽는다 (2026-08-20)
#
# 예전에는 컬럼 번호를 박아 뒀다(`ws.cell(r, 7)` = 쿨타임 …). 그런데 그 뒤 표에
# **`value_04` 컬럼이 새로 생겼다** — 지금 시트는 10칸이다:
#   1 skill_id · 2 skill_name · 3 skill_type · 4~8 value_01~05 · 9 cool_time ·
#   10 skill_icon · 11 skill_explain
#   (2026-08-20 에 value_05 가 또 늘었다 — 이름으로 읽으므로 아무 일도 안 났다.
#    번호로 읽었다면 그때마다 뒤가 전부 밀렸을 것이다.)
#
# ⚠⚠ 그래서 번호로 읽으면 **전부 한 칸씩 밀린다**: 쿨타임 칸이 `value_04` 를 읽고,
#    아이콘 칸이 **숫자(쿨타임)** 를 읽고, 플레이버 칸이 **아이콘 이름**을 읽는다.
#    다행히 그 컬럼이 생긴 뒤로 이 스크립트를 한 번도 안 돌려서 에셋은 아직 멀쩡했다 —
#    <b>다음 실행이 전부 망가뜨릴 상태였다</b>(2026-08-20 에 발견해 고쳤다).
#    `sync_tables_to_assets.py` 의 read_rows 주석이 적어둔 2026-08-13 사고와 같은 종류다.
_SKILL_COL = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(2, c).value
    if v:
        _SKILL_COL[str(v).strip()] = c


def skill_cell(row, field):
    c = _SKILL_COL.get(field)
    return ws.cell(row, c).value if c else None


for _f in ('skill_id', 'skill_name', 'skill_type', 'cool_time', 'skill_icon', 'skill_explain'):
    if _f not in _SKILL_COL:
        raise SystemExit('캐릭터 테이블 Skill 시트에 %s 컬럼이 없습니다.' % _f)

for r in range(4, ws.max_row + 1):
    sid = skill_cell(r, 'skill_id')
    if not sid:
        continue
    sid = int(sid)
    sname = text_of(skill_cell(r, 'skill_name'))   # 셀은 이제 키다 — 문구로 되돌린다
    stype = (skill_cell(r, 'skill_type') or '').strip()
    v1 = num(skill_cell(r, 'value_01'))
    v2 = num(skill_cell(r, 'value_02'))
    v3 = num(skill_cell(r, 'value_03'))
    # ★ value_04 — 시그리드 「가학증」이 네 번째 값을 쓴다(아군 회복량 = 시그리드
    #   현재 체력의 value_04%). 컬럼이 없는 옛 표에서는 0 이 된다.
    v4 = num(skill_cell(r, 'value_04'))
    # ★ value_05 — 2026-08-20 유저가 표에 추가. 시그리드 「가학증」의 후퇴기준 고정값.
    v5 = num(skill_cell(r, 'value_05'))
    cool = num(skill_cell(r, 'cool_time'))
    icon = (skill_cell(r, 'skill_icon') or '').strip()
    flavor = text_of(skill_cell(r, 'skill_explain'))
    effect = text_of(skill_types.get(stype, ''))

    asset_name = 'Skill_%d_%s' % (sid, stype.replace(' ', ''))
    rel = 'Resources/PassiveSkills/%s.asset' % asset_name
    g = guid_for(rel)
    skill_guid_by_id[sid] = g

    body = HEADER.format(script_guid=SCRIPT_GUID_SKILL, name=asset_name)
    body += "  skillId: %d\n" % sid
    # ★ 스트링 키 (2026-08-12) — 화면 문구의 정본은 '스트링 키 테이블.xlsx' 다.
    #   키 형식은 Tools/gen_string_table.py 의 규칙과 반드시 같아야 한다.
    #   아래 리터럴(skillName·flavorText·effectTemplate)은 키를 못 찾았을 때의 폴백으로 남긴다.
    body += "  nameKey: %s\n" % yaml_str('skill_name_%d' % sid)
    body += "  flavorKey: %s\n" % yaml_str('skill_explain_%d' % sid)
    body += "  effectKey: %s\n" % yaml_str('skill_type_desc_%s' % stype if stype else '')
    body += "  skillName: %s\n" % yaml_str(sname)
    body += "  skillType: %s\n" % yaml_str(stype)
    body += "  value01: %s\n" % v1
    body += "  value02: %s\n" % v2
    body += "  value03: %s\n" % v3
    body += "  value04: %s\n" % v4
    body += "  value05: %s\n" % v5
    body += "  coolTime: %s\n" % cool
    body += "  iconName: %s\n" % yaml_str(icon)
    body += "  flavorText: %s\n" % yaml_str(flavor)
    body += "  effectTemplate: %s\n" % yaml_str(effect)

    p = os.path.join(OUT_SKILL, asset_name + '.asset')
    with open(p, 'w', encoding='utf-8', newline='\n') as f:
        f.write(body)
    with open(p + '.meta', 'w', encoding='utf-8', newline='\n') as f:
        f.write(ASSET_META.format(guid=g))
    made += 1
print('패시브 스킬 에셋:', made)

# ---- Character + first_Stat 시트 → CharacterDefinitionSO ----
wc = wb['Character']
wsst = wb['first_Stat']
stats_by_id = {}
STAT_COLS = ['hp', 'melee_atk', 'ranged_atk', 'accuracy', 'critical', 'magic',
             'cure', 'atk_speed', 'movement_speed', 'def', 'hp_recovery', 'resistance']
for r in range(4, wsst.max_row + 1):
    cid = wsst.cell(r, 1).value
    if not cid:
        continue
    vals = {STAT_COLS[i]: num(wsst.cell(r, 2 + i).value) for i in range(len(STAT_COLS))}
    stats_by_id[int(cid)] = vals

# ⚠ Character 시트는 <b>필드명으로</b> 읽는다 — 2026-08-13 에 `character_name_EG` 컬럼을
#   지우면서 뒤 컬럼(skill_01~03 · illust)이 전부 한 칸씩 밀렸다. 위치로 읽으면 스킬 자리에
#   일러스트 이름이 들어가고도 "생성 완료"만 찍힌다.
_CHAR_COL = {}
for c in range(1, wc.max_column + 1):
    v = wc.cell(2, c).value
    if v is not None and str(v).strip():
        _CHAR_COL[str(v).strip()] = c


def char_cell(row, field):
    c = _CHAR_COL.get(field)
    return wc.cell(row, c).value if c else None


made = 0
skipped = []
for r in range(4, wc.max_row + 1):
    cid = wc.cell(r, 1).value
    if not cid:
        continue
    cid = int(cid)
    cname = text_of(char_cell(r, 'character_name'))   # 셀은 이제 키다 — 문구로 되돌린다
    # 영어 이름은 스트링 키 테이블의 en 칸이 정본이다(위 english_name 주석).
    cname_en = english_name('character_name_%d' % cid)
    sk = [char_cell(r, 'skill_01'), char_cell(r, 'skill_02'), char_cell(r, 'skill_03')]
    illust = (char_cell(r, 'illust') or '').strip()

    if not cname_en:
        raise SystemExit(
            '스트링 키 테이블에 character_name_%d 의 영어(en) 이름이 없습니다. '
            '에셋 파일 이름에 쓰이므로 비면 guid 가 바뀌어 참조가 끊깁니다 — '
            '스트링 키 테이블을 먼저 채우세요.' % cid)

    if cid in EXCLUDE_CHARACTER_IDS:
        skipped.append(cname)
        continue

    st = stats_by_id.get(cid, {})
    asset_name = 'Character_%d_%s' % (cid, cname_en)
    rel = 'Resources/Characters/%s.asset' % asset_name
    g = guid_for(rel)

    body = HEADER.format(script_guid=SCRIPT_GUID_CHAR, name=asset_name)
    body += "  characterId: %d\n" % cid
    # ★ 스트링 키 (2026-08-12) — 아래 characterName 은 폴백용으로만 남긴다.
    body += "  nameKey: %s\n" % yaml_str('character_name_%d' % cid)
    body += "  characterName: %s\n" % yaml_str(cname)
    body += "  characterNameEn: %s\n" % yaml_str(cname_en)
    # ★ 칭호 (2026-08-19) — 이름과 같은 규칙이다: 키가 정본, 리터럴은 폴백.
    #   표에 칭호가 비어 있으면 키도 비워 둔다 → 상세 카드의 칭호 칸이 빈칸으로 남는다
    #   (유저 확정: "칭호 해금이 되지 않았을 때는 칭호칸 비워놔").
    ctitle = text_of(char_cell(r, 'character_title'))
    body += "  titleKey: %s\n" % yaml_str('character_title_%d' % cid if ctitle else '')
    body += "  title: %s\n" % yaml_str(ctitle)
    body += "  illustName: %s\n" % yaml_str(illust)
    body += "  skinAssetName: %s\n" % yaml_str(skin_asset_name(cid, cname_en))
    body += "  stats:\n"
    body += "    hp: %d\n" % st.get('hp', 5)
    body += "    attack: %d\n" % st.get('melee_atk', 5)
    body += "    defense: %d\n" % st.get('def', 5)
    body += "    regen: %d\n" % st.get('hp_recovery', 5)
    body += "    rangedAttack: %d\n" % st.get('ranged_atk', 5)
    body += "    magic: %d\n" % st.get('magic', 5)
    body += "    cure: %d\n" % st.get('cure', 5)
    body += "    accuracy: %d\n" % st.get('accuracy', 5)
    body += "    critical: %d\n" % st.get('critical', 5)
    body += "    attackSpeed: %d\n" % st.get('atk_speed', 5)
    body += "    moveSpeed: %d\n" % st.get('movement_speed', 5)
    body += "    resistance: %d\n" % st.get('resistance', 50)
    # 역할 — 기본은 Auto(0) 라 CharacterRole 이 능력치에서 역산한다(82-8절).
    # 표에 없는 개념이라, 인물별 고정이 필요한 경우만 ROLE_OVERRIDE 로 덮는다.
    ap, pp = ROLE_OVERRIDE.get(cid, (0, 0))
    body += "  attackPreset: %d\n" % ap
    body += "  positionPreset: %d\n" % pp
    body += "  passives:\n"
    for s in sk:
        if s and int(s) in skill_guid_by_id:
            body += "  - {fileID: 11400000, guid: %s, type: 2}\n" % skill_guid_by_id[int(s)]
        else:
            body += "  - {fileID: 0}\n"
    # 패시브 해금 조건은 여기에 쓰지 않는다 —
    # 씬의 GameSystems > PassiveUnlockConfig 에서 인스펙터로 관리한다(유저 확정 2026-08-11).
    body += "  narrative: %s\n" % yaml_str(NARRATIVE.get(cid, ''))

    p = os.path.join(OUT_CHAR, asset_name + '.asset')
    with open(p, 'w', encoding='utf-8', newline='\n') as f:
        f.write(body)
    with open(p + '.meta', 'w', encoding='utf-8', newline='\n') as f:
        f.write(ASSET_META.format(guid=g))
    made += 1
    print('  캐릭터', cname, '스탯', st)

print('캐릭터 정의 에셋:', made, '/ 제외:', skipped)
