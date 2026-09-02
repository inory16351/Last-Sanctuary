# -*- coding: utf-8 -*-
"""사건 <b>서사</b> 대사의 7개 언어를 표에 넣는다 (2026-09-02).

■ 무엇
  「언어추가」 커밋이 남긴 빈칸 184개 중, 정형문이 아닌 <b>서사 129줄</b>
  (<c>event_script</c> 43 · <c>event_result_script</c> 86)이 대상이다.
  정형문 55줄(<c>event_result_effect</c>)은 <c>table_update_20260902_multilang_effects.py</c> 가
  <b>조립</b>해서 채웠다 — 이쪽은 조립이 안 되므로 문장을 직접 지었다.

■ 자료는 <c>ml_data_events_*.py</c> 에 따로 둔다
  한 파일에 다 넣으면 수천 줄이 되어 고칠 때 어디를 보는지 알 수 없다.
  이 스크립트는 <b>넣는 규칙만</b> 갖고, 문장은 자료 파일이 갖는다.

■ ★ 이미 있는 번역은 <b>덮지 않는다</b>
  이 파일의 다른 형제들(`gen_string_table.py` · `help_string_merge.py`)과 같은 규칙이다.
  사람이 다듬은 문장을 스크립트가 되돌리면 «고쳤는데 되돌아간다» 가 된다.

■ 다음
    py -3 Tools/gen_string_table.py   →   py -3 Tools/link_string_keys.py
"""
import importlib
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET, DATA_ROW0 = 'string', 4
LANGS = ['es', 'fr', 'de', 'ja', 'ru', 'pt', 'pl']

#: 자료 파일 — 있는 것만 읽는다(나눠 쓰는 중에도 돌려볼 수 있게).
MODULES = ['ml_data_events_a', 'ml_data_events_b', 'ml_data_events_c',
           'ml_data_events_d', 'ml_data_events_e', 'ml_data_events_f']


def load():
    data = {}
    used = []
    for name in MODULES:
        try:
            m = importlib.import_module(name)
        except ImportError:
            continue
        dup = set(m.DATA) & set(data)
        if dup:
            raise SystemExit('⚠ 키가 두 자료 파일에 겹칩니다: %s' % sorted(dup)[:5])
        data.update(m.DATA)
        used.append('%s(%d)' % (name, len(m.DATA)))
    print('자료: ' + ' · '.join(used) + '  → 총 %d키' % len(data))
    return data


def main():
    data = load()
    if not data:
        raise SystemExit('⚠ 자료 파일이 하나도 없습니다.')

    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]
    fields = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    C = {n: i + 1 for i, n in enumerate(fields)}
    for L in LANGS:
        if L not in C:
            raise SystemExit('⚠ «%s» 열이 없습니다.' % L)

    where = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = str(ws.cell(r, 1).value or '').strip()
        if k:
            where[k] = r

    missing = [k for k in data if k not in where]
    if missing:
        raise SystemExit('⚠ 표에 없는 키 %d개: %s' % (len(missing), missing[:5]))

    filled, kept, short = 0, 0, []
    for k, langs in data.items():
        r = where[k]
        en = str(ws.cell(r, C['en']).value or '')
        # ⚠ 줄 수가 영어와 다르면 <b>자리표가 어긋난다</b> — 넣기 전에 센다.
        want_lines = en.replace('\\n', '\n').count('\n')
        for L in LANGS:
            if L not in langs:
                raise SystemExit('⚠ %s 에 «%s» 가 없습니다.' % (k, L))
            if str(ws.cell(r, C[L]).value or '').strip():
                kept += 1
                continue
            val = langs[L]
            if val.count('\n') != want_lines:
                short.append('%s/%s (en %d줄 · %s %d줄)'
                             % (k, L, want_lines + 1, L, val.count('\n') + 1))
            ws.cell(r, C[L]).value = val
            filled += 1

    if short:
        print('⚠ 줄 수가 영어와 다른 칸 %d:' % len(short))
        for s in short[:10]:
            print('   ! %s' % s)

    if not filled:
        print('채울 칸이 없습니다 (이미 %d칸 차 있음).' % kept)
        return 0

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.20260902c.bak')
    wb.save(STRING_XLSX)
    print('저장: %s (백업 .20260902c.bak)' % os.path.basename(STRING_XLSX))
    print('  채운 칸 %d · 이미 있어 건너뛴 칸 %d' % (filled, kept))
    return 0


if __name__ == '__main__':
    sys.exit(main())
