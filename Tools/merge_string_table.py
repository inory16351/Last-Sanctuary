# -*- coding: utf-8 -*-
"""스트링 키 테이블 **두 판본을 합친다** — 두 PC 가 같은 파일을 따로 고쳤을 때 (2026-08-20 신설).

**왜 필요했나** — 볼트가 두 PC 에서 동시에 갱신됐고 GitHub Desktop 이 한쪽 작업을
`stash` 로 밀어 넣은 뒤 병합해 버렸다(유저: *"지금 볼트 머지 이상하게 됐는데"*).

    이쪽 PC   : 신규 캐릭터 5인(9007~9011)의 이름·스킬·설명 키 수십 개
    저쪽 PC   : 웨이브 보스 <b>레기미아</b>(120006)의 이름·칭호·스킬 키 6개

`.xlsx` 는 **바이너리라 git 이 못 합친다** — 한쪽을 고르면 다른 쪽이 통째로 사라진다.
그래서 «키 단위» 로 합치는 이 스크립트를 만들었다.

규칙은 하나다: **base 에 없는 키만 other 에서 가져온다.**
  · base 에 이미 있는 키는 <b>손대지 않는다</b>(사람이 다듬은 번역을 덮지 않는다 —
    `gen_string_table.py` 의 merge 규칙과 같은 판단).
  · 순서는 base 뒤에 붙인다. 스트링 테이블은 <b>행 순서에 의미가 없다</b>
    (하이퍼링크가 «정의된 이름» 을 가리키므로 행이 밀려도 안 깨진다 — 51-11절).

사용법:
    py -3 Tools/merge_string_table.py <base.xlsx> <other.xlsx> [-o 출력.xlsx]

⚠ openpyxl 로 쓴다 — 하이퍼링크·정의된 이름이 날아가지만, 마지막에
  `gen_string_table.py` → `link_string_keys.py` 를 다시 돌리면 전부 복구된다
  (그 두 스크립트가 «다시 만들 수 있는 것» 이라 이 순서가 성립한다).
"""

import argparse
import os
import shutil
import sys

import openpyxl

SHEET = "string"
DATA_ROW0 = 4


def read_keys(path):
    """{키: (행번호, kr, en, source, note)} — 앞 5열만 본다."""
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]
    out = {}
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k is None or not str(k).strip():
            continue
        out[str(k).strip()] = (r, ) + tuple(ws.cell(r, c).value for c in range(2, 6))
    return wb, ws, out


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser()
    ap.add_argument("base")
    ap.add_argument("other")
    ap.add_argument("-o", "--out")
    args = ap.parse_args()

    wb, ws, base = read_keys(args.base)
    _wb2, _ws2, other = read_keys(args.other)

    missing = [k for k in other if k not in base]
    print("base %d개 · other %d개 · <b>base 에 없는 키 %d개</b>" % (len(base), len(other), len(missing)))

    row = ws.max_row + 1
    for k in missing:
        _r, kr, en, src, note = other[k]
        ws.cell(row, 1).value = k
        ws.cell(row, 2).value = kr
        ws.cell(row, 3).value = en
        ws.cell(row, 4).value = src
        ws.cell(row, 5).value = note
        print("  + %-34s %s" % (k, kr))
        row += 1

    out = args.out or args.base
    if out == args.base and os.path.isfile(out):
        shutil.copy2(out, out + ".bak")
        print("백업:", out + ".bak")
    wb.save(out)
    print("→", out)

    only_base = [k for k in base if k not in other]
    print("(참고) base 에만 있는 키 %d개 — 그대로 유지된다" % len(only_base))


if __name__ == "__main__":
    main()
