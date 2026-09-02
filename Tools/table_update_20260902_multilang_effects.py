# -*- coding: utf-8 -*-
"""사건 «효과» 줄 55개의 7개 언어를 <b>조립해서</b> 채운다 (2026-09-02).

■ 무엇이 비어 있었나
  「언어추가」 커밋으로 스트링 키 테이블에 es·fr·de·ja·ru·pt·pl 일곱 열이 생겼지만,
  <b>행 560~743 이 연속으로 비어</b> 있었다(사건 대사 184개 · 일곱 언어 전부 같은 구간).
  그중 이 스크립트가 맡는 것은 <c>event_result_effect_*</c> <b>55줄</b>이다.

■ ★★★ 왜 손으로 안 쓰고 «조립» 하나
  이 55줄은 «효과 : A · B» 꼴의 <b>정형문</b>이고, 조각의 종류는 37가지뿐이다.
  손으로 385칸(55 × 7)을 쓰면 «같은 능력치를 줄마다 다르게 부르는» 사고가 반드시 난다 —
  202절에서 영어에 실제로 그 일이 있었다(같은 한국어 · 다른 영어 아홉 건).
  그래서 <b>이미 번역된 300001~300031 에서 뽑은 조각</b>을 사전으로 삼아 붙인다.
  ⚠ 그 사전은 <c>scratchpad/glossary.py</c> 가 표에서 실제로 뽑은 값이다 — 지어낸 것이 아니다.

■ 새로 지은 낱말 (선례가 없던 것)
  Critical Chance · Ranged Attack Power · Healing Received · Sanctuary Health ·
  All Enemies' Attack Power/Health/Defense/Attack Speed · Rooted · Burn ·
  Kill Count · Angel dies · Normal Monsters summoned
  → 능력치 이름은 <c>ui_stat_*</c> 의 번역을 그대로 물려받았고(화면 다른 곳과 같아진다),
    «성역» 은 <c>const_name_10001</c>·<c>ui_nexus_title</c> 의 낱말을 따랐다
    (es Santuario · fr Sanctuaire · de Refugium · ja 聖域 · ru Святилище · pt Santuário · pl Sanktuarium).

■ ⚠ 표기 규약 — 기존 줄에서 그대로 가져왔다
  · 음수는 <b>U+2212 −</b> 다(영어의 ASCII 하이픈이 아니다)
  · 초 표기: es/fr/de/pt/pl «(240 s)» · ja «（240秒）» · ru «(240 с)»
  · 구분자: 「 · 」 (ja 만 「・」)
  · ru 는 조각을 <b>소문자</b>로 잇는다(기존 줄이 그렇다). 「Порча」만 대문자다.
  · 유물 이름은 <c>relic_name_*</c> 의 번역을 쓴다 — 영어 쪽이 효과 줄과 유물 창에서
    서로 다른 이름을 쓰고 있지만(«Priceless Grace» ↔ «Grace Beyond Price»),
    <b>플레이어가 유물 창에서 보는 이름</b>에 맞추는 편이 맞다.

■ 다음
    py -3 Tools/gen_string_table.py   →   py -3 Tools/link_string_keys.py
"""
import os
import re
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

STRING_XLSX = os.path.join(TABLE_DIR, '스트링 키 테이블.xlsx')
SHEET, DATA_ROW0 = 'string', 4
LANGS = ['es', 'fr', 'de', 'ja', 'ru', 'pt', 'pl']

MINUS = '−'

PREFIX = {'es': 'Efecto: ', 'fr': 'Effet : ', 'de': 'Effekt: ', 'ja': '効果 : ',
          'ru': 'Эффект: ', 'pt': 'Efeito: ', 'pl': 'Efekt: '}
JOIN = {L: ' · ' for L in LANGS}
JOIN['ja'] = '・'

#: 능력치 이름 — 효과 줄에서 쓰는 형태(기존 줄에서 뽑았고, 없던 것은 ui_stat_* 를 물려받았다)
STAT = {
    'Accuracy':            dict(es='Precisión', fr='Précision', de='Genauigkeit', ja='命中率',
                                ru='точность', pt='Precisão', pl='Celność'),
    'Critical Chance':     dict(es='Probabilidad crítica', fr='Chance critique', de='Kritische Chance',
                                ja='クリティカル率', ru='шанс крит. удара', pt='Chance crítica',
                                pl='Szansa na cios krytyczny'),
    'Resistance':          dict(es='Resistencia', fr='Résistance', de='Widerstand', ja='抵抗力',
                                ru='сопротивление', pt='Resistência', pl='Odporność'),
    'Attack Speed':        dict(es='Velocidad de ataque', fr="Vitesse d'attaque", de='Angriffstempo',
                                ja='攻撃速度', ru='скорость атаки', pt='Velocidade de ataque',
                                pl='Szybkość ataku'),
    'Defense':             dict(es='Defensa', fr='Défense', de='Verteidigung', ja='防御力',
                                ru='защита', pt='Defesa', pl='Obrona'),
    'Vision':              dict(es='Visión', fr='Vision', de='Sicht', ja='視界',
                                ru='обзор', pt='Visão', pl='Pole widzenia'),
    'Movement Speed':      dict(es='Velocidad de movimiento', fr='Vitesse de déplacement',
                                de='Bewegungstempo', ja='移動速度', ru='скорость передвижения',
                                pt='Velocidade de movimento', pl='Szybkość ruchu'),
    'Melee Attack Power':  dict(es='Ataque cuerpo a cuerpo', fr='Attaque de mêlée', de='Nahkampfangriff',
                                ja='近接攻撃力', ru='ближняя атака', pt='Ataque corpo a corpo',
                                pl='Atak wręcz'),
    'Ranged Attack Power': dict(es='Ataque a distancia', fr='Attaque à distance', de='Fernkampfangriff',
                                ja='遠距離攻撃力', ru='дальняя атака', pt='Ataque à distância',
                                pl='Atak dystansowy'),
    'Magic Attack Power':  dict(es='Ataque mágico', fr='Attaque magique', de='Magieangriff',
                                ja='魔法攻撃力', ru='магическая атака', pt='Ataque mágico',
                                pl='Atak magiczny'),
    'Attack Power':        dict(es='Ataque', fr='Attaque', de='Angriff', ja='攻撃力',
                                ru='атака', pt='Ataque', pl='Atak'),
    'Healing Power':       dict(es='Curación', fr='Soins', de='Heilung', ja='回復力',
                                ru='исцеление', pt='Cura', pl='Leczenie'),
    'Healing Received':    dict(es='Curación recibida', fr='Soins reçus', de='Erhaltene Heilung',
                                ja='受ける回復量', ru='получаемое исцеление', pt='Cura recebida',
                                pl='Otrzymywane leczenie'),
    'Health':              dict(es='Salud', fr='Santé', de='Leben', ja='体力',
                                ru='здоровье', pt='Vida', pl='Zdrowie'),
}

#: «천사 전체의 X» · «적 전체의 X» 를 만드는 틀. {s} 자리에 STAT 이 들어간다.
OF_ANGELS = {'es': '{s} de todos los Ángeles', 'fr': '{s} de tous les Anges',
             'de': '{s} aller Engel', 'ja': '天使全体の{s}',
             'ru': '{s} всех Ангелов', 'pt': '{s} de todos os Anjos',
             'pl': '{s} wszystkich Aniołów'}
OF_ENEMIES = {'es': '{s} de todos los enemigos', 'fr': '{s} de tous les ennemis',
              'de': '{s} aller Gegner', 'ja': '敵全体の{s}',
              'ru': '{s} всех врагов', 'pt': '{s} de todos os inimigos',
              'pl': '{s} wszystkich wrogów'}

#: 초 표기
SEC = {'es': '({n} s)', 'fr': '({n} s)', 'de': '({n} s)', 'ja': '（{n}秒）',
       'ru': '({n} с)', 'pt': '({n} s)', 'pl': '({n} s)'}

#: 유물 이름 — relic_name_* 의 번역 그대로
RELIC = {
    'Red Thread':             dict(es='Hilo Rojo', fr='Fil Rouge', de='Roter Faden', ja='赤い糸',
                                   ru='Красная Нить', pt='Fio Vermelho', pl='Czerwona Nić'),
    'Priceless Grace':        dict(es='Gracia Sin Precio', fr='Grâce Sans Prix', de='Gnade Ohne Preis',
                                   ja='値の付かない恩恵', ru='Милость Без Цены', pt='Graça Sem Preço',
                                   pl='Łaska Bez Ceny'),
    'Cool Fever Relief':      dict(es='El Frescor de la Fiebre Rota', fr='La Fraîcheur de la Fièvre Tombée',
                                   de='Das Kühle Fieberbrechen', ja='冷ややかな解熱',
                                   ru='Прохладный Спад Жара', pt='O Frescor da Febre Quebrada',
                                   pl='Chłodny Spadek Gorączki'),
    'Memory of the Way Home': dict(es='Memoria del Lugar al que Volver', fr='Mémoire du Lieu où Revenir',
                                   de='Erinnerung an den Ort der Heimkehr', ja='帰る場所の記憶',
                                   ru='Память о Месте, Куда Возвращаться', pt='Memória do Lugar para Onde Voltar',
                                   pl='Pamięć Miejsca Powrotu'),
    'Self-Mended Flesh':      dict(es='Carne Remendada por Sí Misma', fr="Chair Recousue d'Elle-même",
                                   de='Selbstgeheiltes Fleisch', ja='自ら埋めた肉',
                                   ru='Плоть, Заросшая Сама', pt='Carne Remendada por Si',
                                   pl='Ciało Zrośnięte Samo'),
    'Swollen Lymph Node':     dict(es='Ganglio Linfático Hinchado', fr='Ganglion Lymphatique Enflé',
                                   de='Geschwollener Lymphknoten', ja='腫れたリンパ節',
                                   ru='Вспухший Лимфоузел', pt='Nabrzmiały'  # 아래에서 pt 를 바로잡는다
                                   , pl='Nabrzmiały Węzeł Chłonny'),
}
RELIC['Swollen Lymph Node']['pt'] = 'Gânglio Linfático Inchado'

#: 유물 등급 — relic_grade_* 그대로. ⚠ 효과 줄에서는 «(Común)» 처럼 괄호 안에 들어가는데,
#:   기존 줄이 pl 만 남성형(zwykły/rzadki)을 쓴다 — relikt(남성명사)에 맞춘 것이다.
GRADE = {
    'Common': dict(es='Común', fr='Commune', de='Gewöhnlich', ja='一般', ru='обычная',
                   pt='Comum', pl='zwykły'),
    'Rare':   dict(es='Rara', fr='Rare', de='Selten', ja='レア', ru='редкая',
                   pt='Rara', pl='rzadki'),
    'Epic':   dict(es='Épica', fr='Épique', de='Episch', ja='エピック', ru='эпическая',
                   pt='Épica', pl='epicki'),
}

SANCTUARY = dict(es='Santuario', fr='Sanctuaire', de='Refugium', ja='聖域',
                 ru='Святилище', pt='Santuário', pl='Sanktuarium')


def num(tok):
    """영어의 ASCII 하이픈 음수를 표의 규약(U+2212)으로 바꾼다."""
    return tok.replace('-', MINUS)


# ── 조각 하나를 옮긴다 ────────────────────────────────────────────────────
def seg(en, L):
    s = en.strip()

    def stat_of(who, name, rest):
        base = (OF_ANGELS if who == 'A' else OF_ENEMIES)[L].format(s=STAT[name][L])
        return base + rest

    m = re.match(r"^All (Angels|Enemies)' (.+?) ([+\-−][\d.]+)% \((\d+)s\)$", s)
    if m:
        who = 'A' if m.group(1) == 'Angels' else 'E'
        # ⚠ 일본어는 «%» 와 «（» 사이를 <b>붙여 쓴다</b> — 기존 줄이 「+8%（240秒）」다.
        gap = '' if L == 'ja' else ' '
        return stat_of(who, m.group(2),
                       ' %s%%%s%s' % (num(m.group(3)), gap, SEC[L].format(n=m.group(4))))

    m = re.match(r"^All (Angels|Enemies)' Health ([+\-−][\d.]+)% \(based on Max Health\)$", s)
    if m:
        who = 'A' if m.group(1) == 'Angels' else 'E'
        tail = {'es': ' %s%% (según la salud máxima)', 'fr': ' %s%% (sur la santé maximale)',
                'de': ' %s%% (bezogen auf das maximale Leben)', 'ja': ' %s%%（最大体力基準）',
                'ru': ' %s%% (от максимума)', 'pt': ' %s%% (com base na vida máxima)',
                'pl': ' %s%% (od maksymalnego)'}[L] % num(m.group(2))
        return stat_of(who, 'Health', tail)

    m = re.match(r"^All Angels' Health restored \((\d+)% of Max Health\)$", s)
    if m:
        n = m.group(1)
        return {'es': 'Salud de todos los Ángeles restaurada (%s%% de la salud máxima)' % n,
                'fr': 'Santé de tous les Anges restaurée (%s%% de la santé maximale)' % n,
                'de': 'Leben aller Engel wiederhergestellt (%s%% des maximalen Lebens)' % n,
                'ja': '天使全体の体力回復（最大体力の%s%%）' % n,
                'ru': 'здоровье всех Ангелов восстановлено (%s%% от макс.)' % n,
                'pt': 'Vida de todos os Anjos restaurada (%s%% da vida máxima)' % n,
                'pl': 'Zdrowie wszystkich Aniołów przywrócone (%s%% maks. zdrowia)' % n}[L]

    m = re.match(r"^All Angels' Shield \((\d+)% of Max Health, (\d+)s\)$", s)
    if m:
        p, t = m.group(1), m.group(2)
        return {'es': 'Escudo para todos los Ángeles (%s%% de la salud máxima, %s s)' % (p, t),
                'fr': 'Bouclier pour tous les Anges (%s%% de la santé maximale, %s s)' % (p, t),
                'de': 'Schild für alle Engel (%s%% des maximalen Lebens, %s s)' % (p, t),
                'ja': '天使全体に保護膜（最大体力の%s%%、%s秒）' % (p, t),
                'ru': 'щит всем Ангелам (%s%% от макс. здоровья, %s с)' % (p, t),
                'pt': 'Escudo para todos os Anjos (%s%% da vida máxima, %s s)' % (p, t),
                'pl': 'Tarcza dla wszystkich Aniołów (%s%% maks. zdrowia, %s s)' % (p, t)}[L]

    m = re.match(r"^All Angels' corruption ([+\-−][\d.]+)$", s)
    if m:
        n = num(m.group(1))
        return {'es': 'Corrupción de todos los Ángeles %s' % n,
                'fr': 'Corruption de tous les Anges %s' % n,
                'de': 'Verderbnis aller Engel %s' % n,
                'ja': '天使全体の浸食 %s' % n,
                'ru': 'Порча всех Ангелов %s' % n,
                'pt': 'Corrupção de todos os Anjos %s' % n,
                'pl': 'Skażenie wszystkich Aniołów %s' % n}[L]

    m = re.match(r"^(\d+) Angel's corruption ([+\-−][\d.]+)$", s)
    if m:
        c, n = m.group(1), num(m.group(2))
        return {'es': 'Corrupción de %s Ángel %s' % (c, n),
                'fr': "Corruption d'%s Ange %s" % (c, n),
                'de': 'Verderbnis von %s Engel %s' % (c, n),
                'ja': '天使%s名の浸食 %s' % (c, n),
                'ru': 'Порча %s Ангела %s' % (c, n),
                'pt': 'Corrupção de %s Anjo %s' % (c, n),
                'pl': 'Skażenie %s Anioła %s' % (c, n)}[L]

    m = re.match(r"^Energy ([+\-−][\d.]+)$", s)
    if m:
        n = num(m.group(1))
        return {'es': 'Energía %s', 'fr': 'Énergie %s', 'de': 'Energie %s', 'ja': 'エネルギー %s',
                'ru': 'энергия %s', 'pt': 'Energia %s', 'pl': 'Energia %s'}[L] % n

    m = re.match(r"^Sanctuary Health ([+\-−][\d.]+)%$", s)
    if m:
        n = num(m.group(1))
        sa = SANCTUARY[L]
        return {'es': 'Salud del %s %s%%' % (sa, n),
                'fr': 'Santé du %s %s%%' % (sa, n),
                'de': 'Leben des %ss %s%%' % (sa, n),
                'ja': '%sの体力 %s%%' % (sa, n),
                'ru': 'здоровье %s %s%%' % ('Святилища', n),
                'pt': 'Vida do %s %s%%' % (sa, n),
                'pl': 'Zdrowie %s %s%%' % ('Sanktuarium', n)}[L]

    m = re.match(r'^Relic "(.+?)" obtained \((Common|Rare|Epic)\)$', s)
    if m:
        rn, gr = RELIC[m.group(1)][L], GRADE[m.group(2)][L]
        return {'es': 'Reliquia «%s» obtenida (%s)' % (rn, gr),
                'fr': 'Relique « %s » obtenue (%s)' % (rn, gr),
                'de': 'Relikt „%s“ erhalten (%s)' % (rn, gr),
                'ja': '遺物「%s」獲得（%s）' % (rn, gr),
                'ru': 'получена реликвия «%s» (%s)' % (rn, gr),
                'pt': 'Relíquia "%s" obtida (%s)' % (rn, gr),
                'pl': 'Zdobyto relikt „%s” (%s)' % (rn, gr)}[L]

    m = re.match(r'^(\d+) Normal Monsters summoned in the current wave$', s)
    if m:
        n = m.group(1)
        return {'es': '%s monstruos normales invocados en la oleada actual' % n,
                'fr': '%s monstres normaux invoqués dans la vague actuelle' % n,
                'de': '%s normale Monster in der aktuellen Welle beschworen' % n,
                'ja': '現在のウェーブに一般モンスター%s体 召喚' % n,
                'ru': 'в текущей волне призвано обычных монстров: %s' % n,
                'pt': '%s monstros normais invocados na onda atual' % n,
                'pl': 'Przyzwano %s zwykłych potworów w obecnej fali' % n}[L]

    m = re.match(r"^(\d+) Angel's Mental Error cured$", s)
    if m:
        n = m.group(1)
        return {'es': 'Trastorno mental de %s Ángel curado' % n,
                'fr': "Trouble mental d'%s Ange soigné" % n,
                'de': 'Geistesstörung von %s Engel geheilt' % n,
                'ja': '天使%s名の精神異常解除' % n,
                'ru': 'расстройство рассудка у %s Ангела снято' % n,
                'pt': 'Transtorno mental de %s Anjo curado' % n,
                'pl': 'Usunięto zaburzenie psychiczne %s Anioła' % n}[L]

    m = re.match(r"^(\d+) random Angel's Kill Count ([+\-−][\d.]+)$", s)
    if m:
        c, n = m.group(1), num(m.group(2))
        return {'es': 'Bajas de %s Ángel al azar %s' % (c, n),
                'fr': "Éliminations d'%s Ange au hasard %s" % (c, n),
                'de': 'Abschüsse von %s zufälligen Engel %s' % (c, n),
                'ja': 'ランダムな天使%s名の処置数 %s' % (c, n),
                'ru': 'убийства случайного Ангела %s' % n,
                'pt': 'Abates de %s Anjo aleatório %s' % (c, n),
                'pl': 'Pokonani przeciwnicy losowego Anioła %s' % n}[L]

    m = re.match(r'^(\d+) random Angel gains ([+\-−][\d.]+) to (\d+) stat \(Permanent\)$', s)
    if m:
        c, n, k = m.group(1), num(m.group(2)), m.group(3)
        return {'es': '%s Ángel al azar gana %s en %s característica (permanente)' % (c, n, k),
                'fr': '%s Ange au hasard gagne %s à %s caractéristique (permanent)' % (c, n, k),
                'de': '%s zufälliger Engel erhält %s auf %s Wert (dauerhaft)' % (c, n, k),
                'ja': 'ランダムな天使%s名の能力値%s種 %s（永久）' % (c, k, n),
                'ru': 'случайный Ангел получает %s к %s характеристике (навсегда)' % (n, k),
                'pt': '%s Anjo aleatório ganha %s em %s atributo (permanente)' % (c, n, k),
                'pl': 'Losowy Anioł zyskuje %s do %s atrybutu (na stałe)' % (n, k)}[L]

    m = re.match(r'^(\d+) Angel dies \((\d+)% chance\)$', s)
    if m:
        c, p = m.group(1), m.group(2)
        return {'es': '%s Ángel muere (%s%% de probabilidad)' % (c, p),
                'fr': '%s Ange meurt (%s%% de chance)' % (c, p),
                'de': '%s Engel stirbt (%s%% Chance)' % (c, p),
                'ja': '天使%s名 死亡（確率%s%%）' % (c, p),
                'ru': 'погибает %s Ангел (шанс %s%%)' % (c, p),
                'pt': '%s Anjo morre (%s%% de chance)' % (c, p),
                'pl': 'Ginie %s Anioł (%s%% szansy)' % (c, p)}[L]

    m = re.match(r'^All Enemies Rooted for (\d+)s$', s)
    if m:
        n = m.group(1)
        return {'es': 'Todos los enemigos inmovilizados %s s' % n,
                'fr': 'Tous les ennemis immobilisés %s s' % n,
                'de': 'Alle Gegner %s s festgewurzelt' % n,
                'ja': '敵全体を%s秒間 束縛' % n,
                'ru': 'все враги обездвижены на %s с' % n,
                'pt': 'Todos os inimigos imobilizados por %s s' % n,
                'pl': 'Wszyscy wrogowie unieruchomieni na %s s' % n}[L]

    m = re.match(r'^All Enemies Burn \((\d+)% of Max Health per second, (\d+)s\)$', s)
    if m:
        p, t = m.group(1), m.group(2)
        return {'es': 'Todos los enemigos arden (%s%% de la salud máxima por segundo, %s s)' % (p, t),
                'fr': 'Tous les ennemis brûlent (%s%% de la santé maximale par seconde, %s s)' % (p, t),
                'de': 'Alle Gegner brennen (%s%% des maximalen Lebens pro Sekunde, %s s)' % (p, t),
                'ja': '敵全体を燃焼（毎秒 最大体力の%s%%、%s秒）' % (p, t),
                'ru': 'все враги горят (%s%% от макс. здоровья в секунду, %s с)' % (p, t),
                'pt': 'Todos os inimigos queimam (%s%% da vida máxima por segundo, %s s)' % (p, t),
                'pl': 'Wszyscy wrogowie płoną (%s%% maks. zdrowia na sekundę, %s s)' % (p, t)}[L]

    raise SystemExit('⚠ 옮길 틀이 없는 조각: %r' % s)


def translate(en, L):
    body = en.strip()
    assert body.startswith('Effect: '), body
    parts = [p.strip() for p in body[len('Effect: '):].split(' · ')]
    out = [seg(p, L) for p in parts]
    # ⚠ 첫 조각만 대문자로 올린다 — ru 는 기존 줄이 소문자로 잇는다(「Порча」는 그대로).
    first = out[0]
    if L == 'ru' and first[:1].islower():
        first = first[:1].upper() + first[1:]
    return PREFIX[L] + JOIN[L].join([first] + out[1:])


def main():
    wb = openpyxl.load_workbook(STRING_XLSX)
    ws = wb[SHEET]
    fields = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    C = {n: i + 1 for i, n in enumerate(fields)}
    for L in LANGS:
        if L not in C:
            raise SystemExit('⚠ «%s» 열이 없습니다 — 표에 다국어 열이 아직 없습니다.' % L)

    filled, skipped = 0, 0
    for r in range(DATA_ROW0, ws.max_row + 1):
        k = str(ws.cell(r, 1).value or '')
        if not k.startswith('event_result_effect'):
            continue
        en = str(ws.cell(r, C['en']).value or '').strip()
        if not en:
            continue
        for L in LANGS:
            cur = str(ws.cell(r, C[L]).value or '').strip()
            if cur:
                skipped += 1          # ⚠ 이미 있는 번역은 <b>절대 덮지 않는다</b>
                continue
            ws.cell(r, C[L]).value = translate(en, L)
            filled += 1

    if not filled:
        print('채울 칸이 없습니다 (이미 %d칸 차 있음).' % skipped)
        return 0

    shutil.copyfile(STRING_XLSX, STRING_XLSX + '.20260902b.bak')
    wb.save(STRING_XLSX)
    print('저장: %s (백업 .20260902b.bak)' % os.path.basename(STRING_XLSX))
    print('  채운 칸 %d · 이미 있어 건너뛴 칸 %d' % (filled, skipped))
    return 0


if __name__ == '__main__':
    sys.exit(main())
