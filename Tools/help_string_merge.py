# -*- coding: utf-8 -*-
"""도움말 표의 `StringKeys` 시트를 «스트링 키 테이블.xlsx» 에 <b>병합</b>한다 (2026-08-24).

진행상황 140-6절이 «아직 안 한 것» 1번으로 적어 둔 그 일이다:
    ① `StringKeys` 시트를 스트링 키 테이블에 병합 → `gen_string_table.py` 재실행

★★ <b>help_* 키의 정본은 «도움말 표» 다</b>
------------------------------------------
문구를 두 곳에서 고칠 수 있으면 반드시 어긋난다. 그래서 규칙을 하나로 못박는다 —
<b>`help_` 로 시작하는 키는 도움말 표가 정본</b>이고, 이 스크립트가 매번 스트링 키 테이블로
<b>덮어쓴다</b>. (다른 키는 손대지 않는다.)

  ⚠ 그러니 도움말 문구를 다듬을 때는 <b>스트링 키 테이블이 아니라 도움말 표</b>를 고칠 것.
    스트링 키 테이블에서 고치면 다음 병합에 되돌아간다.
  · 반대로 하려면 ``--keep`` — 이미 있는 키는 그대로 두고 <b>새 키만</b> 붙인다
    (`gen_string_table.py`·`merge_string_table.py` 의 «기존 우선» 규칙과 같아진다).

⚠ openpyxl 로 쓰지만 하이퍼링크가 날아갈 걱정은 없다 — 뒤이어 도는
  `gen_string_table.py` 가 표를 <b>통째로 새로 쓰고</b>, `link_string_keys.py` 가
  하이퍼링크를 다시 만든다. 그래서 이 순서가 성립한다.

사용법:
    py -3 Tools/help_string_merge.py              도움말 표를 정본으로 덮어쓴다(권장)
    py -3 Tools/help_string_merge.py --keep       새 키만 붙인다
    py -3 Tools/help_string_merge.py --dry        무엇이 바뀌는지만 찍고 저장하지 않는다

다음:
    py -3 Tools/gen_string_table.py     (TSV 내보내기)
    py -3 Tools/link_string_keys.py     (하이퍼링크 재생성)
"""

import argparse
import io
import os
import shutil
import sys

import openpyxl

from vault_path import TABLE_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HELP_XLSX = os.path.join(TABLE_DIR, "Last_Sanctuary_도움말테이블_Ver01.xlsx")
STRING_XLSX = os.path.join(TABLE_DIR, "스트링 키 테이블.xlsx")

HELP_SHEET = "StringKeys"
HELP_ROW0 = 2          # 도움말 표는 머리글 <b>한 줄</b>이다

STRING_SHEET = "string"
STRING_ROW0 = 4        # 스트링 키 테이블은 머리글 <b>세 줄</b>이다(한글/필드/자료형)

PREFIX = "help_"
SOURCE_LABEL = "도움말 표"


def norm(v):
    return "" if v is None else str(v).strip()


def read_help():
    """[(key, kr, en, note)] — 도움말 표의 StringKeys 시트 그대로."""
    if not os.path.isfile(HELP_XLSX):
        raise SystemExit("⚠ 도움말 표가 없습니다: %s" % HELP_XLSX)

    wb = openpyxl.load_workbook(HELP_XLSX, data_only=True)
    if HELP_SHEET not in wb.sheetnames:
        raise SystemExit("⚠ '%s' 시트가 없습니다 — 시트: %s" % (HELP_SHEET, wb.sheetnames))

    ws = wb[HELP_SHEET]
    out = []
    for r in range(HELP_ROW0, ws.max_row + 1):
        key = norm(ws.cell(r, 1).value)
        if not key:
            continue
        # ⚠ 줄바꿈은 <b>그대로 옮긴다</b> — 접는 일(\n 리터럴화)은 gen_string_table.py 가
        #   TSV 로 내보낼 때 한다. 여기서 미리 접으면 두 번 접힌다.
        out.append((key, norm(ws.cell(r, 2).value), norm(ws.cell(r, 3).value),
                    norm(ws.cell(r, 4).value)))
    return out


def field_index(ws):
    """2행의 필드명 → 열 번호. gen_string_table.py 와 같은 규약이다."""
    idx = {}
    for c in range(1, ws.max_column + 1):
        f = norm(ws.cell(2, c).value)
        if f:
            idx[f] = c
    return idx


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--keep", action="store_true",
                    help="이미 있는 키는 그대로 두고 새 키만 붙인다")
    ap.add_argument("--dry", action="store_true", help="저장하지 않고 결과만 찍는다")
    args = ap.parse_args()

    rows = read_help()
    print("[도움말 → 스트링] 도움말 표 StringKeys %d행" % len(rows))

    if not os.path.isfile(STRING_XLSX):
        raise SystemExit("⚠ 스트링 키 테이블이 없습니다: %s" % STRING_XLSX)

    wb = openpyxl.load_workbook(STRING_XLSX)
    if STRING_SHEET not in wb.sheetnames:
        raise SystemExit("⚠ '%s' 시트가 없습니다 — 시트: %s" % (STRING_SHEET, wb.sheetnames))

    ws = wb[STRING_SHEET]
    idx = field_index(ws)
    c_key = idx.get("string_key", 1)
    c_kr = idx.get("kr", 2)
    c_en = idx.get("en", 3)
    c_src = idx.get("source", 4)
    c_note = idx.get("note", 5)

    where = {}
    for r in range(STRING_ROW0, ws.max_row + 1):
        k = norm(ws.cell(r, c_key).value)
        if k:
            where[k] = r

    added, changed, kept = [], [], []
    write_row = ws.max_row + 1

    for key, kr, en, note in rows:
        if not key.startswith(PREFIX):
            # 규칙을 어기는 키가 섞여 있으면 <b>조용히 넣지 않는다</b> — 정본이 어디인지가
            # 접두사로 결정되기 때문이다(맨 위 ★★).
            print("  ⚠ help_ 로 시작하지 않아 건너뜁니다: %s" % key)
            continue

        r = where.get(key)
        if r is None:
            ws.cell(write_row, c_key).value = key
            ws.cell(write_row, c_kr).value = kr
            ws.cell(write_row, c_en).value = en
            ws.cell(write_row, c_src).value = SOURCE_LABEL
            ws.cell(write_row, c_note).value = note
            where[key] = write_row
            write_row += 1
            added.append(key)
            continue

        if args.keep:
            kept.append(key)
            continue

        before = norm(ws.cell(r, c_kr).value)
        ws.cell(r, c_kr).value = kr
        # ⚠ en 은 <b>비어 있을 때만</b> 넣는다 — 유저가 스트링 키 테이블에서 영어를 채웠을 수
        #   있고, 도움말 표의 en 열은 «자리만 비워 둔다» 가 확정사항 ④ 다.
        if en and not norm(ws.cell(r, c_en).value):
            ws.cell(r, c_en).value = en
        ws.cell(r, c_src).value = SOURCE_LABEL
        if note:
            ws.cell(r, c_note).value = note
        if before != kr:
            changed.append(key)

    print("  새로 붙인 키 %d개 · 문구를 덮어쓴 키 %d개 · 그대로 둔 키 %d개"
          % (len(added), len(changed), len(kept)))
    for k in added:
        print("    + %s" % k)
    for k in changed:
        print("    ~ %s" % k)

    if args.dry:
        print("  (--dry — 저장하지 않았습니다)")
        return 0

    if not added and not changed:
        print("  바뀐 것이 없어 저장하지 않았습니다.")
        return 0

    shutil.copy2(STRING_XLSX, STRING_XLSX + ".bak")
    wb.save(STRING_XLSX)
    print("  저장: %s  (백업 %s)" % (os.path.basename(STRING_XLSX),
                                     os.path.basename(STRING_XLSX) + ".bak"))
    print("  다음: py -3 Tools/gen_string_table.py  →  py -3 Tools/link_string_keys.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
